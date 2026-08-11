"""Render MIS pack facts into downloadable Excel / PDF / PPT bytes (ADR-014)."""
from __future__ import annotations

import io
import re
import zipfile
from typing import Any
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from app.modules.office_ai.models import utcnow


def _dims(fact: dict[str, Any]) -> dict[str, Any]:
    dims = fact.get("dimensions")
    return dims if isinstance(dims, dict) else {}


def _amount(fact: dict[str, Any]) -> float | None:
    raw = fact.get("amount_decimal")
    if raw is None:
        raw = fact.get("value")
    try:
        return float(raw)
    except (TypeError, ValueError):
        return None


def _fmt_num(value: float | None) -> str:
    if value is None:
        return "—"
    return f"{value:,.2f}"


def _safe_filename(pack: dict[str, Any], export_format: str) -> str:
    period = re.sub(r"[^A-Za-z0-9._-]+", "_", str(pack.get("period") or "period"))[:40]
    key = re.sub(r"[^A-Za-z0-9._-]+", "_", str(pack.get("pack_key") or "mis"))[:40]
    ext = {"excel": "xlsx", "pdf_summary": "pdf", "ppt": "pptx"}.get(export_format, "bin")
    return f"MIS_{key}_{period}.{ext}"


def _group_facts(facts: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {
        "kpi": [],
        "pnl_line": [],
        "bs_line": [],
        "cash_summary": [],
        "aging_bucket": [],
        "other": [],
    }
    for fact in facts:
        et = str(fact.get("entity_type") or "").strip().lower()
        if et in grouped:
            grouped[et].append(fact)
        else:
            grouped["other"].append(fact)
    return grouped


def build_excel_bytes(*, pack: dict[str, Any], facts: list[dict[str, Any]]) -> bytes:
    grouped = _group_facts(facts)
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "OfficeMitra CA Analysis Pack"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A2"] = f"Pack: {pack.get('pack_key') or ''}"
    summary["A3"] = f"Period: {pack.get('period') or ''}"
    summary["A4"] = f"Status: {pack.get('status') or ''}"
    summary["A5"] = f"Generated: {utcnow().isoformat()}"
    summary["A6"] = "Source: imported MIS facts (not AI estimates)"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    def write_sheet(title: str, headers: list[str], rows: list[list[Any]]) -> None:
        ws = wb.create_sheet(title[:31])
        for col, label in enumerate(headers, start=1):
            cell = ws.cell(1, col, label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left")
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(r_idx, c_idx, value)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    kpi_rows = []
    for fact in grouped["kpi"]:
        dims = _dims(fact)
        kpi_rows.append([dims.get("kpi") or fact.get("source_id"), _amount(fact), dims.get("unit") or ""])
    write_sheet("KPIs", ["KPI", "Value", "Unit"], kpi_rows)

    pnl_rows = []
    for fact in grouped["pnl_line"]:
        dims = _dims(fact)
        if dims.get("trend"):
            continue
        pnl_rows.append([dims.get("line") or "", _amount(fact), fact.get("period") or pack.get("period") or ""])
    write_sheet("PnL", ["Line", "Amount", "Period"], pnl_rows)

    bs_rows = []
    for fact in grouped["bs_line"]:
        dims = _dims(fact)
        bs_rows.append([dims.get("line") or "", _amount(fact)])
    write_sheet("BalanceSheet", ["Line", "Amount"], bs_rows)

    cash_rows = []
    for fact in grouped["cash_summary"]:
        dims = _dims(fact)
        cash_rows.append([dims.get("line") or "", _amount(fact)])
    write_sheet("CashFlow", ["Activity", "Amount"], cash_rows)

    ageing_rows = []
    for fact in grouped["aging_bucket"]:
        dims = _dims(fact)
        ageing_rows.append([dims.get("side") or "", dims.get("bucket") or "", _amount(fact)])
    write_sheet("Ageing", ["Side", "Bucket", "Amount"], ageing_rows)

    all_rows = []
    for fact in facts:
        dims = _dims(fact)
        all_rows.append(
            [
                fact.get("entity_type") or "",
                fact.get("period") or "",
                _amount(fact),
                fact.get("source_system") or "",
                str(dims),
            ]
        )
    write_sheet("AllFacts", ["Entity", "Period", "Amount/Value", "Source", "Dimensions"], all_rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_summary_bytes(*, pack: dict[str, Any], facts: list[dict[str, Any]]) -> bytes:
    grouped = _group_facts(facts)
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    y = height - 48

    def line(text: str, *, size: int = 11, bold: bool = False, gap: int = 16) -> None:
        nonlocal y
        if y < 48:
            c.showPage()
            y = height - 48
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        # reportlab Helvetica is Latin-1; replace unsupported glyphs
        safe = str(text).encode("latin-1", "replace").decode("latin-1")
        c.drawString(40, y, safe[:110])
        y -= gap

    line("OfficeMitra CA Analysis Pack — CEO summary", size=14, bold=True, gap=22)
    line(f"Pack: {pack.get('pack_key') or ''}  |  Period: {pack.get('period') or ''}")
    line(f"Status: {pack.get('status') or ''}  |  Generated: {utcnow().isoformat()}")
    line("Figures derived from imported MIS facts (not AI estimates).", size=9, gap=20)

    line("Key KPIs", bold=True, gap=18)
    for fact in grouped["kpi"][:12]:
        dims = _dims(fact)
        line(f"- {dims.get('kpi') or fact.get('source_id')}: {_fmt_num(_amount(fact))} {dims.get('unit') or ''}")

    y -= 8
    line("P&L snapshot", bold=True, gap=18)
    for fact in grouped["pnl_line"]:
        dims = _dims(fact)
        if dims.get("trend"):
            continue
        line(f"- {dims.get('line') or ''}: {_fmt_num(_amount(fact))}")

    y -= 8
    line("Ageing", bold=True, gap=18)
    for fact in grouped["aging_bucket"][:12]:
        dims = _dims(fact)
        line(f"- {dims.get('side')}/{dims.get('bucket')}: {_fmt_num(_amount(fact))}")

    c.showPage()
    c.save()
    return buf.getvalue()


def _pptx_slide_xml(title: str, bullets: list[str]) -> str:
    texts = [f"<a:t>{xml_escape(title)}</a:t>"]
    title_body = f"""
      <p:sp>
        <p:nvSpPr><p:cNvPr id="2" name="Title"/><p:cNvSpPr/><p:nvPr/></p:nvSpPr>
        <p:spPr>
          <a:xfrm><a:off x="457200" y="274320"/><a:ext cx="8229600" cy="914400"/></a:xfrm>
        </p:spPr>
        <p:txBody>
          <a:bodyPr/><a:lstStyle/>
          <a:p><a:r><a:rPr lang="en-US" sz="2800" b="1"/>{texts[0]}</a:r></a:p>
        </p:txBody>
      </p:sp>
    """
    bullet_paras = []
    for bullet in bullets[:10]:
        bullet_paras.append(
            f"""<a:p><a:r><a:rPr lang="en-US" sz="1600"/><a:t>{xml_escape(bullet)}</a:t></a:r></a:p>"""
        )
    body = f"""
      <p:sp>
        <p:nvSpPr><p:cNvPr id="3" name="Body"/><p:cNvSpPr/><p:nvPr/></p:nvSpPr>
        <p:spPr>
          <a:xfrm><a:off x="457200" y="1371600"/><a:ext cx="8229600" cy="4114800"/></a:xfrm>
        </p:spPr>
        <p:txBody>
          <a:bodyPr/><a:lstStyle/>
          {''.join(bullet_paras) or '<a:p><a:endParaRPr lang="en-US"/></a:p>'}
        </p:txBody>
      </p:sp>
    """
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
 xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">
  <p:cSld><p:spTree>
    <p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr>
    <p:grpSpPr/>
    {title_body}
    {body}
  </p:spTree></p:cSld>
  <p:clrMapOvr><a:masterClrMapping/></p:clrMapOvr>
</p:sld>
"""


def build_ppt_bytes(*, pack: dict[str, Any], facts: list[dict[str, Any]]) -> bytes:
    """Build a minimal OOXML PPTX (no python-pptx dependency)."""
    grouped = _group_facts(facts)
    kpi_lines = []
    for fact in grouped["kpi"][:8]:
        dims = _dims(fact)
        kpi_lines.append(f"{dims.get('kpi') or fact.get('source_id')}: {_fmt_num(_amount(fact))} {dims.get('unit') or ''}".strip())

    pnl_lines = []
    for fact in grouped["pnl_line"]:
        dims = _dims(fact)
        if dims.get("trend"):
            continue
        pnl_lines.append(f"{dims.get('line')}: {_fmt_num(_amount(fact))}")

    ageing_lines = []
    for fact in grouped["aging_bucket"][:10]:
        dims = _dims(fact)
        ageing_lines.append(f"{dims.get('side')} {dims.get('bucket')}: {_fmt_num(_amount(fact))}")

    slides = [
        (
            "CA Analysis Pack",
            [
                f"Pack: {pack.get('pack_key') or ''}",
                f"Period: {pack.get('period') or ''}",
                f"Status: {pack.get('status') or ''}",
                "Derived from imported MIS facts (not AI estimates)",
            ],
        ),
        ("Key KPIs", kpi_lines or ["No KPI facts"]),
        ("P&L snapshot", pnl_lines or ["No P&L facts"]),
        ("Ageing", ageing_lines or ["No ageing facts"]),
    ]

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(
            "[Content_Types].xml",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/ppt/presentation.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml"/>
"""
            + "".join(
                f'  <Override PartName="/ppt/slides/slide{i}.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>'
                for i in range(1, len(slides) + 1)
            )
            + "\n</Types>\n",
        )
        zf.writestr(
            "_rels/.rels",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="ppt/presentation.xml"/>
</Relationships>
""",
        )
        slide_rels = "".join(
            f'<Relationship Id="rId{i}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" Target="slides/slide{i}.xml"/>'
            for i in range(1, len(slides) + 1)
        )
        zf.writestr(
            "ppt/_rels/presentation.xml.rels",
            f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
{slide_rels}
</Relationships>
""",
        )
        sld_ids = "".join(
            f'<p:sldId id="{255 + i}" r:id="rId{i}"/>' for i in range(1, len(slides) + 1)
        )
        zf.writestr(
            "ppt/presentation.xml",
            f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:presentation xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
 xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">
  <p:sldIdLst>{sld_ids}</p:sldIdLst>
  <p:sldSz cx="9144000" cy="6858000"/>
  <p:notesSz cx="6858000" cy="9144000"/>
</p:presentation>
""",
        )
        for idx, (title, bullets) in enumerate(slides, start=1):
            zf.writestr(f"ppt/slides/slide{idx}.xml", _pptx_slide_xml(title, bullets))
    return buf.getvalue()


def render_mis_export(
    *,
    pack: dict[str, Any],
    facts: list[dict[str, Any]],
    export_format: str,
) -> tuple[bytes, str, str]:
    """Return (content_bytes, filename, content_type)."""
    fmt = str(export_format or "").strip().lower()
    filename = _safe_filename(pack, fmt)
    if fmt == "excel":
        return build_excel_bytes(pack=pack, facts=facts), filename, (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    if fmt == "pdf_summary":
        return build_pdf_summary_bytes(pack=pack, facts=facts), filename, "application/pdf"
    if fmt == "ppt":
        return build_ppt_bytes(pack=pack, facts=facts), filename, (
            "application/vnd.openxmlformats-officedocument.presentationml.presentation"
        )
    raise ValueError(f"Unsupported export format: {export_format}")
