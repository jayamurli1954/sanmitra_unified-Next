from __future__ import annotations

import json
import re
from decimal import Decimal
from io import BytesIO
from typing import Any, Iterable

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook

from app.modules.office_ai.models import MIS_ENTITY_TYPES


def _normalize_header(value: Any) -> str:
    key = str(value or "").strip().lower()
    key = re.sub(r"[^a-z0-9]+", "_", key).strip("_")
    return key


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def _coerce_decimal_string(value: Any) -> str | None:
    if value is None or value == "":
        return None
    try:
        # Convert through Decimal to avoid binary float artifacts.
        return format(Decimal(str(value)), "f")
    except Exception:
        return None


def _parse_dimensions(value: Any) -> dict[str, Any]:
    if value is None or value == "":
        return {}
    if isinstance(value, dict):
        return dict(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return {}
        # Allow either JSON string or "k=v;k2=v2" (very light fallback).
        try:
            parsed = json.loads(text)
            if isinstance(parsed, dict):
                return parsed
        except Exception:
            pass
        if "=" in text and ";" in text:
            out: dict[str, Any] = {}
            for part in text.split(";"):
                if not part.strip():
                    continue
                k, _, v = part.partition("=")
                out[k.strip()] = v.strip()
            return out
    raise ValueError("dimensions must be a JSON object")


def _choose_sheet(wb, sheet_name: str | None) -> str:
    if sheet_name:
        if sheet_name in wb.sheetnames:
            return sheet_name
        raise HTTPException(status_code=400, detail=f"Sheet not found: {sheet_name}")

    # Template-first: accept either MIS_FACTS / MIS_Facts or first non-empty sheet.
    for name in wb.sheetnames:
        n = str(name or "").strip().lower()
        if "fact" in n:
            return name
    return wb.sheetnames[0] if wb.sheetnames else ""


def _parse_xlsx_rows(content: bytes, *, sheet_name: str | None = None) -> list[dict[str, Any]]:
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = _choose_sheet(wb, sheet_name=sheet_name)
    if not sheet:
        return []
    ws = wb[sheet]

    # Header row = row 1.
    header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_values:
        return []
    headers = [_normalize_header(h) for h in header_values]
    if not any(headers):
        return []

    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row_dict: dict[str, Any] = {}
        any_value = False
        for idx, cell_value in enumerate(values):
            if idx >= len(headers):
                continue
            key = headers[idx]
            if not key:
                continue
            if not _is_blank(cell_value):
                any_value = True
                row_dict[key] = cell_value
        if not any_value:
            continue
        rows.append(row_dict)
    return rows


def _required_columns() -> set[str]:
    # v1 template contract (minimal):
    # - entity_type identifies which MIS fact entity this row creates
    # - source_ref is used for citations
    # - numeric metric column depends on entity_type
    return {"entity_type", "source_ref"}


def _infer_amount_columns(entity_type: str) -> tuple[str, ...]:
    # Support both amount_decimal and amount as aliases.
    if entity_type in {"pnl_line", "bs_line", "cash_summary", "aging_bucket"}:
        return ("amount_decimal", "amount", "amount_minor")
    if entity_type in {"kpi"}:
        return ("value",)
    if entity_type in {"party"}:
        return ("value",)
    return ()


def parse_mis_excel(
    *,
    file: UploadFile,
    sheet_name: str | None,
    max_rows: int = 5000,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """
    Parse + validate a SanMitra CA MIS template Excel.

    Returns (facts_preview, validation_report).
    Facts preview contains normalized dicts compatible with MISFactsInsertRequest.
    """
    content = file.file.read() if hasattr(file.file, "read") else None
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded Excel file is empty")

    rows = _parse_xlsx_rows(content, sheet_name=sheet_name)
    errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    facts: list[dict[str, Any]] = []

    expected = _required_columns()
    if rows:
        # Validate required columns by checking first row keys (template-first).
        first_keys = set(rows[0].keys())
        missing = sorted(expected - first_keys)
        if missing:
            for m in missing:
                errors.append({"row": 1, "column": m, "message": "Missing required column"})

    for idx, raw in enumerate(rows[:max_rows]):
        row_no = idx + 2  # Excel row numbers: header at row 1
        normalized: dict[str, Any] = {}

        entity_type = str(raw.get("entity_type") or "").strip().lower()
        if not entity_type:
            errors.append({"row": row_no, "column": "entity_type", "message": "entity_type is required"})
            continue
        if entity_type not in MIS_ENTITY_TYPES:
            errors.append({"row": row_no, "column": "entity_type", "message": f"Unknown entity_type: {entity_type}"})
            continue

        source_ref = str(raw.get("source_ref") or "").strip()
        if not source_ref:
            errors.append({"row": row_no, "column": "source_ref", "message": "source_ref is required"})
            continue

        normalized["entity_type"] = entity_type
        normalized["fact_id"] = str(raw.get("fact_id") or "").strip() or None
        normalized["period"] = str(raw.get("period") or "").strip() or None
        normalized["as_of"] = str(raw.get("as_of") or "").strip() or None
        normalized["source_system"] = str(raw.get("source_system") or "excel_import").strip().lower() or "excel_import"
        normalized["source_id"] = str(raw.get("source_id") or "").strip() or None
        normalized["source_ref"] = source_ref

        currency = str(raw.get("currency") or "").strip().upper() or "INR"
        normalized["currency"] = currency

        dimensions_raw = raw.get("dimensions_json") or raw.get("dimensions")
        try:
            normalized["dimensions"] = _parse_dimensions(dimensions_raw)
        except Exception as exc:
            errors.append({"row": row_no, "column": "dimensions", "message": str(exc)})
            continue

        # Metric coercion
        amount_decimal = None
        amount_minor = None
        value = None

        if entity_type in {"pnl_line", "bs_line", "cash_summary", "aging_bucket"}:
            if raw.get("amount_decimal") not in (None, ""):
                amount_decimal = _coerce_decimal_string(raw.get("amount_decimal"))
            elif raw.get("amount") not in (None, ""):
                amount_decimal = _coerce_decimal_string(raw.get("amount"))
            elif raw.get("amount_minor") not in (None, ""):
                # amount_minor stored as int minor units.
                try:
                    amount_minor = int(raw.get("amount_minor"))
                except Exception:
                    errors.append({"row": row_no, "column": "amount_minor", "message": "amount_minor must be integer"})
                    continue

            if amount_decimal is None and amount_minor is None:
                errors.append(
                    {
                        "row": row_no,
                        "column": "amount_decimal/amount_minor",
                        "message": "amount_decimal (or amount_minor) is required for this entity_type",
                    }
                )
                continue

        elif entity_type in {"kpi", "party"}:
            if raw.get("value") in (None, ""):
                errors.append({"row": row_no, "column": "value", "message": "value is required for this entity_type"})
                continue
            # Keep as-is-ish; mis_store stores raw `value`.
            value = raw.get("value")

        normalized["amount_decimal"] = amount_decimal
        normalized["amount_minor"] = amount_minor
        normalized["value"] = value

        facts.append(normalized)

    validation_report = {
        "total_rows_in_sheet": len(rows),
        "facts_previewed": len(facts),
        "errors": errors,
        "warnings": warnings,
    }
    return facts, validation_report

