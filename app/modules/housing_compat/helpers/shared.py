"""Shared housing_compat helpers extracted from router.py.

Pure move per docs/operations/LARGE_FILE_MODULARIZATION_PLAN.md.
"""
from __future__ import annotations

import calendar
import csv
import json
import re
import unicodedata
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import uuid4

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.accounting.models import Account, JournalEntry, JournalLine
from app.modules.housing_compat import router as housing_router

def _format_upload_size(limit_bytes: int) -> str:
    if limit_bytes >= 1024 * 1024:
        return f"{limit_bytes / (1024 * 1024):g} MB"
    if limit_bytes >= 1024:
        return f"{limit_bytes / 1024:g} KB"
    return f"{limit_bytes} bytes"


async def _read_housing_upload_with_size_limit(file: UploadFile, limit_bytes: int, feature_name: str) -> bytes:
    data = bytearray()
    while True:
        chunk = await file.read(1024 * 1024)
        if not chunk:
            break
        data.extend(chunk)
        if len(data) > limit_bytes:
            raise HTTPException(
                status_code=413,
                detail=f"{feature_name} exceeds the upload limit of {housing_router._format_upload_size(limit_bytes)}.",
            )
    if not data:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    return bytes(data)




def _safe_file_name(original: str, prefix: str = "doc") -> str:
    name = (original or "").strip()
    stem, dot, ext = name.rpartition(".")
    raw_base = stem if dot else name
    raw_ext = ext.lower() if dot else "bin"
    base = re.sub(r"[^a-zA-Z0-9_-]+", "-", raw_base).strip("-").lower() or prefix
    safe_ext = re.sub(r"[^a-z0-9]+", "", raw_ext) or "bin"
    return f"{prefix}_{base}_{uuid4().hex[:10]}.{safe_ext}"


def _meeting_collections() -> tuple[Any, Any]:
    return housing_router.get_collection("housing_meetings"), housing_router.get_collection("housing_meeting_resolutions")


def _message_collections() -> tuple[Any, Any]:
    return housing_router.get_collection("housing_message_rooms"), housing_router.get_collection("housing_messages")


def _complaints_collection() -> Any:
    return housing_router.get_collection("housing_complaints")


def _visitors_collection() -> Any:
    return housing_router.get_collection("housing_visitor_entries")


def _asset_response(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": row.get("id"),
        "tenant_id": row.get("tenant_id"),
        "app_key": row.get("app_key"),
        "asset_code": row.get("asset_code") or row.get("account_code") or "",
        "name": row.get("name") or "",
        "category": row.get("category") or "other",
        "account_code": housing_router._canonical_gruha_account_code(row.get("account_code") or "16003"),
        "quantity": housing_router._safe_int(row.get("quantity"), default=1),
        "location": row.get("location") or "",
        "status": row.get("status") or "Active",
        "acquisition_type": row.get("acquisition_type") or "builder_handover",
        "handover_date": row.get("handover_date"),
        "purchase_date": row.get("purchase_date"),
        "original_cost": housing_router._safe_float(row.get("original_cost")),
        "depreciation_method": row.get("depreciation_method") or "straight_line",
        "depreciation_rate": housing_router._safe_float(row.get("depreciation_rate")),
        "useful_life_years": housing_router._safe_int(row.get("useful_life_years"), default=0),
        "residual_value": housing_router._safe_float(row.get("residual_value")),
        "amc_vendor": row.get("amc_vendor") or "",
        "amc_expiry": row.get("amc_expiry"),
        "insurance_policy_no": row.get("insurance_policy_no") or "",
        "insurance_expiry": row.get("insurance_expiry"),
        "vendor_name": row.get("vendor_name") or "",
        "invoice_no": row.get("invoice_no") or "",
        "notes": row.get("notes") or "",
        "is_scrapped": bool(row.get("is_scrapped") or False),
        "journal_entry_id": row.get("journal_entry_id"),
        "accounting_posting_status": row.get("accounting_posting_status") or "not_posted",
        "accounting_posted_at": row.get("accounting_posted_at"),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
    }


async def _next_asset_code(*, tenant_id: str, app_key: str, category: str) -> str:
    prefix = re.sub(r"[^A-Z0-9]+", "", str(category or "ASSET").upper())[:3] or "AST"
    count = await housing_router.get_collection("housing_assets").count_documents({"tenant_id": tenant_id, "app_key": app_key})
    return f"{prefix}-{count + 1:04d}"



async def _get_meeting_or_404(*, tenant_id: str, app_key: str, meeting_id: str) -> dict[str, Any]:
    meetings, _ = housing_router._meeting_collections()
    row = await meetings.find_one({"tenant_id": tenant_id, "app_key": app_key, "id": meeting_id})
    if not row:
        raise HTTPException(status_code=404, detail="Meeting not found")
    return row


async def _get_or_create_message_room(
    *,
    tenant_id: str,
    app_key: str,
    name: str,
    room_type: str,
    description: str,
    audience_type: str = "public",
    allowed_flat_numbers: list[str] | None = None,
) -> dict[str, Any]:
    rooms_col, _ = housing_router._message_collections()
    room = await rooms_col.find_one({"tenant_id": tenant_id, "app_key": app_key, "type": room_type})
    if room:
        return room

    now = datetime.now(timezone.utc).isoformat()
    room = {
        "id": str(uuid4()),
        "tenant_id": tenant_id,
        "app_key": app_key,
        "name": name,
        "type": room_type,
        "description": description,
        "audience_type": audience_type,
        "allowed_flat_numbers": housing_router._normalize_flat_numbers(allowed_flat_numbers or []),
        "allowed_member_ids": [],
        "created_by": "system",
        "created_at": now,
        "updated_at": now,
        "last_message_at": None,
    }
    await rooms_col.insert_one(room)
    return room


def _build_meeting_notice_message(meeting: dict[str, Any], eligible_members: int) -> str:
    agenda_items = meeting.get("agenda_items") or []
    agenda_lines: list[str] = []
    if isinstance(agenda_items, list):
        for item in agenda_items:
            if not isinstance(item, dict):
                continue
            title = str(item.get("item_title") or "").strip()
            description = str(item.get("item_description") or "").strip()
            if title and description:
                agenda_lines.append(f"- {title}: {description}")
            elif title:
                agenda_lines.append(f"- {title}")

    lines = [
        f"Meeting Notice: {meeting.get('meeting_title') or 'Society Meeting'}",
        f"Type: {meeting.get('meeting_type') or 'Meeting'}",
        f"Date: {meeting.get('meeting_date') or 'To be confirmed'}",
        f"Time: {meeting.get('meeting_time') or 'To be confirmed'}",
        f"Venue: {meeting.get('venue') or 'To be confirmed'}",
        f"Eligible members: {eligible_members}",
    ]
    if agenda_lines:
        lines.extend(["", "Agenda:", *agenda_lines])
    return "\n".join(lines)


def _sanitize_mongo_doc(doc: dict[str, Any]) -> dict[str, Any]:
    return {k: v for k, v in doc.items() if k != "_id"}


def _normalize_status(value: Any) -> str:
    status = str(value or "SCHEDULED").strip().upper()
    return status if status in {"SCHEDULED", "COMPLETED", "CANCELLED"} else "SCHEDULED"


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _round_money(value: float) -> float:
    return round(float(value or 0.0) + 0.0000001, 2)


def _as_decimal_money(value: Any) -> Decimal:
    return Decimal(str(housing_router._round_money(housing_router._safe_float(value)))).quantize(Decimal("0.01"))


def _month_name(month: int) -> str:
    names = [
        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July",
        "August",
        "September",
        "October",
        "November",
        "December",
    ]
    return names[month - 1] if 1 <= month <= 12 else str(month)


def _split_charge(total: float, index: int, count: int, *, weight: float = 1.0, total_weight: float = 0.0) -> float:
    if total <= 0 or count <= 0:
        return 0.0
    if total_weight > 0:
        return housing_router._round_money(total * weight / total_weight)
    base = housing_router._round_money(total / count)
    if index == count - 1:
        return housing_router._round_money(total - (base * (count - 1)))
    return base


async def _count_eligible_members(*, tenant_id: str, app_key: str) -> int:
    rows = await housing_router.get_collection("housing_members").find(
        {"tenant_id": tenant_id, "app_key": app_key},
        {"status": 1},
    ).to_list(length=5000)
    count = 0
    for row in rows:
        status = str(row.get("status") or "active").strip().lower()
        if status in {"moved_out", "inactive", "closed", "rejected"}:
            continue
        count += 1
    return count


def _normalize_member_ids(value: Any) -> list[str]:
    raw_values: list[Any]
    if isinstance(value, list):
        raw_values = value
    elif isinstance(value, str):
        raw_values = re.split(r"[,;\n]+", value)
    else:
        raw_values = []
    seen: set[str] = set()
    member_ids: list[str] = []
    for item in raw_values:
        member_id = str(item or "").strip()
        if member_id and member_id not in seen:
            seen.add(member_id)
            member_ids.append(member_id)
    return member_ids


async def _resolve_eligible_meeting_members(
    *, tenant_id: str, app_key: str, member_ids: list[str] | None = None, flat_numbers: list[str] | None = None
) -> list[dict[str, Any]]:
    member_ids = housing_router._normalize_member_ids(member_ids or [])
    flat_numbers = housing_router._normalize_flat_numbers(flat_numbers or [])
    clauses: list[dict[str, Any]] = []
    if member_ids:
        clauses.append({"id": {"$in": member_ids}})
    if flat_numbers:
        clauses.append({"flat_number": {"$in": flat_numbers}})
    if not clauses:
        return []

    query: dict[str, Any] = {"tenant_id": tenant_id, "app_key": app_key}
    query["$or"] = clauses
    rows = await housing_router.get_collection("housing_members").find(
        query,
        {"id": 1, "flat_number": 1, "status": 1, "name": 1, "email": 1},
    ).to_list(length=5000)
    eligible: list[dict[str, Any]] = []
    for row in rows:
        status = str(row.get("status") or "active").strip().lower()
        if status in {"moved_out", "inactive", "closed", "rejected"}:
            continue
        eligible.append(row)
    return eligible


async def _meeting_eligible_count(*, tenant_id: str, app_key: str, meeting: dict[str, Any]) -> int:
    member_ids = housing_router._normalize_member_ids(meeting.get("eligible_member_ids") or [])
    flat_numbers = housing_router._normalize_flat_numbers(meeting.get("eligible_flat_numbers") or [])
    if member_ids or flat_numbers:
        members = await housing_router._resolve_eligible_meeting_members(
            tenant_id=tenant_id,
            app_key=app_key,
            member_ids=member_ids,
            flat_numbers=flat_numbers,
        )
        return len(members) or len(member_ids) or len(flat_numbers)
    return await housing_router._count_eligible_members(tenant_id=tenant_id, app_key=app_key)


def _meeting_stats(
    meeting_row: dict[str, Any], *, eligible_hint: int = 0
) -> dict[str, int | bool]:
    attendance = meeting_row.get("attendance") or []
    if not isinstance(attendance, list):
        attendance = []
    total_members_present = sum(
        1
        for item in attendance
        if str((item or {}).get("status") or "").strip().lower() in {"present", "proxy"}
    )
    total_members_eligible = housing_router._safe_int(meeting_row.get("total_members_eligible"), default=0)
    if total_members_eligible <= 0 and eligible_hint > 0:
        total_members_eligible = eligible_hint
    quorum_required = housing_router._safe_int(meeting_row.get("quorum_required"), default=0)
    quorum_met = bool(meeting_row.get("quorum_met") or False)
    if quorum_required > 0:
        quorum_met = total_members_present >= quorum_required
    return {
        "total_members_present": total_members_present,
        "total_members_eligible": total_members_eligible,
        "quorum_required": quorum_required,
        "quorum_met": quorum_met,
    }


def _build_simple_pdf(
    lines: list[str], *, title: str | None = None, top_margin: float = 60.0
) -> bytes:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - max(top_margin, 40.0)
    if title:
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(48, y, housing_router._sanitize_pdf_text(title))
        y -= 24
    pdf.setFont("Helvetica", 11)
    for raw in lines:
        line = housing_router._sanitize_pdf_text(raw)
        if not line:
            y -= 12
            continue
        if y < 56:
            pdf.showPage()
            pdf.setFont("Helvetica", 11)
            y = height - 56
        pdf.drawString(48, y, line[:140])
        y -= 16
    pdf.showPage()
    pdf.save()
    return buffer.getvalue()


async def _get_society_branding(*, tenant_id: str, app_key: str) -> dict[str, str]:
    settings = await housing_router.get_collection("housing_society_settings").find_one(
        {"tenant_id": tenant_id, "app_key": app_key}
    ) or {}
    logo_url = str(settings.get("logo_url") or "").strip()
    logo_name = ""
    if logo_url:
        logo_name = logo_url.rstrip("/").split("/")[-1]
    if not logo_name:
        logo_doc = await housing_router.get_collection("housing_documents").find_one(
            {"tenant_id": tenant_id, "app_key": app_key, "document_type": "society_logo"},
            sort=[("updated_at", -1)],
        )
        if logo_doc:
            logo_name = str(logo_doc.get("stored_name") or "").strip()
    logo_bytes = b""
    if logo_name:
        logo_doc = await housing_router.get_collection("housing_documents").find_one(
            {"tenant_id": tenant_id, "app_key": app_key, "stored_name": logo_name},
            {"content": 1},
        )
        logo_bytes = bytes(logo_doc.get("content") or b"") if logo_doc else b""

    return {
        "society_name": str(settings.get("society_name") or "GruhaMitra Society").strip(),
        "society_address": str(settings.get("society_address") or "").strip(),
        "city": str(settings.get("city") or "").strip(),
        "state": str(settings.get("state") or "").strip(),
        "pin_code": str(settings.get("pin_code") or "").strip(),
        "contact_email": str(settings.get("contact_email") or "").strip(),
        "contact_phone": str(settings.get("contact_phone") or "").strip(),
        "logo_bytes": logo_bytes,
    }


def _maintenance_collections() -> tuple[Any, Any]:
    return housing_router.get_collection("housing_maintenance_bills"), housing_router.get_collection("housing_maintenance_reversals")


async def _list_maintenance_bills(*, tenant_id: str, app_key: str, month: int, year: int) -> list[dict[str, Any]]:
    bills_col, _ = housing_router._maintenance_collections()
    rows = (
        await bills_col.find({"tenant_id": tenant_id, "app_key": app_key, "month": month, "year": year})
        .sort([("flat_number", 1), ("created_at", 1)])
        .to_list(length=5000)
    )
    return [housing_router._sanitize_mongo_doc(row) for row in rows]


async def _flat_occupants_map(*, tenant_id: str, app_key: str) -> dict[str, int]:
    members = await housing_router.get_collection("housing_members").find(
        {"tenant_id": tenant_id, "app_key": app_key},
        {"flat_number": 1, "status": 1, "is_primary": 1, "total_occupants": 1},
    ).to_list(length=5000)
    occupants: dict[str, int] = {}
    for member in members:
        status = str(member.get("status") or "active").strip().lower()
        if status in {"moved_out", "inactive", "closed", "rejected"}:
            continue
        flat_number = str(member.get("flat_number") or "").strip().upper()
        if not flat_number:
            continue
        count = max(0, housing_router._safe_int(member.get("total_occupants"), 1))
        if member.get("is_primary") or flat_number not in occupants:
            occupants[flat_number] = count
    return occupants


def _month_date_range(month: int, year: int) -> tuple[date, date]:
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_day)


def _expense_period_labels(month: int, year: int) -> set[str]:
    full = calendar.month_name[month]
    short = calendar.month_abbr[month]
    short_year = f"{year % 100:02d}"
    return {
        f"{full} {year}".lower(),
        f"{full}, {year}".lower(),
        f"{short} {year}".lower(),
        f"{short}, {year}".lower(),
        f"{short}-{year}".lower(),
        f"{month:02d}-{year}".lower(),
        f"{year}-{month:02d}".lower(),
        f"{month:02d}/{year}".lower(),
        f"{full} {short_year}".lower(),
        f"{full}, {short_year}".lower(),
        f"{full}-{short_year}".lower(),
        f"{short} {short_year}".lower(),
        f"{short}. {short_year}".lower(),
        f"{short}, {short_year}".lower(),
        f"{short}-{short_year}".lower(),
        f"{month:02d}-{short_year}".lower(),
        f"{month:02d}/{short_year}".lower(),
    }


def _matches_expense_period(txn: dict[str, Any], *, month: int, year: int) -> bool:
    labels = housing_router._expense_period_labels(month, year)
    explicit = " ".join(
        str(txn.get(key) or "")
        for key in ("expense_month", "expense_for_month", "period", "billing_month")
    ).lower()
    if explicit and any(label in explicit for label in labels):
        return True
    text_parts = [str(txn.get("description") or ""), str(txn.get("narration") or "")]
    for line in txn.get("lines") or []:
        text_parts.append(str((line or {}).get("description") or ""))
    text = " ".join(text_parts).lower()
    return any(label in text for label in labels)


def _is_water_expense_account(code: Any, name: Any, *context: Any) -> bool:
    text = " ".join(str(value or "") for value in (code, name, *context)).lower()
    return any(token in text for token in ("water", "tanker", "bwssb", "borewell"))


async def _expense_accounts_for_period(
    session: AsyncSession,
    *,
    tenant_id: str,
    app_key: str,
    month: int,
    year: int,
) -> list[dict[str, Any]]:
    from_date, to_date = housing_router._month_date_range(month, year)
    stmt = (
        select(
            Account.code.label("account_code"),
            Account.name.label("account_name"),
            func.coalesce(func.sum(JournalLine.debit), 0).label("debit_total"),
            func.coalesce(func.sum(JournalLine.credit), 0).label("credit_total"),
            func.count(JournalLine.id).label("transaction_count"),
        )
        .join(JournalLine, JournalLine.account_id == Account.id)
        .join(JournalEntry, JournalEntry.id == JournalLine.journal_id)
        .where(
            and_(
                Account.tenant_id == tenant_id,
                Account.app_key == app_key,
                Account.accounting_entity_id == "primary",
                Account.type == "expense",
                JournalEntry.tenant_id == tenant_id,
                JournalEntry.app_key == app_key,
                JournalEntry.accounting_entity_id == "primary",
                JournalEntry.entry_date >= from_date,
                JournalEntry.entry_date <= to_date,
            )
        )
        .group_by(Account.code, Account.name)
        .order_by(Account.code.asc().nullslast(), Account.name.asc())
    )
    rows = (await session.execute(stmt)).all()
    accounts_by_key: dict[str, dict[str, Any]] = {}
    for row in rows:
        amount = housing_router._round_money(float(row.debit_total or 0) - float(row.credit_total or 0))
        if amount <= 0:
            continue
        code = str(row.account_code or "")
        is_water = housing_router._is_water_expense_account(row.account_code, row.account_name)
        accounts_by_key[f"{code}:{'water' if is_water else 'fixed'}"] = {
            "account_code": row.account_code,
            "account_name": row.account_name,
            "total_amount": amount,
            "transaction_count": int(row.transaction_count or 0),
            "is_water": is_water,
        }

    txns = await housing_router.get_collection("mb_transactions").find(
        {"tenant_id": tenant_id, "app_key": app_key, "voucher_type": "payment"}
    ).to_list(length=5000)
    period_txns = [txn for txn in txns if housing_router._matches_expense_period(txn, month=month, year=year)]
    mongo_codes = sorted(
        {
            str((line or {}).get("account_code") or "")
            for txn in period_txns
            for line in (txn.get("lines") or [])
            if housing_router._safe_float((line or {}).get("debit")) > 0 and str((line or {}).get("account_code") or "")
        }
    )
    names_by_code: dict[str, str] = {}
    if mongo_codes:
        account_rows = (
            await session.execute(
                select(Account.code, Account.name).where(
                    Account.tenant_id == tenant_id,
                    Account.app_key == app_key,
                    Account.accounting_entity_id == "primary",
                    Account.code.in_(mongo_codes),
                )
            )
        ).all()
        names_by_code = {str(code): str(name) for code, name in account_rows}

    for txn in period_txns:
        txn_context = " ".join(
            str(txn.get(key) or "")
            for key in ("description", "narration", "paid_to", "vendor_name", "reference", "voucher_number")
        )
        for line in txn.get("lines") or []:
            debit = housing_router._safe_float((line or {}).get("debit"))
            code = str((line or {}).get("account_code") or "")
            if debit <= 0 or not code:
                continue
            line_description = str((line or {}).get("description") or "")
            name = names_by_code.get(code) or line_description or code
            is_water = housing_router._is_water_expense_account(code, name, line_description, txn_context)
            account_key = f"{code}:{'water' if is_water else 'fixed'}"
            existing = accounts_by_key.get(account_key)
            if existing:
                # If the SQL date query already counted the same voucher, do not double count.
                existing_refs = existing.setdefault("_mongo_refs", set())
                ref = str(txn.get("voucher_number") or txn.get("journal_entry_id") or txn.get("id"))
                if ref in existing_refs:
                    continue
                existing_refs.add(ref)
                existing["total_amount"] = housing_router._round_money(housing_router._safe_float(existing.get("total_amount")) + debit)
                existing["transaction_count"] = housing_router._safe_int(existing.get("transaction_count")) + 1
            else:
                accounts_by_key[account_key] = {
                    "account_code": code,
                    "account_name": name,
                    "total_amount": housing_router._round_money(debit),
                    "transaction_count": 1,
                    "is_water": is_water,
                    "_mongo_refs": {str(txn.get("voucher_number") or txn.get("journal_entry_id") or txn.get("id"))},
                }

    accounts = []
    for row in accounts_by_key.values():
        row.pop("_mongo_refs", None)
        accounts.append(row)
    return sorted(accounts, key=lambda item: (str(item.get("account_code") or ""), str(item.get("account_name") or "")))


async def _build_maintenance_bills(
    *,
    tenant_id: str,
    app_key: str,
    payload: dict[str, Any],
    current_user: dict[str, Any],
    session: AsyncSession,
    replace_existing: bool = True,
) -> dict[str, Any]:
    month = housing_router._safe_int(payload.get("month"))
    year = housing_router._safe_int(payload.get("year"))
    if month < 1 or month > 12 or year < 2000:
        raise HTTPException(status_code=422, detail="Valid month and year are required")

    flats = await housing_router.list_flats(tenant_id=tenant_id, app_key=app_key)
    if not flats:
        raise HTTPException(status_code=400, detail="No flats found. Please add flats before generating bills.")

    flat_occupants = await housing_router._flat_occupants_map(tenant_id=tenant_id, app_key=app_key)
    billable_flats = [
        flat
        for flat in flats
        if str(flat.get("flat_number") or "").strip().upper() in flat_occupants
    ]
    if not billable_flats:
        raise HTTPException(
            status_code=400,
            detail="No active onboarded members are assigned to flats. Onboard or assign members before generating bills.",
        )

    bills_col, _ = housing_router._maintenance_collections()
    if replace_existing:
        existing_posted = await bills_col.count_documents(
            {"tenant_id": tenant_id, "app_key": app_key, "month": month, "year": year, "is_posted": True}
        )
        if existing_posted:
            raise HTTPException(
                status_code=409,
                detail=f"Posted bills already exist for {housing_router._month_name(month)} {year}. Reverse them before regenerating.",
            )
        await bills_col.delete_many({"tenant_id": tenant_id, "app_key": app_key, "month": month, "year": year})

    count = len(billable_flats)
    total_area = sum(housing_router._safe_float(flat.get("area_sqft")) for flat in billable_flats)
    adjusted_inmates = payload.get("adjusted_inmates") or {}
    inmate_counts = [
        max(
            0,
            housing_router._safe_int(
                adjusted_inmates.get(str(flat.get("id")))
                or adjusted_inmates.get(str(flat.get("flat_number")))
                or flat_occupants.get(str(flat.get("flat_number") or "").strip().upper()),
                0,
            ),
        )
        for flat in billable_flats
    ]
    total_inmates = sum(inmate_counts) or count

    settings = await housing_router.get_society_settings(tenant_id=tenant_id, app_key=app_key)
    sqft_rate = (
        housing_router._safe_float(payload.get("override_sqft_rate"))
        if payload.get("override_sqft_rate") not in (None, "")
        else housing_router._safe_float(settings.get("maintenance_rate_sqft"))
    )
    flat_maintenance_rate = housing_router._safe_float(settings.get("maintenance_rate_flat"))
    expense_accounts = await housing_router._expense_accounts_for_period(
        session,
        tenant_id=tenant_id,
        app_key=app_key,
        month=month,
        year=year,
    )
    non_water_expense_codes = [
        str(row.get("account_code"))
        for row in expense_accounts
        if row.get("account_code") and not row.get("is_water")
    ]
    selected_fixed_codes = [str(code) for code in (payload.get("selected_fixed_expense_codes") or [])]
    if not selected_fixed_codes:
        selected_fixed_codes = non_water_expense_codes
    ledger_water_total = housing_router._round_money(
        sum(housing_router._safe_float(row.get("total_amount")) for row in expense_accounts if row.get("is_water"))
    )
    selected_fixed_total = housing_router._round_money(
        sum(
            housing_router._safe_float(row.get("total_amount"))
            for row in expense_accounts
            if not row.get("is_water") and str(row.get("account_code")) in selected_fixed_codes
        )
    )

    total_water = (
        housing_router._safe_float(payload.get("override_water_charges"))
        if payload.get("override_water_charges") not in (None, "")
        else ledger_water_total
    )
    water_rate_per_person_exact = (total_water / total_inmates) if total_water and total_inmates else 0.0
    water_rate_per_person = round(water_rate_per_person_exact, 4)
    total_fixed = (
        housing_router._safe_float(payload.get("override_fixed_expenses"))
        if payload.get("override_fixed_expenses") not in (None, "")
        else selected_fixed_total
    )
    sinking_rate = housing_router._safe_float(settings.get("sinking_fund_rate"))
    repair_rate = housing_router._safe_float(settings.get("repair_fund_rate"))
    association_rate = housing_router._safe_float(settings.get("association_fund_rate"))
    corpus_rate = housing_router._safe_float(settings.get("corpus_fund_rate"))
    total_sinking = (
        housing_router._safe_float(payload.get("override_sinking_fund"))
        if payload.get("override_sinking_fund") not in (None, "")
        else housing_router._round_money(sinking_rate * count)
    )
    total_repair = (
        housing_router._safe_float(payload.get("override_repair_fund"))
        if payload.get("override_repair_fund") not in (None, "")
        else housing_router._round_money(repair_rate * count)
    )
    total_association = housing_router._round_money(association_rate * count)
    total_corpus = (
        housing_router._safe_float(payload.get("override_corpus_fund"))
        if payload.get("override_corpus_fund") not in (None, "")
        else housing_router._round_money(corpus_rate * count)
    )
    fixed_method = str(payload.get("fixed_calculation_method") or "equal").lower()
    sinking_method = str(payload.get("sinking_calculation_method") or "equal").lower()
    repair_method = str(payload.get("repair_fund_calculation_method") or "equal").lower()
    corpus_method = str(payload.get("corpus_fund_calculation_method") or "equal").lower()

    now = datetime.now(timezone.utc).isoformat()
    docs: list[dict[str, Any]] = []
    for idx, flat in enumerate(billable_flats):
        flat_area = housing_router._safe_float(flat.get("area_sqft"))
        inmates = inmate_counts[idx]
        maintenance = housing_router._round_money((flat_area * sqft_rate if sqft_rate > 0 else 0.0) + flat_maintenance_rate)
        water = housing_router._round_money(water_rate_per_person_exact * inmates)

        def fund_share(total: float, method: str) -> float:
            use_area = method in {"sqft", "area", "area_sqft"}
            return housing_router._split_charge(total, idx, count, weight=flat_area, total_weight=total_area if use_area else 0.0)

        fixed = fund_share(total_fixed, fixed_method)
        sinking = fund_share(total_sinking, sinking_method)
        repair = fund_share(total_repair, repair_method)
        association = housing_router._split_charge(total_association, idx, count)
        corpus = fund_share(total_corpus, corpus_method)
        total_amount = housing_router._round_money(maintenance + water + fixed + sinking + repair + association + corpus)
        bill_id = str(uuid4())
        doc = {
            "id": bill_id,
            "tenant_id": tenant_id,
            "app_key": app_key,
            "flat_id": flat.get("id"),
            "flat_number": flat.get("flat_number"),
            "month": month,
            "year": year,
            "amount": total_amount,
            "maintenance_amount": maintenance,
            "water_amount": water,
            "fixed_amount": fixed,
            "sinking_fund_amount": sinking,
            "repair_fund_amount": repair,
            "association_fund_amount": association,
            "corpus_fund_amount": corpus,
            "arrears_amount": 0.0,
            "late_fee_amount": 0.0,
            "status": "generated",
            "is_posted": False,
            "posting_status": "not_posted",
            "auto_post_requested": bool(payload.get("auto_post_to_accounting")),
            "breakdown": {
                "maintenance_sqft": maintenance,
                "sqft_calculation": (
                    f"{flat_area:g} sq.ft x {sqft_rate:g}"
                    + (f" + flat charge {flat_maintenance_rate:,.2f}" if flat_maintenance_rate else "")
                    if maintenance
                    else ""
                ),
                "water_charges": water,
                "water_per_person_rate": water_rate_per_person,
                "inmates_used": inmates,
                "total_inmates": total_inmates,
                "water_calculation": (
                    f"Water expenses {total_water:,.2f} / {total_inmates} residents = "
                    f"{water_rate_per_person:,.4f} x {inmates} resident(s)"
                    if water and total_inmates
                    else ""
                ),
                "water_source_total": ledger_water_total,
                "fixed_expenses": fixed,
                "fixed_expenses_calculation": f"{fixed_method.title()} share of selected expenses {total_fixed:,.2f}" if fixed else "",
                "selected_fixed_expense_codes": selected_fixed_codes,
                "selected_fixed_expenses_total": selected_fixed_total,
                "sinking_fund": sinking,
                "sinking_fund_calculation": f"Per flat setting {sinking_rate:,.2f}" if sinking_rate else f"{sinking_method.title()} share" if sinking else "",
                "repair_fund": repair,
                "repair_fund_calculation": f"Per flat setting {repair_rate:,.2f}" if repair_rate else f"{repair_method.title()} share" if repair else "",
                "association_fund": association,
                "association_fund_calculation": f"Per flat setting {association_rate:,.2f}" if association_rate else "",
                "corpus_fund": corpus,
                "corpus_fund_calculation": f"Per flat setting {corpus_rate:,.2f}" if corpus_rate else f"{corpus_method.title()} share" if corpus else "",
            },
            "created_by": str(current_user.get("sub") or "system"),
            "created_at": now,
            "updated_at": now,
        }
        docs.append(doc)

    if docs:
        await bills_col.insert_many(docs)

    return {
        "month": month,
        "year": year,
        "total_bills_generated": len(docs),
        "total_amount": housing_router._round_money(sum(housing_router._safe_float(doc.get("amount")) for doc in docs)),
        "ledger_water_total": ledger_water_total,
        "selected_fixed_expenses_total": selected_fixed_total,
        "bills": [housing_router._sanitize_mongo_doc(doc) for doc in docs],
    }





def _draw_pdf_header(pdf: canvas.Canvas, *, title: str, branding: dict[str, Any]) -> float:
    page_w, page_h = A4
    margin_x = 42
    y_top = page_h - 38
    logo_bytes = branding.get("logo_bytes") or b""
    if logo_bytes:
        try:
            img = ImageReader(BytesIO(logo_bytes))
            pdf.drawImage(img, margin_x, y_top - 48, width=42, height=42, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass
    text_x = margin_x + 52
    society_name = str(branding.get("society_name") or "GruhaMitra Society").strip()
    address_parts = [
        str(branding.get("society_address") or "").strip(),
        ", ".join(
            [p for p in [branding.get("city"), branding.get("state"), branding.get("pin_code")] if str(p or "").strip()]
        ),
    ]
    contact_parts = [p for p in [branding.get("contact_phone"), branding.get("contact_email")] if str(p or "").strip()]

    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(text_x, y_top, society_name[:75])
    pdf.setFont("Helvetica", 9)
    y = y_top - 14
    for part in address_parts:
        part_text = str(part or "").strip()
        if part_text:
            pdf.drawString(text_x, y, part_text[:110])
            y -= 11
    if contact_parts:
        pdf.drawString(text_x, y, " | ".join(contact_parts)[:110])
        y -= 11

    pdf.setLineWidth(0.8)
    pdf.line(margin_x, y - 2, page_w - margin_x, y - 2)
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(margin_x, y - 20, title[:95])
    return y - 42


def _build_branded_pdf(*, title: str, lines: list[str], branding: dict[str, Any]) -> bytes:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    page_w, page_h = A4
    margin_x = 42
    y = housing_router._draw_pdf_header(pdf, title=title, branding=branding)
    pdf.setFont("Helvetica", 10.5)
    for raw in lines:
        line = housing_router._sanitize_pdf_text(raw)
        if not line:
            y -= 10
            continue
        if y < 56:
            pdf.showPage()
            y = housing_router._draw_pdf_header(pdf, title=title, branding=branding)
            pdf.setFont("Helvetica", 10.5)
        pdf.drawString(margin_x, y, line[:140])
        y -= 14
    pdf.setFont("Helvetica-Oblique", 8.5)
    pdf.drawString(
        margin_x,
        34,
        f"Generated by GruhaMitra on {datetime.now(timezone.utc).strftime('%d-%m-%Y %H:%M UTC')}",
    )
    pdf.save()
    return buffer.getvalue()


def _sanitize_pdf_text(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    # Drop common mojibake/replacement/control characters that render as black boxes.
    text = (
        text.replace("\ufffd", " ")
        .replace("\u25a0", " ")
        .replace("\u25aa", " ")
        .replace("\u200b", "")
        .replace("\u200c", "")
        .replace("\u200d", "")
        .replace("\ufeff", "")
    )
    text = re.sub(r"[\x00-\x1F\x7F]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _normalize_header(value: Any) -> str:
    key = str(value or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "_", key).strip("_")


def _truthy(value: Any, *, default: bool = False) -> bool:
    if value is None:
        return default
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y"}:
        return True
    if text in {"0", "false", "no", "n"}:
        return False
    return default


def _parse_member_move_in(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    try:
        parsed = datetime.fromisoformat(text)
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
    except ValueError:
        return None


def _row_get(row: dict[str, Any], keys: list[str], default: Any = None) -> Any:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return default


def _coerce_bulk_member_payload(row: dict[str, Any], row_no: int) -> MemberCreateRequest:
    name = str(housing_router._row_get(row, ["name", "full_name"]) or "").strip()
    phone = str(housing_router._row_get(row, ["phone_number", "phone", "mobile", "mobile_number"]) or "").strip()
    email_raw = housing_router._row_get(row, ["email", "email_address"], None)
    email = str(email_raw).strip().lower() if email_raw not in (None, "") else None
    flat = str(housing_router._row_get(row, ["flat_number", "flat", "flat_no", "unit", "unit_number"]) or "").strip().upper()
    member_type = str(housing_router._row_get(row, ["member_type", "type"], "owner") or "owner").strip().lower()
    move_in_date = housing_router._parse_member_move_in(housing_router._row_get(row, ["move_in_date", "move_in", "move_in_dt"], None))
    occupants_raw = housing_router._row_get(row, ["total_occupants", "occupants", "family_count"], 1)
    try:
        total_occupants = int(str(occupants_raw).strip())
    except Exception:
        total_occupants = 1
    occupation_raw = housing_router._row_get(row, ["occupation", "profession"], None)
    occupation = str(occupation_raw).strip() if occupation_raw not in (None, "") else None
    is_primary = housing_router._truthy(housing_router._row_get(row, ["is_primary", "primary"], True), default=True)
    is_mobile_public = housing_router._truthy(housing_router._row_get(row, ["is_mobile_public", "mobile_public"], False), default=False)

    try:
        return MemberCreateRequest(
            name=name,
            phone_number=phone,
            email=email,
            flat_number=flat,
            member_type=member_type,  # type: ignore[arg-type]
            move_in_date=move_in_date,
            total_occupants=total_occupants,
            is_primary=is_primary,
            occupation=occupation,
            is_mobile_public=is_mobile_public,
        )
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Row {row_no}: {exc}") from exc


def _parse_csv_rows(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(text.splitlines())
    rows: list[dict[str, Any]] = []
    for raw in reader:
        normalized: dict[str, Any] = {}
        for key, value in (raw or {}).items():
            normalized[housing_router._normalize_header(key)] = value
        if not any(str(v or "").strip() for v in normalized.values()):
            continue
        rows.append(normalized)
    return rows


def _parse_xlsx_rows(content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active
    header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_values:
        return []
    headers = [housing_router._normalize_header(h) for h in header_values]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        item: dict[str, Any] = {}
        for idx, value in enumerate(values):
            if idx >= len(headers):
                continue
            key = headers[idx]
            if key:
                item[key] = value
        if not any(str(v or "").strip() for v in item.values()):
            continue
        rows.append(item)
    return rows


def _can_manage_chat_rooms(role: Any) -> bool:
    normalized = str(role or "").strip().lower()
    return normalized in {
        "admin",
        "tenant_admin",
        "super_admin",
        "secretary",
        "society_admin",
        "societyadmin",
        "chairman",
        "chairperson",
        "president",
        "treasurer",
        "committee",
        "committee_member",
        "management_committee",
        "mc",
        "owner",
        "manager",
    }


def _normalize_flat_number(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def _normalize_flat_numbers(value: Any) -> list[str]:
    raw_values: list[Any]
    if isinstance(value, list):
        raw_values = value
    elif isinstance(value, str):
        raw_values = re.split(r"[,;\n]+", value)
    else:
        raw_values = []
    seen: set[str] = set()
    flats: list[str] = []
    for item in raw_values:
        flat = housing_router._normalize_flat_number(item)
        if flat and flat not in seen:
            seen.add(flat)
            flats.append(flat)
    return flats


def _room_response(row: dict[str, Any]) -> dict[str, Any]:
    audience_type = str(row.get("audience_type") or "public").strip().lower()
    allowed_flat_numbers = housing_router._normalize_flat_numbers(row.get("allowed_flat_numbers") or [])
    allowed_member_ids = [str(v).strip() for v in (row.get("allowed_member_ids") or []) if str(v or "").strip()]
    return {
        "id": row.get("id"),
        "name": row.get("name") or "General",
        "type": row.get("type") or "general",
        "description": row.get("description") or "",
        "audience_type": audience_type,
        "allowed_flat_numbers": allowed_flat_numbers,
        "allowed_member_ids": allowed_member_ids,
        "created_by": row.get("created_by"),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
        "last_message_at": row.get("last_message_at"),
    }


def _message_response(row: dict[str, Any]) -> dict[str, Any]:
    attachments = row.get("attachments") or []
    if not isinstance(attachments, list):
        attachments = []
    return {
        "id": row.get("id"),
        "room_id": row.get("room_id"),
        "sender_id": row.get("sender_id"),
        "sender_name": row.get("sender_name") or "User",
        "content": row.get("content") or "",
        "message_type": row.get("message_type") or "text",
        "attachments": [
            {
                "id": item.get("id"),
                "file_name": item.get("file_name") or "attachment",
                "content_type": item.get("content_type") or "application/octet-stream",
                "size": item.get("size") or 0,
                "download_url": item.get("download_url"),
            }
            for item in attachments
            if isinstance(item, dict)
        ],
        "expires_at": row.get("expires_at"),
        "retention_days": row.get("retention_days"),
        "created_at": row.get("created_at"),
    }


def _message_retention_days(value: Any) -> int:
    days = housing_router._safe_int(value, default=30)
    if days <= 0:
        days = 30
    return min(days, 90)


async def _current_user_message_audience(
    *, current_user: dict[str, Any], tenant_id: str, app_key: str
) -> dict[str, set[str]]:
    flat_candidates = {
        housing_router._normalize_flat_number(current_user.get(key))
        for key in ("flat_number", "flat_no", "unit_number", "unit_no")
    }
    member_candidates = {
        str(current_user.get(key) or "").strip()
        for key in ("sub", "id", "user_id", "email")
        if str(current_user.get(key) or "").strip()
    }
    flat_candidates.discard("")

    member_query_terms = [value for value in member_candidates if value]
    if member_query_terms:
        members = await housing_router.get_collection("housing_members").find(
            {
                "tenant_id": tenant_id,
                "app_key": app_key,
                "$or": [
                    {"user_id": {"$in": member_query_terms}},
                    {"id": {"$in": member_query_terms}},
                    {"email": {"$in": member_query_terms}},
                    {"member_id": {"$in": member_query_terms}},
                ],
            },
            {"flat_number": 1, "id": 1, "user_id": 1, "email": 1, "member_id": 1},
        ).to_list(length=50)
        for member in members:
            flat = housing_router._normalize_flat_number(member.get("flat_number"))
            if flat:
                flat_candidates.add(flat)
            for key in ("id", "user_id", "email", "member_id"):
                value = str(member.get(key) or "").strip()
                if value:
                    member_candidates.add(value)

    return {"flats": flat_candidates, "members": member_candidates}


async def _can_access_message_room(
    *, room: dict[str, Any], current_user: dict[str, Any], tenant_id: str, app_key: str
) -> bool:
    if housing_router._can_manage_chat_rooms(current_user.get("role")):
        return True
    audience_type = str(room.get("audience_type") or "public").strip().lower()
    if audience_type in {"", "public", "all"}:
        return True

    audience = await housing_router._current_user_message_audience(current_user=current_user, tenant_id=tenant_id, app_key=app_key)
    allowed_flats = set(housing_router._normalize_flat_numbers(room.get("allowed_flat_numbers") or []))
    if allowed_flats and audience["flats"].intersection(allowed_flats):
        return True

    allowed_members = {str(v).strip() for v in (room.get("allowed_member_ids") or []) if str(v or "").strip()}
    if allowed_members and audience["members"].intersection(allowed_members):
        return True
    return False


def _can_manage_complaints(role: Any) -> bool:
    normalized = str(role or "").strip().lower()
    return normalized in {"admin", "super_admin", "secretary"}


def _can_manage_visitors(role: Any) -> bool:
    normalized = str(role or "").strip().lower()
    return normalized in {
        "admin",
        "super_admin",
        "tenant_admin",
        "secretary",
        "chairman",
        "security",
        "security_guard",
        "guard",
        "gate",
        "watchman",
    }


def _can_manage_facilities(role: Any) -> bool:
    normalized = str(role or "").strip().lower()
    return normalized in {"admin", "super_admin", "tenant_admin", "secretary", "chairman", "committee"}


def _facility_booking_response(row: dict[str, Any]) -> dict[str, Any]:
    cleaned = housing_router._sanitize_mongo_doc(row)
    cleaned["amount"] = housing_router._as_decimal_money(cleaned.get("amount") or 0)
    cleaned["deposit_amount"] = housing_router._as_decimal_money(cleaned.get("deposit_amount") or 0)
    cleaned.setdefault("payment_status", "not_required")
    cleaned.setdefault("status", "pending")
    return cleaned


def _facility_confirmation_message(row: dict[str, Any]) -> str:
    return (
        f"{row.get('facility_name') or 'Facility'} booking confirmed for Flat {row.get('flat_number') or 'N/A'} "
        f"from {housing_router.format_date_time_for_display(row.get('start_time'))} to {housing_router.format_date_time_for_display(row.get('end_time'))}."
    )


def format_date_time_for_display(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y %H:%M")
    return str(value or "N/A")


async def _facility_booking_access_query(
    *, current_user: dict[str, Any], tenant_id: str, app_key: str
) -> dict[str, Any]:
    if housing_router._can_manage_facilities(current_user.get("role")):
        return {}
    audience = await housing_router._current_user_message_audience(current_user=current_user, tenant_id=tenant_id, app_key=app_key)
    flats = sorted(value for value in audience["flats"] if value)
    user_terms = sorted(value for value in audience["members"] if value)
    clauses: list[dict[str, Any]] = []
    if flats:
        clauses.append({"flat_number": {"$in": flats}})
    if user_terms:
        clauses.append({"created_by": {"$in": user_terms}})
    if not clauses:
        raise HTTPException(status_code=403, detail="Resident flat context is required for facility booking access")
    return {"$or": clauses}


async def _resolve_booking_flat(
    *, requested_flat: str | None, current_user: dict[str, Any], tenant_id: str, app_key: str
) -> str:
    normalized = housing_router._normalize_flat_number(requested_flat)
    if housing_router._can_manage_facilities(current_user.get("role")):
        if not normalized:
            raise HTTPException(status_code=422, detail="Flat number is required")
        return normalized

    audience = await housing_router._current_user_message_audience(current_user=current_user, tenant_id=tenant_id, app_key=app_key)
    flats = sorted(value for value in audience["flats"] if value)
    if normalized and normalized not in flats:
        raise HTTPException(status_code=403, detail="Residents can book facilities only for their own flat")
    if normalized:
        return normalized
    if not flats:
        raise HTTPException(status_code=403, detail="Resident flat context is required for facility booking")
    return flats[0]


async def _find_overlapping_facility_booking(
    *, tenant_id: str, app_key: str, facility_id: str, start_time: datetime, end_time: datetime, exclude_id: str | None = None
) -> dict[str, Any] | None:
    query: dict[str, Any] = {
        "tenant_id": tenant_id,
        "app_key": app_key,
        "facility_id": facility_id,
        "status": {"$in": ["pending", "approved"]},
        "start_time": {"$lt": end_time},
        "end_time": {"$gt": start_time},
    }
    if exclude_id:
        query["id"] = {"$ne": exclude_id}
    return await housing_router.get_collection("housing_facility_bookings").find_one(query)


def _visitor_response(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": row.get("id"),
        "tenant_id": row.get("tenant_id"),
        "app_key": row.get("app_key"),
        "flat_number": row.get("flat_number"),
        "visitor_type": row.get("visitor_type") or "guest",
        "visitor_name": row.get("visitor_name") or "",
        "phone_number": row.get("phone_number") or "",
        "vehicle_type": row.get("vehicle_type") or "none",
        "vehicle_number": row.get("vehicle_number") or "",
        "purpose": row.get("purpose") or "",
        "vendor_name": row.get("vendor_name") or "",
        "passcode": row.get("passcode") or "",
        "status": row.get("status") or "pending",
        "approval_by": row.get("approval_by"),
        "approved_at": row.get("approved_at"),
        "rejected_reason": row.get("rejected_reason"),
        "checked_in_at": row.get("checked_in_at"),
        "checked_out_at": row.get("checked_out_at"),
        "created_by": row.get("created_by"),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
    }


async def _visitor_access_query(
    *, current_user: dict[str, Any], tenant_id: str, app_key: str
) -> dict[str, Any]:
    if housing_router._can_manage_visitors(current_user.get("role")):
        return {}
    audience = await housing_router._current_user_message_audience(current_user=current_user, tenant_id=tenant_id, app_key=app_key)
    flats = sorted(value for value in audience["flats"] if value)
    user_terms = sorted(value for value in audience["members"] if value)
    clauses: list[dict[str, Any]] = []
    if flats:
        clauses.append({"flat_number": {"$in": flats}})
    if user_terms:
        clauses.append({"created_by": {"$in": user_terms}})
        clauses.append({"approval_by": {"$in": user_terms}})
    if not clauses:
        raise HTTPException(status_code=403, detail="Resident flat context is required for visitor access")
    return {"$or": clauses}


async def _get_visitor_or_404(
    *, visitor_id: str, current_user: dict[str, Any], tenant_id: str, app_key: str
) -> dict[str, Any]:
    query: dict[str, Any] = {"tenant_id": tenant_id, "app_key": app_key, "id": visitor_id}
    access_query = await housing_router._visitor_access_query(current_user=current_user, tenant_id=tenant_id, app_key=app_key)
    if access_query:
        query.update(access_query)
    row = await housing_router._visitors_collection().find_one(query)
    if not row:
        raise HTTPException(status_code=404, detail="Visitor entry not found")
    return row

import asyncio

def _send_web_push_sync(subscription_info: dict, data: str, private_key: str) -> None:
    try:
        pywebpush.webpush(
            subscription_info=subscription_info,
            data=data,
            vapid_private_key=private_key,
            vapid_claims={"sub": "mailto:support@sanmitra.net"}
        )
    except Exception:
        pass


async def _send_web_push_notification(
    tenant_id: str, app_key: str, flat_number: str, title: str, body: str, visitor_id: str
) -> None:
    app_settings = housing_router.get_settings()
    sub_col = housing_router.get_collection("housing_web_push_subscriptions")
    cursor = sub_col.find({"tenant_id": tenant_id, "app_key": app_key, "flat_number": flat_number})
    subscriptions = await cursor.to_list(length=100)
    if not subscriptions:
        return

    payload = {
        "title": title,
        "body": body,
        "data": {
            "visitor_id": visitor_id,
            "flat_number": flat_number,
            "type": "visitor_approval",
        },
    }
    data_str = json.dumps(payload)

    for sub_doc in subscriptions:
        if not app_settings.VAPID_PRIVATE_KEY:
            continue
        await asyncio.to_thread(
            _send_web_push_sync,
            sub_doc["subscription"],
            data_str,
            app_settings.VAPID_PRIVATE_KEY
        )

BILLING_JOBS = "housing_billing_jobs"


async def _run_billing_job(
    *,
    job_id: str,
    tenant_id: str,
    app_key: str,
    payload: dict[str, Any],
    created_by: str,
) -> None:
    """Background task: executes bill generation and updates the job record."""
    jobs_col = housing_router.get_collection(BILLING_JOBS)
    try:
        await jobs_col.update_one(
            {"id": job_id},
            {"$set": {"status": "running", "started_at": datetime.now(timezone.utc).isoformat()}},
        )
        session_factory = housing_router.get_session_factory()
        async with session_factory() as session:
            result = await housing_router._build_maintenance_bills(
                tenant_id=tenant_id,
                app_key=app_key,
                payload=payload,
                current_user={"sub": created_by},
                session=session,
                replace_existing=True,
            )
        await jobs_col.update_one(
            {"id": job_id},
            {"$set": {
                "status": "completed",
                "completed_at": datetime.now(timezone.utc).isoformat(),
                "total_bills": result.get("total_bills_generated", 0),
                "total_amount": result.get("total_amount", 0),
            }},
        )
    except HTTPException as exc:
        await jobs_col.update_one(
            {"id": job_id},
            {"$set": {
                "status": "failed",
                "failed_at": datetime.now(timezone.utc).isoformat(),
                "error": str(exc.detail),
            }},
        )
    except Exception as exc:
        await jobs_col.update_one(
            {"id": job_id},
            {"$set": {
                "status": "failed",
                "failed_at": datetime.now(timezone.utc).isoformat(),
                "error": str(exc),
            }},
        )

