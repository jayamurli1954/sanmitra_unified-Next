from __future__ import annotations

import calendar
import csv
import json
import re
import unicodedata
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import uuid4

from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, Query, UploadFile
from fastapi.responses import Response, StreamingResponse
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.accounting.models import Account, JournalEntry, JournalLine
from app.accounting.schemas import JournalLineIn, JournalPostRequest
from app.accounting.service import AccountingNotFoundError, AccountingValidationError, post_journal_entry
from app.core.auth.dependencies import get_current_user
from app.core.tenants.app_resolvers import resolve_gruha_tenant
from app.core.tenants.context import resolve_app_key, resolve_tenant_id
from app.db.mongo import get_collection
from app.db.postgres import get_async_session
from app.modules.housing_compat.schemas import (
    ApproveJoinRequest,
    ArrearsResponse,
    ArrearsTransferRequest,
    CompleteResidentRegistrationRequest,
    CompleteResidentRegistrationResponse,
    DamageClaimCreate,
    FinancialYearCloseRequest,
    FinancialYearCreateRequest,
    FinancialYearResponse,
    FinalBillResponse,
    FlatCreateRequest,
    FlatResponse,
    FlatTransferRequest,
    FlatUpdateRequest,
    MemberCreateRequest,
    MemberChecklistResponse,
    MemberChecklistUpdate,
    MemberResponse,
    MemberUpdateRequest,
    MembershipResponse,
    PublicJoinRequestCreate,
    PublicJoinRequestResponse,
    RejectJoinRequest,
    SocietySettingsResponse,
    SocietySettingsUpdate,
    SocietySearchItem,
)
from app.modules.housing_compat.service import (
    approve_join_request,
    calculate_final_bill,
    complete_resident_registration,
    create_member,
    get_member_checklist,
    create_public_join_request,
    create_flat,
    create_financial_year,
    generate_ndc,
    get_society,
    get_society_settings,
    get_flat,
    get_active_financial_year,
    list_flats,
    list_financial_years,
    list_join_requests,
    list_members,
    list_my_memberships,
    list_personal_arrears,
    list_society_units,
    raise_damage_claim,
    reject_join_request,
    provisional_close_financial_year,
    save_society_settings,
    search_societies,
    transfer_flat_to_flat,
    transfer_to_arrears,
    update_member_checklist,
    update_member,
    update_flat,
    final_close_financial_year,
)

router = APIRouter(tags=["housing-compat"])


def _safe_file_name(original: str, prefix: str = "doc") -> str:
    name = (original or "").strip()
    stem, dot, ext = name.rpartition(".")
    raw_base = stem if dot else name
    raw_ext = ext.lower() if dot else "bin"
    base = re.sub(r"[^a-zA-Z0-9_-]+", "-", raw_base).strip("-").lower() or prefix
    safe_ext = re.sub(r"[^a-z0-9]+", "", raw_ext) or "bin"
    return f"{prefix}_{base}_{uuid4().hex[:10]}.{safe_ext}"


def _meeting_collections() -> tuple[Any, Any]:
    return get_collection("housing_meetings"), get_collection("housing_meeting_resolutions")


def _message_collections() -> tuple[Any, Any]:
    return get_collection("housing_message_rooms"), get_collection("housing_messages")


def _complaints_collection() -> Any:
    return get_collection("housing_complaints")


def _asset_response(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": row.get("id"),
        "tenant_id": row.get("tenant_id"),
        "app_key": row.get("app_key"),
        "asset_code": row.get("asset_code") or row.get("account_code") or "",
        "name": row.get("name") or "",
        "category": row.get("category") or "other",
        "account_code": row.get("account_code") or "1500",
        "quantity": _safe_int(row.get("quantity"), default=1),
        "location": row.get("location") or "",
        "status": row.get("status") or "Active",
        "acquisition_type": row.get("acquisition_type") or "builder_handover",
        "handover_date": row.get("handover_date"),
        "purchase_date": row.get("purchase_date"),
        "original_cost": _safe_float(row.get("original_cost")),
        "depreciation_method": row.get("depreciation_method") or "straight_line",
        "depreciation_rate": _safe_float(row.get("depreciation_rate")),
        "useful_life_years": _safe_int(row.get("useful_life_years"), default=0),
        "residual_value": _safe_float(row.get("residual_value")),
        "amc_vendor": row.get("amc_vendor") or "",
        "amc_expiry": row.get("amc_expiry"),
        "insurance_policy_no": row.get("insurance_policy_no") or "",
        "insurance_expiry": row.get("insurance_expiry"),
        "vendor_name": row.get("vendor_name") or "",
        "invoice_no": row.get("invoice_no") or "",
        "notes": row.get("notes") or "",
        "is_scrapped": bool(row.get("is_scrapped") or False),
        "journal_entry_id": row.get("journal_entry_id"),
        "accounting_posting_status": row.get("accounting_posting_status") or "not_posted",
        "accounting_posted_at": row.get("accounting_posted_at"),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
    }


async def _next_asset_code(*, tenant_id: str, app_key: str, category: str) -> str:
    prefix = re.sub(r"[^A-Z0-9]+", "", str(category or "ASSET").upper())[:3] or "AST"
    count = await get_collection("housing_assets").count_documents({"tenant_id": tenant_id, "app_key": app_key})
    return f"{prefix}-{count + 1:04d}"


async def _find_account_by_code(
    session: AsyncSession,
    *,
    tenant_id: str,
    app_key: str,
    code: str,
    account_type: str | None = None,
    accounting_entity_id: str = "primary",
) -> Account | None:
    conditions = [
        Account.tenant_id == tenant_id,
        Account.app_key == app_key,
        Account.accounting_entity_id == accounting_entity_id,
        Account.code == str(code).strip(),
    ]
    if account_type:
        conditions.append(Account.type == account_type)
    result = await session.execute(select(Account).where(*conditions))
    return result.scalar_one_or_none()


async def _find_corpus_account(
    session: AsyncSession,
    *,
    tenant_id: str,
    app_key: str,
    accounting_entity_id: str = "primary",
) -> Account | None:
    scoped = [
        Account.tenant_id == tenant_id,
        Account.app_key == app_key,
        Account.accounting_entity_id == accounting_entity_id,
        Account.type == "equity",
    ]
    by_name = await session.execute(select(Account).where(*scoped, func.lower(Account.name) == "corpus fund"))
    account = by_name.scalar_one_or_none()
    if account is not None:
        return account

    for code in ("3000", "3010"):
        by_code = await session.execute(select(Account).where(*scoped, Account.code == code, Account.name.ilike("%corpus%")))
        account = by_code.scalar_one_or_none()
        if account is not None:
            return account
    return None


def _asset_accounting_date(row: dict[str, Any]) -> date:
    value = row.get("purchase_date") or row.get("handover_date")
    if value:
        try:
            return date.fromisoformat(str(value)[:10])
        except ValueError:
            pass
    return datetime.now(timezone.utc).date()


async def _post_asset_capitalization_journal(
    session: AsyncSession,
    *,
    tenant_id: str,
    app_key: str,
    asset: dict[str, Any],
    created_by: str,
) -> JournalEntry | None:
    amount = Decimal(str(_safe_float(asset.get("original_cost")))).quantize(Decimal("0.01"))
    if amount <= 0:
        return None
    if asset.get("acquisition_type") != "builder_handover":
        return None

    asset_account = await _find_account_by_code(
        session,
        tenant_id=tenant_id,
        app_key=app_key,
        code=str(asset.get("account_code") or ""),
        account_type="asset",
    )
    if asset_account is None and str(asset.get("account_code") or "") == "1500":
        asset_account = await _find_account_by_code(
            session,
            tenant_id=tenant_id,
            app_key=app_key,
            code="1320",
            account_type="asset",
        )
    if asset_account is None:
        raise HTTPException(
            status_code=422,
            detail=(
                f"Asset account {asset.get('account_code') or ''} was not found in this tenant's Chart of Accounts. "
                "Initialize the Chart of Accounts or choose an existing asset account."
            ),
        )

    corpus_account = await _find_corpus_account(session, tenant_id=tenant_id, app_key=app_key)
    if corpus_account is None:
        raise HTTPException(status_code=422, detail="Corpus Fund account was not found in this tenant's Chart of Accounts.")

    try:
        entry, _created = await post_journal_entry(
            session,
            tenant_id=tenant_id,
            app_key=app_key,
            accounting_entity_id="primary",
            created_by=created_by,
            payload=JournalPostRequest(
                entry_date=_asset_accounting_date(asset),
                description=f"Asset capitalized: {asset.get('name') or asset.get('asset_code')}",
                reference=str(asset.get("asset_code") or asset.get("id") or "")[:120],
                lines=[
                    JournalLineIn(account_id=int(asset_account.id), debit=amount, credit=Decimal("0.00")),
                    JournalLineIn(account_id=int(corpus_account.id), debit=Decimal("0.00"), credit=amount),
                ],
            ),
            idempotency_key=f"housing-asset:{asset.get('id')}:capitalization",
        )
    except (AccountingNotFoundError, AccountingValidationError) as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return entry


async def _get_meeting_or_404(*, tenant_id: str, app_key: str, meeting_id: str) -> dict[str, Any]:
    meetings, _ = _meeting_collections()
    row = await meetings.find_one({"tenant_id": tenant_id, "app_key": app_key, "id": meeting_id})
    if not row:
        raise HTTPException(status_code=404, detail="Meeting not found")
    return row


async def _get_or_create_message_room(
    *,
    tenant_id: str,
    app_key: str,
    name: str,
    room_type: str,
    description: str,
    audience_type: str = "public",
    allowed_flat_numbers: list[str] | None = None,
) -> dict[str, Any]:
    rooms_col, _ = _message_collections()
    room = await rooms_col.find_one({"tenant_id": tenant_id, "app_key": app_key, "type": room_type})
    if room:
        return room

    now = datetime.now(timezone.utc).isoformat()
    room = {
        "id": str(uuid4()),
        "tenant_id": tenant_id,
        "app_key": app_key,
        "name": name,
        "type": room_type,
        "description": description,
        "audience_type": audience_type,
        "allowed_flat_numbers": _normalize_flat_numbers(allowed_flat_numbers or []),
        "allowed_member_ids": [],
        "created_by": "system",
        "created_at": now,
        "updated_at": now,
        "last_message_at": None,
    }
    await rooms_col.insert_one(room)
    return room


def _build_meeting_notice_message(meeting: dict[str, Any], eligible_members: int) -> str:
    agenda_items = meeting.get("agenda_items") or []
    agenda_lines: list[str] = []
    if isinstance(agenda_items, list):
        for item in agenda_items:
            if not isinstance(item, dict):
                continue
            title = str(item.get("item_title") or "").strip()
            description = str(item.get("item_description") or "").strip()
            if title and description:
                agenda_lines.append(f"- {title}: {description}")
            elif title:
                agenda_lines.append(f"- {title}")

    lines = [
        f"Meeting Notice: {meeting.get('meeting_title') or 'Society Meeting'}",
        f"Type: {meeting.get('meeting_type') or 'Meeting'}",
        f"Date: {meeting.get('meeting_date') or 'To be confirmed'}",
        f"Time: {meeting.get('meeting_time') or 'To be confirmed'}",
        f"Venue: {meeting.get('venue') or 'To be confirmed'}",
        f"Eligible members: {eligible_members}",
    ]
    if agenda_lines:
        lines.extend(["", "Agenda:", *agenda_lines])
    return "\n".join(lines)


def _sanitize_mongo_doc(doc: dict[str, Any]) -> dict[str, Any]:
    return {k: v for k, v in doc.items() if k != "_id"}


def _normalize_status(value: Any) -> str:
    status = str(value or "SCHEDULED").strip().upper()
    return status if status in {"SCHEDULED", "COMPLETED", "CANCELLED"} else "SCHEDULED"


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _round_money(value: float) -> float:
    return round(float(value or 0.0) + 0.0000001, 2)


def _as_decimal_money(value: Any) -> Decimal:
    return Decimal(str(_round_money(_safe_float(value)))).quantize(Decimal("0.01"))


def _month_name(month: int) -> str:
    names = [
        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July",
        "August",
        "September",
        "October",
        "November",
        "December",
    ]
    return names[month - 1] if 1 <= month <= 12 else str(month)


def _split_charge(total: float, index: int, count: int, *, weight: float = 1.0, total_weight: float = 0.0) -> float:
    if total <= 0 or count <= 0:
        return 0.0
    if total_weight > 0:
        return _round_money(total * weight / total_weight)
    base = _round_money(total / count)
    if index == count - 1:
        return _round_money(total - (base * (count - 1)))
    return base


async def _count_eligible_members(*, tenant_id: str, app_key: str) -> int:
    rows = await get_collection("housing_members").find(
        {"tenant_id": tenant_id, "app_key": app_key},
        {"status": 1},
    ).to_list(length=5000)
    count = 0
    for row in rows:
        status = str(row.get("status") or "active").strip().lower()
        if status in {"moved_out", "inactive", "closed", "rejected"}:
            continue
        count += 1
    return count


def _normalize_member_ids(value: Any) -> list[str]:
    raw_values: list[Any]
    if isinstance(value, list):
        raw_values = value
    elif isinstance(value, str):
        raw_values = re.split(r"[,;\n]+", value)
    else:
        raw_values = []
    seen: set[str] = set()
    member_ids: list[str] = []
    for item in raw_values:
        member_id = str(item or "").strip()
        if member_id and member_id not in seen:
            seen.add(member_id)
            member_ids.append(member_id)
    return member_ids


async def _resolve_eligible_meeting_members(
    *, tenant_id: str, app_key: str, member_ids: list[str] | None = None, flat_numbers: list[str] | None = None
) -> list[dict[str, Any]]:
    member_ids = _normalize_member_ids(member_ids or [])
    flat_numbers = _normalize_flat_numbers(flat_numbers or [])
    clauses: list[dict[str, Any]] = []
    if member_ids:
        clauses.append({"id": {"$in": member_ids}})
    if flat_numbers:
        clauses.append({"flat_number": {"$in": flat_numbers}})
    if not clauses:
        return []

    query: dict[str, Any] = {"tenant_id": tenant_id, "app_key": app_key}
    query["$or"] = clauses
    rows = await get_collection("housing_members").find(
        query,
        {"id": 1, "flat_number": 1, "status": 1, "name": 1, "email": 1},
    ).to_list(length=5000)
    eligible: list[dict[str, Any]] = []
    for row in rows:
        status = str(row.get("status") or "active").strip().lower()
        if status in {"moved_out", "inactive", "closed", "rejected"}:
            continue
        eligible.append(row)
    return eligible


async def _meeting_eligible_count(*, tenant_id: str, app_key: str, meeting: dict[str, Any]) -> int:
    member_ids = _normalize_member_ids(meeting.get("eligible_member_ids") or [])
    flat_numbers = _normalize_flat_numbers(meeting.get("eligible_flat_numbers") or [])
    if member_ids or flat_numbers:
        members = await _resolve_eligible_meeting_members(
            tenant_id=tenant_id,
            app_key=app_key,
            member_ids=member_ids,
            flat_numbers=flat_numbers,
        )
        return len(members) or len(member_ids) or len(flat_numbers)
    return await _count_eligible_members(tenant_id=tenant_id, app_key=app_key)


def _meeting_stats(
    meeting_row: dict[str, Any], *, eligible_hint: int = 0
) -> dict[str, int | bool]:
    attendance = meeting_row.get("attendance") or []
    if not isinstance(attendance, list):
        attendance = []
    total_members_present = sum(
        1
        for item in attendance
        if str((item or {}).get("status") or "").strip().lower() in {"present", "proxy"}
    )
    total_members_eligible = _safe_int(meeting_row.get("total_members_eligible"), default=0)
    if total_members_eligible <= 0 and eligible_hint > 0:
        total_members_eligible = eligible_hint
    quorum_required = _safe_int(meeting_row.get("quorum_required"), default=0)
    quorum_met = bool(meeting_row.get("quorum_met") or False)
    if quorum_required > 0:
        quorum_met = total_members_present >= quorum_required
    return {
        "total_members_present": total_members_present,
        "total_members_eligible": total_members_eligible,
        "quorum_required": quorum_required,
        "quorum_met": quorum_met,
    }


def _build_simple_pdf(
    lines: list[str], *, title: str | None = None, top_margin: float = 60.0
) -> bytes:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - max(top_margin, 40.0)
    if title:
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(48, y, _sanitize_pdf_text(title))
        y -= 24
    pdf.setFont("Helvetica", 11)
    for raw in lines:
        line = _sanitize_pdf_text(raw)
        if not line:
            y -= 12
            continue
        if y < 56:
            pdf.showPage()
            pdf.setFont("Helvetica", 11)
            y = height - 56
        pdf.drawString(48, y, line[:140])
        y -= 16
    pdf.showPage()
    pdf.save()
    return buffer.getvalue()


async def _get_society_branding(*, tenant_id: str, app_key: str) -> dict[str, str]:
    settings = await get_collection("housing_society_settings").find_one(
        {"tenant_id": tenant_id, "app_key": app_key}
    ) or {}
    logo_url = str(settings.get("logo_url") or "").strip()
    logo_name = ""
    if logo_url:
        logo_name = logo_url.rstrip("/").split("/")[-1]
    if not logo_name:
        logo_doc = await get_collection("housing_documents").find_one(
            {"tenant_id": tenant_id, "app_key": app_key, "document_type": "society_logo"},
            sort=[("updated_at", -1)],
        )
        if logo_doc:
            logo_name = str(logo_doc.get("stored_name") or "").strip()
    logo_bytes = b""
    if logo_name:
        logo_doc = await get_collection("housing_documents").find_one(
            {"tenant_id": tenant_id, "app_key": app_key, "stored_name": logo_name},
            {"content": 1},
        )
        logo_bytes = bytes(logo_doc.get("content") or b"") if logo_doc else b""

    return {
        "society_name": str(settings.get("society_name") or "GruhaMitra Society").strip(),
        "society_address": str(settings.get("society_address") or "").strip(),
        "city": str(settings.get("city") or "").strip(),
        "state": str(settings.get("state") or "").strip(),
        "pin_code": str(settings.get("pin_code") or "").strip(),
        "contact_email": str(settings.get("contact_email") or "").strip(),
        "contact_phone": str(settings.get("contact_phone") or "").strip(),
        "logo_bytes": logo_bytes,
    }


def _maintenance_collections() -> tuple[Any, Any]:
    return get_collection("housing_maintenance_bills"), get_collection("housing_maintenance_reversals")


async def _list_maintenance_bills(*, tenant_id: str, app_key: str, month: int, year: int) -> list[dict[str, Any]]:
    bills_col, _ = _maintenance_collections()
    rows = (
        await bills_col.find({"tenant_id": tenant_id, "app_key": app_key, "month": month, "year": year})
        .sort([("flat_number", 1), ("created_at", 1)])
        .to_list(length=5000)
    )
    return [_sanitize_mongo_doc(row) for row in rows]


async def _flat_occupants_map(*, tenant_id: str, app_key: str) -> dict[str, int]:
    members = await get_collection("housing_members").find(
        {"tenant_id": tenant_id, "app_key": app_key},
        {"flat_number": 1, "status": 1, "is_primary": 1, "total_occupants": 1},
    ).to_list(length=5000)
    occupants: dict[str, int] = {}
    for member in members:
        status = str(member.get("status") or "active").strip().lower()
        if status in {"moved_out", "inactive", "closed", "rejected"}:
            continue
        flat_number = str(member.get("flat_number") or "").strip().upper()
        if not flat_number:
            continue
        count = max(0, _safe_int(member.get("total_occupants"), 1))
        if member.get("is_primary") or flat_number not in occupants:
            occupants[flat_number] = count
    return occupants


def _month_date_range(month: int, year: int) -> tuple[date, date]:
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_day)


def _expense_period_labels(month: int, year: int) -> set[str]:
    full = calendar.month_name[month]
    short = calendar.month_abbr[month]
    return {
        f"{full} {year}".lower(),
        f"{full}, {year}".lower(),
        f"{short} {year}".lower(),
        f"{short}, {year}".lower(),
        f"{short}-{year}".lower(),
        f"{month:02d}-{year}".lower(),
        f"{year}-{month:02d}".lower(),
        f"{month:02d}/{year}".lower(),
    }


def _matches_expense_period(txn: dict[str, Any], *, month: int, year: int) -> bool:
    labels = _expense_period_labels(month, year)
    explicit = " ".join(
        str(txn.get(key) or "")
        for key in ("expense_month", "expense_for_month", "period", "billing_month")
    ).lower()
    if explicit and any(label in explicit for label in labels):
        return True
    text_parts = [str(txn.get("description") or ""), str(txn.get("narration") or "")]
    for line in txn.get("lines") or []:
        text_parts.append(str((line or {}).get("description") or ""))
    text = " ".join(text_parts).lower()
    return any(label in text for label in labels)


def _is_water_expense_account(code: Any, name: Any) -> bool:
    text = f"{code or ''} {name or ''}".lower()
    return any(token in text for token in ("water", "tanker", "bwssb", "borewell"))


async def _expense_accounts_for_period(
    session: AsyncSession,
    *,
    tenant_id: str,
    app_key: str,
    month: int,
    year: int,
) -> list[dict[str, Any]]:
    from_date, to_date = _month_date_range(month, year)
    stmt = (
        select(
            Account.code.label("account_code"),
            Account.name.label("account_name"),
            func.coalesce(func.sum(JournalLine.debit), 0).label("debit_total"),
            func.coalesce(func.sum(JournalLine.credit), 0).label("credit_total"),
            func.count(JournalLine.id).label("transaction_count"),
        )
        .join(JournalLine, JournalLine.account_id == Account.id)
        .join(JournalEntry, JournalEntry.id == JournalLine.journal_id)
        .where(
            and_(
                Account.tenant_id == tenant_id,
                Account.app_key == app_key,
                Account.accounting_entity_id == "primary",
                Account.type == "expense",
                JournalEntry.tenant_id == tenant_id,
                JournalEntry.app_key == app_key,
                JournalEntry.accounting_entity_id == "primary",
                JournalEntry.entry_date >= from_date,
                JournalEntry.entry_date <= to_date,
            )
        )
        .group_by(Account.code, Account.name)
        .order_by(Account.code.asc().nullslast(), Account.name.asc())
    )
    rows = (await session.execute(stmt)).all()
    accounts_by_code: dict[str, dict[str, Any]] = {}
    for row in rows:
        amount = _round_money(float(row.debit_total or 0) - float(row.credit_total or 0))
        if amount <= 0:
            continue
        code = str(row.account_code or "")
        accounts_by_code[code] = {
            "account_code": row.account_code,
            "account_name": row.account_name,
            "total_amount": amount,
            "transaction_count": int(row.transaction_count or 0),
            "is_water": _is_water_expense_account(row.account_code, row.account_name),
        }

    txns = await get_collection("mb_transactions").find(
        {"tenant_id": tenant_id, "app_key": app_key, "voucher_type": "payment"}
    ).to_list(length=5000)
    period_txns = [txn for txn in txns if _matches_expense_period(txn, month=month, year=year)]
    mongo_codes = sorted(
        {
            str((line or {}).get("account_code") or "")
            for txn in period_txns
            for line in (txn.get("lines") or [])
            if _safe_float((line or {}).get("debit")) > 0 and str((line or {}).get("account_code") or "")
        }
    )
    names_by_code: dict[str, str] = {}
    if mongo_codes:
        account_rows = (
            await session.execute(
                select(Account.code, Account.name).where(
                    Account.tenant_id == tenant_id,
                    Account.app_key == app_key,
                    Account.accounting_entity_id == "primary",
                    Account.code.in_(mongo_codes),
                )
            )
        ).all()
        names_by_code = {str(code): str(name) for code, name in account_rows}

    for txn in period_txns:
        for line in txn.get("lines") or []:
            debit = _safe_float((line or {}).get("debit"))
            code = str((line or {}).get("account_code") or "")
            if debit <= 0 or not code:
                continue
            name = names_by_code.get(code) or str((line or {}).get("description") or code)
            existing = accounts_by_code.get(code)
            if existing:
                # If the SQL date query already counted the same voucher, do not double count.
                existing_refs = existing.setdefault("_mongo_refs", set())
                ref = str(txn.get("voucher_number") or txn.get("journal_entry_id") or txn.get("id"))
                if ref in existing_refs:
                    continue
                existing_refs.add(ref)
                existing["total_amount"] = _round_money(_safe_float(existing.get("total_amount")) + debit)
                existing["transaction_count"] = _safe_int(existing.get("transaction_count")) + 1
            else:
                accounts_by_code[code] = {
                    "account_code": code,
                    "account_name": name,
                    "total_amount": _round_money(debit),
                    "transaction_count": 1,
                    "is_water": _is_water_expense_account(code, name),
                    "_mongo_refs": {str(txn.get("voucher_number") or txn.get("journal_entry_id") or txn.get("id"))},
                }

    accounts = []
    for row in accounts_by_code.values():
        row.pop("_mongo_refs", None)
        accounts.append(row)
    return sorted(accounts, key=lambda item: (str(item.get("account_code") or ""), str(item.get("account_name") or "")))


async def _build_maintenance_bills(
    *,
    tenant_id: str,
    app_key: str,
    payload: dict[str, Any],
    current_user: dict[str, Any],
    session: AsyncSession,
    replace_existing: bool = True,
) -> dict[str, Any]:
    month = _safe_int(payload.get("month"))
    year = _safe_int(payload.get("year"))
    if month < 1 or month > 12 or year < 2000:
        raise HTTPException(status_code=422, detail="Valid month and year are required")

    flats = await list_flats(tenant_id=tenant_id, app_key=app_key)
    if not flats:
        raise HTTPException(status_code=400, detail="No flats found. Please add flats before generating bills.")

    bills_col, _ = _maintenance_collections()
    if replace_existing:
        existing_posted = await bills_col.count_documents(
            {"tenant_id": tenant_id, "app_key": app_key, "month": month, "year": year, "is_posted": True}
        )
        if existing_posted:
            raise HTTPException(
                status_code=409,
                detail=f"Posted bills already exist for {_month_name(month)} {year}. Reverse them before regenerating.",
            )
        await bills_col.delete_many({"tenant_id": tenant_id, "app_key": app_key, "month": month, "year": year})

    count = len(flats)
    total_area = sum(_safe_float(flat.get("area_sqft")) for flat in flats)
    flat_occupants = await _flat_occupants_map(tenant_id=tenant_id, app_key=app_key)
    adjusted_inmates = payload.get("adjusted_inmates") or {}
    inmate_counts = [
        max(
            0,
            _safe_int(
                adjusted_inmates.get(str(flat.get("id")))
                or adjusted_inmates.get(str(flat.get("flat_number")))
                or flat_occupants.get(str(flat.get("flat_number") or "").strip().upper())
                or 1,
                1,
            ),
        )
        for flat in flats
    ]
    total_inmates = sum(inmate_counts) or count

    sqft_rate = _safe_float(payload.get("override_sqft_rate"))
    expense_accounts = await _expense_accounts_for_period(
        session,
        tenant_id=tenant_id,
        app_key=app_key,
        month=month,
        year=year,
    )
    expense_by_code = {str(row.get("account_code")): row for row in expense_accounts}
    selected_fixed_codes = [str(code) for code in (payload.get("selected_fixed_expense_codes") or [])]
    ledger_water_total = _round_money(
        sum(_safe_float(row.get("total_amount")) for row in expense_accounts if row.get("is_water"))
    )
    selected_fixed_total = _round_money(
        sum(
            _safe_float(expense_by_code[code].get("total_amount"))
            for code in selected_fixed_codes
            if code in expense_by_code and not expense_by_code[code].get("is_water")
        )
    )

    total_water = (
        _safe_float(payload.get("override_water_charges"))
        if payload.get("override_water_charges") not in (None, "")
        else ledger_water_total
    )
    water_rate_per_person = _round_money(total_water / total_inmates) if total_water and total_inmates else 0.0
    total_fixed = (
        _safe_float(payload.get("override_fixed_expenses"))
        if payload.get("override_fixed_expenses") not in (None, "")
        else selected_fixed_total
    )
    settings = await get_society_settings(tenant_id=tenant_id, app_key=app_key)
    sinking_rate = _safe_float(settings.get("sinking_fund_rate"))
    repair_rate = _safe_float(settings.get("repair_fund_rate"))
    association_rate = _safe_float(settings.get("association_fund_rate"))
    corpus_rate = _safe_float(settings.get("corpus_fund_rate"))
    total_sinking = (
        _safe_float(payload.get("override_sinking_fund"))
        if payload.get("override_sinking_fund") not in (None, "")
        else _round_money(sinking_rate * count)
    )
    total_repair = (
        _safe_float(payload.get("override_repair_fund"))
        if payload.get("override_repair_fund") not in (None, "")
        else _round_money(repair_rate * count)
    )
    total_association = _round_money(association_rate * count)
    total_corpus = (
        _safe_float(payload.get("override_corpus_fund"))
        if payload.get("override_corpus_fund") not in (None, "")
        else _round_money(corpus_rate * count)
    )
    fixed_method = str(payload.get("fixed_calculation_method") or "equal").lower()
    sinking_method = str(payload.get("sinking_calculation_method") or "equal").lower()
    repair_method = str(payload.get("repair_fund_calculation_method") or "equal").lower()
    corpus_method = str(payload.get("corpus_fund_calculation_method") or "equal").lower()

    now = datetime.now(timezone.utc).isoformat()
    docs: list[dict[str, Any]] = []
    for idx, flat in enumerate(flats):
        flat_area = _safe_float(flat.get("area_sqft"))
        inmates = inmate_counts[idx]
        maintenance = _round_money(flat_area * sqft_rate) if sqft_rate > 0 else 0.0
        water = _round_money(water_rate_per_person * inmates)

        def fund_share(total: float, method: str) -> float:
            use_area = method in {"sqft", "area", "area_sqft"}
            return _split_charge(total, idx, count, weight=flat_area, total_weight=total_area if use_area else 0.0)

        fixed = fund_share(total_fixed, fixed_method)
        sinking = fund_share(total_sinking, sinking_method)
        repair = fund_share(total_repair, repair_method)
        association = _split_charge(total_association, idx, count)
        corpus = fund_share(total_corpus, corpus_method)
        total_amount = _round_money(maintenance + water + fixed + sinking + repair + association + corpus)
        bill_id = str(uuid4())
        doc = {
            "id": bill_id,
            "tenant_id": tenant_id,
            "app_key": app_key,
            "flat_id": flat.get("id"),
            "flat_number": flat.get("flat_number"),
            "month": month,
            "year": year,
            "amount": total_amount,
            "maintenance_amount": maintenance,
            "water_amount": water,
            "fixed_amount": fixed,
            "sinking_fund_amount": sinking,
            "repair_fund_amount": repair,
            "association_fund_amount": association,
            "corpus_fund_amount": corpus,
            "arrears_amount": 0.0,
            "late_fee_amount": 0.0,
            "status": "generated",
            "is_posted": False,
            "posting_status": "not_posted",
            "auto_post_requested": bool(payload.get("auto_post_to_accounting")),
            "breakdown": {
                "maintenance_sqft": maintenance,
                "sqft_calculation": f"{flat_area:g} sq.ft x {sqft_rate:g}" if maintenance else "",
                "water_charges": water,
                "water_per_person_rate": water_rate_per_person,
                "inmates_used": inmates,
                "total_inmates": total_inmates,
                "water_calculation": (
                    f"Water expenses {total_water:,.2f} / {total_inmates} residents = "
                    f"{water_rate_per_person:,.2f} x {inmates} resident(s)"
                    if water and total_inmates
                    else ""
                ),
                "water_source_total": ledger_water_total,
                "fixed_expenses": fixed,
                "fixed_expenses_calculation": f"{fixed_method.title()} share of selected expenses {total_fixed:,.2f}" if fixed else "",
                "selected_fixed_expense_codes": selected_fixed_codes,
                "selected_fixed_expenses_total": selected_fixed_total,
                "sinking_fund": sinking,
                "sinking_fund_calculation": f"Per flat setting {sinking_rate:,.2f}" if sinking_rate else f"{sinking_method.title()} share" if sinking else "",
                "repair_fund": repair,
                "repair_fund_calculation": f"Per flat setting {repair_rate:,.2f}" if repair_rate else f"{repair_method.title()} share" if repair else "",
                "association_fund": association,
                "association_fund_calculation": f"Per flat setting {association_rate:,.2f}" if association_rate else "",
                "corpus_fund": corpus,
                "corpus_fund_calculation": f"Per flat setting {corpus_rate:,.2f}" if corpus_rate else f"{corpus_method.title()} share" if corpus else "",
            },
            "created_by": str(current_user.get("sub") or "system"),
            "created_at": now,
            "updated_at": now,
        }
        docs.append(doc)

    if docs:
        await bills_col.insert_many(docs)

    return {
        "month": month,
        "year": year,
        "total_bills_generated": len(docs),
        "total_amount": _round_money(sum(_safe_float(doc.get("amount")) for doc in docs)),
        "ledger_water_total": ledger_water_total,
        "selected_fixed_expenses_total": selected_fixed_total,
        "bills": [_sanitize_mongo_doc(doc) for doc in docs],
    }


async def _account_ids_by_code(
    session: AsyncSession,
    *,
    tenant_id: str,
    app_key: str,
    codes: set[str],
) -> dict[str, int]:
    rows = (
        await session.execute(
            select(Account.code, Account.id).where(
                Account.tenant_id == tenant_id,
                Account.app_key == app_key,
                Account.accounting_entity_id == "primary",
                Account.code.in_(codes),
            )
        )
    ).all()
    found = {str(code): int(account_id) for code, account_id in rows}
    missing = sorted(codes - set(found))
    if missing:
        raise HTTPException(status_code=400, detail=f"Required accounting accounts missing: {', '.join(missing)}")
    return found


def _bill_credit_components(bill: dict[str, Any]) -> list[tuple[str, Decimal]]:
    maintenance = _as_decimal_money(bill.get("maintenance_amount")) + _as_decimal_money(bill.get("fixed_amount"))
    extra_charges = _as_decimal_money(bill.get("extra_charges_amount"))
    components = [
        ("41001", maintenance + extra_charges),
        ("41004", _as_decimal_money(bill.get("water_amount"))),
        ("31011", _as_decimal_money(bill.get("sinking_fund_amount"))),
        ("31012", _as_decimal_money(bill.get("repair_fund_amount"))),
        ("31010", _as_decimal_money(bill.get("association_fund_amount"))),
        ("31004", _as_decimal_money(bill.get("corpus_fund_amount"))),
    ]
    return [(code, amount) for code, amount in components if amount > 0]


async def _post_maintenance_bill_to_accounting(
    session: AsyncSession,
    *,
    tenant_id: str,
    app_key: str,
    bill: dict[str, Any],
    current_user: dict[str, Any],
) -> int:
    month = _safe_int(bill.get("month"))
    year = _safe_int(bill.get("year"))
    _from_date, entry_date = _month_date_range(month, year)
    credit_components = _bill_credit_components(bill)
    if not credit_components:
        raise HTTPException(status_code=400, detail=f"Bill {bill.get('flat_number')} has no billable components")

    total = _as_decimal_money(bill.get("amount"))
    credit_total = sum((amount for _code, amount in credit_components), Decimal("0.00"))
    if total != credit_total:
        total = credit_total

    required_codes = {"13001", *(code for code, _amount in credit_components)}
    account_ids = await _account_ids_by_code(session, tenant_id=tenant_id, app_key=app_key, codes=required_codes)
    lines = [JournalLineIn(account_id=account_ids["13001"], debit=total, credit=Decimal("0.00"))]
    lines.extend(
        JournalLineIn(account_id=account_ids[code], debit=Decimal("0.00"), credit=amount)
        for code, amount in credit_components
    )
    reference = f"MBILL-{year}-{month:02d}-{bill.get('flat_number')}"
    try:
        entry, _created = await post_journal_entry(
            session,
            app_key=app_key,
            tenant_id=tenant_id,
            accounting_entity_id="primary",
            created_by=str(current_user.get("sub") or "system"),
            payload=JournalPostRequest(
                entry_date=entry_date,
                description=f"Monthly maintenance bill for {bill.get('flat_number')} - {_month_name(month)} {year}",
                reference=reference,
                lines=lines,
            ),
            idempotency_key=f"gruhamitra:{tenant_id}:{app_key}:maintenance-bill:{bill.get('id')}",
        )
    except (AccountingValidationError, AccountingNotFoundError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return int(entry.id)


async def _reverse_maintenance_bill_accounting(
    session: AsyncSession,
    *,
    tenant_id: str,
    app_key: str,
    bill: dict[str, Any],
    reason: str,
    current_user: dict[str, Any],
) -> int | None:
    original_journal_id = bill.get("journal_entry_id")
    if not original_journal_id:
        return None

    rows = (
        await session.execute(
            select(JournalLine.account_id, JournalLine.debit, JournalLine.credit, JournalEntry.entry_date)
            .join(JournalEntry, JournalEntry.id == JournalLine.journal_id)
            .where(
                JournalEntry.tenant_id == tenant_id,
                JournalEntry.app_key == app_key,
                JournalEntry.accounting_entity_id == "primary",
                JournalEntry.id == int(original_journal_id),
            )
            .order_by(JournalLine.id.asc())
        )
    ).all()
    if not rows:
        raise HTTPException(status_code=400, detail="Original bill journal entry not found for reversal")

    lines = [
        JournalLineIn(
            account_id=int(row.account_id),
            debit=Decimal(row.credit or 0).quantize(Decimal("0.01")),
            credit=Decimal(row.debit or 0).quantize(Decimal("0.01")),
        )
        for row in rows
    ]
    month = _safe_int(bill.get("month"))
    year = _safe_int(bill.get("year"))
    _from_date, entry_date = _month_date_range(month, year)
    try:
        entry, _created = await post_journal_entry(
            session,
            app_key=app_key,
            tenant_id=tenant_id,
            accounting_entity_id="primary",
            created_by=str(current_user.get("sub") or "system"),
            payload=JournalPostRequest(
                entry_date=entry_date,
                description=f"Reversal of maintenance bill for {bill.get('flat_number')} - {_month_name(month)} {year}: {reason}",
                reference=f"REV-MBILL-{year}-{month:02d}-{bill.get('flat_number')}",
                lines=lines,
            ),
            idempotency_key=f"gruhamitra:{tenant_id}:{app_key}:maintenance-bill-reversal:{bill.get('id')}",
        )
    except (AccountingValidationError, AccountingNotFoundError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return int(entry.id)


def _draw_pdf_header(pdf: canvas.Canvas, *, title: str, branding: dict[str, Any]) -> float:
    page_w, page_h = A4
    margin_x = 42
    y_top = page_h - 38
    logo_bytes = branding.get("logo_bytes") or b""
    if logo_bytes:
        try:
            img = ImageReader(BytesIO(logo_bytes))
            pdf.drawImage(img, margin_x, y_top - 48, width=42, height=42, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass
    text_x = margin_x + 52
    society_name = str(branding.get("society_name") or "GruhaMitra Society").strip()
    address_parts = [
        str(branding.get("society_address") or "").strip(),
        ", ".join(
            [p for p in [branding.get("city"), branding.get("state"), branding.get("pin_code")] if str(p or "").strip()]
        ),
    ]
    contact_parts = [p for p in [branding.get("contact_phone"), branding.get("contact_email")] if str(p or "").strip()]

    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(text_x, y_top, society_name[:75])
    pdf.setFont("Helvetica", 9)
    y = y_top - 14
    for part in address_parts:
        part_text = str(part or "").strip()
        if part_text:
            pdf.drawString(text_x, y, part_text[:110])
            y -= 11
    if contact_parts:
        pdf.drawString(text_x, y, " | ".join(contact_parts)[:110])
        y -= 11

    pdf.setLineWidth(0.8)
    pdf.line(margin_x, y - 2, page_w - margin_x, y - 2)
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(margin_x, y - 20, title[:95])
    return y - 42


def _build_branded_pdf(*, title: str, lines: list[str], branding: dict[str, Any]) -> bytes:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    page_w, page_h = A4
    margin_x = 42
    y = _draw_pdf_header(pdf, title=title, branding=branding)
    pdf.setFont("Helvetica", 10.5)
    for raw in lines:
        line = _sanitize_pdf_text(raw)
        if not line:
            y -= 10
            continue
        if y < 56:
            pdf.showPage()
            y = _draw_pdf_header(pdf, title=title, branding=branding)
            pdf.setFont("Helvetica", 10.5)
        pdf.drawString(margin_x, y, line[:140])
        y -= 14
    pdf.setFont("Helvetica-Oblique", 8.5)
    pdf.drawString(
        margin_x,
        34,
        f"Generated by GruhaMitra on {datetime.now(timezone.utc).strftime('%d-%m-%Y %H:%M UTC')}",
    )
    pdf.save()
    return buffer.getvalue()


def _sanitize_pdf_text(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    # Drop common mojibake/replacement/control characters that render as black boxes.
    text = (
        text.replace("\ufffd", " ")
        .replace("\u25a0", " ")
        .replace("\u25aa", " ")
        .replace("\u200b", "")
        .replace("\u200c", "")
        .replace("\u200d", "")
        .replace("\ufeff", "")
    )
    text = re.sub(r"[\x00-\x1F\x7F]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _normalize_header(value: Any) -> str:
    key = str(value or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "_", key).strip("_")


def _truthy(value: Any, *, default: bool = False) -> bool:
    if value is None:
        return default
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y"}:
        return True
    if text in {"0", "false", "no", "n"}:
        return False
    return default


def _parse_member_move_in(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    try:
        parsed = datetime.fromisoformat(text)
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
    except ValueError:
        return None


def _row_get(row: dict[str, Any], keys: list[str], default: Any = None) -> Any:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return default


def _coerce_bulk_member_payload(row: dict[str, Any], row_no: int) -> MemberCreateRequest:
    name = str(_row_get(row, ["name", "full_name"]) or "").strip()
    phone = str(_row_get(row, ["phone_number", "phone", "mobile", "mobile_number"]) or "").strip()
    email_raw = _row_get(row, ["email", "email_address"], None)
    email = str(email_raw).strip().lower() if email_raw not in (None, "") else None
    flat = str(_row_get(row, ["flat_number", "flat", "flat_no", "unit", "unit_number"]) or "").strip().upper()
    member_type = str(_row_get(row, ["member_type", "type"], "owner") or "owner").strip().lower()
    move_in_date = _parse_member_move_in(_row_get(row, ["move_in_date", "move_in", "move_in_dt"], None))
    occupants_raw = _row_get(row, ["total_occupants", "occupants", "family_count"], 1)
    try:
        total_occupants = int(str(occupants_raw).strip())
    except Exception:
        total_occupants = 1
    occupation_raw = _row_get(row, ["occupation", "profession"], None)
    occupation = str(occupation_raw).strip() if occupation_raw not in (None, "") else None
    is_primary = _truthy(_row_get(row, ["is_primary", "primary"], True), default=True)
    is_mobile_public = _truthy(_row_get(row, ["is_mobile_public", "mobile_public"], False), default=False)

    try:
        return MemberCreateRequest(
            name=name,
            phone_number=phone,
            email=email,
            flat_number=flat,
            member_type=member_type,  # type: ignore[arg-type]
            move_in_date=move_in_date,
            total_occupants=total_occupants,
            is_primary=is_primary,
            occupation=occupation,
            is_mobile_public=is_mobile_public,
        )
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Row {row_no}: {exc}") from exc


def _parse_csv_rows(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(text.splitlines())
    rows: list[dict[str, Any]] = []
    for raw in reader:
        normalized: dict[str, Any] = {}
        for key, value in (raw or {}).items():
            normalized[_normalize_header(key)] = value
        if not any(str(v or "").strip() for v in normalized.values()):
            continue
        rows.append(normalized)
    return rows


def _parse_xlsx_rows(content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active
    header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_values:
        return []
    headers = [_normalize_header(h) for h in header_values]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        item: dict[str, Any] = {}
        for idx, value in enumerate(values):
            if idx >= len(headers):
                continue
            key = headers[idx]
            if key:
                item[key] = value
        if not any(str(v or "").strip() for v in item.values()):
            continue
        rows.append(item)
    return rows


def _can_manage_chat_rooms(role: Any) -> bool:
    normalized = str(role or "").strip().lower()
    return normalized in {
        "admin",
        "tenant_admin",
        "super_admin",
        "secretary",
        "society_admin",
        "societyadmin",
        "chairman",
        "chairperson",
        "president",
        "treasurer",
        "committee",
        "committee_member",
        "management_committee",
        "mc",
        "owner",
        "manager",
    }


def _normalize_flat_number(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def _normalize_flat_numbers(value: Any) -> list[str]:
    raw_values: list[Any]
    if isinstance(value, list):
        raw_values = value
    elif isinstance(value, str):
        raw_values = re.split(r"[,;\n]+", value)
    else:
        raw_values = []
    seen: set[str] = set()
    flats: list[str] = []
    for item in raw_values:
        flat = _normalize_flat_number(item)
        if flat and flat not in seen:
            seen.add(flat)
            flats.append(flat)
    return flats


def _room_response(row: dict[str, Any]) -> dict[str, Any]:
    audience_type = str(row.get("audience_type") or "public").strip().lower()
    allowed_flat_numbers = _normalize_flat_numbers(row.get("allowed_flat_numbers") or [])
    allowed_member_ids = [str(v).strip() for v in (row.get("allowed_member_ids") or []) if str(v or "").strip()]
    return {
        "id": row.get("id"),
        "name": row.get("name") or "General",
        "type": row.get("type") or "general",
        "description": row.get("description") or "",
        "audience_type": audience_type,
        "allowed_flat_numbers": allowed_flat_numbers,
        "allowed_member_ids": allowed_member_ids,
        "created_by": row.get("created_by"),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
        "last_message_at": row.get("last_message_at"),
    }


def _message_response(row: dict[str, Any]) -> dict[str, Any]:
    attachments = row.get("attachments") or []
    if not isinstance(attachments, list):
        attachments = []
    return {
        "id": row.get("id"),
        "room_id": row.get("room_id"),
        "sender_id": row.get("sender_id"),
        "sender_name": row.get("sender_name") or "User",
        "content": row.get("content") or "",
        "message_type": row.get("message_type") or "text",
        "attachments": [
            {
                "id": item.get("id"),
                "file_name": item.get("file_name") or "attachment",
                "content_type": item.get("content_type") or "application/octet-stream",
                "size": item.get("size") or 0,
                "download_url": item.get("download_url"),
            }
            for item in attachments
            if isinstance(item, dict)
        ],
        "expires_at": row.get("expires_at"),
        "retention_days": row.get("retention_days"),
        "created_at": row.get("created_at"),
    }


def _message_retention_days(value: Any) -> int:
    days = _safe_int(value, default=30)
    if days <= 0:
        days = 30
    return min(days, 90)


async def _current_user_message_audience(
    *, current_user: dict[str, Any], tenant_id: str, app_key: str
) -> dict[str, set[str]]:
    flat_candidates = {
        _normalize_flat_number(current_user.get(key))
        for key in ("flat_number", "flat_no", "unit_number", "unit_no")
    }
    member_candidates = {
        str(current_user.get(key) or "").strip()
        for key in ("sub", "id", "user_id", "email")
        if str(current_user.get(key) or "").strip()
    }
    flat_candidates.discard("")

    member_query_terms = [value for value in member_candidates if value]
    if member_query_terms:
        members = await get_collection("housing_members").find(
            {
                "tenant_id": tenant_id,
                "app_key": app_key,
                "$or": [
                    {"user_id": {"$in": member_query_terms}},
                    {"id": {"$in": member_query_terms}},
                    {"email": {"$in": member_query_terms}},
                    {"member_id": {"$in": member_query_terms}},
                ],
            },
            {"flat_number": 1, "id": 1, "user_id": 1, "email": 1, "member_id": 1},
        ).to_list(length=50)
        for member in members:
            flat = _normalize_flat_number(member.get("flat_number"))
            if flat:
                flat_candidates.add(flat)
            for key in ("id", "user_id", "email", "member_id"):
                value = str(member.get(key) or "").strip()
                if value:
                    member_candidates.add(value)

    return {"flats": flat_candidates, "members": member_candidates}


async def _can_access_message_room(
    *, room: dict[str, Any], current_user: dict[str, Any], tenant_id: str, app_key: str
) -> bool:
    if _can_manage_chat_rooms(current_user.get("role")):
        return True
    audience_type = str(room.get("audience_type") or "public").strip().lower()
    if audience_type in {"", "public", "all"}:
        return True

    audience = await _current_user_message_audience(current_user=current_user, tenant_id=tenant_id, app_key=app_key)
    allowed_flats = set(_normalize_flat_numbers(room.get("allowed_flat_numbers") or []))
    if allowed_flats and audience["flats"].intersection(allowed_flats):
        return True

    allowed_members = {str(v).strip() for v in (room.get("allowed_member_ids") or []) if str(v or "").strip()}
    if allowed_members and audience["members"].intersection(allowed_members):
        return True
    return False


def _can_manage_complaints(role: Any) -> bool:
    normalized = str(role or "").strip().lower()
    return normalized in {"admin", "super_admin", "secretary"}


@router.get("/complaints/")
async def complaints_list(
    status: str | None = Query(default=None),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_context = resolve_gruha_tenant(
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
        operation="maintenance bill regeneration",
    )
    tenant_id = tenant_context.tenant_id
    app_key = tenant_context.app_key
    query: dict[str, Any] = {"tenant_id": tenant_id, "app_key": app_key}
    if status:
        query["status"] = str(status).strip().lower()

    role = str(current_user.get("role") or "").strip().lower()
    user_sub = str(current_user.get("sub") or "").strip()
    if not _can_manage_complaints(role):
        query["$or"] = [{"scope": "common_area"}, {"created_by": user_sub}]

    col = _complaints_collection()
    rows = await col.find(query).sort("created_at", -1).to_list(length=1000)
    return [
        {
            "id": row.get("id"),
            "title": row.get("title") or "",
            "description": row.get("description") or "",
            "type": row.get("type") or "other",
            "priority": row.get("priority") or "medium",
            "scope": row.get("scope") or "individual",
            "status": row.get("status") or "open",
            "user_name": row.get("user_name") or "Resident",
            "flat_number": row.get("flat_number") or "N/A",
            "created_by": row.get("created_by"),
            "created_at": row.get("created_at"),
            "updated_at": row.get("updated_at"),
            "resolved_at": row.get("resolved_at"),
        }
        for row in rows
    ]


@router.post("/complaints/")
async def complaints_create(
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())

    title = str(payload.get("title") or "").strip()
    description = str(payload.get("description") or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="title is required")
    if not description:
        raise HTTPException(status_code=400, detail="description is required")

    complaint_type = str(payload.get("type") or "other").strip().lower() or "other"
    priority = str(payload.get("priority") or "medium").strip().lower() or "medium"
    scope = str(payload.get("scope") or "individual").strip().lower() or "individual"
    if priority not in {"low", "medium", "high"}:
        priority = "medium"
    if scope not in {"individual", "common_area"}:
        scope = "individual"

    now = datetime.now(timezone.utc).isoformat()
    creator_email = str(current_user.get("email") or "").strip()
    creator_sub = str(current_user.get("sub") or "system")
    user_name = (
        str(current_user.get("full_name") or "").strip()
        or creator_email
        or "Resident"
    )
    flat_number = str(current_user.get("flat_number") or "").strip() or "N/A"
    doc = {
        "id": str(uuid4()),
        "tenant_id": tenant_id,
        "app_key": app_key,
        "title": title,
        "description": description,
        "type": complaint_type,
        "priority": priority,
        "scope": scope,
        "status": "open",
        "user_name": user_name,
        "flat_number": flat_number,
        "created_by": creator_sub,
        "created_at": now,
        "updated_at": now,
        "resolved_at": None,
    }
    await _complaints_collection().insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@router.patch("/complaints/{complaint_id}")
async def complaints_update(
    complaint_id: str,
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    col = _complaints_collection()
    row = await col.find_one({"tenant_id": tenant_id, "app_key": app_key, "id": complaint_id})
    if not row:
        raise HTTPException(status_code=404, detail="Complaint not found")

    role = str(current_user.get("role") or "").strip().lower()
    user_sub = str(current_user.get("sub") or "").strip()
    allowed_status = {"open", "in_progress", "resolved", "closed"}
    requested_status = str(payload.get("status") or "").strip().lower()
    if requested_status and requested_status not in allowed_status:
        raise HTTPException(status_code=400, detail="Invalid status")

    can_manage = _can_manage_complaints(role)
    if not can_manage:
        if row.get("created_by") != user_sub:
            raise HTTPException(status_code=403, detail="Only creator or admin can update this complaint")
        if requested_status and requested_status != "closed":
            raise HTTPException(status_code=403, detail="Residents can only close their own complaint")

    update_doc: dict[str, Any] = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if requested_status:
        update_doc["status"] = requested_status
        if requested_status in {"resolved", "closed"}:
            update_doc["resolved_at"] = update_doc["updated_at"]

    if len(update_doc) == 1:
        raise HTTPException(status_code=400, detail="No valid fields to update")

    await col.update_one({"tenant_id": tenant_id, "app_key": app_key, "id": complaint_id}, {"$set": update_doc})
    updated = await col.find_one({"tenant_id": tenant_id, "app_key": app_key, "id": complaint_id})
    if not updated:
        raise HTTPException(status_code=404, detail="Complaint not found")
    return {k: v for k, v in updated.items() if k != "_id"}


@router.get("/maintenance/bills")
async def maintenance_list_bills(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2000),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    return await _list_maintenance_bills(tenant_id=tenant_id, app_key=app_key, month=month, year=year)


@router.get("/maintenance/expense-accounts-for-period")
async def maintenance_expense_accounts_for_period(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2000),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    return await _expense_accounts_for_period(session, tenant_id=tenant_id, app_key=app_key, month=month, year=year)


@router.post("/maintenance/generate-bills")
async def maintenance_generate_bills(
    payload: dict[str, Any],
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_context = resolve_gruha_tenant(
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
        operation="maintenance bill generation",
    )
    tenant_id = tenant_context.tenant_id
    app_key = tenant_context.app_key
    return await _build_maintenance_bills(
        tenant_id=tenant_id,
        app_key=app_key,
        payload=payload,
        current_user=current_user,
        session=session,
        replace_existing=True,
    )


@router.post("/maintenance/post-bills")
async def maintenance_post_bills(
    payload: dict[str, Any],
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_context = resolve_gruha_tenant(
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
        operation="maintenance bill posting",
    )
    tenant_id = tenant_context.tenant_id
    app_key = tenant_context.app_key
    month = _safe_int(payload.get("month"))
    year = _safe_int(payload.get("year"))
    if month < 1 or month > 12 or year < 2000:
        raise HTTPException(status_code=422, detail="Valid month and year are required")
    bills_col, _ = _maintenance_collections()
    bills = await bills_col.find(
        {
            "tenant_id": tenant_id,
            "app_key": app_key,
            "month": month,
            "year": year,
            "status": {"$ne": "reversed"},
            "is_posted": {"$ne": True},
        }
    ).to_list(length=5000)
    if not bills:
        return {"month": month, "year": year, "total_bills_generated": 0, "posted_journal_entries": []}

    now = datetime.now(timezone.utc).isoformat()
    posted_entries: list[int] = []
    for bill in bills:
        journal_entry_id = await _post_maintenance_bill_to_accounting(
            session,
            tenant_id=tenant_id,
            app_key=app_key,
            bill=bill,
            current_user=current_user,
        )
        posted_entries.append(journal_entry_id)
        await bills_col.update_one(
            {"tenant_id": tenant_id, "app_key": app_key, "id": bill.get("id")},
            {
                "$set": {
                    "is_posted": True,
                    "posting_status": "posted",
                    "status": "posted",
                    "journal_entry_id": journal_entry_id,
                    "posted_at": now,
                    "updated_at": now,
                }
            },
        )
    return {
        "month": month,
        "year": year,
        "total_bills_generated": len(posted_entries),
        "posted_journal_entries": posted_entries,
    }


@router.post("/maintenance/add-extra-charge")
async def maintenance_add_extra_charge(
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_context = resolve_gruha_tenant(
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
        operation="maintenance extra charge",
    )
    tenant_id = tenant_context.tenant_id
    app_key = tenant_context.app_key
    bill_id = str(payload.get("bill_id") or "").strip()
    amount = _round_money(_safe_float(payload.get("amount")))
    description = str(payload.get("description") or "").strip()
    if not bill_id:
        raise HTTPException(status_code=422, detail="bill_id is required")
    if amount <= 0:
        raise HTTPException(status_code=422, detail="Amount must be greater than zero")
    if not description:
        raise HTTPException(status_code=422, detail="Description is required")

    bills_col, _ = _maintenance_collections()
    bill = await bills_col.find_one({"tenant_id": tenant_id, "app_key": app_key, "id": bill_id})
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    if bill.get("is_posted") or str(bill.get("status") or "").lower() == "posted":
        raise HTTPException(status_code=409, detail="Posted bills cannot be edited. Reverse and regenerate before adding charges.")
    if str(bill.get("status") or "").lower() == "reversed":
        raise HTTPException(status_code=409, detail="Cannot add charges to a reversed bill")

    now = datetime.now(timezone.utc).isoformat()
    charge = {
        "id": str(uuid4()),
        "name": str(payload.get("name") or "Damage / Extra Charge").strip(),
        "amount": amount,
        "description": description,
        "calculation": description,
        "created_by": str(current_user.get("sub") or "system"),
        "created_at": now,
    }
    existing_extra = _safe_float(bill.get("extra_charges_amount"))
    existing_amount = _safe_float(bill.get("amount"))
    await bills_col.update_one(
        {"tenant_id": tenant_id, "app_key": app_key, "id": bill_id},
        {
            "$inc": {"amount": amount, "extra_charges_amount": amount},
            "$push": {"breakdown.supplementary_charges": charge},
            "$set": {"updated_at": now},
        },
    )
    updated = await bills_col.find_one({"tenant_id": tenant_id, "app_key": app_key, "id": bill_id})
    if updated:
        updated["amount"] = _round_money(existing_amount + amount)
        updated["extra_charges_amount"] = _round_money(existing_extra + amount)
    return _sanitize_mongo_doc(updated or bill)


@router.post("/maintenance/reverse-bill")
async def maintenance_reverse_bill(
    payload: dict[str, Any],
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_context = resolve_gruha_tenant(
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
        operation="maintenance bill reversal",
    )
    tenant_id = tenant_context.tenant_id
    app_key = tenant_context.app_key
    bill_id = str(payload.get("bill_id") or "").strip()
    if not bill_id:
        raise HTTPException(status_code=422, detail="bill_id is required")
    reason = str(payload.get("reversal_reason") or "").strip()
    if len(reason) < 10:
        raise HTTPException(status_code=422, detail="Reversal reason must be at least 10 characters")
    bills_col, reversals_col = _maintenance_collections()
    row = await bills_col.find_one({"tenant_id": tenant_id, "app_key": app_key, "id": bill_id})
    if not row:
        raise HTTPException(status_code=404, detail="Bill not found")
    now = datetime.now(timezone.utc).isoformat()
    reversal_journal_entry_id = None
    if row.get("is_posted"):
        reversal_journal_entry_id = await _reverse_maintenance_bill_accounting(
            session,
            tenant_id=tenant_id,
            app_key=app_key,
            bill=row,
            reason=reason,
            current_user=current_user,
        )
    await bills_col.update_one(
        {"tenant_id": tenant_id, "app_key": app_key, "id": bill_id},
        {
            "$set": {
                "status": "reversed",
                "is_posted": False,
                "posting_status": "reversed",
                "reversal_reason": reason,
                "reversal_journal_entry_id": reversal_journal_entry_id,
                "reversed_at": now,
                "updated_at": now,
            }
        },
    )
    reversal_doc = {
        "id": str(uuid4()),
        "tenant_id": tenant_id,
        "app_key": app_key,
        "bill_id": bill_id,
        "flat_id": row.get("flat_id"),
        "flat_number": row.get("flat_number"),
        "month": row.get("month"),
        "year": row.get("year"),
        "amount": row.get("amount"),
        "reason": reason,
        "committee_approval": payload.get("committee_approval"),
        "reversal_journal_entry_id": reversal_journal_entry_id,
        "created_by": str(current_user.get("sub") or "system"),
        "created_at": now,
    }
    await reversals_col.insert_one(reversal_doc)
    return {
        "status": "reversed",
        "reversal_journal_entry_id": reversal_journal_entry_id,
        "bill": _sanitize_mongo_doc({**row, "status": "reversed", "is_posted": False}),
    }


@router.post("/maintenance/regenerate-bill")
async def maintenance_regenerate_bill(
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_context = resolve_gruha_tenant(
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
        operation="maintenance bill regeneration",
    )
    tenant_id = tenant_context.tenant_id
    app_key = tenant_context.app_key
    flat_id = str(payload.get("flat_id") or "").strip()
    month = _safe_int(payload.get("month"))
    year = _safe_int(payload.get("year"))
    flat = await get_collection("housing_flats").find_one({"tenant_id": tenant_id, "app_key": app_key, "id": flat_id})
    if not flat:
        raise HTTPException(status_code=404, detail="Flat not found")
    bills_col, _ = _maintenance_collections()
    previous = await bills_col.find_one(
        {"tenant_id": tenant_id, "app_key": app_key, "flat_id": flat_id, "month": month, "year": year},
        sort=[("updated_at", -1), ("created_at", -1)],
    )
    now = datetime.now(timezone.utc).isoformat()
    breakdown = dict((previous or {}).get("breakdown") or {})
    corrected_occupants = _safe_int(payload.get("corrected_occupants"), 0)
    water_rate = _safe_float(breakdown.get("water_per_person_rate"))
    maintenance = (
        _safe_float(payload.get("override_maintenance"))
        if payload.get("override_maintenance") not in (None, "")
        else _safe_float((previous or {}).get("maintenance_amount"))
    )
    if payload.get("override_water") not in (None, ""):
        water = _safe_float(payload.get("override_water"))
    elif corrected_occupants > 0 and water_rate > 0:
        water = _round_money(water_rate * corrected_occupants)
    else:
        water = _safe_float((previous or {}).get("water_amount"))
    fixed = (
        _safe_float(payload.get("override_fixed"))
        if payload.get("override_fixed") not in (None, "")
        else _safe_float((previous or {}).get("fixed_amount"))
    )
    sinking = (
        _safe_float(payload.get("override_sinking"))
        if payload.get("override_sinking") not in (None, "")
        else _safe_float((previous or {}).get("sinking_fund_amount"))
    )
    repair = (
        _safe_float(payload.get("override_repair"))
        if payload.get("override_repair") not in (None, "")
        else _safe_float((previous or {}).get("repair_fund_amount"))
    )
    association = _safe_float((previous or {}).get("association_fund_amount"))
    corpus = (
        _safe_float(payload.get("override_corpus"))
        if payload.get("override_corpus") not in (None, "")
        else _safe_float((previous or {}).get("corpus_fund_amount"))
    )
    amount = _round_money(maintenance + water + fixed + sinking + repair + association + corpus)
    if corrected_occupants > 0 and water_rate > 0:
        breakdown["inmates_used"] = corrected_occupants
        breakdown["water_charges"] = water
        breakdown["water_calculation"] = f"Corrected: {water_rate:,.2f} x {corrected_occupants} resident(s)"
    doc = {
        "id": str(uuid4()),
        "tenant_id": tenant_id,
        "app_key": app_key,
        "flat_id": flat_id,
        "flat_number": flat.get("flat_number"),
        "month": month,
        "year": year,
        "amount": amount,
        "maintenance_amount": maintenance,
        "water_amount": water,
        "fixed_amount": fixed,
        "sinking_fund_amount": sinking,
        "repair_fund_amount": repair,
        "association_fund_amount": association,
        "corpus_fund_amount": corpus,
        "status": "generated",
        "is_posted": False,
        "posting_status": "not_posted",
        "notes": payload.get("notes"),
        "breakdown": {
            **breakdown,
            "maintenance_sqft": maintenance,
            "water_charges": water,
            "fixed_expenses": fixed,
            "sinking_fund": sinking,
            "repair_fund": repair,
            "association_fund": association,
            "corpus_fund": corpus,
        },
        "created_by": str(current_user.get("sub") or "system"),
        "created_at": now,
        "updated_at": now,
    }
    await bills_col.insert_one(doc)
    return _sanitize_mongo_doc(doc)


@router.get("/messages/rooms")
async def messages_list_rooms(
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    rooms_col, _ = _message_collections()
    await _get_or_create_message_room(
        tenant_id=tenant_id,
        app_key=app_key,
        name="General",
        room_type="general",
        description="General society messages",
    )
    await _get_or_create_message_room(
        tenant_id=tenant_id,
        app_key=app_key,
        name="Meeting Notices",
        room_type="meeting_notices",
        description="Official meeting notices for eligible society members",
    )
    rows = (
        await rooms_col.find({"tenant_id": tenant_id, "app_key": app_key})
        .sort([("updated_at", -1), ("created_at", -1)])
        .to_list(length=500)
    )
    visible_rooms: list[dict[str, Any]] = []
    for row in rows:
        if await _can_access_message_room(room=row, current_user=current_user, tenant_id=tenant_id, app_key=app_key):
            visible_rooms.append(_room_response(row))
    return visible_rooms


@router.get("/assets")
@router.get("/assets/")
async def assets_list(
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    rows = (
        await get_collection("housing_assets")
        .find({"tenant_id": tenant_id, "app_key": app_key})
        .sort([("created_at", -1)])
        .to_list(length=1000)
    )
    return [_asset_response(row) for row in rows]


@router.post("/assets")
@router.post("/assets/")
async def assets_create(
    payload: dict[str, Any],
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    name = str(payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=422, detail="Asset name is required")
    category = str(payload.get("category") or "other").strip().lower() or "other"
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid4()),
        "tenant_id": tenant_id,
        "app_key": app_key,
        "asset_code": await _next_asset_code(tenant_id=tenant_id, app_key=app_key, category=category),
        "name": name,
        "category": category,
        "account_code": str(payload.get("account_code") or "1500").strip(),
        "quantity": max(1, _safe_int(payload.get("quantity"), default=1)),
        "location": str(payload.get("location") or "").strip(),
        "status": str(payload.get("status") or "Active").strip() or "Active",
        "acquisition_type": str(payload.get("acquisition_type") or "builder_handover").strip(),
        "handover_date": payload.get("handover_date"),
        "purchase_date": payload.get("purchase_date"),
        "original_cost": _safe_float(payload.get("original_cost")),
        "depreciation_method": str(payload.get("depreciation_method") or "straight_line").strip(),
        "depreciation_rate": _safe_float(payload.get("depreciation_rate")),
        "useful_life_years": _safe_int(payload.get("useful_life_years"), default=0),
        "residual_value": _safe_float(payload.get("residual_value")),
        "amc_vendor": str(payload.get("amc_vendor") or "").strip(),
        "amc_expiry": payload.get("amc_expiry"),
        "insurance_policy_no": str(payload.get("insurance_policy_no") or "").strip(),
        "insurance_expiry": payload.get("insurance_expiry"),
        "vendor_name": str(payload.get("vendor_name") or "").strip(),
        "invoice_no": str(payload.get("invoice_no") or "").strip(),
        "notes": str(payload.get("notes") or "").strip(),
        "is_scrapped": False,
        "created_by": str(current_user.get("sub") or "system"),
        "created_at": now,
        "updated_at": now,
    }
    if doc["acquisition_type"] == "builder_handover" and doc["original_cost"] > 0:
        entry = await _post_asset_capitalization_journal(
            session,
            tenant_id=tenant_id,
            app_key=app_key,
            asset=doc,
            created_by=str(current_user.get("sub") or "system"),
        )
        if entry is not None:
            doc["journal_entry_id"] = entry.id
            doc["accounting_posting_status"] = "posted"
            doc["accounting_posted_at"] = now
    elif doc["acquisition_type"] == "society_purchase" and doc["original_cost"] > 0:
        doc["accounting_posting_status"] = "pending_payment_account"
    else:
        doc["accounting_posting_status"] = "not_required"
    await get_collection("housing_assets").insert_one(doc)
    return _asset_response(doc)


@router.get("/assets/{asset_id}")
async def assets_get(
    asset_id: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    row = await get_collection("housing_assets").find_one({"tenant_id": tenant_id, "app_key": app_key, "id": asset_id})
    if not row:
        raise HTTPException(status_code=404, detail="Asset not found")
    return _asset_response(row)


@router.post("/assets/{asset_id}/post-accounting")
async def assets_post_accounting(
    asset_id: str,
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    assets = get_collection("housing_assets")
    row = await assets.find_one({"tenant_id": tenant_id, "app_key": app_key, "id": asset_id})
    if not row:
        raise HTTPException(status_code=404, detail="Asset not found")
    if row.get("journal_entry_id"):
        return _asset_response(row)

    entry = await _post_asset_capitalization_journal(
        session,
        tenant_id=tenant_id,
        app_key=app_key,
        asset=row,
        created_by=str(current_user.get("sub") or "system"),
    )
    if entry is None:
        raise HTTPException(status_code=422, detail="This asset does not require automatic accounting posting.")

    now = datetime.now(timezone.utc).isoformat()
    await assets.update_one(
        {"tenant_id": tenant_id, "app_key": app_key, "id": asset_id},
        {
            "$set": {
                "journal_entry_id": entry.id,
                "accounting_posting_status": "posted",
                "accounting_posted_at": now,
                "updated_at": now,
            }
        },
    )
    row.update(
        {
            "journal_entry_id": entry.id,
            "accounting_posting_status": "posted",
            "accounting_posted_at": now,
            "updated_at": now,
        }
    )
    return _asset_response(row)


@router.post("/assets/{asset_id}/scrap")
async def assets_scrap(
    asset_id: str,
    scrapping_reason: str = Query(default=""),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    now = datetime.now(timezone.utc).isoformat()
    result = await get_collection("housing_assets").update_one(
        {"tenant_id": tenant_id, "app_key": app_key, "id": asset_id},
        {
            "$set": {
                "status": "Scrapped",
                "is_scrapped": True,
                "scrapping_reason": scrapping_reason,
                "scrapped_at": now,
                "updated_at": now,
            }
        },
    )
    if getattr(result, "matched_count", 0) == 0:
        raise HTTPException(status_code=404, detail="Asset not found")
    row = await get_collection("housing_assets").find_one({"tenant_id": tenant_id, "app_key": app_key, "id": asset_id})
    return _asset_response(row or {})


@router.post("/messages/rooms")
async def messages_create_room(
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    if not _can_manage_chat_rooms(current_user.get("role")):
        raise HTTPException(status_code=403, detail="Only admin/secretary can create chat rooms")

    name = str(payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")

    room_type = str(payload.get("type") or "general").strip().lower() or "general"
    description = str(payload.get("description") or "").strip()
    audience_type = str(payload.get("audience_type") or "public").strip().lower() or "public"
    if audience_type not in {"public", "flats"}:
        raise HTTPException(status_code=400, detail="audience_type must be public or flats")
    allowed_flat_numbers = _normalize_flat_numbers(
        payload.get("allowed_flat_numbers") or payload.get("flat_numbers") or []
    )
    if audience_type == "flats" and not allowed_flat_numbers:
        raise HTTPException(status_code=400, detail="Select at least one flat for a restricted room")
    now = datetime.now(timezone.utc).isoformat()
    room_id = str(uuid4())
    doc = {
        "id": room_id,
        "tenant_id": tenant_id,
        "app_key": app_key,
        "name": name,
        "type": room_type,
        "description": description,
        "audience_type": audience_type,
        "allowed_flat_numbers": allowed_flat_numbers if audience_type == "flats" else [],
        "allowed_member_ids": [],
        "created_by": str(current_user.get("sub") or "system"),
        "created_at": now,
        "updated_at": now,
        "last_message_at": None,
    }
    rooms_col, _ = _message_collections()
    await rooms_col.insert_one(doc)
    return _room_response(doc)


@router.get("/messages/rooms/{room_id}/messages")
async def messages_list_for_room(
    room_id: str,
    limit: int = Query(default=100, ge=1, le=500),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    rooms_col, messages_col = _message_collections()
    room = await rooms_col.find_one({"tenant_id": tenant_id, "app_key": app_key, "id": room_id})
    if not room:
        raise HTTPException(status_code=404, detail="Room not found")
    if not await _can_access_message_room(room=room, current_user=current_user, tenant_id=tenant_id, app_key=app_key):
        raise HTTPException(status_code=404, detail="Room not found")
    now = datetime.now(timezone.utc).isoformat()
    await messages_col.delete_many({"tenant_id": tenant_id, "app_key": app_key, "room_id": room_id, "expires_at": {"$lte": now}})
    rows = (
        await messages_col.find(
            {
                "tenant_id": tenant_id,
                "app_key": app_key,
                "room_id": room_id,
                "$or": [{"expires_at": {"$exists": False}}, {"expires_at": {"$gt": now}}],
            }
        )
        .sort("created_at", 1)
        .to_list(length=limit)
    )
    return [_message_response(row) for row in rows]


@router.get("/messages/rooms/{room_id}/messages/{message_id}/attachments/{attachment_id}")
async def messages_download_attachment(
    room_id: str,
    message_id: str,
    attachment_id: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    rooms_col, messages_col = _message_collections()
    room = await rooms_col.find_one({"tenant_id": tenant_id, "app_key": app_key, "id": room_id})
    if not room:
        raise HTTPException(status_code=404, detail="Room not found")
    if not await _can_access_message_room(room=room, current_user=current_user, tenant_id=tenant_id, app_key=app_key):
        raise HTTPException(status_code=404, detail="Room not found")
    message = await messages_col.find_one({"tenant_id": tenant_id, "app_key": app_key, "room_id": room_id, "id": message_id})
    if not message:
        raise HTTPException(status_code=404, detail="Attachment not found")
    expires_at = str(message.get("expires_at") or "")
    if expires_at and expires_at <= datetime.now(timezone.utc).isoformat():
        await messages_col.delete_one({"tenant_id": tenant_id, "app_key": app_key, "room_id": room_id, "id": message_id})
        raise HTTPException(status_code=404, detail="Attachment expired")
    for item in message.get("attachments") or []:
        if isinstance(item, dict) and item.get("id") == attachment_id:
            headers = {"Content-Disposition": f'attachment; filename="{_safe_file_name(item.get("file_name") or "attachment", "attachment")}"'}
            return Response(content=bytes(item.get("content") or b""), media_type=item.get("content_type") or "application/octet-stream", headers=headers)
    raise HTTPException(status_code=404, detail="Attachment not found")


@router.post("/messages/rooms/{room_id}/messages")
async def messages_send_to_room(
    room_id: str,
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    rooms_col, messages_col = _message_collections()
    room = await rooms_col.find_one({"tenant_id": tenant_id, "app_key": app_key, "id": room_id})
    if not room:
        raise HTTPException(status_code=404, detail="Room not found")
    if not await _can_access_message_room(room=room, current_user=current_user, tenant_id=tenant_id, app_key=app_key):
        raise HTTPException(status_code=404, detail="Room not found")

    content = str(payload.get("text") or payload.get("content") or "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="text is required")

    now = datetime.now(timezone.utc).isoformat()
    retention_days = _message_retention_days(payload.get("retention_days"))
    expires_at = (datetime.now(timezone.utc) + timedelta(days=retention_days)).isoformat()
    sender_name = (
        str(current_user.get("full_name") or "").strip()
        or str(current_user.get("email") or "").strip()
        or "User"
    )
    message_doc = {
        "id": str(uuid4()),
        "tenant_id": tenant_id,
        "app_key": app_key,
        "room_id": room_id,
        "sender_id": str(current_user.get("sub") or "system"),
        "sender_name": sender_name,
        "content": content,
        "message_type": "text",
        "attachments": [],
        "retention_days": retention_days,
        "expires_at": expires_at,
        "created_at": now,
    }
    await messages_col.insert_one(message_doc)
    await rooms_col.update_one(
        {"tenant_id": tenant_id, "app_key": app_key, "id": room_id},
        {"$set": {"updated_at": now, "last_message_at": now}},
    )
    return _message_response(message_doc)


@router.post("/messages/rooms/{room_id}/messages/with-attachment")
async def messages_send_to_room_with_attachment(
    room_id: str,
    text: str = Form(default=""),
    retention_days: int = Form(default=30),
    file: UploadFile | None = File(default=None),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    rooms_col, messages_col = _message_collections()
    room = await rooms_col.find_one({"tenant_id": tenant_id, "app_key": app_key, "id": room_id})
    if not room:
        raise HTTPException(status_code=404, detail="Room not found")
    if not await _can_access_message_room(room=room, current_user=current_user, tenant_id=tenant_id, app_key=app_key):
        raise HTTPException(status_code=404, detail="Room not found")

    content = str(text or "").strip()
    attachment_items: list[dict[str, Any]] = []
    if file is not None and file.filename:
        raw = await file.read()
        if len(raw) > 10 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="Attachment must be 10 MB or smaller")
        attachment_id = str(uuid4())
        attachment_items.append(
            {
                "id": attachment_id,
                "file_name": file.filename,
                "content_type": file.content_type or "application/octet-stream",
                "size": len(raw),
                "content": raw,
                "download_url": f"/api/v1/messages/rooms/{room_id}/messages/{{message_id}}/attachments/{attachment_id}",
            }
        )
    if not content and not attachment_items:
        raise HTTPException(status_code=400, detail="message text or attachment is required")

    now_dt = datetime.now(timezone.utc)
    now = now_dt.isoformat()
    retention = _message_retention_days(retention_days)
    message_id = str(uuid4())
    message_doc = {
        "id": message_id,
        "tenant_id": tenant_id,
        "app_key": app_key,
        "room_id": room_id,
        "sender_id": str(current_user.get("sub") or "system"),
        "sender_name": str(current_user.get("full_name") or current_user.get("email") or "User"),
        "content": content,
        "message_type": "attachment" if attachment_items else "text",
        "attachments": attachment_items,
        "retention_days": retention,
        "expires_at": (now_dt + timedelta(days=retention)).isoformat(),
        "created_at": now,
    }
    for item in attachment_items:
        item["download_url"] = item["download_url"].replace("{message_id}", message_id)
    await messages_col.insert_one(message_doc)
    await rooms_col.update_one(
        {"tenant_id": tenant_id, "app_key": app_key, "id": room_id},
        {"$set": {"updated_at": now, "last_message_at": now}},
    )
    return _message_response(message_doc)


@router.post("/society/upload-logo")
async def society_upload_logo(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())

    allowed = {"image/png", "image/jpeg", "image/jpg"}
    content_type = (file.content_type or "").lower()
    if content_type not in allowed:
        raise HTTPException(status_code=400, detail="Only PNG and JPG files are allowed")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Logo file must be less than 2MB")

    stored_name = _safe_file_name(file.filename or "logo.png", prefix="logo")
    now = datetime.now(timezone.utc).isoformat()
    await get_collection("housing_documents").insert_one(
        {
            "tenant_id": tenant_id,
            "app_key": app_key,
            "stored_name": stored_name,
            "original_name": file.filename or stored_name,
            "content_type": content_type,
            "document_type": "society_logo",
            "size_bytes": len(content),
            "content": content,
            "created_at": now,
            "updated_at": now,
        }
    )

    logo_url = f"/society/documents/{stored_name}"
    return {"logo_url": logo_url, "file_name": stored_name, "content_type": content_type, "size_bytes": len(content)}


@router.post("/society/upload-document")
async def society_upload_document(
    file: UploadFile = File(...),
    document_type: str = "other",
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Document file must be less than 10MB")

    stored_name = _safe_file_name(file.filename or "document.bin", prefix="doc")
    now = datetime.now(timezone.utc).isoformat()
    await get_collection("housing_documents").insert_one(
        {
            "tenant_id": tenant_id,
            "app_key": app_key,
            "stored_name": stored_name,
            "original_name": file.filename or stored_name,
            "content_type": file.content_type or "application/octet-stream",
            "document_type": (document_type or "other").strip().lower(),
            "size_bytes": len(content),
            "content": content,
            "created_at": now,
            "updated_at": now,
        }
    )

    url = f"/society/documents/{stored_name}"
    return {"url": url, "file_name": file.filename or stored_name, "stored_name": stored_name, "size_bytes": len(content)}


@router.get("/society/documents/{file_name}")
async def society_download_document(
    file_name: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    doc = await get_collection("housing_documents").find_one(
        {"tenant_id": tenant_id, "app_key": app_key, "stored_name": file_name}
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    content = doc.get("content")
    if content is None:
        raise HTTPException(status_code=404, detail="Document content missing")
    content_type = str(doc.get("content_type") or "application/octet-stream")
    original_name = str(doc.get("original_name") or file_name)
    headers = {"Content-Disposition": f'inline; filename="{original_name}"'}
    return StreamingResponse(BytesIO(bytes(content)), media_type=content_type, headers=headers)


@router.delete("/society/documents/{file_name}")
async def society_delete_document(
    file_name: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    result = await get_collection("housing_documents").delete_one(
        {"tenant_id": tenant_id, "app_key": app_key, "stored_name": file_name}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"status": "deleted", "file_name": file_name}


@router.post("/attachments/upload/{journal_entry_id}")
async def attachment_upload(
    journal_entry_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    content = await file.read()
    now = datetime.now(timezone.utc).isoformat()
    attachment_id = str(uuid4())
    stored_name = _safe_file_name(file.filename or "attachment", "attachment")
    await get_collection("housing_documents").insert_one(
        {
            "id": attachment_id,
            "tenant_id": tenant_id,
            "app_key": app_key,
            "document_type": "journal_attachment",
            "journal_entry_id": str(journal_entry_id),
            "stored_name": stored_name,
            "file_name": file.filename or stored_name,
            "content_type": file.content_type or "application/octet-stream",
            "content": content,
            "size_bytes": len(content),
            "created_at": now,
            "updated_at": now,
        }
    )
    return {"id": attachment_id, "journal_entry_id": str(journal_entry_id), "file_name": file.filename or stored_name}


@router.get("/attachments/journal/{journal_entry_id}")
async def attachment_list(
    journal_entry_id: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    rows = await get_collection("housing_documents").find(
        {"tenant_id": tenant_id, "app_key": app_key, "document_type": "journal_attachment", "journal_entry_id": str(journal_entry_id)},
        {"content": 0},
    ).sort("created_at", -1).to_list(length=500)
    return [
        {
            "id": row.get("id"),
            "journal_entry_id": row.get("journal_entry_id"),
            "file_name": row.get("file_name") or row.get("stored_name"),
            "content_type": row.get("content_type"),
            "size_bytes": row.get("size_bytes") or 0,
            "created_at": row.get("created_at"),
        }
        for row in rows
    ]


@router.get("/attachments/{attachment_id}")
async def attachment_download(
    attachment_id: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    doc = await get_collection("housing_documents").find_one(
        {"tenant_id": tenant_id, "app_key": app_key, "document_type": "journal_attachment", "id": attachment_id}
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Attachment not found")
    headers = {"Content-Disposition": f'attachment; filename="{_safe_file_name(doc.get("file_name") or "attachment", "attachment")}"'}
    return Response(content=bytes(doc.get("content") or b""), media_type=doc.get("content_type") or "application/octet-stream", headers=headers)


@router.delete("/attachments/{attachment_id}")
async def attachment_delete(
    attachment_id: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    result = await get_collection("housing_documents").delete_one(
        {"tenant_id": tenant_id, "app_key": app_key, "document_type": "journal_attachment", "id": attachment_id}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Attachment not found")
    return {"status": "deleted"}


@router.post("/resources/files/upload")
async def resource_file_upload(
    file: UploadFile = File(...),
    category: str = Form(default="general"),
    description: str | None = Form(default=None),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    content = await file.read()
    now = datetime.now(timezone.utc).isoformat()
    file_id = str(uuid4())
    stored_name = _safe_file_name(file.filename or "resource", "resource")
    await get_collection("housing_documents").insert_one(
        {
            "id": file_id,
            "tenant_id": tenant_id,
            "app_key": app_key,
            "document_type": "resource_file",
            "category": (category or "general").strip() or "general",
            "description": (description or "").strip(),
            "stored_name": stored_name,
            "file_name": file.filename or stored_name,
            "content_type": file.content_type or "application/octet-stream",
            "content": content,
            "size_bytes": len(content),
            "created_at": now,
            "updated_at": now,
        }
    )
    return {"id": file_id, "file_name": file.filename or stored_name, "category": category, "size_bytes": len(content)}


@router.get("/resources/files")
async def resource_files_list(
    category: str | None = None,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    query = {"tenant_id": tenant_id, "app_key": app_key, "document_type": "resource_file"}
    if category:
        query["category"] = category
    rows = await get_collection("housing_documents").find(query, {"content": 0}).sort("created_at", -1).to_list(length=500)
    return [
        {
            "id": row.get("id"),
            "file_name": row.get("file_name") or row.get("stored_name"),
            "category": row.get("category") or "general",
            "description": row.get("description") or "",
            "content_type": row.get("content_type"),
            "size_bytes": row.get("size_bytes") or 0,
            "created_at": row.get("created_at"),
        }
        for row in rows
    ]


@router.get("/resources/files/{file_id}")
@router.get("/resources/files/{file_id}/download")
async def resource_file_download(
    file_id: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    doc = await get_collection("housing_documents").find_one(
        {"tenant_id": tenant_id, "app_key": app_key, "document_type": "resource_file", "id": file_id}
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Resource file not found")
    headers = {"Content-Disposition": f'attachment; filename="{_safe_file_name(doc.get("file_name") or "resource", "resource")}"'}
    return Response(content=bytes(doc.get("content") or b""), media_type=doc.get("content_type") or "application/octet-stream", headers=headers)


@router.delete("/resources/files/{file_id}")
async def resource_file_delete(
    file_id: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    result = await get_collection("housing_documents").delete_one(
        {"tenant_id": tenant_id, "app_key": app_key, "document_type": "resource_file", "id": file_id}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Resource file not found")
    return {"status": "deleted"}


@router.get("/meetings")
async def meetings_list(
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    meetings, _ = _meeting_collections()
    rows = await meetings.find({"tenant_id": tenant_id, "app_key": app_key}).sort("created_at", -1).to_list(length=500)
    return [
        {
            "id": row.get("id"),
            "meeting_title": row.get("meeting_title") or "Meeting",
            "meeting_type": row.get("meeting_type") or "MC",
            "meeting_date": row.get("meeting_date") or "",
            "meeting_time": row.get("meeting_time") or "",
            "venue": row.get("venue") or "",
            "status": _normalize_status(row.get("status")),
            "notice_sent": bool(row.get("notice_sent") or False),
        }
        for row in rows
    ]


@router.post("/meetings")
async def meetings_create(
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    title = str(payload.get("meeting_title") or "").strip()
    meeting_date = str(payload.get("meeting_date") or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="meeting_title is required")
    if not meeting_date:
        raise HTTPException(status_code=400, detail="meeting_date is required")

    now = datetime.now(timezone.utc).isoformat()
    agenda_items = payload.get("agenda_items")
    if not isinstance(agenda_items, list):
        agenda_items = []
    normalized_agenda: list[dict[str, Any]] = []
    for idx, item in enumerate(agenda_items):
        if not isinstance(item, dict):
            continue
        item_title = str(item.get("item_title") or "").strip()
        item_description = str(item.get("item_description") or "").strip()
        if not item_title and not item_description:
            continue
        normalized_agenda.append(
            {
                "item_number": int(item.get("item_number") or (idx + 1)),
                "item_title": item_title,
                "item_description": item_description,
            }
        )

    eligible_member_ids = _normalize_member_ids(payload.get("eligible_member_ids") or [])
    selected_members = await _resolve_eligible_meeting_members(
        tenant_id=tenant_id,
        app_key=app_key,
        member_ids=eligible_member_ids,
    )
    eligible_flat_numbers = _normalize_flat_numbers([member.get("flat_number") for member in selected_members])
    eligible_default = len(selected_members) if eligible_member_ids else await _count_eligible_members(tenant_id=tenant_id, app_key=app_key)
    total_members_eligible = len(selected_members) if eligible_member_ids else (
        _safe_int(payload.get("total_members_eligible"), default=0) or eligible_default
    )
    quorum_required = _safe_int(payload.get("quorum_required"), default=0)
    notice_room_id = str(payload.get("notice_room_id") or "").strip()
    if notice_room_id:
        rooms_col, _ = _message_collections()
        notice_room = await rooms_col.find_one({"tenant_id": tenant_id, "app_key": app_key, "id": notice_room_id})
        if not notice_room:
            raise HTTPException(status_code=400, detail="Notice room not found")

    meeting_id = str(uuid4())
    meeting_doc = {
        "id": meeting_id,
        "tenant_id": tenant_id,
        "app_key": app_key,
        "meeting_title": title,
        "meeting_type": str(payload.get("meeting_type") or "MC"),
        "meeting_date": meeting_date,
        "meeting_time": str(payload.get("meeting_time") or ""),
        "venue": str(payload.get("venue") or ""),
        "notice_sent_to": str(payload.get("notice_sent_to") or "all_members"),
        "notice_room_id": notice_room_id or None,
        "eligible_member_ids": [str(member.get("id")) for member in selected_members if member.get("id")] if eligible_member_ids else [],
        "eligible_flat_numbers": eligible_flat_numbers,
        "agenda_items": normalized_agenda,
        "agenda": str(payload.get("agenda") or ""),
        "minutes_text": "",
        "status": _normalize_status(payload.get("status")),
        "total_members_eligible": total_members_eligible,
        "total_members_present": 0,
        "quorum_required": quorum_required,
        "quorum_met": bool(payload.get("quorum_met") or False),
        "notice_sent": False,
        "notice_sent_at": None,
        "last_change_action": None,
        "last_change_reason": None,
        "change_log": [],
        "attendance": [],
        "created_by": str(current_user.get("sub") or "system"),
        "created_at": now,
        "updated_at": now,
    }
    if quorum_required > 0:
        meeting_doc["quorum_met"] = False
    meetings, _ = _meeting_collections()
    await meetings.insert_one(meeting_doc)
    return _sanitize_mongo_doc(meeting_doc)


@router.get("/meetings/{meeting_id}")
async def meetings_get(
    meeting_id: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    row = await _get_meeting_or_404(tenant_id=tenant_id, app_key=app_key, meeting_id=meeting_id)
    eligible_hint = await _meeting_eligible_count(tenant_id=tenant_id, app_key=app_key, meeting=row)
    meeting = _sanitize_mongo_doc(row)
    meeting.update(_meeting_stats(row, eligible_hint=eligible_hint))
    return meeting


@router.get("/meetings/{meeting_id}/details")
async def meetings_details(
    meeting_id: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    row = await _get_meeting_or_404(tenant_id=tenant_id, app_key=app_key, meeting_id=meeting_id)
    _, resolutions_col = _meeting_collections()
    resolutions = await resolutions_col.find({"tenant_id": tenant_id, "app_key": app_key, "meeting_id": meeting_id}).sort("created_at", -1).to_list(length=500)
    eligible_hint = await _meeting_eligible_count(tenant_id=tenant_id, app_key=app_key, meeting=row)
    meeting = _sanitize_mongo_doc(row)
    meeting.update(_meeting_stats(row, eligible_hint=eligible_hint))
    return {
        "meeting": meeting,
        "agenda_items": row.get("agenda_items") or [],
        "attendance": row.get("attendance") or [],
        "resolutions": [_sanitize_mongo_doc(r) for r in resolutions],
    }


@router.patch("/meetings/{meeting_id}")
async def meetings_update(
    meeting_id: str,
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    await _get_meeting_or_404(tenant_id=tenant_id, app_key=app_key, meeting_id=meeting_id)
    allowed = {
        "meeting_title",
        "meeting_type",
        "meeting_date",
        "meeting_time",
        "venue",
        "status",
        "agenda",
        "total_members_eligible",
        "quorum_required",
        "quorum_met",
        "notice_room_id",
        "eligible_member_ids",
        "change_action",
        "change_reason",
    }
    existing = await _get_meeting_or_404(tenant_id=tenant_id, app_key=app_key, meeting_id=meeting_id)
    update_doc = {k: v for k, v in payload.items() if k in allowed}
    change_action = str(update_doc.get("change_action") or "").strip().lower() or "general_update"
    change_reason = str(update_doc.get("change_reason") or "").strip()
    if change_action in {"cancel", "postpone", "prepone"} and not change_reason:
        raise HTTPException(status_code=400, detail="change_reason is required for cancel/postpone/prepone")
    if change_action == "cancel":
        update_doc["status"] = "CANCELLED"
    if "status" in update_doc:
        update_doc["status"] = _normalize_status(update_doc.get("status"))
    if "total_members_eligible" in update_doc:
        update_doc["total_members_eligible"] = _safe_int(update_doc.get("total_members_eligible"), default=0)
    if "quorum_required" in update_doc:
        update_doc["quorum_required"] = _safe_int(update_doc.get("quorum_required"), default=0)
    if "eligible_member_ids" in update_doc:
        selected_ids = _normalize_member_ids(update_doc.get("eligible_member_ids") or [])
        selected_members = await _resolve_eligible_meeting_members(
            tenant_id=tenant_id,
            app_key=app_key,
            member_ids=selected_ids,
        )
        update_doc["eligible_member_ids"] = [
            str(member.get("id")) for member in selected_members if member.get("id")
        ] if selected_ids else []
        update_doc["eligible_flat_numbers"] = _normalize_flat_numbers(
            [member.get("flat_number") for member in selected_members]
        )
        update_doc["total_members_eligible"] = len(selected_members) if selected_ids else await _count_eligible_members(
            tenant_id=tenant_id,
            app_key=app_key,
        )
    if "notice_room_id" in update_doc:
        notice_room_id = str(update_doc.get("notice_room_id") or "").strip()
        if notice_room_id:
            rooms_col, _ = _message_collections()
            notice_room = await rooms_col.find_one({"tenant_id": tenant_id, "app_key": app_key, "id": notice_room_id})
            if not notice_room:
                raise HTTPException(status_code=400, detail="Notice room not found")
            update_doc["notice_room_id"] = notice_room_id
        else:
            update_doc["notice_room_id"] = None
    now = datetime.now(timezone.utc).isoformat()
    if change_reason:
        update_doc["last_change_reason"] = change_reason
    update_doc["last_change_action"] = change_action
    update_doc["updated_at"] = now
    if any(key in update_doc for key in {"meeting_date", "meeting_time", "status"}) or change_reason:
        previous_log = existing.get("change_log") or []
        if not isinstance(previous_log, list):
            previous_log = []
        change_entry = {
            "changed_at": now,
            "changed_by": str(current_user.get("sub") or "system"),
            "action": change_action,
            "reason": change_reason or None,
            "previous": {
                "meeting_date": existing.get("meeting_date"),
                "meeting_time": existing.get("meeting_time"),
                "status": existing.get("status"),
            },
            "current": {
                "meeting_date": update_doc.get("meeting_date", existing.get("meeting_date")),
                "meeting_time": update_doc.get("meeting_time", existing.get("meeting_time")),
                "status": update_doc.get("status", existing.get("status")),
            },
        }
        update_doc["change_log"] = [change_entry, *previous_log][:50]
    meetings, _ = _meeting_collections()
    await meetings.update_one({"tenant_id": tenant_id, "app_key": app_key, "id": meeting_id}, {"$set": update_doc})
    row = await _get_meeting_or_404(tenant_id=tenant_id, app_key=app_key, meeting_id=meeting_id)
    eligible_hint = await _meeting_eligible_count(tenant_id=tenant_id, app_key=app_key, meeting=row)
    meeting = _sanitize_mongo_doc(row)
    meeting.update(_meeting_stats(row, eligible_hint=eligible_hint))
    return meeting


@router.delete("/meetings/{meeting_id}")
async def meetings_delete(
    meeting_id: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    meetings, resolutions_col = _meeting_collections()
    result = await meetings.delete_one({"tenant_id": tenant_id, "app_key": app_key, "id": meeting_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Meeting not found")
    await resolutions_col.delete_many({"tenant_id": tenant_id, "app_key": app_key, "meeting_id": meeting_id})
    return {"status": "deleted", "id": meeting_id}


@router.post("/meetings/{meeting_id}/attendance")
async def meetings_attendance(
    meeting_id: str,
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    row = await _get_meeting_or_404(tenant_id=tenant_id, app_key=app_key, meeting_id=meeting_id)
    attendees = payload.get("attendees")
    if not isinstance(attendees, list):
        attendees = []
    normalized: list[dict[str, Any]] = []
    for item in attendees:
        if not isinstance(item, dict):
            continue
        member_id = str(item.get("member_id") or "").strip()
        status = str(item.get("status") or "absent").strip().lower()
        if not member_id:
            continue
        if status not in {"present", "absent", "proxy"}:
            status = "absent"
        normalized.append({"member_id": member_id, "status": status})
    now = datetime.now(timezone.utc).isoformat()
    eligible_hint = await _meeting_eligible_count(tenant_id=tenant_id, app_key=app_key, meeting=row)
    stats = _meeting_stats({**row, "attendance": normalized}, eligible_hint=eligible_hint)
    update_set: dict[str, Any] = {
        "attendance": normalized,
        "updated_at": now,
        "total_members_present": stats["total_members_present"],
        "quorum_met": stats["quorum_met"],
    }
    if _safe_int(row.get("total_members_eligible"), default=0) <= 0 and stats["total_members_eligible"] > 0:
        update_set["total_members_eligible"] = stats["total_members_eligible"]
    meetings, _ = _meeting_collections()
    await meetings.update_one(
        {"tenant_id": tenant_id, "app_key": app_key, "id": meeting_id},
        {"$set": update_set},
    )
    updated = {**row, **update_set}
    return {"meeting_id": meeting_id, "attendance": normalized, "meeting": _sanitize_mongo_doc(updated)}


@router.post("/meetings/{meeting_id}/minutes")
async def meetings_minutes(
    meeting_id: str,
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    await _get_meeting_or_404(tenant_id=tenant_id, app_key=app_key, meeting_id=meeting_id)
    minutes_text = str(payload.get("minutes_text") or "")
    now = datetime.now(timezone.utc).isoformat()
    meetings, _ = _meeting_collections()
    await meetings.update_one(
        {"tenant_id": tenant_id, "app_key": app_key, "id": meeting_id},
        {"$set": {"minutes_text": minutes_text, "updated_at": now}},
    )
    row = await _get_meeting_or_404(tenant_id=tenant_id, app_key=app_key, meeting_id=meeting_id)
    return {"meeting": _sanitize_mongo_doc(row)}


@router.post("/meetings/{meeting_id}/resolutions")
async def meetings_resolution_create(
    meeting_id: str,
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    await _get_meeting_or_404(tenant_id=tenant_id, app_key=app_key, meeting_id=meeting_id)
    title = str(payload.get("resolution_title") or "").strip()
    text = str(payload.get("resolution_text") or "").strip()
    if not title or not text:
        raise HTTPException(status_code=400, detail="resolution_title and resolution_text are required")
    now = datetime.now(timezone.utc).isoformat()
    resolution = {
        "id": str(uuid4()),
        "tenant_id": tenant_id,
        "app_key": app_key,
        "meeting_id": meeting_id,
        "resolution_title": title,
        "resolution_text": text,
        "resolution_type": str(payload.get("resolution_type") or "ordinary"),
        "proposed_by_id": str(payload.get("proposed_by_id") or ""),
        "seconded_by_id": str(payload.get("seconded_by_id") or ""),
        "votes_for": int(payload.get("votes_for") or 0),
        "votes_against": int(payload.get("votes_against") or 0),
        "votes_abstain": int(payload.get("votes_abstain") or 0),
        "result": str(payload.get("result") or "passed"),
        "created_by": str(current_user.get("sub") or "system"),
        "created_at": now,
        "updated_at": now,
    }
    _, resolutions_col = _meeting_collections()
    await resolutions_col.insert_one(resolution)
    return _sanitize_mongo_doc(resolution)


@router.post("/meetings/{meeting_id}/send-notice")
async def meetings_send_notice(
    meeting_id: str,
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    meeting = await _get_meeting_or_404(tenant_id=tenant_id, app_key=app_key, meeting_id=meeting_id)
    now = datetime.now(timezone.utc).isoformat()
    eligible_members = await _meeting_eligible_count(tenant_id=tenant_id, app_key=app_key, meeting=meeting)
    rooms_col, messages_col = _message_collections()
    target_room_id = str(payload.get("room_id") or meeting.get("notice_room_id") or "").strip()
    if target_room_id:
        notice_room = await rooms_col.find_one({"tenant_id": tenant_id, "app_key": app_key, "id": target_room_id})
        if not notice_room:
            raise HTTPException(status_code=404, detail="Notice room not found")
    elif _normalize_flat_numbers(meeting.get("eligible_flat_numbers") or []):
        notice_room = await _get_or_create_message_room(
            tenant_id=tenant_id,
            app_key=app_key,
            name=f"Notice: {str(meeting.get('meeting_title') or 'Meeting')[:60]}",
            room_type=f"meeting_notice_{meeting_id}",
            description="Restricted meeting notice room for selected eligible members",
            audience_type="flats",
            allowed_flat_numbers=_normalize_flat_numbers(meeting.get("eligible_flat_numbers") or []),
        )
    else:
        notice_room = await _get_or_create_message_room(
            tenant_id=tenant_id,
            app_key=app_key,
            name="Meeting Notices",
            room_type="meeting_notices",
            description="Official meeting notices for eligible society members",
        )

    notice_message_id = str(meeting.get("notice_message_id") or "").strip()
    if notice_message_id:
        notice_message = await messages_col.find_one(
            {"tenant_id": tenant_id, "app_key": app_key, "room_id": notice_room["id"], "id": notice_message_id}
        )
    else:
        notice_message = None

    if not notice_message:
        notice_message = {
            "id": str(uuid4()),
            "tenant_id": tenant_id,
            "app_key": app_key,
            "room_id": notice_room["id"],
            "sender_id": str(current_user.get("sub") or "system"),
            "sender_name": str(current_user.get("name") or current_user.get("email") or "Society Admin"),
            "content": _build_meeting_notice_message(meeting, eligible_members),
            "message_type": "meeting_notice",
            "meeting_id": meeting_id,
            "eligible_members": eligible_members,
            "created_at": now,
        }
        await messages_col.insert_one(notice_message)
        await rooms_col.update_one(
            {"tenant_id": tenant_id, "app_key": app_key, "id": notice_room["id"]},
            {"$set": {"updated_at": now, "last_message_at": now}},
        )

    meetings, _ = _meeting_collections()
    await meetings.update_one(
        {"tenant_id": tenant_id, "app_key": app_key, "id": meeting_id},
        {
            "$set": {
                "notice_sent": True,
                "notice_sent_at": now,
                "notice_room_id": notice_room["id"],
                "notice_room_name": notice_room.get("name") or "Meeting Notices",
                "notice_message_id": notice_message["id"],
                "notice_delivery": {
                    "message_board": "posted",
                    "eligible_members": eligible_members,
                },
                "notice_channels": {
                    "message_board": True,
                    "email": bool(payload.get("send_email", False)),
                    "sms": bool(payload.get("send_sms", False)),
                },
                "updated_at": now,
            }
        },
    )
    row = await _get_meeting_or_404(tenant_id=tenant_id, app_key=app_key, meeting_id=meeting_id)
    return {
        "meeting": _sanitize_mongo_doc(row),
        "status": "notice_posted",
        "message_room": _sanitize_mongo_doc(notice_room),
        "message": _sanitize_mongo_doc(notice_message),
        "eligible_members": eligible_members,
    }


@router.post("/member-onboarding/", response_model=MemberResponse)
async def member_onboarding_create(
    payload: MemberCreateRequest,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    doc = await create_member(tenant_id=tenant_id, app_key=app_key, payload=payload)
    return MemberResponse(**doc)


@router.get("/member-onboarding/template")
async def member_onboarding_template_download():
    csv_text = (
        "name,phone_number,email,flat_number,member_type,move_in_date,total_occupants,occupation,is_primary,is_mobile_public\n"
        "Raghavan Iyer,9876512340,raghavan.iyer01@gmail.com,A-101,owner,2025-12-01,5,Professional,true,true\n"
        "Lakshmi Narayanan,9876523451,lakshmi.narayanan02@gmail.com,A-102,owner,2026-01-01,4,Homemaker,true,false\n"
        "Suresh Babu,9876534562,suresh.babu03@gmail.com,A-103,tenant,2025-12-01,6,Doctor,true,true\n"
    )
    headers = {"Content-Disposition": 'attachment; filename="members_template.csv"'}
    return StreamingResponse(BytesIO(csv_text.encode("utf-8")), media_type="text/csv", headers=headers)


@router.post("/member-onboarding/bulk-import")
async def member_onboarding_bulk_import(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    file_name = str(file.filename or "").strip().lower()
    content_type = str(file.content_type or "").strip().lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File must be less than 5MB")

    is_csv = file_name.endswith(".csv") or "csv" in content_type
    is_excel = file_name.endswith(".xlsx") or "spreadsheetml" in content_type
    if not (is_csv or is_excel):
        raise HTTPException(status_code=400, detail="Only CSV or XLSX files are supported")

    try:
        rows = _parse_csv_rows(content) if is_csv else _parse_xlsx_rows(content)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {exc}") from exc

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in uploaded file")

    created = 0
    failed = 0
    errors: list[dict[str, Any]] = []
    created_members: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        try:
            payload = _coerce_bulk_member_payload(row, idx)
            doc = await create_member(tenant_id=tenant_id, app_key=app_key, payload=payload)
            created += 1
            created_members.append(
                {
                    "id": str(doc.get("id") or ""),
                    "flat_number": str(doc.get("flat_number") or ""),
                    "name": str(doc.get("name") or ""),
                }
            )
        except HTTPException as exc:
            failed += 1
            errors.append({"row": idx, "error": str(exc.detail)})
        except Exception as exc:
            failed += 1
            errors.append({"row": idx, "error": str(exc)})

    return {
        "status": "completed",
        "total_rows": len(rows),
        "created": created,
        "failed": failed,
        "created_members": created_members[:100],
        "errors": errors[:200],
    }


@router.get("/onboarding-imports/templates/{kind}.csv")
async def onboarding_import_template(kind: str):
    if kind == "flats":
        csv_text = "flat_number,block,floor,status,area_sqft,bedrooms,parking_slots\nA-101,A,1,vacant,1000,2,P1\n"
    else:
        csv_text = "name,phone_number,email,flat_number,member_type,status,total_occupants\nResident,9999999999,resident@example.com,A-101,owner,active,1\n"
    return Response(content=csv_text, media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="{kind}.csv"'})


@router.post("/onboarding-imports/demo/import")
async def onboarding_import_demo(current_user: dict = Depends(get_current_user)):
    return {"status": "ok", "message": "Demo import hook acknowledged", "tenant_id": current_user.get("tenant_id")}


@router.post("/onboarding-imports/import/flats")
async def onboarding_import_flats(
    file: UploadFile = File(...),
    replace_existing: bool = Query(default=False),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    content = (await file.read()).decode("utf-8-sig", "replace")
    rows = list(csv.DictReader(content.splitlines()))
    if replace_existing:
        await get_collection("housing_flats").delete_many({"tenant_id": tenant_id, "app_key": app_key})
    imported = 0
    errors: list[dict[str, Any]] = []
    for index, item in enumerate(rows, start=2):
        flat_number = str(item.get("flat_number") or item.get("unit") or "").strip()
        if not flat_number:
            errors.append({"row": index, "error": "flat_number is required"})
            continue
        try:
            await create_flat(
                tenant_id=tenant_id,
                app_key=app_key,
                payload=FlatCreateRequest(
                    flat_number=flat_number,
                    block=(item.get("block") or None),
                    floor=int(item["floor"]) if str(item.get("floor") or "").strip() else None,
                    status=(item.get("status") or "vacant"),
                    area_sqft=float(item["area_sqft"]) if str(item.get("area_sqft") or "").strip() else None,
                    bedrooms=int(item["bedrooms"]) if str(item.get("bedrooms") or "").strip() else None,
                    parking_slots=(item.get("parking_slots") or None),
                ),
            )
            imported += 1
        except Exception as exc:
            errors.append({"row": index, "error": str(exc)})
    return {"status": "ok", "imported": imported, "errors": errors, "total_rows": len(rows)}


@router.post("/onboarding-imports/import/members")
async def onboarding_import_members(
    file: UploadFile = File(...),
    update_existing: bool = Query(default=False),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    result = await member_onboarding_bulk_import(file=file, current_user=current_user, x_tenant_id=x_tenant_id)
    result["update_existing"] = update_existing
    return result


@router.get("/member-onboarding", response_model=list[MemberResponse])
@router.get("/member-onboarding/", response_model=list[MemberResponse])
async def member_onboarding_list(
    status_filter: str | None = None,
    flat_number: str | None = None,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    rows = await list_members(tenant_id=tenant_id, app_key=app_key, status_filter=status_filter, flat_number=flat_number)
    return [MemberResponse(**row) for row in rows]


@router.get("/member-onboarding/my-profile", response_model=MemberResponse)
async def member_onboarding_my_profile(
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    email = str(current_user.get("email") or "").strip().lower()
    rows = await list_members(tenant_id=tenant_id, app_key=app_key, status_filter=None, flat_number=None)
    for row in rows:
        if str(row.get("email") or "").strip().lower() == email:
            return MemberResponse(**row)
    raise HTTPException(status_code=404, detail="Member profile not found")


@router.get("/member-onboarding/debug")
async def member_onboarding_debug(
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    rows = await list_members(tenant_id=tenant_id, app_key=app_key, status_filter=None, flat_number=None)
    return {"tenant_id": tenant_id, "app_key": app_key, "member_count": len(rows)}


@router.patch("/member-onboarding/{member_id}", response_model=MemberResponse)
@router.patch("/member-onboarding/{member_id}/", response_model=MemberResponse)
async def member_onboarding_update(
    member_id: str,
    payload: MemberUpdateRequest,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    row = await update_member(tenant_id=tenant_id, app_key=app_key, member_id=member_id, payload=payload)
    return MemberResponse(**row)


@router.get("/member-onboarding/{member_id}/checklist", response_model=MemberChecklistResponse)
@router.get("/member-onboarding/{member_id}/checklist/", response_model=MemberChecklistResponse)
async def member_onboarding_checklist_get(
    member_id: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    row = await get_member_checklist(tenant_id=tenant_id, app_key=app_key, member_id=member_id)
    return MemberChecklistResponse(**row)


@router.patch("/member-onboarding/{member_id}/checklist", response_model=MemberChecklistResponse)
@router.patch("/member-onboarding/{member_id}/checklist/", response_model=MemberChecklistResponse)
async def member_onboarding_checklist_update(
    member_id: str,
    payload: MemberChecklistUpdate,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    row = await update_member_checklist(
        tenant_id=tenant_id,
        app_key=app_key,
        member_id=member_id,
        payload=payload,
        updated_by=str(current_user.get("sub") or "system"),
    )
    return MemberChecklistResponse(**row)


@router.get("/settings/society", response_model=SocietySettingsResponse)
async def society_settings_get(
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    row = await get_society_settings(tenant_id=tenant_id, app_key=app_key)
    return SocietySettingsResponse(**row)


@router.patch("/settings/society", response_model=SocietySettingsResponse)
async def society_settings_patch(
    payload: SocietySettingsUpdate,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    row = await save_society_settings(tenant_id=tenant_id, app_key=app_key, payload=payload)
    return SocietySettingsResponse(**row)


@router.get("/flats", response_model=list[FlatResponse])
@router.get("/flats/", response_model=list[FlatResponse])
async def flats_list(
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    rows = await list_flats(tenant_id=tenant_id, app_key=app_key)
    occupants = await _flat_occupants_map(tenant_id=tenant_id, app_key=app_key)
    for row in rows:
        row["occupants"] = occupants.get(str(row.get("flat_number") or "").strip().upper(), 0)
    return [FlatResponse(**row) for row in rows]


@router.get("/flats/{flat_id}", response_model=FlatResponse)
async def flats_get(
    flat_id: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    row = await get_flat(tenant_id=tenant_id, app_key=app_key, flat_id=flat_id)
    return FlatResponse(**row)


@router.post("/flats/", response_model=FlatResponse)
async def flats_create(
    payload: FlatCreateRequest,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    row = await create_flat(tenant_id=tenant_id, app_key=app_key, payload=payload)
    return FlatResponse(**row)


@router.put("/flats/{flat_id}", response_model=FlatResponse)
async def flats_update(
    flat_id: str,
    payload: FlatUpdateRequest,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    row = await update_flat(tenant_id=tenant_id, app_key=app_key, flat_id=flat_id, payload=payload)
    return FlatResponse(**row)


@router.get("/financial-years/", response_model=list[FinancialYearResponse])
async def financial_years_list(
    include_closed: bool = True,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    rows = await list_financial_years(tenant_id=tenant_id, app_key=app_key, include_closed=include_closed)
    return [FinancialYearResponse(**row) for row in rows]


@router.get("/financial-years/active", response_model=FinancialYearResponse)
async def financial_year_active(
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    row = await get_active_financial_year(tenant_id=tenant_id, app_key=app_key)
    return FinancialYearResponse(**row)


@router.post("/financial-years/", response_model=FinancialYearResponse)
async def financial_year_create(
    payload: FinancialYearCreateRequest,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    row = await create_financial_year(tenant_id=tenant_id, app_key=app_key, payload=payload)
    return FinancialYearResponse(**row)


@router.post("/financial-years/{year_id}/provisional-close", response_model=FinancialYearResponse)
async def financial_year_provisional_close(
    year_id: str,
    payload: FinancialYearCloseRequest,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    row = await provisional_close_financial_year(tenant_id=tenant_id, app_key=app_key, year_id=year_id, payload=payload)
    return FinancialYearResponse(**row)


@router.post("/financial-years/{year_id}/final-close", response_model=FinancialYearResponse)
async def financial_year_final_close(
    year_id: str,
    payload: FinancialYearCloseRequest,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    row = await final_close_financial_year(tenant_id=tenant_id, app_key=app_key, year_id=year_id, payload=payload)
    return FinancialYearResponse(**row)


@router.post("/v2/societies/{society_id}/join-requests", response_model=PublicJoinRequestResponse)
@router.post("/v2/public/societies/{society_id}/join-requests", response_model=PublicJoinRequestResponse)
async def public_join_request_create(society_id: str, payload: PublicJoinRequestCreate):
    doc = await create_public_join_request(society_id=society_id, payload=payload)
    return PublicJoinRequestResponse(
        membership_id=doc["id"],
        society_id=doc["society_id"],
        status=doc["status"],
        message="Join request submitted. Wait for admin approval before completing registration.",
    )


@router.get("/v2/societies/search", response_model=list[SocietySearchItem])
async def societies_search(q: str | None = None, city: str | None = None, pin_code: str | None = None):
    rows = await search_societies(q=q, city=city, pin_code=pin_code)
    return [SocietySearchItem(**row) for row in rows]


@router.get("/v2/societies/{society_id}", response_model=SocietySearchItem)
@router.get("/society/{society_id}", response_model=SocietySearchItem)
async def society_get(society_id: str):
    row = await get_society(society_id=society_id)
    return SocietySearchItem(**row)


@router.get("/v2/societies/{society_id}/units")
async def society_units_list(society_id: str):
    return await list_society_units(society_id=society_id, app_key="gruhamitra")


@router.get("/v2/me/memberships", response_model=list[MembershipResponse])
async def my_memberships_list(current_user: dict = Depends(get_current_user)):
    app_key = str(current_user.get("app_key") or "gruhamitra")
    rows = await list_my_memberships(email=str(current_user.get("email") or ""), app_key=app_key)
    return [MembershipResponse(**row) for row in rows]


@router.post(
    "/v2/public/residents/complete-registration",
    response_model=CompleteResidentRegistrationResponse,
)
async def resident_complete_registration(payload: CompleteResidentRegistrationRequest):
    row = await complete_resident_registration(payload=payload)
    return CompleteResidentRegistrationResponse(**row)


@router.get("/v2/societies/{society_id}/join-requests", response_model=list[MembershipResponse])
async def join_requests_list(
    society_id: str,
    status_filter: str = "pending",
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    if tenant_id != society_id and current_user.get("role") != "super_admin":
        from fastapi import HTTPException

        raise HTTPException(status_code=403, detail="Access denied to this society")
    app_key = str(current_user.get("app_key") or "gruhamitra")
    rows = await list_join_requests(society_id=society_id, app_key=app_key, status=status_filter)
    return [MembershipResponse(**row) for row in rows]


@router.post("/v2/join-requests/{membership_id}/approve", response_model=MembershipResponse)
async def join_request_approve(
    membership_id: str,
    payload: ApproveJoinRequest,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    row = await approve_join_request(
        membership_id=membership_id,
        approver=str(current_user.get("sub") or "system"),
        tenant_scope=tenant_id,
        app_key=app_key,
        payload=payload,
    )
    return MembershipResponse(**row)


@router.post("/v2/join-requests/{membership_id}/reject", response_model=MembershipResponse)
async def join_request_reject(
    membership_id: str,
    payload: RejectJoinRequest,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    row = await reject_join_request(
        membership_id=membership_id,
        rejector=str(current_user.get("sub") or "system"),
        tenant_scope=tenant_id,
        app_key=app_key,
        payload=payload,
    )
    return MembershipResponse(**row)


@router.post("/move-governance/transfer-to-arrears", response_model=ArrearsResponse)
async def move_transfer_to_arrears(
    payload: ArrearsTransferRequest,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    row = await transfer_to_arrears(tenant_id=tenant_id, app_key=app_key, payload=payload)
    return ArrearsResponse(**row)


@router.post("/move-governance/transfer-flat-to-flat")
async def move_transfer_flat_to_flat(
    payload: FlatTransferRequest,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    return await transfer_flat_to_flat(tenant_id=tenant_id, app_key=app_key, payload=payload)


@router.get("/move-governance/personal-arrears", response_model=list[ArrearsResponse])
async def move_list_personal_arrears(
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    rows = await list_personal_arrears(tenant_id=tenant_id, app_key=app_key)
    return [ArrearsResponse(**row) for row in rows]


@router.get("/move-governance/generate-ndc/{flat_id}")
async def move_generate_ndc(
    flat_id: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    ndc = await generate_ndc(tenant_id=tenant_id, app_key=app_key, flat_id=flat_id)
    issued_at = ndc.get("issued_at")
    if isinstance(issued_at, datetime):
        issued_at_str = issued_at.strftime("%d-%m-%Y %H:%M")
    else:
        issued_at_str = str(issued_at or datetime.now(timezone.utc).strftime("%d-%m-%Y %H:%M"))
    pdf_bytes = _build_simple_pdf(
        title="No Dues Certificate (NDC)",
        top_margin=130.0,
        lines=[
            f"Flat Number: {flat_id}",
            f"Certificate Status: {ndc.get('status', 'issued')}",
            f"Issued At: {issued_at_str}",
            "",
            "This certifies that no outstanding dues were found for the above flat",
            "as on the issue date and time recorded in this certificate.",
        ],
    )
    headers = {"Content-Disposition": f'attachment; filename="NDC_{flat_id}.pdf"'}
    return StreamingResponse(BytesIO(pdf_bytes), media_type="application/pdf", headers=headers)


@router.get("/move-governance/police-verification-form/{member_id}")
@router.get("/move-governance/police-verification-form/{member_id}/")
async def move_police_verification_form(
    member_id: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    member = await get_collection("housing_members").find_one({"tenant_id": tenant_id, "app_key": app_key, "id": member_id})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    member_type = str(member.get("member_type") or "").strip().lower()
    current_occupier = "Tenant" if member_type == "tenant" else "Owner"
    owner_for_flat = None
    if member_type == "tenant":
        owner_for_flat = await get_collection("housing_members").find_one(
            {
                "tenant_id": tenant_id,
                "app_key": app_key,
                "flat_number": member.get("flat_number"),
                "member_type": "owner",
            },
            sort=[("status", 1), ("updated_at", -1), ("created_at", -1)],
        )

    move_in = member.get("move_in_date")
    if isinstance(move_in, datetime):
        move_in_str = move_in.strftime("%d-%m-%Y")
    else:
        move_in_str = str(move_in or "N/A")
    pdf_bytes = _build_simple_pdf(
        title="Owner / Tenant Police Verification Form",
        top_margin=130.0,
        lines=[
            "To,",
            "The Station House Officer,",
            "Local Police Station",
            "",
            "Subject: Police Verification Request",
            "",
            f"Flat Number: {member.get('flat_number', 'N/A')}",
            f"Current Occupier: {current_occupier}",
            f"Resident Name: {member.get('name', 'N/A')}",
            f"Member Type: {str(member.get('member_type') or '').title() or 'N/A'}",
            f"Occupation: {member.get('occupation') or 'N/A'}",
            f"Phone: {member.get('phone_number', 'N/A')}",
            f"Email: {member.get('email', 'N/A')}",
            f"Move-In Date: {move_in_str}",
            f"Total Occupants: {member.get('total_occupants') or 'N/A'}",
            "",
            "Owner Details:",
            f"Owner Name: {(owner_for_flat or member).get('name', 'N/A')}",
            f"Owner Phone: {(owner_for_flat or member).get('phone_number', 'N/A')}",
            f"Owner Email: {(owner_for_flat or member).get('email', 'N/A')}",
            "",
            "Submitted for verification and society records.",
            "",
            "Signature (Resident): ___________________    Date: __________",
            "Signature (Society Office): _____________    Date: __________",
        ],
    )
    safe_member = re.sub(r"[^A-Za-z0-9_-]+", "_", str(member.get("name") or member_id)).strip("_") or member_id
    headers = {"Content-Disposition": f'attachment; filename="Police_Verification_{safe_member}.pdf"'}
    return StreamingResponse(BytesIO(pdf_bytes), media_type="application/pdf", headers=headers)


@router.get("/move-governance/tenant-id-form/{member_id}")
@router.get("/move-governance/tenant-id-form/{member_id}/")
async def move_tenant_id_form(
    member_id: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    member = await get_collection("housing_members").find_one({"tenant_id": tenant_id, "app_key": app_key, "id": member_id})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    move_in = member.get("move_in_date")
    if isinstance(move_in, datetime):
        move_in_str = move_in.strftime("%d-%m-%Y")
    else:
        move_in_str = str(move_in or "N/A")
    member_kind = str(member.get("member_type") or "").strip().lower()
    id_title = "Tenant ID Form" if member_kind == "tenant" else "Owner ID Form"
    pdf_bytes = _build_simple_pdf(
        title=id_title,
        top_margin=130.0,
        lines=[
            f"Flat Number: {member.get('flat_number', 'N/A')}",
            f"Name: {member.get('name', 'N/A')}",
            f"Member Type: {str(member.get('member_type') or '').title() or 'N/A'}",
            f"Move-In Date: {move_in_str}",
            f"Phone: {member.get('phone_number', 'N/A')}",
            f"Email: {member.get('email', 'N/A')}",
            f"Total Occupants: {member.get('total_occupants') or 'N/A'}",
            "",
            "ID / Address Proof Details:",
            "Government ID Type: ______________________",
            "Government ID Number: ____________________",
            "Permanent Address: _______________________",
            "",
            "Emergency Contact Name: __________________",
            "Emergency Contact Number: _______________",
            "",
            "Resident Signature: ______________________",
            "Society Office Seal & Signature: _________",
        ],
    )
    safe_member = re.sub(r"[^A-Za-z0-9_-]+", "_", str(member.get("name") or member_id)).strip("_") or member_id
    headers = {"Content-Disposition": f'attachment; filename="Tenant_ID_{safe_member}.pdf"'}
    return StreamingResponse(BytesIO(pdf_bytes), media_type="application/pdf", headers=headers)


@router.get("/move-governance/calculate-final-bill/{flat_id}", response_model=FinalBillResponse)
async def move_calculate_final_bill(
    flat_id: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    row = await calculate_final_bill(tenant_id=tenant_id, app_key=app_key, flat_id=flat_id)
    return FinalBillResponse(**row)


@router.post("/move-governance/damage-claim")
async def move_damage_claim(
    payload: DamageClaimCreate,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = str(current_user.get("app_key") or "gruhamitra")
    return await raise_damage_claim(tenant_id=tenant_id, app_key=app_key, payload=payload)


@router.get("/database/backups")
async def database_backups_list(
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    col = get_collection("housing_backups")
    rows = (
        await col.find({"tenant_id": tenant_id, "app_key": app_key})
        .sort("created_at", -1)
        .limit(25)
        .to_list(length=25)
    )
    backups: list[dict[str, Any]] = []
    for row in rows:
        size_kb = row.get("size_kb")
        if size_kb is None:
            size_kb = 1
        backups.append(
            {
                "filename": str(row.get("filename") or ""),
                "created_at": row.get("created_at") or datetime.now(timezone.utc).isoformat(),
                "size_kb": int(size_kb),
            }
        )
    return backups


@router.post("/database/backup")
async def database_backup_create(
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    now = datetime.now(timezone.utc)
    created_at = now.isoformat()
    filename = f"{tenant_id}_{now.strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:8]}.json"
    payload = {
        "tenant_id": tenant_id,
        "app_key": app_key,
        "created_at": created_at,
        "kind": "metadata-only-backup",
        "note": "Compatibility backup created from GruhaMitra settings page.",
    }
    encoded_payload = json.dumps(payload, ensure_ascii=True).encode("utf-8")
    size_kb = max(1, len(encoded_payload) // 1024)
    doc = {
        "tenant_id": tenant_id,
        "app_key": app_key,
        "filename": filename,
        "created_at": created_at,
        "size_kb": size_kb,
        "payload": payload,
    }
    await get_collection("housing_backups").insert_one(doc)
    return {"message": "Backup created successfully", "filename": filename, "created_at": created_at, "size_kb": size_kb}


@router.post("/database/restore")
async def database_backup_restore(
    filename: str = Query(..., min_length=1),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    doc = await get_collection("housing_backups").find_one({"tenant_id": tenant_id, "app_key": app_key, "filename": filename})
    if not doc:
        raise HTTPException(status_code=404, detail="Backup file not found")
    return {"message": f"Restore prepared for '{filename}'. Restart backend to apply restored data."}


@router.get("/database/backups/{filename}/download")
async def database_backup_download(
    filename: str,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "gruhamitra").strip())
    doc = await get_collection("housing_backups").find_one({"tenant_id": tenant_id, "app_key": app_key, "filename": filename})
    if not doc:
        raise HTTPException(status_code=404, detail="Backup file not found")

    payload = doc.get("payload") or {
        "tenant_id": tenant_id,
        "app_key": app_key,
        "filename": filename,
        "created_at": doc.get("created_at"),
        "kind": "metadata-only-backup",
    }
    content = json.dumps(payload, ensure_ascii=True, indent=2).encode("utf-8")
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(BytesIO(content), media_type="application/octet-stream", headers=headers)
