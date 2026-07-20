
# ════════════════════════════════════════════════════════════════════════
# SECTION: MODULE HEADER + IMPORTS
# NOTE   : 9K-line MandirMitra FastAPI router. Use Ctrl+F '# SECTION:' to navigate.
# Split trigger: when a second developer joins or file exceeds 12K lines.
# ════════════════════════════════════════════════════════════════════════

from __future__ import annotations

import calendar
import json
import logging
import os
import re
from urllib.parse import urlencode
from datetime import date, datetime, timedelta, timezone
from io import BytesIO, StringIO
import csv
import httpx
from functools import lru_cache
from pathlib import Path
from typing import Any
from uuid import uuid4
from xml.sax.saxutils import escape
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, File, Header, HTTPException, Query, Request, Response, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from sqlalchemy import and_, func, select, update
from sqlalchemy.ext.asyncio import AsyncSession
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, A5
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.utils import ImageReader

try:
    from PIL import Image as PILImage
    from PIL import ImageDraw, ImageFont, features as pil_features
except Exception:
    PILImage = None
    ImageDraw = None
    ImageFont = None
    pil_features = None

try:
    from weasyprint import HTML
except Exception:
    HTML = None

from app.core.auth.dependencies import get_current_user
from app.core.audit.service import log_audit_event
from app.core.modules.dependencies import require_enabled_module
from app.core.permissions.rbac import Role, require_roles
from app.core.rate_limiting import limiter
from app.core.tenants.app_resolvers import resolve_mandir_tenant
from app.core.tenants.context import resolve_app_key, resolve_tenant_id
from app.db.mongo import get_collection
from app.db.postgres import get_async_session
from app.accounting.models.entities import Account, JournalEntry, JournalLine
from app.accounting.service import (
    AccountingValidationError,
    create_account,
    get_accounts_payable,
    get_accounts_receivable,
    get_balance_sheet,
    get_cost_centre_ledger_pl,
    get_journal_drilldown,
    get_ledger_lines,
    get_profit_loss,
    get_receipts_payments,
    get_trial_balance,
    list_accounts,
    post_journal_entry,
    reverse_journal_entry,
    validate_cash_balance_for_journal_lines,
)
from app.accounting.schemas import JournalPostRequest, JournalLineIn
from app.modules.business.dimensions import create_dimension, deactivate_dimension
from app.modules.mandir_compat.report_helpers import (
    accounts_payable_report,
    accounts_receivable_report,
    balance_sheet_report,
    bank_book_report,
    cash_book_report,
    category_income_report,
    day_book_report,
    detailed_donation_report,
    detailed_seva_report,
    donation_category_wise_report,
    donation_daily_report,
    donation_monthly_report,
    journal_entries_report,
    ledger_report,
    posted_donations,
    posted_sevas,
    profit_loss_report,
    receipts_payments_report,
    seva_schedule_report,
    top_donors_report,
    trial_balance_report,
)
from app.modules.mandir_compat.donation_compliance import (
    classify_donation_compliance,
    compliance_public_fields,
    donation_compliance_config_view,
    donation_compliance_receipt_note,
    mask_pan,
    validate_donation_compliance_config,
)
from app.modules.mandir_compat.schemas import (
    MandirFirstLoginOnboardingRequest,
    MandirFirstLoginOnboardingResponse,
)
from app.modules.mandir_compat.service import (
    create_mandir_first_login_onboarding,
    ensure_temple_numeric_id,
    list_mandir_temples,
    resolve_tenant_by_temple_id,
)
from app.services.panchang import PanchangService

router = APIRouter(tags=["mandir-compat"])
_MANDIR_WRITE_ROUTE_DEPS = [
    Depends(require_enabled_module("temple")),
    Depends(require_roles([Role.operator, Role.accountant, Role.tenant_admin, Role.super_admin])),
]
_MANDIR_ADMIN_ROUTE_DEPS = [
    Depends(require_enabled_module("temple")),
    Depends(require_roles([Role.tenant_admin, Role.super_admin])),
]
MANDIR_COMPAT_DATA_DIR = Path(__file__).resolve().parent / "data"
MANDIR_LEGACY_COA_PATH = MANDIR_COMPAT_DATA_DIR / "legacy_mandir_coa.json"
logger = logging.getLogger(__name__)

_MANDIR_COUNTERS_COLLECTION = "mandir_counters"
_MANDIR_RECEIPT_WIDTH = 7
_MANDIR_RECEIPT_PREFIX_BY_KIND = {
    "donation": "DON",
    "seva": "SEV",
}
_MANDIR_JOURNAL_ENTRY_PREFIX = "JE"
_DEFAULT_PUBLIC_DONATION_CATEGORIES = [
    {"id": "general", "name": "General Donation", "description": "General donation to the temple"},
    {"id": "annadanam", "name": "Annadanam", "description": "Sponsoring food/meals for devotees"},
    {"id": "construction", "name": "Construction Fund", "description": "Temple construction and renovation"},
    {"id": "corpus", "name": "Corpus Fund", "description": "Temple corpus fund"},
    {"id": "vastra_seva", "name": "Vastra Seva", "description": "Clothing and decoration for deity"},
    {"id": "nitya_puja", "name": "Nitya Puja", "description": "Daily worship sponsorship"},
]

_PANCHANG_DEFAULT_LOCATION = {
    "name": "Bengaluru",
    "state": "Karnataka",
    "country": "India",
    "lat": 12.9716,
    "lon": 77.5946,
    "timezone": "Asia/Kolkata",
    "display": "Bengaluru, Karnataka",
}

_PANCHANG_CITY_OPTIONS: tuple[dict[str, Any], ...] = (
    _PANCHANG_DEFAULT_LOCATION,
    {"name": "Mysuru", "state": "Karnataka", "country": "India", "lat": 12.2958, "lon": 76.6394, "timezone": "Asia/Kolkata", "display": "Mysuru, Karnataka"},
    {"name": "Mangaluru", "state": "Karnataka", "country": "India", "lat": 12.9141, "lon": 74.8560, "timezone": "Asia/Kolkata", "display": "Mangaluru, Karnataka"},
    {"name": "Udupi", "state": "Karnataka", "country": "India", "lat": 13.3409, "lon": 74.7421, "timezone": "Asia/Kolkata", "display": "Udupi, Karnataka"},
    {"name": "Chennai", "state": "Tamil Nadu", "country": "India", "lat": 13.0827, "lon": 80.2707, "timezone": "Asia/Kolkata", "display": "Chennai, Tamil Nadu"},
    {"name": "Coimbatore", "state": "Tamil Nadu", "country": "India", "lat": 11.0168, "lon": 76.9558, "timezone": "Asia/Kolkata", "display": "Coimbatore, Tamil Nadu"},
    {"name": "Madurai", "state": "Tamil Nadu", "country": "India", "lat": 9.9252, "lon": 78.1198, "timezone": "Asia/Kolkata", "display": "Madurai, Tamil Nadu"},
    {"name": "Hyderabad", "state": "Telangana", "country": "India", "lat": 17.3850, "lon": 78.4867, "timezone": "Asia/Kolkata", "display": "Hyderabad, Telangana"},
    {"name": "Tirupati", "state": "Andhra Pradesh", "country": "India", "lat": 13.6288, "lon": 79.4192, "timezone": "Asia/Kolkata", "display": "Tirupati, Andhra Pradesh"},
    {"name": "Mumbai", "state": "Maharashtra", "country": "India", "lat": 19.0760, "lon": 72.8777, "timezone": "Asia/Kolkata", "display": "Mumbai, Maharashtra"},
    {"name": "Pune", "state": "Maharashtra", "country": "India", "lat": 18.5204, "lon": 73.8567, "timezone": "Asia/Kolkata", "display": "Pune, Maharashtra"},
    {"name": "New Delhi", "state": "Delhi", "country": "India", "lat": 28.6139, "lon": 77.2090, "timezone": "Asia/Kolkata", "display": "New Delhi, Delhi"},
    {"name": "Kolkata", "state": "West Bengal", "country": "India", "lat": 22.5726, "lon": 88.3639, "timezone": "Asia/Kolkata", "display": "Kolkata, West Bengal"},
    {"name": "Ahmedabad", "state": "Gujarat", "country": "India", "lat": 23.0225, "lon": 72.5714, "timezone": "Asia/Kolkata", "display": "Ahmedabad, Gujarat"},
    {"name": "Varanasi", "state": "Uttar Pradesh", "country": "India", "lat": 25.3176, "lon": 82.9739, "timezone": "Asia/Kolkata", "display": "Varanasi, Uttar Pradesh"},
)


_MANDIR_CANONICAL_INCOME_CODES: dict[str, tuple[str, str]] = {
    'general donation': ('44001', 'General Donations'),
    'donation income': ('44001', 'General Donations'),
    'general donations': ('44001', 'General Donations'),
    'hundi collections': ('44002', 'Hundi Collections'),
    'specific purpose donation': ('44003', 'Specific Purpose Donations'),
    'specific purpose donations': ('44003', 'Specific Purpose Donations'),
    'in-kind donation income': ('44004', 'In-Kind Donation Income'),
    'in kind donation income': ('44004', 'In-Kind Donation Income'),
    'sponsorship': ('45001', 'Sponsorship Income'),
    'sponsorship income': ('45001', 'Sponsorship Income'),
    'in-kind sponsorship income': ('45002', 'In-Kind Sponsorship Income'),
    'in kind sponsorship income': ('45002', 'In-Kind Sponsorship Income'),
    'seva booking revenue': ('42002', 'Seva Income - General'),
    'pooja revenue': ('42002', 'Seva Income - General'),
    'seva income': ('42002', 'Seva Income - General'),
    'seva income - general': ('42002', 'Seva Income - General'),
}


from app.modules.mandir_compat.helpers.account_categories import (
    _MANDIR_INCOME_BUCKET_ALIASES,
    _MANDIR_INCOME_LEGACY_CODES,
    _MANDIR_LEGACY_ACCOUNT_CODE_MAP,
    _MANDIR_SPONSORSHIP_CATEGORY_MARKERS,
    _MANDIR_UTR_REFERENCE_PATTERN,
    _is_mandir_sponsorship_category,
    _mandir_cash_income_category,
    _mandir_in_kind_debit_account_target,
    _mandir_in_kind_income_category,
    _mandir_income_bucket_for_account,
    _normalize_income_category,
    _normalize_mandir_account_code,
    _normalize_public_payment_utr_reference,
)

# ════════════════════════════════════════════════════════════════════════
# SECTION: ACCOUNT CODE + CATEGORY HELPERS
# NOTE   : moved to helpers/account_categories.py; imported above.
# ════════════════════════════════════════════════════════════════════════

# SECTION: ASYNC ACCOUNT RESOLVERS
# NOTE   : _normalize_mandir_income_accounts, _resolve_mandir_income_account, _resolve_or_create_mandir_account, _mandir_inventory_accounting_enabled, _resolve_mandir_in_kind_debit_account, _resolve_mandir_payment_account_id
# ════════════════════════════════════════════════════════════════════════

async def _normalize_mandir_income_accounts(session: AsyncSession, tenant_id: str) -> dict[str, int]:
    canonical_targets = {
        'donation': ('44001', 'General Donations'),
        'seva': ('42002', 'Seva Income - General'),
    }

    accounts = await list_accounts(session, tenant_id=tenant_id)
    income_accounts = [acc for acc in accounts if str(acc.type or '').strip().lower() == 'income']

    canonical_by_bucket: dict[str, Account] = {}
    dirty = False
    remapped_lines = 0

    for bucket, (target_code, target_name) in canonical_targets.items():
        canonical = next((acc for acc in income_accounts if str(acc.code or '').strip() == target_code), None)

        if canonical is None:
            candidate = next(
                (
                    acc
                    for acc in income_accounts
                    if _mandir_income_bucket_for_account(acc.name, acc.code) == bucket
                ),
                None,
            )
            if candidate is not None:
                candidate.code = target_code
                candidate.name = target_name
                candidate.type = 'income'
                candidate.classification = 'nominal'
                canonical = candidate
                dirty = True
            else:
                canonical = await create_account(
                    session,
                    tenant_id=tenant_id,
                    code=target_code,
                    name=target_name,
                    account_type='income',
                    classification='nominal',
                    is_cash_bank=False,
                    is_receivable=False,
                    is_payable=False,
                )
                accounts = await list_accounts(session, tenant_id=tenant_id)
                income_accounts = [acc for acc in accounts if str(acc.type or '').strip().lower() == 'income']

        canonical_by_bucket[bucket] = canonical

    for bucket, canonical in canonical_by_bucket.items():
        duplicate_ids = [
            int(acc.id)
            for acc in income_accounts
            if int(acc.id) != int(canonical.id)
            and _mandir_income_bucket_for_account(acc.name, acc.code) == bucket
        ]
        if not duplicate_ids:
            continue

        tenant_journal_ids = select(JournalEntry.id).where(JournalEntry.tenant_id == tenant_id)
        remap_stmt = (
            update(JournalLine)
            .where(
                JournalLine.account_id.in_(duplicate_ids),
                JournalLine.journal_id.in_(tenant_journal_ids),
            )
            .values(account_id=int(canonical.id))
        )
        result = await session.execute(remap_stmt)
        changed = int(result.rowcount or 0)
        if changed > 0:
            remapped_lines += changed
            dirty = True

    if dirty:
        await session.commit()

    return {'remapped_lines': remapped_lines}


async def _resolve_mandir_income_account(session: AsyncSession, tenant_id: str, category_name: str) -> int:
    normalized_category = _normalize_income_category(category_name)
    preferred_code, preferred_name = _MANDIR_CANONICAL_INCOME_CODES.get(
        normalized_category,
        ('42002', 'Seva Income - General') if any(token in normalized_category for token in ('seva', 'pooja')) else ('44001', 'General Donations'),
    )

    await _normalize_mandir_income_accounts(session, tenant_id)

    accounts = await list_accounts(session, tenant_id=tenant_id)
    for acc in accounts:
        if str(acc.type or '').strip().lower() == 'income' and str(acc.code or '').strip() == preferred_code:
            return int(acc.id)

    new_acc = await create_account(
        session,
        tenant_id=tenant_id,
        code=preferred_code,
        name=preferred_name,
        account_type='income',
        classification='nominal',
        is_cash_bank=False,
        is_receivable=False,
        is_payable=False,
    )
    return int(new_acc.id)


async def _resolve_or_create_mandir_account(
    session: AsyncSession,
    tenant_id: str,
    *,
    code: str,
    name: str,
    account_type: str,
    classification: str,
) -> int:
    accounts = await list_accounts(session, tenant_id=tenant_id)
    for acc in accounts:
        if str(acc.code or "").strip() == str(code).strip():
            return int(acc.id)
    new_acc = await create_account(
        session,
        tenant_id=tenant_id,
        code=code,
        name=name,
        account_type=account_type,
        classification=classification,
        is_cash_bank=False,
        is_receivable=False,
        is_payable=False,
    )
    return int(new_acc.id)


async def _mandir_inventory_accounting_enabled(tenant_id: str, app_key: str) -> bool:
    query = {"tenant_id": tenant_id, "app_key": app_key}
    doc = await get_collection("mandir_temples").find_one(query) or {}
    return bool(doc.get("module_inventory_enabled", False))


async def _resolve_mandir_in_kind_debit_account(
    session: AsyncSession,
    tenant_id: str,
    payload: dict[str, Any],
    category_name: Any,
    *,
    app_key: str = "mandirmitra",
) -> int:
    inventory_enabled = await _mandir_inventory_accounting_enabled(tenant_id, app_key)
    code, name, account_type = _mandir_in_kind_debit_account_target(
        payload,
        category_name,
        inventory_accounting_enabled=inventory_enabled,
    )
    classification = "nominal" if account_type == "expense" else "real"
    return await _resolve_or_create_mandir_account(
        session,
        tenant_id,
        code=code,
        name=name,
        account_type=account_type,
        classification=classification,
    )


async def _resolve_mandir_payment_account_id(
    session: AsyncSession,
    tenant_id: str,
    raw_account_id: Any,
    payment_mode: str | None,
) -> int | None:
    raw_value = str(raw_account_id).strip() if raw_account_id is not None else ""

    if raw_value:
        maybe_id = _safe_optional_int(raw_value)
        if maybe_id:
            by_id_stmt = select(Account.id).where(
                Account.tenant_id == tenant_id,
                Account.id == maybe_id,
            )
            by_id = (await session.execute(by_id_stmt)).scalar_one_or_none()
            if by_id is not None:
                return int(by_id)

        code_candidate = raw_value
        if " - " in raw_value:
            code_candidate = raw_value.split(" - ", 1)[0].strip()
        code_candidate = _normalize_mandir_account_code(code_candidate)

        if code_candidate.isdigit():
            by_code_stmt = select(Account.id).where(
                Account.tenant_id == tenant_id,
                Account.code == code_candidate,
            )
            by_code = (await session.execute(by_code_stmt)).scalar_one_or_none()
            if by_code is not None:
                return int(by_code)

    accounts = await list_accounts(session, tenant_id=tenant_id)
    mode = str(payment_mode or "").strip().lower()

    if mode == "cash":
        for preferred_code in ("11001",):
            preferred = next(
                (
                    acc
                    for acc in accounts
                    if acc.is_cash_bank and str(acc.code or "").strip() == preferred_code
                ),
                None,
            )
            if preferred is not None:
                return int(preferred.id)
        for acc in accounts:
            if acc.is_cash_bank and "cash" in str(acc.name).lower():
                return int(acc.id)
    elif mode == "bank":
        for preferred_code in ("12001",):
            preferred = next(
                (
                    acc
                    for acc in accounts
                    if acc.is_cash_bank and str(acc.code or "").strip() == preferred_code
                ),
                None,
            )
            if preferred is not None:
                return int(preferred.id)
        for acc in accounts:
            if acc.is_cash_bank and "bank" in str(acc.name).lower():
                return int(acc.id)

    for acc in accounts:
        if acc.is_cash_bank:
            return int(acc.id)

    return None


def _mandir_actor_id(current_user: dict[str, Any]) -> str:
    return str(
        current_user.get("sub")
        or current_user.get("user_id")
        or current_user.get("email")
        or "mandir_compat_system"
    )



# ════════════════════════════════════════════════════════════════════════
# SECTION: RECEIPT CANCELLATION HELPERS
# NOTE   : _mandir_actor_id, _mandir_receipt_cancellation_metadata, _reverse_mandir_source_journal, _cancel_mandir_receipt_source
# ════════════════════════════════════════════════════════════════════════

def _mandir_receipt_cancellation_metadata(payload: dict[str, Any] | None, current_user: dict[str, Any]) -> dict[str, Any]:
    payload = payload if isinstance(payload, dict) else {}
    reason = str(
        payload.get("reason")
        or payload.get("cancellation_reason")
        or payload.get("refund_reason")
        or ""
    ).strip()
    if len(reason) < 3:
        raise HTTPException(status_code=400, detail="Cancellation reason is required")
    now = datetime.now(timezone.utc).isoformat()
    refund_mode = _safe_optional_str(payload.get("refund_mode") or payload.get("refund_payment_mode"))
    refund_reference = _safe_optional_str(payload.get("refund_reference") or payload.get("refund_utr") or payload.get("refund_transaction_reference"))
    refund_date = _safe_optional_str(payload.get("refund_date")) or now[:10]
    return {
        "status": "reversed",
        "cancellation_reason": reason,
        "refund_mode": refund_mode,
        "refund_reference": refund_reference,
        "refund_date": refund_date if refund_mode or refund_reference else None,
        "reversed_at": now,
        "cancelled_at": now,
        "cancelled_by": _mandir_actor_id(current_user),
        "updated_at": now,
    }


async def _reverse_mandir_source_journal(
    session: AsyncSession,
    *,
    tenant_id: str,
    app_key: str,
    source_key: str,
    reason: str,
    current_user: dict[str, Any],
) -> tuple[JournalEntry, bool]:
    stmt = select(JournalEntry).where(
        JournalEntry.tenant_id == tenant_id,
        JournalEntry.app_key == app_key,
        JournalEntry.accounting_entity_id == "primary",
        JournalEntry.idempotency_key == source_key,
    )
    original = (await session.execute(stmt)).scalar_one_or_none()
    if original is None:
        raise HTTPException(status_code=404, detail="Original accounting journal was not found for this receipt")
    try:
        return await reverse_journal_entry(
            session=session,
            tenant_id=tenant_id,
            app_key=app_key,
            accounting_entity_id="primary",
            journal_id=int(original.id),
            created_by=_mandir_actor_id(current_user),
            reason=reason,
            idempotency_key=f"{source_key}_receipt_reversal",
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to reverse receipt journal: {exc}") from exc


async def _cancel_mandir_receipt_source(
    *,
    source_kind: str,
    source_id: str,
    collection_name: str,
    id_field: str,
    idempotency_prefix: str,
    payload: dict[str, Any] | None,
    session: AsyncSession,
    current_user: dict[str, Any],
    x_tenant_id: str | None,
    x_app_key: str | None,
) -> dict[str, Any]:
    tenant_context = resolve_mandir_tenant(
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
        operation=f"{source_kind} receipt cancellation",
    )
    tenant_id = tenant_context.tenant_id
    app_key = tenant_context.app_key
    collection = get_collection(collection_name)
    existing = await collection.find_one({id_field: str(source_id), "tenant_id": tenant_id, "app_key": app_key})
    if existing is None:
        raise HTTPException(status_code=404, detail=f"{source_kind.title()} receipt not found")

    current_status = str(existing.get("status") or "").strip().lower()
    if current_status in {"cancelled", "reversed"}:
        view = _mandir_donation_view(existing) if source_kind == "donation" else _mandir_seva_booking_view(existing)
        view["_idempotent"] = True
        return view

    patch = _mandir_receipt_cancellation_metadata(payload, current_user)
    if source_kind == "donation" and current_status == "pending_valuation":
        patch["valuation_status"] = "cancelled"
        await collection.update_one(
            {id_field: str(source_id), "tenant_id": tenant_id, "app_key": app_key},
            {"$set": patch}, upsert=False,
        )
        return _mandir_donation_view({**existing, **patch})

    inventory_reversal = None
    movement_collection = None
    if source_kind == "donation" and existing.get("inventory_movement_id"):
        movement_collection = get_collection("mandir_inventory_movements")
        receipt_movement = await movement_collection.find_one(
            {"id": str(existing["inventory_movement_id"]), "tenant_id": tenant_id, "app_key": app_key, "status": "posted"}
        )
        item = await get_collection("mandir_inventory_items").find_one(
            {"id": str(existing.get("inventory_item_id") or ""), "tenant_id": tenant_id, "app_key": app_key, "is_active": True}
        )
        if receipt_movement is None or item is None:
            raise HTTPException(status_code=409, detail="Inventory receipt evidence is unavailable for donation reversal")
        available = await _mandir_inventory_item_balance(tenant_id=tenant_id, app_key=app_key, item=item)
        receipt_quantity = Decimal(str(receipt_movement.get("quantity") or "0"))
        if available < receipt_quantity:
            raise HTTPException(status_code=409, detail="Cannot reverse donation after its inventory has been consumed")
        reversal_movement_id = f"donation-reversal-{source_id}"
        inventory_reversal = await movement_collection.find_one(
            {"id": reversal_movement_id, "tenant_id": tenant_id, "app_key": app_key}
        )
        if inventory_reversal is None:
            inventory_reversal = {
                "id": reversal_movement_id, "tenant_id": tenant_id, "app_key": app_key,
                "item_id": receipt_movement.get("item_id"), "item_name": receipt_movement.get("item_name"),
                "movement_type": "receipt_reversal", "quantity": str(receipt_quantity),
                "unit_value": str(receipt_movement.get("unit_value") or "0"),
                "total_value": str(receipt_movement.get("total_value") or "0"),
                "movement_date": datetime.now(timezone.utc).date().isoformat(),
                "source_type": "in_kind_donation_reversal", "source_id": str(source_id),
                "status": "pending_accounting", "created_by": _mandir_actor_id(current_user),
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await movement_collection.insert_one(inventory_reversal)
    reversal_entry, _created = await _reverse_mandir_source_journal(
        session,
        tenant_id=tenant_id,
        app_key=app_key,
        source_key=f"{idempotency_prefix}{source_id}",
        reason=str(patch["cancellation_reason"]),
        current_user=current_user,
    )
    patch["reversal_journal_id"] = int(reversal_entry.id)
    patch["reversal_idempotency_key"] = f"{idempotency_prefix}{source_id}_receipt_reversal"
    if inventory_reversal is not None and movement_collection is not None:
        await movement_collection.update_one(
            {"id": inventory_reversal["id"], "tenant_id": tenant_id, "app_key": app_key},
            {"$set": {
                "status": "posted", "journal_entry_id": int(reversal_entry.id),
                "posted_at": datetime.now(timezone.utc).isoformat(),
            }},
            upsert=False,
        )
        patch["inventory_reversal_movement_id"] = inventory_reversal["id"]

    await collection.update_one(
        {id_field: str(source_id), "tenant_id": tenant_id, "app_key": app_key},
        {"$set": patch},
        upsert=False,
    )
    try:
        await log_audit_event(
            tenant_id=tenant_id,
            user_id=_mandir_actor_id(current_user),
            product=app_key,
            action=f"{source_kind}_receipt_cancelled",
            entity_type=f"mandir_{source_kind}_receipt",
            entity_id=str(source_id),
            old_value={"status": existing.get("status"), "receipt_number": existing.get("receipt_number")},
            new_value={
                "status": patch["status"],
                "reason": patch["cancellation_reason"],
                "reversal_journal_id": patch["reversal_journal_id"],
                "refund_mode": patch.get("refund_mode"),
                "refund_reference": patch.get("refund_reference"),
            },
        )
    except Exception:
        logger.warning("Failed to write audit event for %s receipt cancellation %s", source_kind, source_id, exc_info=True)

    updated = {**existing, **patch}
    return _mandir_donation_view(updated) if source_kind == "donation" else _mandir_seva_booking_view(updated)


# Refund-request routes moved to routes/refunds.py
# (docs/operations/LARGE_FILE_MODULARIZATION_PLAN.md); registered via import at end of module.


# Default COA seed + legacy import helpers moved to helpers/*.py
# (docs/operations/LARGE_FILE_MODULARIZATION_PLAN.md); re-exported at end of module.

# ════════════════════════════════════════════════════════════════════════
# SECTION: SAFE TYPE COERCIONS
# NOTE   : _safe_float, _safe_optional_float, _safe_optional_int, _safe_bool, _safe_optional_str, _parse_opening_balance_decimal
# ════════════════════════════════════════════════════════════════════════

def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except Exception:
        return default


def _safe_optional_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except Exception:
        return None


def _safe_optional_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except Exception:
        return None


def _safe_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return value != 0

    raw = str(value).strip().lower()
    if raw in {"true", "1", "yes", "y", "on"}:
        return True
    if raw in {"false", "0", "no", "n", "off", ""}:
        return False
    return default


def _safe_optional_str(value: Any) -> str | None:
    if value is None:
        return None
    raw = str(value).strip()
    return raw if raw else None


def _parse_opening_balance_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    cleaned = raw.replace(',', '')
    if cleaned.startswith('(') and cleaned.endswith(')'):
        cleaned = f"-{cleaned[1:-1]}"
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError, TypeError):
        raise ValueError(f"Invalid amount value: {value}")



# ════════════════════════════════════════════════════════════════════════
# SECTION: OPENING BALANCE HELPERS
# NOTE   : _parse_opening_balance_rows, _find_or_create_opening_balance_offset_account, _current_opening_balance_net
# ════════════════════════════════════════════════════════════════════════

async def _parse_opening_balance_rows(file: UploadFile) -> list[dict[str, Any]]:
    filename = str(file.filename or '').strip()
    if not filename:
        raise HTTPException(status_code=400, detail='File name is required')

    suffix = Path(filename).suffix.lower()
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail='Uploaded file is empty')

    if suffix == '.csv':
        try:
            text = raw.decode('utf-8-sig')
        except UnicodeDecodeError as exc:
            raise HTTPException(status_code=400, detail='Unable to decode CSV file as UTF-8') from exc
        rows = list(csv.DictReader(StringIO(text)))
        return [row for row in rows if isinstance(row, dict)]

    if suffix in {'.xlsx', '.xlsm'}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise HTTPException(
                status_code=500,
                detail='XLSX import support is unavailable on server (missing openpyxl dependency)',
            ) from exc

        try:
            workbook = load_workbook(BytesIO(raw), data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail='Unable to read XLSX workbook') from exc

        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []

        headers = [str(cell).strip() if cell is not None else '' for cell in values[0]]
        parsed_rows: list[dict[str, Any]] = []
        for row in values[1:]:
            if not any(cell not in (None, '') for cell in row):
                continue
            item: dict[str, Any] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                if index < len(row):
                    item[header] = row[index]
            if item:
                parsed_rows.append(item)
        return parsed_rows

    raise HTTPException(status_code=400, detail='Unsupported file format. Use .csv or .xlsx')


async def _find_or_create_opening_balance_offset_account(session: AsyncSession, tenant_id: str) -> Account:
    stmt = select(Account).where(
        Account.tenant_id == tenant_id,
        Account.code == '33001',
    )
    existing = (await session.execute(stmt)).scalar_one_or_none()
    if existing is not None:
        return existing

    try:
        return await create_account(
            session,
            tenant_id=tenant_id,
            code='33001',
            name='Opening Balance',
            account_type='equity',
            classification='real',
            is_cash_bank=False,
            is_receivable=False,
            is_payable=False,
        )
    except IntegrityError:
        await session.rollback()
        existing = (await session.execute(stmt)).scalar_one_or_none()
        if existing is None:
            raise
        return existing


async def _current_opening_balance_net(
    session: AsyncSession,
    *,
    tenant_id: str,
    account_id: int,
    reference: str,
) -> Decimal:
    stmt = (
        select(
            func.coalesce(func.sum(JournalLine.debit), 0).label('debit_total'),
            func.coalesce(func.sum(JournalLine.credit), 0).label('credit_total'),
        )
        .join(JournalEntry, JournalEntry.id == JournalLine.journal_id)
        .where(
            JournalEntry.tenant_id == tenant_id,
            JournalEntry.reference == reference,
            JournalLine.account_id == account_id,
        )
    )
    row = (await session.execute(stmt)).one()
    debit_total = Decimal(str(row.debit_total or 0))
    credit_total = Decimal(str(row.credit_total or 0))
    return debit_total - credit_total


# Legacy COA import helpers moved to helpers/legacy_coa.py

def _sanitize_mongo_doc(doc: dict[str, Any]) -> dict[str, Any]:
    row = dict(doc or {})
    # ObjectId is not JSON serializable; hide Mongo internals from API clients.
    row.pop("_id", None)
    return row



# ════════════════════════════════════════════════════════════════════════
# SECTION: SEQUENCE NUMBERS
# NOTE   : _format_mandir_receipt_number, _format_mandir_sequence_number, _next_receipt_number, _next_journal_entry_number, _receipt_number_for_donation, _sanitize_mongo_doc
# ════════════════════════════════════════════════════════════════════════

def _format_mandir_receipt_number(prefix: str, sequence: int) -> str:
    normalized_prefix = str(prefix or "").strip().upper()
    if normalized_prefix not in {"DON", "SEV"}:
        raise ValueError(f"Unsupported MandirMitra receipt prefix: {prefix!r}")
    if int(sequence) < 1:
        raise ValueError("Receipt sequence must be positive")
    return f"{normalized_prefix}-{int(sequence):0{_MANDIR_RECEIPT_WIDTH}d}"


def _format_mandir_sequence_number(prefix: str, sequence: int) -> str:
    normalized_prefix = str(prefix or "").strip().upper()
    if not normalized_prefix:
        raise ValueError("Sequence prefix is required")
    if int(sequence) < 1:
        raise ValueError("Sequence must be positive")
    return f"{normalized_prefix}-{int(sequence):0{_MANDIR_RECEIPT_WIDTH}d}"


async def _next_receipt_number(
    *,
    tenant_id: str,
    app_key: str,
    receipt_kind: str,
    receipt_date: Any = None,
) -> str:
    prefix = _MANDIR_RECEIPT_PREFIX_BY_KIND.get(str(receipt_kind or "").strip().lower())
    if not prefix:
        raise ValueError(f"Unsupported MandirMitra receipt kind: {receipt_kind!r}")

    now = datetime.now(timezone.utc).isoformat()
    counter_id = f"{app_key}:{tenant_id}:receipt:{prefix}"
    counters = get_collection(_MANDIR_COUNTERS_COLLECTION)
    result = await counters.find_one_and_update(
        {"_id": counter_id},
        {
            "$inc": {"seq": 1},
            "$set": {"updated_at": now},
            "$setOnInsert": {
                "app_key": app_key,
                "tenant_id": tenant_id,
                "prefix": prefix,
                "kind": receipt_kind,
                "created_at": now,
            },
        },
        upsert=True,
        return_document=True,
    )
    sequence = int((result or {}).get("seq") or 1)
    return _format_mandir_receipt_number(prefix, sequence)


async def _next_journal_entry_number(*, tenant_id: str, app_key: str) -> str:
    now = datetime.now(timezone.utc).isoformat()
    counter_id = f"{app_key}:{tenant_id}:journal-entry:{_MANDIR_JOURNAL_ENTRY_PREFIX}"
    counters = get_collection(_MANDIR_COUNTERS_COLLECTION)
    result = await counters.find_one_and_update(
        {"_id": counter_id},
        {
            "$inc": {"seq": 1},
            "$set": {"updated_at": now},
            "$setOnInsert": {
                "app_key": app_key,
                "tenant_id": tenant_id,
                "prefix": _MANDIR_JOURNAL_ENTRY_PREFIX,
                "kind": "journal_entry",
                "created_at": now,
            },
        },
        upsert=True,
        return_document=True,
    )
    sequence = int((result or {}).get("seq") or 1)
    return _format_mandir_sequence_number(_MANDIR_JOURNAL_ENTRY_PREFIX, sequence)



def _receipt_number_for_donation(doc: dict[str, Any]) -> str:
    existing = str(doc.get("receipt_number") or "").strip()
    if existing:
        return existing

    donation_id = str(doc.get("donation_id") or doc.get("id") or "").strip()
    if donation_id:
        return f"DON-{donation_id[:8].upper()}"

    return f"DON-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"



# ════════════════════════════════════════════════════════════════════════
# SECTION: DONATION VIEW + ROW FILTERING
# NOTE   : _mandir_donation_view, _mandir_row_date_text, _mandir_row_matches_search, _mandir_filter_rows
# ════════════════════════════════════════════════════════════════════════

def _mandir_donation_view(doc: dict[str, Any]) -> dict[str, Any]:
    row = _sanitize_mongo_doc(doc)
    row.pop("donor_pan", None)
    row.update(compliance_public_fields(doc))
    donation_id = str(row.get("donation_id") or row.get("id") or "").strip()
    if donation_id and not str(row.get("id") or "").strip():
        row["id"] = donation_id
    if donation_id and not str(row.get("donation_id") or "").strip():
        row["donation_id"] = donation_id

    receipt_number = _receipt_number_for_donation(row)
    row["receipt_number"] = receipt_number
    if donation_id:
        row["receipt_pdf_url"] = f"/api/v1/donations/{donation_id}/receipt/pdf"
    if not row.get("donation_date") and row.get("created_at"):
        row["donation_date"] = row.get("created_at")
    return row


def _mandir_row_date_text(row: dict[str, Any], fields: tuple[str, ...]) -> str:
    for field in fields:
        value = str(row.get(field) or "").strip()
        if value:
            return value[:10]
    return ""


def _mandir_row_matches_search(row: dict[str, Any], query: str, fields: tuple[str, ...]) -> bool:
    normalized_query = query.strip().lower()
    if not normalized_query:
        return True
    return any(normalized_query in str(row.get(field) or "").lower() for field in fields)


def _mandir_filter_rows(
    rows: list[dict[str, Any]],
    *,
    q: str | None,
    from_date: date | None,
    to_date: date | None,
    date_fields: tuple[str, ...],
    search_fields: tuple[str, ...],
) -> list[dict[str, Any]]:
    filtered: list[dict[str, Any]] = []
    from_text = from_date.isoformat() if from_date else None
    to_text = to_date.isoformat() if to_date else None
    for row in rows:
        row_date = _mandir_row_date_text(row, date_fields)
        if from_text and row_date and row_date < from_text:
            continue
        if to_text and row_date and row_date > to_text:
            continue
        if q and not _mandir_row_matches_search(row, q, search_fields):
            continue
        filtered.append(row)
    return filtered



# ════════════════════════════════════════════════════════════════════════
# SECTION: DONATION RECEIPT PDF (ReportLab)
# NOTE   : _generate_donation_receipt_pdf_bytes
# ════════════════════════════════════════════════════════════════════════


# ════════════════════════════════════════════════════════════════════════
# SECTION: DONATION RECEIPT PDF (ReportLab)
# NOTE   : PDF builders moved to receipt_pdf.py; re-exported at end of module.
# SECTION: SEVA RECEIPT PDF (ReportLab) + SEVA BOOKING VIEW
# NOTE   : _receipt_number_for_seva and _mandir_seva_booking_view remain below.
# ════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════
# SECTION: SEVA RECEIPT PDF (ReportLab) + SEVA BOOKING VIEW
# NOTE   : _receipt_number_for_seva, _mandir_seva_booking_view, _generate_seva_receipt_pdf_bytes
# ════════════════════════════════════════════════════════════════════════

def _receipt_number_for_seva(doc: dict[str, Any]) -> str:
    existing = str(doc.get("receipt_number") or "").strip()
    if existing:
        return existing

    booking_id = str(doc.get("id") or doc.get("booking_id") or "").strip()
    if booking_id:
        return f"SEV-{booking_id[:8].upper()}"

    return f"SEV-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"


def _mandir_seva_booking_view(doc: dict[str, Any]) -> dict[str, Any]:
    row = _sanitize_mongo_doc(doc)
    booking_id = str(row.get("id") or row.get("booking_id") or "").strip()
    if booking_id and not str(row.get("id") or "").strip():
        row["id"] = booking_id
    receipt_number = _receipt_number_for_seva(row)
    row["receipt_number"] = receipt_number
    if booking_id:
        row["receipt_pdf_url"] = f"/api/v1/sevas/bookings/{booking_id}/receipt/pdf"
    return row



# ════════════════════════════════════════════════════════════════════════
# SECTION: PINCODE + DATE-WINDOW HELPERS
# NOTE   : _normalize_pincode, _lookup_pincode_city_state, _to_positive_int
# ════════════════════════════════════════════════════════════════════════

def _normalize_pincode(value: Any) -> str:
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    return digits[:6]


MANDIR_PINCODE_FALLBACKS: dict[str, tuple[str, str]] = {
    "560001": ("Bengaluru", "Karnataka"),
    "560002": ("Bengaluru", "Karnataka"),
    "560003": ("Bengaluru", "Karnataka"),
    "560004": ("Bengaluru", "Karnataka"),
    "560011": ("Bengaluru", "Karnataka"),
    "560019": ("Bengaluru", "Karnataka"),
    "560070": ("Bengaluru", "Karnataka"),
    "560078": ("Bengaluru", "Karnataka"),
    "575001": ("Mangaluru", "Karnataka"),
    "575002": ("Mangaluru", "Karnataka"),
    "575003": ("Mangaluru", "Karnataka"),
    "575004": ("Mangaluru", "Karnataka"),
    "575005": ("Mangaluru", "Karnataka"),
    "600004": ("Chennai", "Tamil Nadu"),
    "600017": ("Chennai", "Tamil Nadu"),
    "600020": ("Chennai", "Tamil Nadu"),
    "600028": ("Chennai", "Tamil Nadu"),
    "500034": ("Hyderabad", "Telangana"),
    "530017": ("Visakhapatnam", "Andhra Pradesh"),
    "641002": ("Coimbatore", "Tamil Nadu"),
}


async def _lookup_pincode_city_state(pincode: str) -> tuple[str | None, str | None]:
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            response = await client.get(f"https://api.postalpincode.in/pincode/{pincode}")
        response.raise_for_status()
        payload = response.json()
    except Exception:
        return MANDIR_PINCODE_FALLBACKS.get(pincode, (None, None))

    if not isinstance(payload, list) or not payload:
        return MANDIR_PINCODE_FALLBACKS.get(pincode, (None, None))

    first = payload[0] if isinstance(payload[0], dict) else {}
    if str(first.get("Status") or "").strip().lower() != "success":
        return MANDIR_PINCODE_FALLBACKS.get(pincode, (None, None))

    offices = first.get("PostOffice")
    if not isinstance(offices, list) or not offices:
        return MANDIR_PINCODE_FALLBACKS.get(pincode, (None, None))

    primary = offices[0] if isinstance(offices[0], dict) else {}
    city = str(primary.get("District") or primary.get("Taluk") or primary.get("Name") or "").strip() or None
    state = str(primary.get("State") or "").strip() or None
    return (city, state) if city and state else MANDIR_PINCODE_FALLBACKS.get(pincode, (city, state))


def _to_positive_int(value: Any) -> int | None:
    parsed = _safe_optional_int(value)
    if parsed is None or parsed <= 0:
        return None
    return parsed


_SEVA_ALLOWED_CATEGORIES = {
    "abhisheka",
    "alankara",
    "pooja",
    "archana",
    "vahana_seva",
    "special",
    "festival",
}
_SEVA_ALLOWED_AVAILABILITY = {
    "daily",
    "weekday",
    "weekend",
    "specific_day",
    "except_day",
    "festival_only",
}



# ════════════════════════════════════════════════════════════════════════
# SECTION: SEVA BOOKING HELPERS
# NOTE   : _normalize_seva_category/availability/day, _today_weekday, _parse_booking_date, _validate_seva_booking_date, _count_seva_bookings_for_date, _validate_seva_booking_capacity, _compute_seva_available_today, _resolve_report_date_window, _resolve_export_window
# ════════════════════════════════════════════════════════════════════════

def _normalize_seva_category(value: Any) -> str:
    candidate = str(value or "pooja").strip().lower()
    return candidate if candidate in _SEVA_ALLOWED_CATEGORIES else "pooja"


def _normalize_seva_availability(value: Any) -> str:
    candidate = str(value or "daily").strip().lower()
    return candidate if candidate in _SEVA_ALLOWED_AVAILABILITY else "daily"


def _normalize_seva_day(value: Any) -> int | None:
    parsed = _safe_optional_int(value)
    if parsed is None:
        return None
    if 0 <= parsed <= 6:
        return parsed
    return None


_IST_TIMEZONE = ZoneInfo("Asia/Kolkata")


def _today_weekday_js_index() -> int:
    # JavaScript Date.getDay convention: Sunday=0 ... Saturday=6.
    return (datetime.now(_IST_TIMEZONE).weekday() + 1) % 7


def _weekday_js_index_for_date(value: date) -> int:
    # JavaScript Date.getDay convention: Sunday=0 ... Saturday=6.
    return (value.weekday() + 1) % 7


def _parse_booking_date(value: Any) -> date | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    normalized = raw.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(normalized).date()
    except Exception:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(raw[:10], fmt).date()
            except Exception:
                continue
    return None


def _validate_seva_booking_date(seva_doc: dict[str, Any] | None, booking_date: date) -> None:
    if not seva_doc:
        return

    target_day = _weekday_js_index_for_date(booking_date)
    specific_day = _normalize_seva_day(seva_doc.get("specific_day"))
    except_day = _normalize_seva_day(seva_doc.get("except_day"))
    day_names = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

    if specific_day is not None and target_day != specific_day:
        raise HTTPException(
            status_code=400,
            detail=f"This seva is available only on {day_names[specific_day]}. Please select a {day_names[specific_day]} date.",
        )
    if except_day is not None and target_day == except_day:
        raise HTTPException(
            status_code=400,
            detail=f"This seva is not available on {day_names[except_day]}. Please select another date.",
        )

    availability = _normalize_seva_availability(seva_doc.get("availability"))
    if availability == "weekday" and target_day not in {1, 2, 3, 4, 5}:
        raise HTTPException(status_code=400, detail="This seva is available only on weekdays.")
    if availability == "weekend" and target_day not in {0, 6}:
        raise HTTPException(status_code=400, detail="This seva is available only on weekends.")
    if availability == "festival_only":
        raise HTTPException(status_code=400, detail="This seva is available only on configured festival dates.")


async def _count_seva_bookings_for_date(
    *,
    tenant_id: str,
    app_key: str,
    seva_id: str,
    booking_date: date,
) -> int:
    booking_date_text = booking_date.isoformat()
    docs = await get_collection("mandir_seva_bookings").find(
        {
            "tenant_id": tenant_id,
            "app_key": app_key,
            "seva_id": str(seva_id),
            "booking_date": booking_date_text,
        }
    ).to_list(length=5000)
    return sum(
        1
        for doc in docs
        if str(doc.get("status") or "").strip().lower() not in {"cancelled", "canceled"}
    )


async def _validate_seva_booking_capacity(
    seva_doc: dict[str, Any] | None,
    *,
    tenant_id: str,
    app_key: str,
    seva_id: str,
    booking_date: date,
) -> tuple[int | None, int, int | None]:
    max_bookings = _safe_optional_int((seva_doc or {}).get("max_bookings_per_day"))
    booked_count = await _count_seva_bookings_for_date(
        tenant_id=tenant_id,
        app_key=app_key,
        seva_id=seva_id,
        booking_date=booking_date,
    )
    if max_bookings is None or max_bookings <= 0:
        return None, booked_count, None

    slots_left = max(max_bookings - booked_count, 0)
    if slots_left <= 0:
        raise HTTPException(
            status_code=400,
            detail=f"This seva is fully booked for {booking_date.strftime('%d-%m-%Y')}. Please select another date.",
        )
    return max_bookings, booked_count, slots_left


def _compute_seva_available_today(row: dict[str, Any]) -> bool:
    if not _safe_bool(row.get("is_active"), True):
        return False

    slots_left = _safe_optional_int(row.get("bookings_available"))
    if slots_left is not None and slots_left <= 0:
        return False

    today = _today_weekday_js_index()
    specific_day = _normalize_seva_day(row.get("specific_day"))
    except_day = _normalize_seva_day(row.get("except_day"))

    # Explicit day constraints are authoritative even if availability is stale.
    if specific_day is not None:
        return specific_day == today
    if except_day is not None:
        return except_day != today

    availability = _normalize_seva_availability(row.get("availability"))
    if availability == "weekday":
        return 1 <= today <= 5
    if availability == "weekend":
        return today in {0, 6}
    if availability == "festival_only":
        return False
    return True

def _resolve_report_date_window(
    *,
    from_date: date | None,
    to_date: date | None,
    single_date: date | None = None,
    month: int | None = None,
    year: int | None = None,
) -> tuple[date, date]:
    if single_date is not None:
        return single_date, single_date

    if from_date is not None and to_date is not None:
        if from_date > to_date:
            raise HTTPException(status_code=422, detail="from_date cannot be greater than to_date")
        return from_date, to_date

    if month is not None or year is not None:
        resolved_year = year or datetime.now(timezone.utc).year
        if month is None:
            start = date(resolved_year, 1, 1)
            end = date(resolved_year, 12, 31)
            return start, end

        _, last_day = calendar.monthrange(resolved_year, month)
        start = date(resolved_year, month, 1)
        end = date(resolved_year, month, last_day)
        return start, end

    if from_date is not None and to_date is None:
        return from_date, from_date
    if to_date is not None and from_date is None:
        return to_date, to_date

    raise HTTPException(
        status_code=422,
        detail="Provide either date, from_date/to_date, or month/year query parameters",
    )


def _resolve_export_window(
    *,
    from_date: date | None,
    to_date: date | None,
    date_from: date | None,
    date_to: date | None,
) -> tuple[date, date]:
    start = from_date or date_from
    end = to_date or date_to
    if start is None or end is None:
        raise HTTPException(status_code=422, detail="from_date/to_date (or date_from/date_to) are required")
    if start > end:
        raise HTTPException(status_code=422, detail="from_date cannot be greater than to_date")
    return start, end



# ════════════════════════════════════════════════════════════════════════
# SECTION: DASHBOARD STATS + SEVA BUILDERS
# NOTE   : _dashboard_posted_stats, _canonical_seva_name, _build_seva_item, _build_seva_patch, _serialize_seva_doc, _seva_import_template_csv, _normalize_phone, _parse_iso_datetime
# ════════════════════════════════════════════════════════════════════════

async def _dashboard_posted_stats(
    *,
    session: AsyncSession,
    tenant_id: str,
    app_key: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    today = datetime.now(timezone.utc).date()
    start_of_year = date(today.year, 1, 1)
    try:
        donations = await posted_donations(
            session,
            tenant_id=tenant_id,
            app_key=app_key,
            from_date=start_of_year,
            to_date=today,
        )
    except Exception as exc:
        logger.warning("Dashboard: failed to fetch posted donations for tenant=%s: %s", tenant_id, exc)
        donations = []

    try:
        sevas = await posted_sevas(
            session,
            tenant_id=tenant_id,
            app_key=app_key,
            from_date=start_of_year,
            to_date=today,
        )
    except Exception as exc:
        logger.warning("Dashboard: failed to fetch posted sevas for tenant=%s: %s", tenant_id, exc)
        sevas = []

    return donations, sevas

def _canonical_seva_name(payload: dict[str, Any]) -> str:
    name = str(payload.get("name_english") or payload.get("name") or payload.get("seva_name") or "Seva").strip()
    return name or "Seva"


def _build_seva_item(payload: dict[str, Any], *, tenant_id: str, app_key: str) -> dict[str, Any]:
    now = datetime.now(timezone.utc).isoformat()
    name = _canonical_seva_name(payload)
    advance_days = _safe_optional_int(payload.get("advance_booking_days"))

    return {
        "id": str(uuid4()),
        "tenant_id": tenant_id,
        "app_key": app_key,
        "name": name,
        "name_english": name,
        "name_kannada": _safe_optional_str(payload.get("name_kannada")) or "",
        "name_sanskrit": _safe_optional_str(payload.get("name_sanskrit")) or "",
        "description": _safe_optional_str(payload.get("description")) or "",
        "category": _normalize_seva_category(payload.get("category")),
        "amount": _safe_float(payload.get("amount"), 0.0),
        "min_amount": _safe_optional_float(payload.get("min_amount")),
        "max_amount": _safe_optional_float(payload.get("max_amount")),
        "availability": _normalize_seva_availability(payload.get("availability")),
        "specific_day": _normalize_seva_day(payload.get("specific_day")),
        "except_day": _normalize_seva_day(payload.get("except_day")),
        "time_slot": _safe_optional_str(payload.get("time_slot")) or "",
        "max_bookings_per_day": _safe_optional_int(payload.get("max_bookings_per_day")),
        "advance_booking_days": advance_days if advance_days and advance_days > 0 else 30,
        "requires_approval": _safe_bool(payload.get("requires_approval"), False),
        "is_active": _safe_bool(payload.get("is_active"), True),
        "quick_ticket_enabled": _safe_bool(payload.get("quick_ticket_enabled"), False),
        "requires_devotee_details": _safe_bool(payload.get("requires_devotee_details"), True),
        "benefits": _safe_optional_str(payload.get("benefits")) or "",
        "instructions": _safe_optional_str(payload.get("instructions")) or "",
        "duration_minutes": _safe_optional_int(payload.get("duration_minutes")),
        "created_at": now,
        "updated_at": now,
    }


def _build_seva_patch(payload: dict[str, Any]) -> dict[str, Any]:
    patch: dict[str, Any] = {}

    if {"name", "name_english", "seva_name"} & payload.keys():
        name = _canonical_seva_name(payload)
        patch["name"] = name
        patch["name_english"] = name

    if "name_kannada" in payload:
        patch["name_kannada"] = _safe_optional_str(payload.get("name_kannada")) or ""
    if "name_sanskrit" in payload:
        patch["name_sanskrit"] = _safe_optional_str(payload.get("name_sanskrit")) or ""
    if "description" in payload:
        patch["description"] = _safe_optional_str(payload.get("description")) or ""
    if "category" in payload:
        patch["category"] = _normalize_seva_category(payload.get("category"))
    if "amount" in payload:
        patch["amount"] = _safe_float(payload.get("amount"), 0.0)
    if "min_amount" in payload:
        patch["min_amount"] = _safe_optional_float(payload.get("min_amount"))
    if "max_amount" in payload:
        patch["max_amount"] = _safe_optional_float(payload.get("max_amount"))
    if "availability" in payload:
        patch["availability"] = _normalize_seva_availability(payload.get("availability"))
    if "specific_day" in payload:
        patch["specific_day"] = _normalize_seva_day(payload.get("specific_day"))
    if "except_day" in payload:
        patch["except_day"] = _normalize_seva_day(payload.get("except_day"))
    if "time_slot" in payload:
        patch["time_slot"] = _safe_optional_str(payload.get("time_slot")) or ""
    if "max_bookings_per_day" in payload:
        patch["max_bookings_per_day"] = _safe_optional_int(payload.get("max_bookings_per_day"))
    if "advance_booking_days" in payload:
        days = _safe_optional_int(payload.get("advance_booking_days"))
        patch["advance_booking_days"] = days if days and days > 0 else 30
    if "requires_approval" in payload:
        patch["requires_approval"] = _safe_bool(payload.get("requires_approval"), False)
    if "is_active" in payload:
        patch["is_active"] = _safe_bool(payload.get("is_active"), True)
    if "quick_ticket_enabled" in payload:
        patch["quick_ticket_enabled"] = _safe_bool(payload.get("quick_ticket_enabled"), False)
    if "requires_devotee_details" in payload:
        patch["requires_devotee_details"] = _safe_bool(payload.get("requires_devotee_details"), True)
    if "benefits" in payload:
        patch["benefits"] = _safe_optional_str(payload.get("benefits")) or ""
    if "instructions" in payload:
        patch["instructions"] = _safe_optional_str(payload.get("instructions")) or ""
    if "duration_minutes" in payload:
        patch["duration_minutes"] = _safe_optional_int(payload.get("duration_minutes"))

    return patch


def _serialize_seva_doc(doc: dict[str, Any]) -> dict[str, Any]:
    row = dict(doc)
    row.pop("_id", None)

    name = str(row.get("name_english") or row.get("name") or row.get("seva_name") or "Seva").strip() or "Seva"
    row["name_english"] = name
    row["name"] = name
    row["category"] = _normalize_seva_category(row.get("category"))
    row["availability"] = _normalize_seva_availability(row.get("availability"))
    row["amount"] = _safe_float(row.get("amount"), 0.0)
    row["min_amount"] = _safe_optional_float(row.get("min_amount"))
    row["max_amount"] = _safe_optional_float(row.get("max_amount"))
    row["specific_day"] = _normalize_seva_day(row.get("specific_day"))
    row["except_day"] = _normalize_seva_day(row.get("except_day"))
    row["max_bookings_per_day"] = _safe_optional_int(row.get("max_bookings_per_day"))
    row["bookings_available"] = _safe_optional_int(row.get("bookings_available"))
    row["duration_minutes"] = _safe_optional_int(row.get("duration_minutes"))
    row["advance_booking_days"] = _safe_optional_int(row.get("advance_booking_days")) or 30
    row["requires_approval"] = _safe_bool(row.get("requires_approval"), False)
    row["is_active"] = _safe_bool(row.get("is_active"), True)
    row["quick_ticket_enabled"] = _safe_bool(row.get("quick_ticket_enabled"), False)
    row["requires_devotee_details"] = _safe_bool(row.get("requires_devotee_details"), True)
    row["is_available_today"] = _compute_seva_available_today(row)
    row["description"] = _safe_optional_str(row.get("description")) or ""
    row["name_kannada"] = _safe_optional_str(row.get("name_kannada")) or ""
    row["name_sanskrit"] = _safe_optional_str(row.get("name_sanskrit")) or ""
    row["time_slot"] = _safe_optional_str(row.get("time_slot")) or ""
    row["benefits"] = _safe_optional_str(row.get("benefits")) or ""
    row["instructions"] = _safe_optional_str(row.get("instructions")) or ""
    row["id"] = str(row.get("id") or row.get("seva_id") or "")

    return row


_SEVA_IMPORT_COLUMNS = [
    "name_english",
    "name_kannada",
    "name_sanskrit",
    "description",
    "category",
    "amount",
    "min_amount",
    "max_amount",
    "availability",
    "specific_day",
    "except_day",
    "time_slot",
    "max_bookings_per_day",
    "advance_booking_days",
    "requires_approval",
    "is_active",
    "benefits",
    "instructions",
    "duration_minutes",
]


def _seva_import_template_csv() -> str:
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=_SEVA_IMPORT_COLUMNS)
    writer.writeheader()
    writer.writerow(
        {
            "name_english": "Daily Archana",
            "name_kannada": "ದೈನಿಕ ಆರ್ಚನೆ",
            "name_sanskrit": "दैनिक आरचना",
            "description": "Daily morning archana seva",
            "category": "archana",
            "amount": "50",
            "min_amount": "",
            "max_amount": "",
            "availability": "daily",
            "specific_day": "",
            "except_day": "",
            "time_slot": "Morning 6:00 AM",
            "max_bookings_per_day": "",
            "advance_booking_days": "30",
            "requires_approval": "false",
            "is_active": "true",
            "benefits": "",
            "instructions": "",
            "duration_minutes": "",
        }
    )
    writer.writerow(
        {
            "name_english": "Sarva Seva",
            "name_kannada": "ಸರ್ವ ಸೇವೆ",
            "name_sanskrit": "सर्व सेवा",
            "description": "Comprehensive daily worship services",
            "category": "pooja",
            "amount": "500",
            "min_amount": "",
            "max_amount": "",
            "availability": "daily",
            "specific_day": "",
            "except_day": "",
            "time_slot": "Daily",
            "max_bookings_per_day": "",
            "advance_booking_days": "30",
            "requires_approval": "false",
            "is_active": "true",
            "benefits": "",
            "instructions": "",
            "duration_minutes": "",
        }
    )
    return output.getvalue()

def _normalize_phone(phone: str | None) -> str:
    digits = "".join(ch for ch in str(phone or "") if ch.isdigit())
    return digits[-10:] if len(digits) > 10 else digits


def _parse_iso_datetime(value: Any) -> datetime | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc)
    except Exception:
        return None



# ════════════════════════════════════════════════════════════════════════
# SECTION: DEVOTEE + UPI HELPERS
# NOTE   : _upi_receipt_number, _mandir_upi_payment_view, _find_devotee_by_phone, _upsert_devotee_from_contribution, _build_upi_intent_uri
# ════════════════════════════════════════════════════════════════════════

def _upi_receipt_number(doc: dict[str, Any]) -> str:
    existing = str(doc.get("receipt_number") or "").strip()
    if existing:
        return existing

    row_id = str(doc.get("id") or "").strip()
    if row_id:
        return f"UPI-{row_id[:8].upper()}"

    return f"UPI-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"


def _mandir_upi_payment_view(doc: dict[str, Any]) -> dict[str, Any]:
    row = _sanitize_mongo_doc(doc)
    row_id = str(row.get("id") or "").strip()
    if row_id:
        row["id"] = row_id

    row["amount"] = _safe_float(row.get("amount"), 0.0)
    row["payment_purpose"] = str(row.get("payment_purpose") or "DONATION").strip().upper()
    row["receipt_number"] = _upi_receipt_number(row)
    if not row.get("payment_datetime"):
        row["payment_datetime"] = row.get("created_at")

    row["sender_phone"] = _normalize_phone(row.get("sender_phone") or row.get("devotee_phone"))
    row["devotee_phone"] = _normalize_phone(row.get("devotee_phone") or row.get("sender_phone"))
    return row


async def _find_devotee_by_phone(
    tenant_id: str,
    app_key: str,
    phone: str,
    temple_id: int | None = None,
) -> dict[str, Any] | None:
    normalized = _normalize_phone(phone)
    if not normalized:
        return None

    scope_query: dict[str, Any] = {"tenant_id": tenant_id, "app_key": app_key}
    if temple_id is not None and temple_id > 0:
        scope_query["temple_id"] = temple_id

    legacy_scope_query: dict[str, Any] | None = None
    if temple_id is not None and temple_id > 0:
        # Older MandirMitra donation/seva-derived devotee data was tenant-scoped
        # but did not persist temple_id. Keep this fallback inside the resolved
        # tenant/app and only match records with no explicit temple assignment.
        legacy_scope_query = {"tenant_id": tenant_id, "app_key": app_key, "temple_id": None}

    devotee_col = get_collection("mandir_devotees")
    docs = await devotee_col.find({**scope_query, "phone": normalized}).limit(1).to_list(length=1)
    if docs:
        return _sanitize_mongo_doc(docs[0])
    if legacy_scope_query is not None:
        docs = await devotee_col.find({**legacy_scope_query, "phone": normalized}).limit(1).to_list(length=1)
        if docs:
            devotee = _sanitize_mongo_doc(docs[0])
            devotee["temple_id"] = temple_id
            return devotee

    donation_col = get_collection("mandir_donations")
    donation_docs = await (
        donation_col.find({**scope_query, "devotee_phone": normalized})
        .sort("created_at", -1)
        .limit(1)
        .to_list(length=1)
    )
    if donation_docs:
        donation = _sanitize_mongo_doc(donation_docs[0])
        snapshot = donation.get("devotee") if isinstance(donation.get("devotee"), dict) else {}
        return {
            "id": str(snapshot.get("id") or donation.get("devotee_id") or f"donation:{donation.get('donation_id') or donation.get('id')}"),
            "tenant_id": tenant_id,
            "app_key": app_key,
            "temple_id": temple_id,
            "name_prefix": donation.get("devotee_prefix") or snapshot.get("name_prefix"),
            "name": donation.get("devotee_name") or snapshot.get("name") or "Devotee",
            "first_name": donation.get("devotee_name") or snapshot.get("first_name") or snapshot.get("name") or "Devotee",
            "last_name": snapshot.get("last_name") or "",
            "phone": normalized,
            "email": donation.get("devotee_email") or snapshot.get("email"),
            "address": donation.get("devotee_address") or snapshot.get("address"),
            "city": donation.get("devotee_city") or snapshot.get("city"),
            "state": donation.get("devotee_state") or snapshot.get("state"),
            "pincode": donation.get("devotee_pincode") or snapshot.get("pincode"),
            "source": "donation",
        }
    if legacy_scope_query is not None:
        donation_docs = await (
            donation_col.find({**legacy_scope_query, "devotee_phone": normalized})
            .sort("created_at", -1)
            .limit(1)
            .to_list(length=1)
        )
        if donation_docs:
            donation = _sanitize_mongo_doc(donation_docs[0])
            snapshot = donation.get("devotee") if isinstance(donation.get("devotee"), dict) else {}
            return {
                "id": str(snapshot.get("id") or donation.get("devotee_id") or f"donation:{donation.get('donation_id') or donation.get('id')}"),
                "tenant_id": tenant_id,
                "app_key": app_key,
                "temple_id": temple_id,
                "name_prefix": donation.get("devotee_prefix") or snapshot.get("name_prefix"),
                "name": donation.get("devotee_name") or snapshot.get("name") or "Devotee",
                "first_name": donation.get("devotee_name") or snapshot.get("first_name") or snapshot.get("name") or "Devotee",
                "last_name": snapshot.get("last_name") or "",
                "phone": normalized,
                "email": donation.get("devotee_email") or snapshot.get("email"),
                "address": donation.get("devotee_address") or snapshot.get("address"),
                "city": donation.get("devotee_city") or snapshot.get("city"),
                "state": donation.get("devotee_state") or snapshot.get("state"),
                "pincode": donation.get("devotee_pincode") or snapshot.get("pincode"),
                "source": "donation",
            }

    seva_col = get_collection("mandir_seva_bookings")
    seva_docs = await (
        seva_col.find({**scope_query, "devotee_phone": normalized})
        .sort("created_at", -1)
        .limit(1)
        .to_list(length=1)
    )
    if seva_docs:
        seva = _sanitize_mongo_doc(seva_docs[0])
        snapshot = seva.get("devotee") if isinstance(seva.get("devotee"), dict) else {}
        return {
            "id": str(snapshot.get("id") or seva.get("devotee_id") or f"seva:{seva.get('booking_id') or seva.get('id')}"),
            "tenant_id": tenant_id,
            "app_key": app_key,
            "temple_id": temple_id,
            "name_prefix": seva.get("devotee_prefix") or snapshot.get("name_prefix"),
            "name": seva.get("devotee_name") or seva.get("devotee_names") or snapshot.get("name") or "Devotee",
            "first_name": seva.get("devotee_name") or seva.get("devotee_names") or snapshot.get("first_name") or snapshot.get("name") or "Devotee",
            "last_name": snapshot.get("last_name") or "",
            "phone": normalized,
            "email": seva.get("devotee_email") or snapshot.get("email"),
            "address": seva.get("devotee_address") or seva.get("address") or snapshot.get("address"),
            "city": seva.get("devotee_city") or seva.get("city") or snapshot.get("city"),
            "state": seva.get("devotee_state") or seva.get("state") or snapshot.get("state"),
            "pincode": seva.get("devotee_pincode") or seva.get("pincode") or snapshot.get("pincode"),
            "source": "seva",
        }
    if legacy_scope_query is not None:
        seva_docs = await (
            seva_col.find({**legacy_scope_query, "devotee_phone": normalized})
            .sort("created_at", -1)
            .limit(1)
            .to_list(length=1)
        )
        if seva_docs:
            seva = _sanitize_mongo_doc(seva_docs[0])
            snapshot = seva.get("devotee") if isinstance(seva.get("devotee"), dict) else {}
            return {
                "id": str(snapshot.get("id") or seva.get("devotee_id") or f"seva:{seva.get('booking_id') or seva.get('id')}"),
                "tenant_id": tenant_id,
                "app_key": app_key,
                "temple_id": temple_id,
                "name_prefix": seva.get("devotee_prefix") or snapshot.get("name_prefix"),
                "name": seva.get("devotee_name") or seva.get("devotee_names") or snapshot.get("name") or "Devotee",
                "first_name": seva.get("devotee_name") or seva.get("devotee_names") or snapshot.get("first_name") or snapshot.get("name") or "Devotee",
                "last_name": snapshot.get("last_name") or "",
                "phone": normalized,
                "email": seva.get("devotee_email") or snapshot.get("email"),
                "address": seva.get("devotee_address") or seva.get("address") or snapshot.get("address"),
                "city": seva.get("devotee_city") or seva.get("city") or snapshot.get("city"),
                "state": seva.get("devotee_state") or seva.get("state") or snapshot.get("state"),
                "pincode": seva.get("devotee_pincode") or seva.get("pincode") or snapshot.get("pincode"),
                "source": "seva",
            }

    return None


async def _upsert_devotee_from_contribution(
    tenant_id: str,
    app_key: str,
    *,
    temple_id: int | None = None,
    phone: str | None,
    name_prefix: str | None = None,
    name: str | None = None,
    email: str | None = None,
    address: str | None = None,
    city: str | None = None,
    state: str | None = None,
    pincode: str | None = None,
) -> None:
    normalized = _normalize_phone(phone)
    if not normalized:
        return

    now = datetime.now(timezone.utc).isoformat()
    devotee_patch: dict[str, Any] = {"updated_at": now}
    if temple_id is not None and temple_id > 0:
        devotee_patch["temple_id"] = temple_id
    for key, value in {
        "name_prefix": name_prefix,
        "name": name,
        "first_name": name,
        "email": email,
        "address": address,
        "city": city,
        "state": state,
        "pincode": pincode,
    }.items():
        if value not in (None, ""):
            devotee_patch[key] = value

    col = get_collection("mandir_devotees")
    query: dict[str, Any] = {"tenant_id": tenant_id, "app_key": app_key, "phone": normalized}
    if temple_id is not None and temple_id > 0:
        query["temple_id"] = temple_id

    insert_doc: dict[str, Any] = {
        "id": str(uuid4()),
        "tenant_id": tenant_id,
        "app_key": app_key,
        "phone": normalized,
        "created_at": now,
    }
    if temple_id is not None and temple_id > 0:
        insert_doc["temple_id"] = temple_id

    await col.update_one(
        query,
        {
            "$setOnInsert": insert_doc,
            "$set": devotee_patch,
        },
        upsert=True,
    )


def _build_upi_intent_uri(
    *,
    upi_id: str,
    payee_name: str,
    amount: float | None,
    note: str | None,
    reference: str | None,
    currency: str = "INR",
) -> str:
    params: list[tuple[str, str]] = [("pa", upi_id), ("pn", payee_name), ("cu", currency)]
    if amount is not None and amount > 0:
        params.append(("am", f"{amount:.2f}"))
    if note:
        params.append(("tn", note))
    if reference:
        params.append(("tr", reference))
    return f"upi://pay?{urlencode(params)}"



# ════════════════════════════════════════════════════════════════════════
# SECTION: PLATFORM + TENANT RESOLVERS
# NOTE   : _is_platform_super_admin, _resolve_tenant_for_mandir_request, _assert_platform_can_write_tenant, _payment_accounts
# ════════════════════════════════════════════════════════════════════════

def _is_platform_super_admin(user: dict[str, Any]) -> bool:
    return bool(user.get("is_superuser")) or str(user.get("role") or "").strip().lower() == "super_admin"


async def _resolve_tenant_for_mandir_request(
    current_user: dict[str, Any],
    x_tenant_id: str | None,
    temple_id: int | None,
    app_key: str = "mandirmitra",
) -> str:
    if temple_id and _is_platform_super_admin(current_user):
        mapped_tenant_id = await resolve_tenant_by_temple_id(temple_id, app_key=app_key)
        if mapped_tenant_id:
            return mapped_tenant_id
    return resolve_tenant_id(current_user, x_tenant_id)


async def _assert_platform_can_write_tenant(
    current_user: dict[str, Any],
    *,
    tenant_id: str,
    app_key: str,
) -> None:
    if not _is_platform_super_admin(current_user):
        return

    doc = await get_collection("mandir_temples").find_one({"tenant_id": tenant_id, "app_key": app_key}) or {}
    if not bool(doc.get("platform_can_write", False)):
        tenant_name = str(doc.get("name") or doc.get("trust_name") or "Selected tenant").strip()
        raise HTTPException(
            status_code=403,
            detail=f"{tenant_name} is read-only for the platform administrator.",
        )


async def _payment_accounts(tenant_id: str, app_key: str) -> dict[str, list[dict[str, Any]]]:
    cash_accounts: list[dict[str, Any]] = []
    bank_accounts: list[dict[str, Any]] = []
    seen_cash_codes: set[str] = set()
    seen_bank_codes: set[str] = set()

    try:
        accounts = get_collection("accounting_accounts")
        await _ensure_default_mandir_accounts(tenant_id, app_key)
        docs = await accounts.find({"tenant_id": tenant_id, "app_key": app_key, "is_active": True}).to_list(length=200)
        for doc in docs:
            item = _mandir_account_view(doc)
            account_code = str(item.get("account_code") or "").strip()
            # Mandir COA uses 5-digit numeric account codes.
            if account_code.isdigit() and len(account_code) < 5:
                continue

            account_type = item["account_type"].lower()
            cash_bank_nature = str(item.get("cash_bank_nature") or "").lower()
            name = str(item.get("account_name") or "").lower()
            if cash_bank_nature == "cash" or account_type in {"cash", "cash_in_hand"} or ("cash" in name and item.get("is_cash_bank")):
                if account_code and account_code in seen_cash_codes:
                    continue
                cash_accounts.append(item)
                if account_code:
                    seen_cash_codes.add(account_code)
            elif cash_bank_nature == "bank" or account_type in {"bank", "bank_account", "current_asset"} or ("bank" in name and item.get("is_cash_bank")):
                if account_code and account_code in seen_bank_codes:
                    continue
                bank_accounts.append(item)
                if account_code:
                    seen_bank_codes.add(account_code)
    except Exception:
        pass
    if not cash_accounts:
        cash_accounts = [{
            "id": "cash-main",
            "account_id": "cash-main",
            "account_code": "11001",
            "account_name": "Cash in Hand - Counter",
            "account_type": "asset",
            "cash_bank_nature": "cash",
            "is_cash_bank": True,
            "is_active": True,
            "sub_accounts": [],
        }]
    return {"cash_accounts": cash_accounts, "bank_accounts": bank_accounts}



# Dashboard stats route moved to routes/dashboard.py
# (docs/operations/LARGE_FILE_MODULARIZATION_PLAN.md); registered via import at end of module.


def _panchang_city_options() -> list[dict[str, Any]]:
    return [dict(city) for city in _PANCHANG_CITY_OPTIONS]


def _safe_panchang_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _resolve_panchang_location(
    settings_doc: dict[str, Any],
    temple_doc: dict[str, Any],
    *,
    city_name: str | None = None,
    latitude: float | str | None = None,
    longitude: float | str | None = None,
) -> tuple[float, float, str]:
    override_lat = _safe_panchang_float(latitude)
    override_lon = _safe_panchang_float(longitude)
    if override_lat is not None and override_lon is not None:
        return override_lat, override_lon, str(city_name or "Selected Location").strip() or "Selected Location"

    settings_lat = _safe_panchang_float(settings_doc.get("latitude"))
    settings_lon = _safe_panchang_float(settings_doc.get("longitude"))
    settings_city = str(settings_doc.get("city_name") or "").strip()
    if settings_lat is not None and settings_lon is not None:
        return settings_lat, settings_lon, settings_city or str(_PANCHANG_DEFAULT_LOCATION["name"])

    temple_lat = _safe_panchang_float(temple_doc.get("latitude"))
    temple_lon = _safe_panchang_float(temple_doc.get("longitude"))
    temple_city = str(temple_doc.get("city") or temple_doc.get("city_name") or "").strip()
    if temple_lat is not None and temple_lon is not None:
        return temple_lat, temple_lon, temple_city or str(_PANCHANG_DEFAULT_LOCATION["name"])

    return (
        float(_PANCHANG_DEFAULT_LOCATION["lat"]),
        float(_PANCHANG_DEFAULT_LOCATION["lon"]),
        str(_PANCHANG_DEFAULT_LOCATION["name"]),
    )



# Panchang (today) route moved to routes/panchang.py
# (docs/operations/LARGE_FILE_MODULARIZATION_PLAN.md); registered via import at end of module.


# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: Donations (GET list / POST / cancel / receipt PDF / reconcile / cleanup / export / daily|monthly reports)
# ROUTES : GET /donations/payment-accounts|categories  GET|POST /donations  GET .../receipt/pdf  POST .../cancel|reconcile-posting  DELETE .../cleanup  GET .../report/daily|monthly  GET .../export/excel|pdf
# ════════════════════════════════════════════════════════════════════════

# Donation read routes moved to routes/donations_read.py
# (docs/operations/LARGE_FILE_MODULARIZATION_PLAN.md); registered via import at end of module.

@router.post("/donations", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
@router.post("/donations/", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
async def create_donation(
    payload: dict[str, Any],
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
    x_temple_id: str | None = Header(default=None, alias="X-Temple-Id"),
):
    tenant_context = resolve_mandir_tenant(
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
        operation="donation creation",
    )
    tenant_id = tenant_context.tenant_id
    app_key = tenant_context.app_key
    temple_id = _to_positive_int(x_temple_id)
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id, raise_on_failure=True)

    donation_id = str(uuid4())
    now = datetime.now(timezone.utc).isoformat()
    devotee_phone = _normalize_phone(payload.get("devotee_phone") or payload.get("phone"))

    amount = _safe_float(payload.get("amount"), 0.0)
    category = str(payload.get("category") or "General Donation")
    payment_mode = str(payload.get("payment_mode") or "Cash").lower()
    donation_type = str(payload.get("donation_type") or "cash").strip().lower() or "cash"
    if donation_type == "in_kind":
        if amount <= 0:
            raise HTTPException(status_code=400, detail="Declared in-kind value must be greater than zero")
        if not _safe_optional_str(payload.get("in_kind_item_name") or payload.get("item_name")):
            raise HTTPException(status_code=400, detail="In-kind item name is required")
        if not _safe_optional_str(payload.get("in_kind_valuation_basis") or payload.get("valuation_basis")):
            raise HTTPException(status_code=400, detail="In-kind valuation basis is required")
    fund = None
    fund_id = str(payload.get("fund_id") or "").strip()
    if fund_id:
        fund = await get_collection("mandir_funds").find_one(
            {"id": fund_id, "tenant_id": tenant_id, "app_key": app_key, "active": True}
        )
        if fund is None:
            raise HTTPException(status_code=404, detail="Active fund not found")
        if not fund.get("accounting_dimension_id"):
            raise HTTPException(status_code=409, detail="Fund accounting dimension is not provisioned")
    festival = None
    festival_id = str(payload.get("festival_id") or "").strip()
    if festival_id:
        festival = await get_collection("mandir_festivals").find_one(
            {"id": festival_id, "tenant_id": tenant_id, "app_key": app_key, "active": True}
        )
        if festival is None:
            raise HTTPException(status_code=404, detail="Active festival not found")
    sponsorship_value = payload.get("is_sponsorship", False)
    is_sponsorship = sponsorship_value is True or str(sponsorship_value).strip().lower() in {"1", "true", "yes"}
    cash_income_category = _mandir_cash_income_category(category)
    if fund and str(fund.get("fund_type") or "").lower() in {"restricted", "corpus"}:
        cash_income_category = "Specific Purpose Donations"
    if is_sponsorship:
        cash_income_category = "Sponsorship Income"
    devotee_prefix = _safe_optional_str(
        payload.get("name_prefix")
        or payload.get("devotee_prefix")
        or payload.get("prefix")
        or payload.get("title")
        or payload.get("salutation")
    )
    devotee_name = str(payload.get("devotee_name") or payload.get("first_name") or "Unknown Devotee").strip() or "Unknown Devotee"
    devotee_address = _safe_optional_str(payload.get("devotee_address") or payload.get("address"))
    devotee_city = _safe_optional_str(payload.get("devotee_city") or payload.get("city"))
    devotee_state = _safe_optional_str(payload.get("devotee_state") or payload.get("state"))
    devotee_pincode = _safe_optional_str(payload.get("devotee_pincode") or payload.get("pincode"))
    raw_payment_account_id = _safe_optional_str(payload.get("bank_account_id") or payload.get("payment_account_id"))
    compliance_config = await get_collection("mandir_donation_compliance_config").find_one(
        {"tenant_id": tenant_id, "app_key": app_key}
    )
    compliance = classify_donation_compliance(
        payload,
        compliance_config,
        amount=Decimal(str(amount)),
        donation_type=donation_type,
        payment_mode=payment_mode,
        donation_date=datetime.now(timezone.utc).date(),
        payment_account_id=raw_payment_account_id,
    )

    donation = {
        "donation_id": donation_id,
        "tenant_id": tenant_id,
        "app_key": app_key,
        "temple_id": temple_id,
        "amount": amount,
        "category": category,
        "donation_type": donation_type,
        "in_kind_item_name": _safe_optional_str(payload.get("in_kind_item_name") or payload.get("item_name")),
        "in_kind_item_type": _safe_optional_str(payload.get("in_kind_item_type") or payload.get("item_type") or payload.get("asset_type")),
        "in_kind_quantity": _safe_optional_str(payload.get("in_kind_quantity") or payload.get("quantity")),
        "in_kind_valuation_basis": _safe_optional_str(payload.get("in_kind_valuation_basis") or payload.get("valuation_basis")),
        "event_name": _safe_optional_str(payload.get("event_name") or payload.get("festival_name")),
        "fund_id": fund_id or None,
        "fund_name": str(fund.get("name")) if fund else None,
        "fund_type": str(fund.get("fund_type")) if fund else None,
        "fund_dimension_id": str(fund.get("accounting_dimension_id")) if fund else None,
        "festival_id": festival_id or None,
        "festival_name": str(festival.get("name")) if festival else None,
        "is_sponsorship": is_sponsorship,
        "income_category": cash_income_category,
        "status": "pending_valuation" if donation_type == "in_kind" else "posted",
        "valuation_status": "pending_approval" if donation_type == "in_kind" else None,
        "inventory_item_id": _safe_optional_str(payload.get("inventory_item_id")),
        "inventory_quantity": _safe_optional_str(payload.get("inventory_quantity")),
        "payment_mode": payload.get("payment_mode") or "Cash",
        "devotee_name": devotee_name,
        "devotee_phone": devotee_phone,
        "devotee_prefix": devotee_prefix,
        "devotee_address": devotee_address,
        "devotee_city": devotee_city,
        "devotee_state": devotee_state,
        "devotee_pincode": devotee_pincode,
        "devotee": {
            "name": devotee_name,
            "name_prefix": devotee_prefix,
            "phone": devotee_phone,
            "email": str(payload.get("email") or "") or None,
            "address": devotee_address,
            "city": devotee_city,
            "state": devotee_state,
            "pincode": devotee_pincode,
        },
        "created_by": _mandir_actor_id(current_user),
        "created_at": now,
        **compliance,
    }

    donation["id"] = donation_id
    donation["receipt_number"] = await _next_receipt_number(
        tenant_id=tenant_id,
        app_key=app_key,
        receipt_kind="donation",
        receipt_date=now,
    )
    donation["receipt_pdf_url"] = f"/api/v1/donations/{donation_id}/receipt/pdf"

    col = get_collection("mandir_donations")
    try:
        await col.insert_one(donation)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to save donation: {exc}") from exc

    # Valued donations and sponsorships must post into accounting; otherwise reports and TB diverge.
    if amount > 0 and donation_type != "in_kind":
        try:
            raw_account_id = raw_payment_account_id
            debit_account_id = await _resolve_mandir_payment_account_id(
                session,
                tenant_id,
                raw_account_id,
                payment_mode,
            )
            if not debit_account_id:
                await col.delete_one({"donation_id": donation_id, "tenant_id": tenant_id, "app_key": app_key})
                raise HTTPException(status_code=400, detail="No valid cash/bank account is configured for donation posting")
            income_category = cash_income_category

            income_acc_id = await _resolve_mandir_income_account(session, tenant_id, income_category)
            journal_payload = JournalPostRequest(
                entry_date=datetime.now(timezone.utc).date(),
                description=f"{category} from {donation['devotee']['name']}",
                reference=donation["receipt_number"],
                source_module="mandirmitra",
                source_document_type="donation",
                source_document_id=donation_id,
                lines=[
                    JournalLineIn(account_id=debit_account_id, debit=Decimal(str(amount)), credit=Decimal("0")),
                    JournalLineIn(
                        account_id=income_acc_id, debit=Decimal("0"), credit=Decimal(str(amount)),
                        cost_center_id=str(fund.get("accounting_dimension_id")) if fund else None,
                    ),
                ],
            )
            await post_journal_entry(
                session=session,
                app_key=app_key,
                tenant_id=tenant_id,
                created_by="mandir_compat_system",
                payload=journal_payload,
                idempotency_key=f"don_{donation_id}",
            )
        except HTTPException:
            raise
        except Exception as exc:
            await col.delete_one({"donation_id": donation_id, "tenant_id": tenant_id, "app_key": app_key})
            raise HTTPException(status_code=500, detail=f"Failed to post donation journal: {exc}") from exc

    try:
        await _upsert_devotee_from_contribution(
            tenant_id,
            app_key,
            temple_id=temple_id,
            phone=devotee_phone,
            name_prefix=devotee_prefix,
            name=devotee_name,
            email=str(payload.get("email") or "") or None,
            address=devotee_address,
            city=devotee_city,
            state=devotee_state,
            pincode=devotee_pincode,
        )
    except Exception as exc:
        logger.warning("Donation saved but devotee upsert failed for tenant=%s phone=%s: %s", tenant_id, devotee_phone, exc)

    return _mandir_donation_view(donation)


@router.post("/donations/{donation_id}/valuation/approve", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)
async def approve_mandir_in_kind_valuation(
    donation_id: str,
    payload: dict[str, Any],
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    context = resolve_mandir_tenant(
        current_user=current_user, x_tenant_id=x_tenant_id, x_app_key=x_app_key,
        operation="in-kind valuation approval",
    )
    collection = get_collection("mandir_donations")
    query = {"donation_id": donation_id, "tenant_id": context.tenant_id, "app_key": context.app_key}
    donation = await collection.find_one(query)
    if donation is None:
        raise HTTPException(status_code=404, detail="Donation not found")
    if str(donation.get("donation_type") or "").lower() != "in_kind":
        raise HTTPException(status_code=409, detail="Only in-kind donations require valuation approval")
    if donation.get("status") == "posted" and donation.get("valuation_status") == "approved":
        return {**_mandir_donation_view(donation), "_idempotent": True}
    if donation.get("valuation_status") != "pending_approval":
        raise HTTPException(status_code=409, detail="Only pending valuations can be approved")
    actor = _mandir_actor_id(current_user)
    if actor == str(donation.get("created_by") or ""):
        raise HTTPException(status_code=409, detail="Valuation maker and approver must be different users")
    try:
        approved_amount = Decimal(str(payload.get("approved_amount") or donation.get("amount") or "0")).quantize(Decimal("0.01"))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Valid approved_amount is required") from exc
    if not approved_amount.is_finite() or approved_amount <= 0:
        raise HTTPException(status_code=400, detail="Approved value must be greater than zero")
    approval_basis = str(payload.get("approval_basis") or "").strip()
    if len(approval_basis) < 3:
        raise HTTPException(status_code=400, detail="Valuation approval basis is required")

    await _ensure_default_mandir_sql_accounts_safe(session, context.tenant_id, raise_on_failure=True)
    inventory_enabled = await _mandir_inventory_accounting_enabled(context.tenant_id, context.app_key)
    account_code, account_name, account_type = _mandir_in_kind_debit_account_target(
        donation, donation.get("category"), inventory_accounting_enabled=inventory_enabled,
    )
    debit_account_id = await _resolve_or_create_mandir_account(
        session, context.tenant_id, code=account_code, name=account_name, account_type=account_type,
        classification="nominal" if account_type == "expense" else "real",
    )
    income_category = (
        "In-Kind Sponsorship Income"
        if donation.get("is_sponsorship")
        else _mandir_in_kind_income_category(donation.get("category"))
    )
    income_account_id = await _resolve_mandir_income_account(session, context.tenant_id, income_category)
    dimension_id = str(donation.get("fund_dimension_id") or "") or None

    movement = None
    movement_collection = get_collection("mandir_inventory_movements")
    if account_code in {"14003", "14004"}:
        item_id = str(payload.get("inventory_item_id") or donation.get("inventory_item_id") or "").strip()
        item = await get_collection("mandir_inventory_items").find_one(
            {"id": item_id, "tenant_id": context.tenant_id, "app_key": context.app_key, "is_active": True}
        )
        if item is None:
            raise HTTPException(status_code=404, detail="Active inventory item is required for inventory-valued donation")
        try:
            quantity = Decimal(str(payload.get("approved_quantity") or donation.get("inventory_quantity") or "0")).quantize(Decimal("0.001"))
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Valid approved_quantity is required") from exc
        if not quantity.is_finite() or quantity <= 0:
            raise HTTPException(status_code=400, detail="Approved inventory quantity must be greater than zero")
        movement = {
            "id": str(uuid4()), "tenant_id": context.tenant_id, "app_key": context.app_key,
            "item_id": item_id, "item_name": item.get("name"), "movement_type": "receipt",
            "quantity": str(quantity), "unit_value": str((approved_amount / quantity).quantize(Decimal("0.01"))),
            "total_value": str(approved_amount), "movement_date": datetime.now(timezone.utc).date().isoformat(),
            "source_type": "in_kind_donation", "source_id": donation_id,
            "status": "pending_accounting", "created_by": actor, "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await movement_collection.insert_one(movement)

    try:
        journal, _created = await post_journal_entry(
            session=session, app_key=context.app_key, tenant_id=context.tenant_id, created_by=actor,
            payload=JournalPostRequest(
                entry_date=datetime.now(timezone.utc).date(),
                description=f"Approved in-kind valuation: {donation.get('in_kind_item_name') or donation.get('category')}",
                reference=str(donation.get("receipt_number")), source_module="mandirmitra",
                source_document_type="in_kind_donation", source_document_id=donation_id,
                lines=[
                    JournalLineIn(
                        account_id=debit_account_id, debit=approved_amount, credit=Decimal("0"),
                        cost_center_id=dimension_id,
                    ),
                    JournalLineIn(
                        account_id=income_account_id, debit=Decimal("0"), credit=approved_amount,
                        cost_center_id=dimension_id,
                    ),
                ],
            ),
            idempotency_key=f"don_{donation_id}",
        )
    except Exception:
        if movement:
            await movement_collection.update_one(
                {"id": movement["id"], "tenant_id": context.tenant_id, "app_key": context.app_key},
                {"$set": {"status": "accounting_failed"}}, upsert=False,
            )
        raise

    now = datetime.now(timezone.utc).isoformat()
    patch = {
        "amount": float(approved_amount), "status": "posted", "valuation_status": "approved",
        "approved_value": str(approved_amount), "valuation_approval_basis": approval_basis,
        "valuation_approved_by": actor, "valuation_approved_at": now,
        "journal_entry_id": int(journal.id), "updated_at": now,
    }
    if movement:
        patch.update({"inventory_item_id": movement["item_id"], "inventory_quantity": movement["quantity"], "inventory_movement_id": movement["id"]})
    try:
        await collection.update_one(query, {"$set": patch}, upsert=False)
        if movement:
            await movement_collection.update_one(
                {"id": movement["id"], "tenant_id": context.tenant_id, "app_key": context.app_key},
                {"$set": {"status": "posted", "journal_entry_id": int(journal.id), "posted_at": now}}, upsert=False,
            )
    except Exception as exc:
        await reverse_journal_entry(
            session=session, tenant_id=context.tenant_id, app_key=context.app_key, accounting_entity_id="primary",
            journal_id=int(journal.id), created_by=actor, reason="Compensate failed in-kind valuation persistence",
            idempotency_key=f"don_{donation_id}_valuation_compensation",
        )
        if movement:
            try:
                await movement_collection.update_one(
                    {"id": movement["id"], "tenant_id": context.tenant_id, "app_key": context.app_key},
                    {"$set": {"status": "compensated", "updated_at": now}}, upsert=False,
                )
            except Exception:
                logger.exception("Failed to mark compensated inventory receipt donation=%s", donation_id)
        try:
            await collection.update_one(
                query, {"$set": {"status": "pending_valuation", "valuation_status": "pending_approval", "updated_at": now}},
                upsert=False,
            )
        except Exception:
            logger.exception("Failed to restore pending valuation state donation=%s", donation_id)
        raise HTTPException(status_code=500, detail="Valuation persistence failed; accounting was reversed") from exc
    return _mandir_donation_view({**donation, **patch})


@router.post("/donations/{donation_id}/cancel", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
async def cancel_donation_receipt(
    donation_id: str,
    payload: dict[str, Any] | None = None,
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    return await _cancel_mandir_receipt_source(
        source_kind="donation",
        source_id=donation_id,
        collection_name="mandir_donations",
        id_field="donation_id",
        idempotency_prefix="don_",
        payload=payload,
        session=session,
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
    )


@router.post("/donations/reconcile-posting", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
async def reconcile_donation_posting(
    limit: int = Query(default=500, ge=1, le=5000),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    """
    Backfill journal entries for legacy donation docs that were saved before posting guardrails.
    """
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id, raise_on_failure=True)

    col = get_collection("mandir_donations")
    try:
        docs = await col.find({"tenant_id": tenant_id, "app_key": app_key}).sort("created_at", -1).limit(limit).to_list(length=limit)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to load donations for reconciliation: {exc}") from exc

    scanned = 0
    posted = 0
    already_posted = 0
    skipped = 0
    errors: list[dict[str, Any]] = []

    for doc in docs:
        scanned += 1
        donation_id = str(doc.get("donation_id") or doc.get("id") or doc.get("_id") or "").strip()
        if not donation_id:
            skipped += 1
            continue

        idempotency_key = f"don_{donation_id}"
        exists_stmt = select(JournalEntry.id).where(
            JournalEntry.tenant_id == tenant_id,
            JournalEntry.idempotency_key == idempotency_key,
        )
        existing_journal_id = (await session.execute(exists_stmt)).scalar_one_or_none()
        if existing_journal_id is not None:
            already_posted += 1
            continue

        amount = _safe_float(doc.get("amount"), 0.0)
        if amount <= 0:
            skipped += 1
            continue

        payment_mode_raw = str(doc.get("payment_mode") or "Cash").strip().lower()
        payment_mode_for_account = "cash" if payment_mode_raw == "cash" else "bank"

        try:
            resolved_account_id = await _resolve_mandir_payment_account_id(
                session,
                tenant_id,
                doc.get("bank_account_id") or doc.get("payment_account_id"),
                payment_mode_for_account,
            )
            if not resolved_account_id:
                resolved_account_id = await _resolve_mandir_payment_account_id(session, tenant_id, None, payment_mode_for_account)
            if not resolved_account_id:
                raise ValueError("No valid cash/bank account is configured for donation posting")

            category = str(doc.get("category") or "General Donation")
            income_acc_id = await _resolve_mandir_income_account(session, tenant_id, "General Donations")
            devotee = doc.get("devotee") if isinstance(doc.get("devotee"), dict) else {}
            devotee_name = str(devotee.get("name") or doc.get("devotee_name") or "Devotee")

            created_raw = str(doc.get("created_at") or "").strip()
            entry_date = datetime.now(timezone.utc).date()
            if created_raw:
                try:
                    entry_date = datetime.fromisoformat(created_raw.replace("Z", "+00:00")).date()
                except Exception:
                    pass

            journal_payload = JournalPostRequest(
                entry_date=entry_date,
                description=f"{category} from {devotee_name}",
                reference=f"DON-{donation_id[:8].upper()}",
                lines=[
                    JournalLineIn(account_id=resolved_account_id, debit=Decimal(str(amount)), credit=Decimal("0")),
                    JournalLineIn(account_id=income_acc_id, debit=Decimal("0"), credit=Decimal(str(amount))),
                ],
            )
            await post_journal_entry(
                session=session,
                tenant_id=tenant_id,
                created_by="mandir_reconcile",
                payload=journal_payload,
                idempotency_key=idempotency_key,
            )
            posted += 1
        except Exception as exc:
            errors.append({"donation_id": donation_id, "error": str(exc)})

    return {
        "status": "ok",
        "tenant_id": tenant_id,
        "app_key": app_key,
        "scanned": scanned,
        "posted": posted,
        "already_posted": already_posted,
        "skipped": skipped,
        "errors": errors[:25],
    }


@router.delete("/donations/cleanup", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)
async def cleanup_donation_entry(
    amount: float = Query(..., gt=0),
    devotee_phone: str = Query(..., min_length=6),
    payment_mode: str | None = Query(default=None),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    normalized_phone = _normalize_phone(devotee_phone)
    normalized_amount = _safe_float(amount, 0.0)
    normalized_mode = str(payment_mode or "").strip().lower() or None

    try:
        col = get_collection("mandir_donations")
        candidates = await col.find(
            {
                "tenant_id": tenant_id,
                "app_key": app_key,
                "amount": normalized_amount,
                "devotee_phone": normalized_phone,
            }
        ).sort("created_at", -1).to_list(length=50)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to search donation entries: {exc}") from exc

    if normalized_mode:
        candidates = [
            row
            for row in candidates
            if str(row.get("payment_mode") or "").strip().lower() == normalized_mode
        ]

    if not candidates:
        raise HTTPException(
            status_code=404,
            detail="Donation entry not found for the provided amount and phone",
        )

    donation = candidates[0]
    donation_id = str(donation.get("donation_id") or "")

    try:
        await col.delete_one(
            {
                "donation_id": donation_id,
                "tenant_id": tenant_id,
                "app_key": app_key,
            }
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to delete donation entry: {exc}") from exc

    journal_deleted = False
    journal_status = "not_found"
    journal_idempotency_key = f"don_{donation_id}" if donation_id else None
    if journal_idempotency_key:
        try:
            journal_stmt = select(JournalEntry).where(
                JournalEntry.tenant_id == tenant_id,
                JournalEntry.idempotency_key == journal_idempotency_key,
            )
            journal_entry = (await session.execute(journal_stmt)).scalar_one_or_none()
            if journal_entry is not None:
                await session.delete(journal_entry)
                await session.commit()
                journal_deleted = True
                journal_status = "deleted"
        except Exception as exc:
            try:
                await session.rollback()
            except Exception:
                pass
            journal_status = f"delete_failed: {exc}"

    return {
        "status": "deleted",
        "matched_count": len(candidates),
        "donation_id": donation_id,
        "amount": normalized_amount,
        "devotee_phone": normalized_phone,
        "payment_mode": donation.get("payment_mode"),
        "journal_deleted": journal_deleted,
        "journal_status": journal_status,
    }


# Devotee routes moved to routes/devotees.py
# (docs/operations/LARGE_FILE_MODULARIZATION_PLAN.md); registered via import at end of module.


# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: Sevas (GET / POST / PUT / DELETE / import / lists / dropdown-options / payment-accounts)
# ROUTES : GET /sevas  POST .../sevas  PUT|DELETE .../sevas/{seva_id}  GET|POST .../import  GET .../priests|dropdown-options|payment-accounts
# ════════════════════════════════════════════════════════════════════════

# Seva routes moved to routes/sevas.py
# (docs/operations/LARGE_FILE_MODULARIZATION_PLAN.md); registered via import at end of module.

# Temple routes moved to routes/temples.py

# Accounts / COA routes moved to routes/accounts.py

# Misc compat routes moved to routes/misc_compat.py

# Hundi routes moved to routes/hundi.py

# Inventory routes moved to routes/inventory.py

def _parse_journal_entry_date(value: Any) -> date:
    if isinstance(value, date):
        return value

    raw = str(value or "").strip()
    if not raw:
        return datetime.now(timezone.utc).date()

    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).date()
    except Exception:
        try:
            return date.fromisoformat(raw[:10])
        except Exception:
            return datetime.now(timezone.utc).date()


async def _resolve_sql_account_for_journal_line(
    session: AsyncSession,
    *,
    tenant_id: str,
    raw_account_id: Any,
) -> Account | None:
    raw_value = str(raw_account_id or "").strip()
    if not raw_value:
        return None

    maybe_id = _safe_optional_int(raw_value)
    if maybe_id is not None:
        by_id_stmt = select(Account).where(
            Account.tenant_id == tenant_id,
            Account.id == maybe_id,
        )
        by_id = (await session.execute(by_id_stmt)).scalar_one_or_none()
        if by_id is not None:
            return by_id

    code_candidate = raw_value.split(" - ", 1)[0].strip()
    normalized_code = _normalize_mandir_account_code(code_candidate)
    if normalized_code:
        by_code_stmt = select(Account).where(
            Account.tenant_id == tenant_id,
            Account.code == normalized_code,
        )
        by_code = (await session.execute(by_code_stmt)).scalar_one_or_none()
        if by_code is not None:
            return by_code

    return None


async def _normalize_mandir_journal_lines(
    session: AsyncSession,
    *,
    tenant_id: str,
    raw_lines: Any,
) -> tuple[list[dict[str, Any]], Decimal, Decimal]:
    if not isinstance(raw_lines, list) or len(raw_lines) < 2:
        raise HTTPException(status_code=400, detail="At least two journal lines are required")

    normalized_lines: list[dict[str, Any]] = []
    total_debit = Decimal("0.00")
    total_credit = Decimal("0.00")

    for index, line in enumerate(raw_lines, start=1):
        if not isinstance(line, dict):
            raise HTTPException(status_code=400, detail=f"Journal line #{index} is invalid")

        account_ref = line.get("account_id")
        account = await _resolve_sql_account_for_journal_line(
            session,
            tenant_id=tenant_id,
            raw_account_id=account_ref,
        )
        if account is None:
            raise HTTPException(status_code=400, detail=f"Invalid account on journal line #{index}")

        try:
            debit_amount = Decimal(str(line.get("debit_amount") or 0)).quantize(Decimal("0.01"))
            credit_amount = Decimal(str(line.get("credit_amount") or 0)).quantize(Decimal("0.01"))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Invalid amount on journal line #{index}") from exc

        if debit_amount < 0 or credit_amount < 0:
            raise HTTPException(status_code=400, detail=f"Amounts cannot be negative on line #{index}")
        if debit_amount == 0 and credit_amount == 0:
            raise HTTPException(status_code=400, detail=f"Either debit or credit is required on line #{index}")
        if debit_amount > 0 and credit_amount > 0:
            raise HTTPException(status_code=400, detail=f"Line #{index} cannot have both debit and credit")

        total_debit += debit_amount
        total_credit += credit_amount

        reference_account_id = _safe_optional_int(account_ref)
        if reference_account_id is None:
            reference_account_id = _safe_optional_int(str(account.code or "").strip()) or int(account.id)

        normalized_lines.append(
            {
                "account_id": reference_account_id,
                "account_code": str(account.code or "").strip(),
                "account_name": str(account.name or "").strip(),
                "ledger_account_id": int(account.id),
                "debit_amount": float(debit_amount),
                "credit_amount": float(credit_amount),
                "description": str(line.get("description") or "").strip(),
            }
        )

    if total_debit <= 0 or total_credit <= 0:
        raise HTTPException(status_code=400, detail="Total debit and credit must be greater than zero")
    if total_debit != total_credit:
        raise HTTPException(status_code=400, detail="Total debit and credit must be equal")

    return normalized_lines, total_debit, total_credit


async def _validate_mandir_journal_cash_balance(
    session: AsyncSession,
    *,
    tenant_id: str,
    app_key: str,
    normalized_lines: list[dict[str, Any]],
) -> None:
    accounting_lines: list[tuple[int, Decimal, Decimal]] = []
    for line in normalized_lines:
        ledger_account_id = _safe_optional_int(line.get("ledger_account_id"))
        if ledger_account_id is None:
            continue
        accounting_lines.append(
            (
                ledger_account_id,
                Decimal(str(line.get("debit_amount") or 0)).quantize(Decimal("0.01")),
                Decimal(str(line.get("credit_amount") or 0)).quantize(Decimal("0.01")),
            )
        )

    try:
        await validate_cash_balance_for_journal_lines(
            session,
            tenant_id=tenant_id,
            app_key=app_key,
            accounting_entity_id="primary",
            normalized_lines=accounting_lines,
        )
    except AccountingValidationError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


def _mandir_journal_entry_view(doc: dict[str, Any]) -> dict[str, Any]:
    row = _sanitize_mongo_doc(doc)
    row["entry_number"] = str(row.get("entry_number") or f"JE-{str(row.get('id') or '')[:8].upper()}")
    row["entry_date"] = str(row.get("entry_date") or datetime.now(timezone.utc).date().isoformat())[:10]
    row["narration"] = str(row.get("narration") or row.get("description") or "").strip()
    row["reference_type"] = str(row.get("reference_type") or "").strip().lower() or None
    row["reference_id"] = _safe_optional_int(row.get("reference_id"))
    row["status"] = str(row.get("status") or "draft").strip().lower() or "draft"
    row["total_debit"] = _safe_float(row.get("total_debit"), 0.0)
    row["total_credit"] = _safe_float(row.get("total_credit"), 0.0)
    row["total_amount"] = _safe_float(
        row.get("total_amount"),
        _safe_float(row.get("total_debit"), 0.0),
    )

    journal_lines: list[dict[str, Any]] = []
    for line in row.get("journal_lines") or []:
        if not isinstance(line, dict):
            continue
        journal_lines.append(
            {
                "account_id": _safe_optional_int(line.get("account_id")),
                "account_code": str(line.get("account_code") or "").strip(),
                "account_name": str(line.get("account_name") or "").strip(),
                "debit_amount": _safe_float(line.get("debit_amount"), 0.0),
                "credit_amount": _safe_float(line.get("credit_amount"), 0.0),
                "description": str(line.get("description") or "").strip(),
            }
        )
    row["journal_lines"] = journal_lines
    return row



# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: Journal entries (GET / POST / drilldown / financial reports / day-book / cash-book / bank-book)
# ROUTES : GET /journal-entries  POST .../journal-entries  GET .../reports/drilldown|balance-sheet|accounts-receivable|payable|ledger|category-income|top-donors|day-book|cash-book|bank-book
# ════════════════════════════════════════════════════════════════════════

@router.get("/journal-entries")
@router.get("/journal-entries/")
async def mandir_journal_entries(
    from_date: date | None = Query(default=None),
    to_date: date | None = Query(default=None),
    reference_type: str | None = Query(default=None),
    limit: int = Query(default=500, ge=1, le=2000),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())

    query: dict[str, Any] = {"tenant_id": tenant_id, "app_key": app_key}
    if reference_type:
        query["reference_type"] = str(reference_type).strip().lower()

    try:
        docs = await get_collection("mandir_journal_entries").find(query).sort("updated_at", -1).limit(limit).to_list(length=limit)
    except Exception:
        docs = []

    rows: list[dict[str, Any]] = []
    for doc in docs:
        view = _mandir_journal_entry_view(doc)
        entry_date = _parse_journal_entry_date(view.get("entry_date"))
        if from_date and entry_date < from_date:
            continue
        if to_date and entry_date > to_date:
            continue
        rows.append(view)

    if rows:
        return rows

    # Backward-compatible fallback: expose posted SQL journals in the same list shape.
    report = await journal_entries_report(session, tenant_id=tenant_id, limit=limit)
    fallback_rows: list[dict[str, Any]] = []
    for item in report.get("items", []):
        entry_date = _parse_journal_entry_date(item.get("entry_date"))
        if from_date and entry_date < from_date:
            continue
        if to_date and entry_date > to_date:
            continue
        fallback_reference_type = str(item.get("reference") or "").split("-", 1)[0].lower() or None
        if reference_type and fallback_reference_type != str(reference_type).strip().lower():
            continue

        fallback_rows.append(
            {
                "id": int(item.get("id")),
                "entry_number": f"JE-{item.get('id')}",
                "entry_date": entry_date.isoformat(),
                "narration": str(item.get("description") or "").strip(),
                "reference_type": fallback_reference_type,
                "reference_id": None,
                "status": "posted",
                "total_amount": _safe_float(item.get("total_debit"), 0.0),
                "total_debit": _safe_float(item.get("total_debit"), 0.0),
                "total_credit": _safe_float(item.get("total_credit"), 0.0),
                "journal_lines": [],
            }
        )

    return fallback_rows


@router.post("/journal-entries", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
@router.post("/journal-entries/", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
async def mandir_create_journal_entry(
    payload: dict[str, Any],
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_context = resolve_mandir_tenant(
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
        operation="journal entry creation",
    )
    tenant_id = tenant_context.tenant_id
    app_key = tenant_context.app_key
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id, raise_on_failure=True)

    normalized_lines, total_debit, total_credit = await _normalize_mandir_journal_lines(
        session,
        tenant_id=tenant_id,
        raw_lines=payload.get("journal_lines"),
    )
    await _validate_mandir_journal_cash_balance(
        session,
        tenant_id=tenant_id,
        app_key=app_key,
        normalized_lines=normalized_lines,
    )

    entry_id = str(uuid4())
    now = datetime.now(timezone.utc).isoformat()
    entry_date = _parse_journal_entry_date(payload.get("entry_date"))
    entry_number = await _next_journal_entry_number(tenant_id=tenant_id, app_key=app_key)

    entry_doc = {
        "id": entry_id,
        "tenant_id": tenant_id,
        "app_key": app_key,
        "entry_number": entry_number,
        "entry_date": entry_date.isoformat(),
        "narration": str(payload.get("narration") or "").strip(),
        "reference_type": str(payload.get("reference_type") or "expense").strip().lower(),
        "reference_id": _safe_optional_int(payload.get("reference_id")),
        "status": "draft",
        "journal_lines": normalized_lines,
        "total_debit": float(total_debit),
        "total_credit": float(total_credit),
        "total_amount": float(total_debit),
        "idempotency_key": f"man_je_{entry_id}",
        "created_by": str(current_user.get("sub") or current_user.get("email") or "system"),
        "created_at": now,
        "updated_at": now,
    }

    await get_collection("mandir_journal_entries").insert_one(entry_doc)
    return _mandir_journal_entry_view(entry_doc)


@router.put("/journal-entries/{entry_id}", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
async def mandir_update_journal_entry(
    entry_id: str,
    payload: dict[str, Any],
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_context = resolve_mandir_tenant(
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
        operation="journal entry update",
    )
    tenant_id = tenant_context.tenant_id
    app_key = tenant_context.app_key

    collection = get_collection("mandir_journal_entries")
    existing = await collection.find_one({"id": str(entry_id), "tenant_id": tenant_id, "app_key": app_key})
    if existing is None:
        raise HTTPException(status_code=404, detail="Journal entry not found")

    if str(existing.get("status") or "").lower() != "draft":
        raise HTTPException(status_code=400, detail="Only draft entries can be edited")

    normalized_lines, total_debit, total_credit = await _normalize_mandir_journal_lines(
        session,
        tenant_id=tenant_id,
        raw_lines=payload.get("journal_lines"),
    )
    await _validate_mandir_journal_cash_balance(
        session,
        tenant_id=tenant_id,
        app_key=app_key,
        normalized_lines=normalized_lines,
    )

    patch = {
        "entry_date": _parse_journal_entry_date(payload.get("entry_date") or existing.get("entry_date")).isoformat(),
        "narration": str(payload.get("narration") or "").strip(),
        "reference_type": str(payload.get("reference_type") or existing.get("reference_type") or "expense").strip().lower(),
        "reference_id": _safe_optional_int(payload.get("reference_id")),
        "journal_lines": normalized_lines,
        "total_debit": float(total_debit),
        "total_credit": float(total_credit),
        "total_amount": float(total_debit),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    await collection.update_one(
        {"id": str(entry_id), "tenant_id": tenant_id, "app_key": app_key},
        {"$set": patch},
        upsert=False,
    )

    updated = {**existing, **patch}
    return _mandir_journal_entry_view(updated)


@router.post("/journal-entries/{entry_id}/post", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
async def mandir_post_journal_entry(
    entry_id: str,
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_context = resolve_mandir_tenant(
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
        operation="journal entry posting",
    )
    tenant_id = tenant_context.tenant_id
    app_key = tenant_context.app_key
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id, raise_on_failure=True)

    collection = get_collection("mandir_journal_entries")
    existing = await collection.find_one({"id": str(entry_id), "tenant_id": tenant_id, "app_key": app_key})
    if existing is None:
        raise HTTPException(status_code=404, detail="Journal entry not found")

    if str(existing.get("status") or "").lower() == "posted":
        return _mandir_journal_entry_view(existing)

    if str(existing.get("status") or "").lower() in {"cancelled", "reversed"}:
        raise HTTPException(status_code=400, detail="Cancelled/reversed entries cannot be posted")

    normalized_lines, total_debit, total_credit = await _normalize_mandir_journal_lines(
        session,
        tenant_id=tenant_id,
        raw_lines=existing.get("journal_lines"),
    )
    await _validate_mandir_journal_cash_balance(
        session,
        tenant_id=tenant_id,
        app_key=app_key,
        normalized_lines=normalized_lines,
    )

    post_lines: list[JournalLineIn] = []
    for line in normalized_lines:
        ledger_account_id = _safe_optional_int(line.get("ledger_account_id"))
        if not ledger_account_id:
            account = await _resolve_sql_account_for_journal_line(
                session,
                tenant_id=tenant_id,
                raw_account_id=line.get("account_id"),
            )
            if account is None:
                raise HTTPException(status_code=400, detail="Unable to resolve account while posting")
            ledger_account_id = int(account.id)

        post_lines.append(
            JournalLineIn(
                account_id=ledger_account_id,
                debit=Decimal(str(line.get("debit_amount") or 0)),
                credit=Decimal(str(line.get("credit_amount") or 0)),
            )
        )

    journal_payload = JournalPostRequest(
        entry_date=_parse_journal_entry_date(existing.get("entry_date")),
        description=str(existing.get("narration") or "").strip(),
        reference=f"{str(existing.get('reference_type') or 'expense').upper()}-{str(existing.get('entry_number') or '')}",
        lines=post_lines,
    )

    try:
        posted_entry, _created = await post_journal_entry(
            session=session,
            tenant_id=tenant_id,
            created_by=str(current_user.get("sub") or current_user.get("email") or "system"),
            payload=journal_payload,
            idempotency_key=str(existing.get("idempotency_key") or f"man_je_{entry_id}"),
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to post journal entry: {exc}") from exc

    patch = {
        "status": "posted",
        "journal_lines": normalized_lines,
        "total_debit": float(total_debit),
        "total_credit": float(total_credit),
        "total_amount": float(total_debit),
        "posted_journal_id": int(posted_entry.id),
        "posted_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await collection.update_one(
        {"id": str(entry_id), "tenant_id": tenant_id, "app_key": app_key},
        {"$set": patch},
        upsert=False,
    )

    return _mandir_journal_entry_view({**existing, **patch})


@router.post("/journal-entries/{entry_id}/cancel", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
async def mandir_cancel_journal_entry(
    entry_id: str,
    payload: dict[str, Any] | None = None,
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_context = resolve_mandir_tenant(
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
        operation="seva booking",
    )
    tenant_id = tenant_context.tenant_id
    app_key = tenant_context.app_key

    collection = get_collection("mandir_journal_entries")
    existing = await collection.find_one({"id": str(entry_id), "tenant_id": tenant_id, "app_key": app_key})
    if existing is None:
        raise HTTPException(status_code=404, detail="Journal entry not found")

    current_status = str(existing.get("status") or "").lower()
    if current_status in {"cancelled", "reversed"}:
        return _mandir_journal_entry_view(existing)

    cancellation_reason = str((payload or {}).get("cancellation_reason") or "").strip() or "Reversal entry"

    if current_status == "draft":
        patch = {
            "status": "cancelled",
            "cancellation_reason": cancellation_reason,
            "cancelled_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await collection.update_one(
            {"id": str(entry_id), "tenant_id": tenant_id, "app_key": app_key},
            {"$set": patch},
            upsert=False,
        )
        return _mandir_journal_entry_view({**existing, **patch})

    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id, raise_on_failure=True)
    normalized_lines, _total_debit, _total_credit = await _normalize_mandir_journal_lines(
        session,
        tenant_id=tenant_id,
        raw_lines=existing.get("journal_lines"),
    )

    reversal_lines: list[JournalLineIn] = []
    for line in normalized_lines:
        ledger_account_id = _safe_optional_int(line.get("ledger_account_id"))
        if not ledger_account_id:
            account = await _resolve_sql_account_for_journal_line(
                session,
                tenant_id=tenant_id,
                raw_account_id=line.get("account_id"),
            )
            if account is None:
                raise HTTPException(status_code=400, detail="Unable to resolve account while reversing")
            ledger_account_id = int(account.id)

        reversal_lines.append(
            JournalLineIn(
                account_id=ledger_account_id,
                debit=Decimal(str(line.get("credit_amount") or 0)),
                credit=Decimal(str(line.get("debit_amount") or 0)),
            )
        )

    reversal_payload = JournalPostRequest(
        entry_date=datetime.now(timezone.utc).date(),
        description=f"Reversal of {existing.get('entry_number')}: {cancellation_reason}",
        reference=f"REV-{existing.get('entry_number')}",
        lines=reversal_lines,
    )

    try:
        reversal_entry, _created = await post_journal_entry(
            session=session,
            tenant_id=tenant_id,
            created_by=str(current_user.get("sub") or current_user.get("email") or "system"),
            payload=reversal_payload,
            idempotency_key=f"{str(existing.get('idempotency_key') or f'man_je_{entry_id}')}_rev",
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to reverse journal entry: {exc}") from exc

    patch = {
        "status": "reversed",
        "cancellation_reason": cancellation_reason,
        "reversed_at": datetime.now(timezone.utc).isoformat(),
        "reversal_journal_id": int(reversal_entry.id),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await collection.update_one(
        {"id": str(entry_id), "tenant_id": tenant_id, "app_key": app_key},
        {"$set": patch},
        upsert=False,
    )

    return _mandir_journal_entry_view({**existing, **patch})


@router.get("/journal-entries/reports/trial-balance")
async def mandir_journal_trial_balance(
    as_of: date,
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id)
    try:
        return await trial_balance_report(session, tenant_id=tenant_id, as_of=as_of)
    except (ConnectionRefusedError, OSError, SQLAlchemyError) as exc:
        logger.exception("Trial balance query failed", extra={"tenant_id": tenant_id, "as_of": as_of.isoformat()})
        raise HTTPException(status_code=503, detail="Accounting database unavailable. Please retry shortly.") from exc


@router.get("/journal-entries/reports/profit-loss")
async def mandir_journal_profit_loss(
    from_date: date = Query(...),
    to_date: date = Query(...),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id)
    return await profit_loss_report(session, tenant_id=tenant_id, from_date=from_date, to_date=to_date)


@router.get("/journal-entries/reports/income-expenditure")
async def mandir_journal_income_expenditure(
    from_date: date = Query(...),
    to_date: date = Query(...),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id)
    return await profit_loss_report(session, tenant_id=tenant_id, from_date=from_date, to_date=to_date)


@router.get("/journal-entries/reports/receipts-payments")
async def mandir_journal_receipts_payments(
    from_date: date = Query(...),
    to_date: date = Query(...),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id)
    return await receipts_payments_report(session, tenant_id=tenant_id, from_date=from_date, to_date=to_date)


@router.get("/journal-entries/reports/drilldown")
async def mandir_journal_drilldown(
    from_date: date = Query(...),
    to_date: date = Query(...),
    level: str = Query(default="month", pattern="^(month|week|day|voucher)$"),
    month: str | None = Query(default=None, pattern=r"^\d{4}-\d{2}$"),
    week_start: date | None = Query(default=None),
    day: date | None = Query(default=None),
    limit: int = Query(default=500, ge=1, le=2000),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    if from_date > to_date:
        raise HTTPException(status_code=422, detail="from_date cannot be greater than to_date")
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id)
    return await get_journal_drilldown(
        session,
        tenant_id=tenant_id,
        app_key=app_key,
        accounting_entity_id="primary",
        from_date=from_date,
        to_date=to_date,
        level=level,
        month=month,
        week_start=week_start,
        day=day,
        limit=limit,
    )


@router.get("/journal-entries/reports/balance-sheet")
async def mandir_journal_balance_sheet(
    as_of: date = Query(...),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id)
    return await balance_sheet_report(session, tenant_id=tenant_id, as_of=as_of)


@router.get("/journal-entries/reports/accounts-receivable")
async def mandir_journal_accounts_receivable(
    as_of: date = Query(...),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id)
    return await accounts_receivable_report(session, tenant_id=tenant_id, as_of=as_of)


@router.get("/journal-entries/reports/accounts-payable")
async def mandir_journal_accounts_payable(
    as_of: date = Query(...),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id)
    return await accounts_payable_report(session, tenant_id=tenant_id, as_of=as_of)


@router.get("/journal-entries/reports/ledger/{account_id}")
async def mandir_journal_ledger(
    account_id: int,
    from_date: date | None = Query(default=None),
    to_date: date | None = Query(default=None),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id)
    return await ledger_report(session, tenant_id=tenant_id, account_id=account_id, from_date=from_date, to_date=to_date)


@router.get("/journal-entries/reports/category-income")
async def mandir_journal_category_income(
    from_date: date = Query(...),
    to_date: date = Query(...),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    return await category_income_report(session, tenant_id=tenant_id, app_key=app_key, from_date=from_date, to_date=to_date)


@router.get("/journal-entries/reports/top-donors")
async def mandir_journal_top_donors(
    from_date: date = Query(...),
    to_date: date = Query(...),
    limit: int = Query(default=10, ge=1, le=100),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    return await top_donors_report(session, tenant_id=tenant_id, app_key=app_key, from_date=from_date, to_date=to_date, limit=limit)


@router.get("/journal-entries/reports/day-book")
async def mandir_journal_day_book(
    date: date = Query(...),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id)
    return await day_book_report(session, tenant_id=tenant_id, date_value=date)


@router.get("/journal-entries/reports/cash-book")
async def mandir_journal_cash_book(
    from_date: date = Query(...),
    to_date: date = Query(...),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id)
    return await cash_book_report(session, tenant_id=tenant_id, from_date=from_date, to_date=to_date)


@router.get("/journal-entries/reports/bank-book/{account_id}")
async def mandir_journal_bank_book(
    account_id: int,
    from_date: date = Query(...),
    to_date: date = Query(...),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id)
    return await bank_book_report(session, tenant_id=tenant_id, account_id=account_id, from_date=from_date, to_date=to_date)



# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: Login + opening balances (template / import)
# ROUTES : POST /login  POST /login/access-token  GET /opening-balances/template  POST .../import
# ════════════════════════════════════════════════════════════════════════

@router.post("/login")
@router.post("/login/access-token")
async def mandir_legacy_login(payload: dict[str, Any], x_app_key: str | None = Header(default=None, alias="X-App-Key")):
    from app.core.auth.service import login_user

    email = str(payload.get("email") or payload.get("username") or "")
    password = str(payload.get("password") or "")
    app_key = resolve_app_key((x_app_key or "mandirmitra").strip())
    access_token, refresh_token = await login_user(email, password, app_key=app_key)
    return {"access_token": access_token, "refresh_token": refresh_token, "token_type": "bearer"}


@router.get("/opening-balances/template")
async def mandir_opening_balances_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Opening Balances"

    headers = ["account_code", "account_name", "opening_balance_debit", "opening_balance_credit"]
    sheet.append(headers)

    header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    example_data = [
        (11001, "Cash", 50000, None),
        (11002, "Bank Account", 100000, None),
        (32001, "General Reserve", None, 150000),
        (21001, "Loan Payable", None, 50000),
    ]

    for row_data in example_data:
        sheet.append(row_data)

    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 25
    sheet.column_dimensions["C"].width = 25
    sheet.column_dimensions["D"].width = 25

    for row in sheet.iter_rows(min_row=2, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Opening_Balance_Template.xlsx"},
    )


@router.post("/opening-balances/import")
async def mandir_opening_balances_import(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_async_session),
    _current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    try:
        tenant_id = resolve_tenant_id(_current_user, x_tenant_id)
        await _ensure_default_mandir_sql_accounts_safe(session, tenant_id)

        rows = await _parse_opening_balance_rows(file)
        if not rows:
            raise HTTPException(status_code=400, detail="Import file is empty")

        sql_accounts = await list_accounts(session, tenant_id=tenant_id)
        accounts_by_code: dict[str, Account] = {}

        def _score_account(acc: Account, normalized_code: str) -> tuple[int, int, int]:
            raw_code = str(acc.code or "").strip()
            return (
                1 if raw_code == normalized_code else 0,
                len(raw_code),
                int(acc.id or 0),
            )

        for acc in sql_accounts:
            normalized_code = _normalize_mandir_account_code(acc.code, account_name=acc.name)
            if not normalized_code:
                continue
            existing = accounts_by_code.get(normalized_code)
            if existing is None or _score_account(acc, normalized_code) > _score_account(existing, normalized_code):
                accounts_by_code[normalized_code] = acc

        opening_offset_account = await _find_or_create_opening_balance_offset_account(session, tenant_id)

        updated_rows: list[dict[str, Any]] = []
        errors: list[dict[str, Any]] = []
        skipped_count = 0

        for row_number, raw_row in enumerate(rows, start=2):
            try:
                row = {
                    str(key or "").strip().lower(): value
                    for key, value in (raw_row or {}).items()
                    if key is not None
                }

                raw_code = row.get("account_code") or row.get("legacy_code") or row.get("code")
                account_name_hint = row.get("account_name") or row.get("name")
                account_code = _normalize_mandir_account_code(raw_code, account_name=account_name_hint)
                if not account_code:
                    raise ValueError("account_code or legacy_code is required")

                account = accounts_by_code.get(account_code)
                if account is None:
                    raise ValueError(f"Account '{account_code}' not found")

                account_type = str(account.type or "").strip().lower()
                if account_type not in {"asset", "liability", "equity"}:
                    raise ValueError("Only balance sheet accounts can have opening balances")

                debit_raw = _parse_opening_balance_decimal(row.get("opening_balance_debit"))
                credit_raw = _parse_opening_balance_decimal(row.get("opening_balance_credit"))
                signed_raw = _parse_opening_balance_decimal(row.get("opening_balance"))

                if debit_raw is None and credit_raw is None and signed_raw is None:
                    raise ValueError("Provide opening_balance_debit/opening_balance_credit or opening_balance")

                debit_amount = max(debit_raw or Decimal("0"), Decimal("0"))
                credit_amount = max(credit_raw or Decimal("0"), Decimal("0"))

                if debit_raw is None and credit_raw is None and signed_raw is not None:
                    if account_type == "asset":
                        debit_amount = max(signed_raw, Decimal("0"))
                        credit_amount = max(-signed_raw, Decimal("0"))
                    else:
                        credit_amount = max(signed_raw, Decimal("0"))
                        debit_amount = max(-signed_raw, Decimal("0"))

                if debit_amount > 0 and credit_amount > 0:
                    raise ValueError("Only one side can be positive for opening balance")

                desired_net = debit_amount - credit_amount
                reference = f"OPENING-{account_code}"
                current_net = await _current_opening_balance_net(
                    session,
                    tenant_id=tenant_id,
                    account_id=int(account.id),
                    reference=reference,
                )
                delta_net = desired_net - current_net

                if delta_net == 0:
                    skipped_count += 1
                    continue

                account_debit = delta_net if delta_net > 0 else Decimal("0")
                account_credit = -delta_net if delta_net < 0 else Decimal("0")

                payload = JournalPostRequest(
                    entry_date=date.today(),
                    reference=reference,
                    description=f"Opening balance import for {account.code} - {account.name}",
                    lines=[
                        JournalLineIn(
                            account_id=int(account.id),
                            debit=account_debit,
                            credit=account_credit,
                        ),
                        JournalLineIn(
                            account_id=int(opening_offset_account.id),
                            debit=account_credit,
                            credit=account_debit,
                        ),
                    ],
                )

                idempotency_key = f"mandir-opening-balance:{account_code}:{str(desired_net)}"
                await post_journal_entry(
                    session,
                    tenant_id=tenant_id,
                    created_by=str(_current_user.get("sub") or _current_user.get("email") or "system"),
                    payload=payload,
                    idempotency_key=idempotency_key,
                )

                updated_rows.append(
                    {
                        "account_code": account_code,
                        "account_name": account.name,
                        "opening_balance_debit": _safe_float(max(desired_net, Decimal("0"))),
                        "opening_balance_credit": _safe_float(max(-desired_net, Decimal("0"))),
                        "applied_delta": _safe_float(delta_net),
                    }
                )
            except Exception as exc:
                errors.append({"row": row_number, "error": str(exc)})

        has_errors = len(errors) > 0
        success = len(updated_rows) > 0 and not has_errors

        if success:
            message = f"✓ Successfully imported {len(updated_rows)} opening balance(s)"
        elif len(updated_rows) > 0:
            message = f"⚠ Partial success: {len(updated_rows)} imported, {len(errors)} failed"
        else:
            message = f"✗ Import failed: All {len(rows)} row(s) had errors"

        return {
            "success": success,
            "status": "success" if success else ("partial" if len(updated_rows) > 0 else "failed"),
            "message": message,
            "processed_count": len(rows),
            "updated_count": len(updated_rows),
            "skipped_count": skipped_count,
            "error_count": len(errors),
            "updated": updated_rows,
            "errors": errors[:200],
        }
    except HTTPException:
        raise
    except Exception as exc:
        return {
            "success": False,
            "status": "failed",
            "message": f"✗ Import failed: {str(exc)}",
            "processed_count": 0,
            "updated_count": 0,
            "skipped_count": 0,
            "error_count": 1,
            "updated": [],
            "errors": [{"row": 0, "error": str(exc)}],
        }



# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: Panchang display settings + panchang on-date
# ROUTES : GET|PUT /panchang/display-settings  GET .../cities  GET /panchang/on-date  GET /panchang/on-date-full
# ════════════════════════════════════════════════════════════════════════

@router.get("/panchang/display-settings")
@router.get("/panchang/display-settings/")
async def mandir_panchang_display_settings(
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    doc = await get_collection("mandir_panchang_settings").find_one({"tenant_id": tenant_id, "app_key": app_key}) or {}
    return {
        # Display mode and language
        "display_mode": str(doc.get("display_mode") or "full"),
        "primary_language": str(doc.get("primary_language") or "English"),
        "show_on_dashboard": bool(doc.get("show_on_dashboard", True)),
        # Panchang limb visibility flags (Panch Angs: Var, Tithi, Nakshatra, Karana, Yoga)
        "show_tithi": bool(doc.get("show_tithi", True)),
        "show_nakshatra": bool(doc.get("show_nakshatra", True)),
        "show_yoga": bool(doc.get("show_yoga", True)),
        "show_karana": bool(doc.get("show_karana", True)),
        # Optional timing displays
        "show_sun_timings": bool(doc.get("show_sun_timings", True)),
        "show_rahu_kaal": bool(doc.get("show_rahu_kaal", True)),
        "show_yamaganda": bool(doc.get("show_yamaganda", True)),
        "show_gulika": bool(doc.get("show_gulika", True)),
        "show_abhijit_muhurat": bool(doc.get("show_abhijit_muhurat", True)),
        # Location settings
        "city_name": doc.get("city_name"),
        "latitude": doc.get("latitude"),
        "longitude": doc.get("longitude"),
        "timezone": doc.get("timezone"),
    }


@router.put("/panchang/display-settings", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)
@router.put("/panchang/display-settings/", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)
async def mandir_panchang_display_settings_update(
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    now = datetime.now(timezone.utc).isoformat()
    allowed = {
        # Display mode and language
        "display_mode",
        "primary_language",
        "show_on_dashboard",
        # Panchang limb visibility flags (Panch Angs)
        "show_tithi",
        "show_nakshatra",
        "show_yoga",
        "show_karana",
        # Optional timing displays
        "show_sun_timings",
        "show_rahu_kaal",
        "show_yamaganda",
        "show_gulika",
        "show_abhijit_muhurat",
        # Location settings
        "city_name",
        "latitude",
        "longitude",
        "timezone",
    }
    patch = {k: payload.get(k) for k in allowed if k in payload}
    patch["updated_at"] = now
    await get_collection("mandir_panchang_settings").update_one(
        {"tenant_id": tenant_id, "app_key": app_key},
        {
            "$set": patch,
            "$setOnInsert": {
                "tenant_id": tenant_id,
                "app_key": app_key,
                "created_at": now,
            },
        },
        upsert=True,
    )
    return await mandir_panchang_display_settings(current_user=current_user, x_tenant_id=x_tenant_id, x_app_key=x_app_key)

@router.get("/panchang/display-settings/cities")
async def mandir_panchang_cities(_current_user: dict = Depends(get_current_user)):
    return {"success": True, "data": _panchang_city_options()}


@router.get("/panchang/on-date")
async def mandir_panchang_on_date(
    target_date: str = Query(...),
    city_name: str | None = Query(default=None),
    latitude: float | None = Query(default=None),
    longitude: float | None = Query(default=None),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    """Get panchang for a specific date for Nakshatra lookup."""
    try:
        # Parse the target date
        target_dt = datetime.strptime(target_date, "%Y-%m-%d")

        tenant_id = resolve_tenant_id(current_user, x_tenant_id)
        app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())

        # Get temple location from MongoDB
        temple_doc = await get_collection("mandir_temples").find_one(
            {"tenant_id": tenant_id, "app_key": app_key}
        )
        if not temple_doc:
            temple_doc = await get_collection("mandir_temples").find_one({"tenant_id": tenant_id})

        # Get panchang display settings for overrides
        settings_doc = await get_collection("mandir_panchang_settings").find_one(
            {"tenant_id": tenant_id, "app_key": app_key}
        ) or {}

        if not temple_doc:
            temple_doc = {}

        latitude, longitude, city = _resolve_panchang_location(
            settings_doc,
            temple_doc,
            city_name=city_name,
            latitude=latitude,
            longitude=longitude,
        )

        # Calculate panchang for the target date
        panchang_service = PanchangService()
        panchang_data = panchang_service.calculate_panchang(target_dt, latitude, longitude, city)

        return {
            "lookup_requested_date": target_date,
            "panchang": panchang_data.get("panchang"),
            "date": panchang_data.get("date"),
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format. Use YYYY-MM-DD: {str(e)}")
    except Exception as e:
        logger.error(f"Error calculating panchang for date {target_date}: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to calculate panchang: {str(e)}"
        )


@router.get("/panchang/on-date-full")
async def mandir_panchang_on_date_full(
    target_date: str = Query(...),
    city_name: str | None = Query(default=None),
    latitude: float | None = Query(default=None),
    longitude: float | None = Query(default=None),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    """Get complete panchang for any date (past/future). Returns full panchang data with all limbs and muhurtas."""
    try:
        # Parse the target date
        target_dt = datetime.strptime(target_date, "%Y-%m-%d")

        tenant_id = resolve_tenant_id(current_user, x_tenant_id)
        app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())

        # Get temple location from MongoDB
        temple_doc = await get_collection("mandir_temples").find_one(
            {"tenant_id": tenant_id, "app_key": app_key}
        )
        if not temple_doc:
            temple_doc = await get_collection("mandir_temples").find_one({"tenant_id": tenant_id})
        if not temple_doc:
            temple_doc = {}

        # Get panchang display settings for overrides
        settings_doc = await get_collection("mandir_panchang_settings").find_one(
            {"tenant_id": tenant_id, "app_key": app_key}
        ) or {}

        latitude, longitude, city = _resolve_panchang_location(
            settings_doc,
            temple_doc,
            city_name=city_name,
            latitude=latitude,
            longitude=longitude,
        )

        # Calculate panchang using Swiss Ephemeris
        panchang_service = PanchangService()
        panchang_data = panchang_service.calculate_panchang(target_dt, latitude, longitude, city)

        return panchang_data
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format. Use YYYY-MM-DD: {str(e)}")
    except Exception as e:
        logger.error(f"Error calculating panchang for date {target_date}: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to calculate panchang: {str(e)}"
        )



# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: Pincode lookup
# ROUTES : GET /pincode/lookup
# ════════════════════════════════════════════════════════════════════════

@router.get("/pincode/lookup")
async def mandir_pincode_lookup(pincode: str = Query(...), _current_user: dict = Depends(get_current_user)):
    normalized = _normalize_pincode(pincode)
    if len(normalized) != 6:
        return {"pincode": normalized, "city": None, "state": None, "country": "India", "found": False}

    city, state = await _lookup_pincode_city_state(normalized)
    found = bool(city and state)

    return {
        "pincode": normalized,
        "city": city if found else None,
        "state": state if found else None,
        "country": "India",
        "found": found,
    }

# ════════════════════════════════════════════════════════════════════════
# Reports routes moved to routes/reports.py
# (docs/operations/LARGE_FILE_MODULARIZATION_PLAN.md); registered via import at end of module.

# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: Role permissions (GET list / PUT / GET assignable)
# ROUTES : GET /role-permissions  PUT .../role-permissions/{role_key}  GET .../assignable
# ════════════════════════════════════════════════════════════════════════

@router.get("/role-permissions")
async def mandir_role_permissions(
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    default_roles = [
        {"role_key": "president", "display_name": "President", "is_enabled": True},
        {"role_key": "secretary", "display_name": "Secretary", "is_enabled": True},
        {"role_key": "treasurer", "display_name": "Treasurer", "is_enabled": True},
        {"role_key": "counter_clerk", "display_name": "Counter Clerk", "is_enabled": True},
        {"role_key": "accounts_clerk", "display_name": "Accounts Clerk", "is_enabled": True},
        {"role_key": "priest_operator", "display_name": "Priest / Temple Operator", "is_enabled": True},
    ]
    docs = await get_collection("mandir_role_permissions").find({"tenant_id": tenant_id, "app_key": app_key}).to_list(length=200)
    by_role = {str(doc.get("role_key") or ""): doc for doc in docs}
    roles = []
    for base in default_roles:
        current = by_role.get(base["role_key"]) or {}
        roles.append(
            {
                "role_key": base["role_key"],
                "display_name": base["display_name"],
                "is_enabled": bool(current.get("is_enabled", base["is_enabled"])),
                "module_permissions": current.get("module_permissions") or {},
                "action_permissions": current.get("action_permissions") or {},
            }
        )
    return {
        "modules": [],
        "actions": [],
        "roles": roles,
        "policy_notice": "Accounting transactions should be reversed with audit reason instead of hard delete.",
    }


@router.put("/role-permissions/{role_key}", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)
async def mandir_role_permissions_update(
    role_key: str,
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "tenant_id": tenant_id,
        "app_key": app_key,
        "role_key": str(role_key).strip().lower(),
        "display_name": str(payload.get("display_name") or role_key).strip(),
        "is_enabled": bool(payload.get("is_enabled", True)),
        "module_permissions": payload.get("module_permissions") or {},
        "action_permissions": payload.get("action_permissions") or {},
        "updated_at": now,
    }
    await get_collection("mandir_role_permissions").update_one(
        {"tenant_id": tenant_id, "app_key": app_key, "role_key": doc["role_key"]},
        {
            "$set": doc,
            "$setOnInsert": {
                "created_at": now,
            },
        },
        upsert=True,
    )
    return {"role": doc}


@router.get("/role-permissions/assignable")
async def mandir_role_permissions_assignable(_current_user: dict = Depends(get_current_user)):
    return {
        "roles": [
            {"role_key": "treasurer", "display_name": "Treasurer"},
            {"role_key": "counter_clerk", "display_name": "Counter Clerk"},
            {"role_key": "accounts_clerk", "display_name": "Accounts Clerk"},
            {"role_key": "priest_operator", "display_name": "Priest / Temple Operator"},
        ]
    }


# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: Setup wizard + Temples management (CRUD / activate / deactivate / onboard / upload / modules config)
# ROUTES : GET /setup-wizard/status  GET /temples  POST /temples/{id}/activate|deactivate|remove  POST /temples/onboard  POST .../upload  GET|PUT .../modules/config
# ════════════════════════════════════════════════════════════════════════

@router.get("/setup-wizard/status")
async def mandir_setup_wizard_status(_current_user: dict = Depends(get_current_user)):
    return {"completed": False, "steps": []}


@router.get("/temples/")
@router.get("/temples")
async def mandir_temples(
    _current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    app_key = resolve_app_key((x_app_key or _current_user.get("app_key") or "mandirmitra").strip())
    if _is_platform_super_admin(_current_user):
        rows = await list_mandir_temples(app_key=app_key, limit=500)
    else:
        tenant_id = resolve_tenant_id(_current_user, x_tenant_id)
        rows = await list_mandir_temples(tenant_id=tenant_id, app_key=app_key, limit=20)

    if rows:
        return [_sanitize_mongo_doc(row) for row in rows]

    if _is_platform_super_admin(_current_user):
        return []

    return []




async def _resolve_temple_target_tenant(
    temple_id: int,
    *,
    current_user: dict,
    x_tenant_id: str | None,
    app_key: str = "mandirmitra",
) -> str:
    target_tenant_id = await resolve_tenant_by_temple_id(temple_id, app_key=app_key)
    if not target_tenant_id:
        raise HTTPException(status_code=404, detail="Temple not found")

    actor_tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    if not _is_platform_super_admin(current_user) and actor_tenant_id != target_tenant_id:
        raise HTTPException(status_code=403, detail="Forbidden for this tenant")

    return target_tenant_id


_MANDIR_TEMPLE_PURGE_COLLECTIONS = [
    "mandir_temples",
    "mandir_donations",
    "mandir_sevas",
    "mandir_seva_bookings",
    "mandir_devotees",
    "mandir_bank_accounts",
    "mandir_bank_statements",
    "mandir_bank_statement_entries",
    "mandir_bank_unmatched_entries",
    "mandir_inventory_items",
    "mandir_journal_entries",
    "mandir_public_payments",
    "mandir_upi_payments",
    "mandir_role_permissions",
    "mandir_panchang_settings",
]

_MANDIR_TEMPLE_BUSINESS_COLLECTIONS = [
    name
    for name in _MANDIR_TEMPLE_PURGE_COLLECTIONS
    if name not in {"mandir_temples", "mandir_role_permissions", "mandir_panchang_settings"}
]


def _mandir_tenant_app_query(tenant_id: str, app_key: str) -> dict[str, Any]:
    query: dict[str, Any] = {"tenant_id": tenant_id}
    if app_key == "mandirmitra":
        query["$or"] = [
            {"app_key": app_key},
            {"app_key": {"$exists": False}},
            {"app_key": None},
            {"app_key": ""},
        ]
    else:
        query["app_key"] = app_key
    return query


async def _mandir_temple_collection_counts(tenant_id: str, app_key: str) -> dict[str, int]:
    query = _mandir_tenant_app_query(tenant_id, app_key)
    counts: dict[str, int] = {}
    for name in _MANDIR_TEMPLE_PURGE_COLLECTIONS:
        try:
            counts[name] = int(await get_collection(name).count_documents(query))
        except Exception:
            counts[name] = 0
    return counts


@router.post("/temples/{temple_id}/activate", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)
async def mandir_activate_temple(
    temple_id: int,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    tenant_id = await _resolve_temple_target_tenant(temple_id, current_user=current_user, x_tenant_id=x_tenant_id, app_key=app_key)
    now = datetime.now(timezone.utc).isoformat()
    await get_collection("mandir_temples").update_one(
        {"tenant_id": tenant_id, "app_key": app_key},
        {"$set": {"is_active": True, "updated_at": now}},
        upsert=False,
    )
    doc = await get_collection("mandir_temples").find_one({"tenant_id": tenant_id, "app_key": app_key}) or {"tenant_id": tenant_id}
    return {"status": "activated", "temple_id": temple_id, "temple": _sanitize_mongo_doc(doc)}


@router.post("/temples/{temple_id}/deactivate", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)
async def mandir_deactivate_temple(
    temple_id: int,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    tenant_id = await _resolve_temple_target_tenant(temple_id, current_user=current_user, x_tenant_id=x_tenant_id, app_key=app_key)
    now = datetime.now(timezone.utc).isoformat()
    await get_collection("mandir_temples").update_one(
        {"tenant_id": tenant_id, "app_key": app_key},
        {"$set": {"is_active": False, "updated_at": now}},
        upsert=False,
    )
    doc = await get_collection("mandir_temples").find_one({"tenant_id": tenant_id, "app_key": app_key}) or {"tenant_id": tenant_id}
    return {"status": "deactivated", "temple_id": temple_id, "temple": _sanitize_mongo_doc(doc)}


@router.get("/temples/{temple_id}/remove-preview")
async def mandir_remove_temple_preview(
    temple_id: int,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    if not _is_platform_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Only platform administrators can preview temple removal")

    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    target_tenant_id = await _resolve_temple_target_tenant(temple_id, current_user=current_user, x_tenant_id=x_tenant_id, app_key=app_key)
    counts = await _mandir_temple_collection_counts(target_tenant_id, app_key)
    business_counts = {name: count for name, count in counts.items() if name in _MANDIR_TEMPLE_BUSINESS_COLLECTIONS and count > 0}
    return {
        "temple_id": temple_id,
        "tenant_id": target_tenant_id,
        "counts": counts,
        "has_business_data": bool(business_counts),
        "business_counts": business_counts,
        "can_remove_placeholder_safely": not business_counts,
    }


@router.delete("/temples/{temple_id}/remove", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)
async def mandir_remove_temple(
    temple_id: int,
    payload: dict[str, Any] | None = None,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    if not _is_platform_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Only platform administrators can remove temples")

    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    target_tenant_id = await _resolve_temple_target_tenant(temple_id, current_user=current_user, x_tenant_id=x_tenant_id, app_key=app_key)
    expected = f"DELETE {temple_id}"
    confirm_text = str((payload or {}).get("confirm_text") or "").strip()
    if confirm_text != expected:
        raise HTTPException(status_code=400, detail=f"Confirmation text mismatch. Expected: {expected}")

    counts = await _mandir_temple_collection_counts(target_tenant_id, app_key)
    business_counts = {name: count for name, count in counts.items() if name in _MANDIR_TEMPLE_BUSINESS_COLLECTIONS and count > 0}
    allow_data_delete = bool((payload or {}).get("allow_data_delete"))
    if business_counts and not allow_data_delete:
        raise HTTPException(
            status_code=409,
            detail={
                "message": "Temple has business data. Review remove-preview before deleting.",
                "tenant_id": target_tenant_id,
                "business_counts": business_counts,
            },
        )

    delete_query = _mandir_tenant_app_query(target_tenant_id, app_key)
    deleted_counts: dict[str, int] = {}
    for name in _MANDIR_TEMPLE_PURGE_COLLECTIONS:
        try:
            result = await get_collection(name).delete_many(delete_query)
            deleted_counts[name] = int(getattr(result, "deleted_count", 0) or 0)
        except Exception:
            deleted_counts[name] = 0

    return {
        "status": "removed",
        "temple_id": temple_id,
        "tenant_id": target_tenant_id,
        "deleted": deleted_counts,
    }

@router.post("/temples/onboard", response_model=MandirFirstLoginOnboardingResponse)
@router.post("/onboarding/first-login", response_model=MandirFirstLoginOnboardingResponse)
async def mandir_temples_onboard(
    payload: MandirFirstLoginOnboardingRequest,
    request: Request,
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
    x_onboarding_token: str | None = Header(default=None, alias="X-Onboarding-Token"),
):
    from app.config import get_settings as _get_settings
    _settings = _get_settings()
    required_secret = _settings.MANDIR_ONBOARDING_SECRET
    if required_secret:
        provided = (x_onboarding_token or "").strip()
        if not provided or provided != required_secret:
            logger.warning(
                "Onboarding attempt rejected: missing/invalid X-Onboarding-Token from %s",
                request.client.host if request.client else "unknown",
            )
            raise HTTPException(status_code=403, detail="Invalid or missing onboarding token")
    else:
        if str(_settings.ENVIRONMENT or "").strip().lower() in {"production", "prod"}:
            raise HTTPException(status_code=503, detail="Mandir onboarding is not configured")
        logger.info(
            "Onboarding endpoint called without secret enforcement (MANDIR_ONBOARDING_SECRET not set). "
            "Set this env var in production to protect this endpoint."
        )
    if not str(x_app_key or "").strip():
        raise HTTPException(status_code=400, detail="X-App-Key header is required")
    app_key = resolve_app_key(str(x_app_key).strip())
    return await create_mandir_first_login_onboarding(payload, app_key=app_key)


@router.post("/temples/upload", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)
async def mandir_temples_upload(_payload: dict[str, Any], _current_user: dict = Depends(get_current_user)):
    return _ok("temples/upload")


@router.get("/compliance/donations/config", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)
async def get_mandir_donation_compliance_config(
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    context = resolve_mandir_tenant(
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
        operation="donation compliance configuration read",
    )
    doc = await get_collection("mandir_donation_compliance_config").find_one(
        {"tenant_id": context.tenant_id, "app_key": context.app_key}
    )
    return donation_compliance_config_view(doc)


@router.put("/compliance/donations/config", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)
async def update_mandir_donation_compliance_config(
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    context = resolve_mandir_tenant(
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
        operation="donation compliance configuration update",
    )
    await _assert_platform_can_write_tenant(
        current_user, tenant_id=context.tenant_id, app_key=context.app_key
    )
    config = validate_donation_compliance_config(payload)
    now = datetime.now(timezone.utc).isoformat()
    stored = {**config, "updated_at": now, "updated_by": _mandir_actor_id(current_user)}
    collection = get_collection("mandir_donation_compliance_config")
    await collection.update_one(
        {"tenant_id": context.tenant_id, "app_key": context.app_key},
        {
            "$set": stored,
            "$setOnInsert": {
                "tenant_id": context.tenant_id,
                "app_key": context.app_key,
                "created_at": now,
            },
        },
        upsert=True,
    )
    try:
        await log_audit_event(
            tenant_id=context.tenant_id,
            user_id=_mandir_actor_id(current_user),
            product=context.app_key,
            action="donation_compliance_config_updated",
            entity_type="mandir_donation_compliance_config",
            entity_id=f"{context.tenant_id}:{context.app_key}",
            old_value=None,
            new_value={
                "enable_80g": config["enable_80g"],
                "enable_fcra": config["enable_fcra"],
                "approval_valid_from": config.get("approval_valid_from"),
                "approval_valid_to": config.get("approval_valid_to"),
                "fcra_valid_from": config.get("fcra_valid_from"),
                "fcra_valid_to": config.get("fcra_valid_to"),
            },
        )
    except Exception:
        logger.warning("Failed to audit donation compliance configuration update", exc_info=True)
    return donation_compliance_config_view(stored)


async def _mandir_compliance_report(
    *, kind: str, tenant_id: str, app_key: str, from_date: date, to_date: date, session: AsyncSession
) -> dict[str, Any]:
    rows = await posted_donations(
        session,
        tenant_id=tenant_id,
        app_key=app_key,
        from_date=from_date,
        to_date=to_date,
    )
    items: list[dict[str, Any]] = []
    status_counts: dict[str, int] = {}
    for row in rows:
        if kind == "80g" and not bool(row.get("request_80g")):
            continue
        if kind == "fcra" and not bool(row.get("is_foreign_contribution")):
            continue
        public = compliance_public_fields(row)
        status_field = "80g_eligibility_status" if kind == "80g" else "fcra_status"
        status = str(public[status_field])
        status_counts[status] = status_counts.get(status, 0) + 1
        items.append({
            "donation_id": row.get("donation_id"),
            "receipt_number": row.get("receipt_number"),
            "date": row.get("date"),
            "amount": row.get("amount"),
            "payment_mode": row.get("payment_mode"),
            "devotee_name": row.get("devotee_name"),
            **public,
        })
    return {
        "report_kind": kind,
        "filing_artifact": False,
        "filing_notice": "Readiness report only; this is not an official government filing or certificate.",
        "from_date": from_date.isoformat(),
        "to_date": to_date.isoformat(),
        "count": len(items),
        "status_counts": status_counts,
        "items": items,
    }


@router.get("/reports/compliance/80g", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)
async def mandir_report_80g_readiness(
    from_date: date = Query(...),
    to_date: date = Query(...),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    if from_date > to_date:
        raise HTTPException(status_code=422, detail="from_date cannot be greater than to_date")
    context = resolve_mandir_tenant(
        current_user=current_user, x_tenant_id=x_tenant_id, x_app_key=x_app_key,
        operation="80G readiness reporting",
    )
    return await _mandir_compliance_report(
        kind="80g", tenant_id=context.tenant_id, app_key=context.app_key,
        from_date=from_date, to_date=to_date, session=session,
    )


@router.get("/reports/compliance/fcra", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)
async def mandir_report_fcra_readiness(
    from_date: date = Query(...),
    to_date: date = Query(...),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    if from_date > to_date:
        raise HTTPException(status_code=422, detail="from_date cannot be greater than to_date")
    context = resolve_mandir_tenant(
        current_user=current_user, x_tenant_id=x_tenant_id, x_app_key=x_app_key,
        operation="FCRA readiness reporting",
    )
    return await _mandir_compliance_report(
        kind="fcra", tenant_id=context.tenant_id, app_key=context.app_key,
        from_date=from_date, to_date=to_date, session=session,
    )


@router.get("/temples/modules/config")
async def mandir_temples_module_config(
    _current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
    temple_id: int | None = Query(default=None),
):
    app_key = resolve_app_key((x_app_key or _current_user.get("app_key") or "mandirmitra").strip())
    tenant_id = await _resolve_tenant_for_mandir_request(_current_user, x_tenant_id, temple_id, app_key)
    col = get_collection("mandir_temples")
    doc = await col.find_one({"tenant_id": tenant_id, "app_key": app_key}) or {}
    compliance_doc = await get_collection("mandir_donation_compliance_config").find_one(
        {"tenant_id": tenant_id, "app_key": app_key}
    )
    compliance = donation_compliance_config_view(compliance_doc)
    return {
        "module_donations_enabled": bool(doc.get("module_donations_enabled", True)),
        "module_sevas_enabled": bool(doc.get("module_sevas_enabled", True)),
        "module_inventory_enabled": bool(doc.get("module_inventory_enabled", False)),
        "module_assets_enabled": bool(doc.get("module_assets_enabled", False)),
        "module_hr_enabled": bool(doc.get("module_hr_enabled", False)),
        "module_hundi_enabled": bool(doc.get("module_hundi_enabled", False)),
        "module_accounting_enabled": bool(doc.get("module_accounting_enabled", True)),
        "module_panchang_enabled": bool(doc.get("module_panchang_enabled", True)),
        "enable_80g": compliance["enable_80g"],
        "enable_fcra": compliance["enable_fcra"],
    }


@router.put("/temples/modules/config", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)
async def mandir_temples_module_config_update(
    payload: dict[str, Any],
    _current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
    temple_id: int | None = Query(default=None),
):
    tenant_id = await _resolve_tenant_for_mandir_request(_current_user, x_tenant_id, temple_id)
    app_key = resolve_app_key((x_app_key or _current_user.get("app_key") or "mandirmitra").strip())
    assigned_temple_id = await ensure_temple_numeric_id(tenant_id, app_key=app_key)
    col = get_collection("mandir_temples")

    allowed_keys = {
        "module_donations_enabled",
        "module_sevas_enabled",
        "module_inventory_enabled",
        "module_assets_enabled",
        "module_hr_enabled",
        "module_hundi_enabled",
        "module_accounting_enabled",
        "module_panchang_enabled",
    }
    update = {key: bool(payload.get(key)) for key in allowed_keys if key in payload}
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    update["id"] = assigned_temple_id
    update["temple_id"] = assigned_temple_id

    await col.update_one(
        {"tenant_id": tenant_id, "app_key": app_key},
        {
            "$set": update,
            "$setOnInsert": {
                "tenant_id": tenant_id,
                "app_key": app_key,
                "created_at": datetime.now(timezone.utc).isoformat(),
            },
        },
        upsert=True,
    )

    doc = await col.find_one({"tenant_id": tenant_id, "app_key": app_key}) or {}
    return {
        "module_donations_enabled": bool(doc.get("module_donations_enabled", True)),
        "module_sevas_enabled": bool(doc.get("module_sevas_enabled", True)),
        "module_inventory_enabled": bool(doc.get("module_inventory_enabled", False)),
        "module_assets_enabled": bool(doc.get("module_assets_enabled", False)),
        "module_hr_enabled": bool(doc.get("module_hr_enabled", False)),
        "module_hundi_enabled": bool(doc.get("module_hundi_enabled", False)),
        "module_accounting_enabled": bool(doc.get("module_accounting_enabled", True)),
        "module_panchang_enabled": bool(doc.get("module_panchang_enabled", True)),
    }


def _mandir_upi_config_view(doc: dict[str, Any], temple_id: int) -> dict[str, Any]:
    upi_id = str(doc.get("upi_id") or "").strip()
    payee_name = str(doc.get("upi_payee_name") or doc.get("trust_name") or doc.get("temple_name") or doc.get("name") or "Temple").strip()
    currency = str(doc.get("upi_currency") or "INR").strip().upper() or "INR"
    return {
        "temple_id": int(temple_id),
        "upi_public_enabled": bool(doc.get("upi_public_enabled", False)),
        "upi_id": upi_id,
        "upi_payee_name": payee_name,
        "upi_currency": currency,
        "upi_qr_note": str(doc.get("upi_qr_note") or "").strip() or None,
        "qr_code_image_url": str(doc.get("qr_code_image_url") or "").strip() or None,
        "admin_whatsapp": str(doc.get("admin_whatsapp") or "").strip() or None,
    }



# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: UPI payments config (GET / PUT)
# ROUTES : GET /upi-payments/config  PUT .../config
# ════════════════════════════════════════════════════════════════════════

@router.get("/upi-payments/config")
async def mandir_upi_payments_config(
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
    temple_id: int | None = Query(default=None),
):
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    tenant_id = await _resolve_tenant_for_mandir_request(current_user, x_tenant_id, temple_id, app_key=app_key)
    assigned_temple_id = await ensure_temple_numeric_id(tenant_id, app_key=app_key)
    doc = await get_collection("mandir_temples").find_one({"tenant_id": tenant_id, "app_key": app_key}) or {}
    return _mandir_upi_config_view(doc, assigned_temple_id)


@router.put("/upi-payments/config", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)
async def mandir_upi_payments_config_update(
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
    temple_id: int | None = Query(default=None),
):
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    tenant_id = await _resolve_tenant_for_mandir_request(current_user, x_tenant_id, temple_id, app_key=app_key)
    assigned_temple_id = await ensure_temple_numeric_id(tenant_id, app_key=app_key)

    upi_id = str(payload.get("upi_id") or "").strip().lower()
    upi_payee_name = str(payload.get("upi_payee_name") or "").strip()
    upi_qr_note = str(payload.get("upi_qr_note") or "").strip()
    upi_currency = str(payload.get("upi_currency") or "INR").strip().upper() or "INR"

    update = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "id": assigned_temple_id,
        "temple_id": assigned_temple_id,
    }
    if "upi_public_enabled" in payload:
        update["upi_public_enabled"] = bool(payload.get("upi_public_enabled"))
    if "upi_id" in payload:
        update["upi_id"] = upi_id or None
    if "upi_payee_name" in payload:
        update["upi_payee_name"] = upi_payee_name or None
    if "upi_qr_note" in payload:
        update["upi_qr_note"] = upi_qr_note or None
    if "upi_currency" in payload:
        update["upi_currency"] = upi_currency
    if "qr_code_image_url" in payload:
        update["qr_code_image_url"] = str(payload.get("qr_code_image_url") or "").strip() or None
    if "admin_whatsapp" in payload:
        update["admin_whatsapp"] = str(payload.get("admin_whatsapp") or "").strip() or None

    col = get_collection("mandir_temples")
    await col.update_one(
        {"tenant_id": tenant_id, "app_key": app_key},
        {
            "$set": update,
            "$setOnInsert": {
                "tenant_id": tenant_id,
                "app_key": app_key,
                "created_at": datetime.now(timezone.utc).isoformat(),
            },
        },
        upsert=True,
    )
    doc = await col.find_one({"tenant_id": tenant_id}) or {}
    return _mandir_upi_config_view(doc, assigned_temple_id)


@router.get("/public/temples/{temple_id}/upi-intent")
async def mandir_public_upi_intent(
    temple_id: int,
    amount: float | None = Query(default=None, ge=0),
    purpose: str | None = Query(default=None),
    reference: str | None = Query(default=None),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    app_key = resolve_app_key((x_app_key or "mandirmitra").strip())
    tenant_id = await resolve_tenant_by_temple_id(temple_id, app_key=app_key)
    if not tenant_id:
        raise HTTPException(status_code=404, detail="Temple not found")

    temples = get_collection("mandir_temples")
    temple_doc = await temples.find_one({"tenant_id": tenant_id, "app_key": app_key}) or {}
    if not temple_doc:
        temple_doc = await temples.find_one({"tenant_id": tenant_id}) or {}

    if not bool(temple_doc.get("upi_public_enabled", False)):
        raise HTTPException(status_code=404, detail="Public UPI is not enabled for this temple")

    upi_id = str(temple_doc.get("upi_id") or "").strip().lower()
    if not upi_id:
        raise HTTPException(status_code=404, detail="Temple UPI ID is not configured")

    payee_name = str(
        temple_doc.get("upi_payee_name")
        or temple_doc.get("trust_name")
        or temple_doc.get("temple_name")
        or temple_doc.get("name")
        or "Temple"
    ).strip()
    currency = str(temple_doc.get("upi_currency") or "INR").strip().upper() or "INR"
    note_text = str(purpose or temple_doc.get("upi_qr_note") or "Temple Offering").strip()
    reference_text = str(reference or "").strip() or None

    intent_uri = _build_upi_intent_uri(
        upi_id=upi_id,
        payee_name=payee_name,
        amount=amount,
        note=note_text,
        reference=reference_text,
        currency=currency,
    )

    return {
        "temple_id": temple_id,
        "upi_id": upi_id,
        "payee_name": payee_name,
        "currency": currency,
        "amount": amount,
        "purpose": note_text,
        "reference": reference_text,
        "intent_uri": intent_uri,
        "qr_payload": intent_uri,
    }


# ---------------------------------------------------------------------------
# VERSION ENDPOINT
# ---------------------------------------------------------------------------


# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: Version + public UPI intent
# ROUTES : GET /mandir/version  GET /public/temples/{temple_id}/upi-intent
# ════════════════════════════════════════════════════════════════════════

@router.get("/mandir/version")
async def mandir_get_version():
    """Get MandirMitra version info. No authentication required."""
    from app.config import get_settings
    from datetime import datetime
    settings = get_settings()
    return {
        "app": "MandirMitra",
        "version": settings.APP_VERSION,
        "released_at": "2026-04-14T00:00:00Z",  # Release date of v1.2.0
        "features": [
            "Quick Ticket Counter mode",
            "Seva renewal reminders (Email + SMS)",
            "Public payment idempotency & audit logging",
            "Rate limiting on public endpoints",
            "Multilingual support (en/kn/hi) for public portal",
            "Indic font rendering for PDFs"
        ]
    }


# ---------------------------------------------------------------------------
# PUBLIC SEVA PAYMENT ENDPOINTS  (no authentication required)
# ---------------------------------------------------------------------------

def _normalize_public_donation_categories(raw: Any, *, fallback_to_default: bool = True) -> list[dict[str, str]]:
    if not isinstance(raw, list):
        return [dict(item) for item in _DEFAULT_PUBLIC_DONATION_CATEGORIES] if fallback_to_default else []

    normalized: list[dict[str, str]] = []
    seen: set[str] = set()
    for index, item in enumerate(raw):
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "").strip()
        if not name:
            continue
        category_id = str(item.get("id") or "").strip().lower()
        if not category_id:
            category_id = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_") or f"category_{index + 1}"
        base_id = category_id
        suffix = 2
        while category_id in seen:
            category_id = f"{base_id}_{suffix}"
            suffix += 1
        seen.add(category_id)
        normalized.append(
            {
                "id": category_id[:80],
                "name": name[:120],
                "description": str(item.get("description") or "").strip()[:300],
            }
        )

    if normalized or not fallback_to_default:
        return normalized
    return [dict(item) for item in _DEFAULT_PUBLIC_DONATION_CATEGORIES]



# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: Public endpoints (temples list / info / sevas / autofill / pincode / donation-categories)
# ROUTES : GET /public/temples  GET .../info  GET .../sevas  GET .../devotee/autofill  GET .../donation-categories  GET /public/location/pincode
# ════════════════════════════════════════════════════════════════════════

@router.get("/public/temples")
async def mandir_public_list_temples(
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    """List temples that have public payments enabled (for temple selector on public page)."""
    app_key = resolve_app_key((x_app_key or "mandirmitra").strip())
    col = get_collection("mandir_temples")
    visibility_query: dict[str, Any] = {
        "$or": [
            {"upi_public_enabled": True},
            {"upi_id": {"$exists": True, "$ne": None, "$ne": ""}},
        ]
    }
    if app_key == "mandirmitra":
        visibility_query["$and"] = [
            {
                "$or": [
                    {"app_key": app_key},
                    {"app_key": {"$exists": False}},
                    {"app_key": None},
                    {"app_key": ""},
                ]
            }
        ]
    else:
        visibility_query["app_key"] = app_key
    docs = await col.find(visibility_query).to_list(length=100)
    result = []
    for doc in docs:
        temple_id = doc.get("temple_id") or doc.get("id")
        if not temple_id:
            continue
        if app_key == "mandirmitra" and not str(doc.get("app_key") or "").strip():
            await col.update_one({"_id": doc.get("_id")}, {"$set": {"app_key": app_key}}, upsert=False)
        result.append({
            "temple_id": int(temple_id),
            "temple_name": str(doc.get("temple_name") or doc.get("name") or ""),
            "trust_name": str(doc.get("trust_name") or ""),
            "city": str(doc.get("city") or ""),
            "state": str(doc.get("state") or ""),
        })
    return result


@router.get("/public/temples/{temple_id}/info")
async def mandir_public_temple_info(
    temple_id: int,
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    app_key = resolve_app_key((x_app_key or "mandirmitra").strip())
    tenant_id = await resolve_tenant_by_temple_id(temple_id, app_key=app_key)
    if not tenant_id:
        raise HTTPException(status_code=404, detail="Temple not found")

    doc = await get_collection("mandir_temples").find_one({"tenant_id": tenant_id, "app_key": app_key}) or {}
    if not doc:
        raise HTTPException(status_code=404, detail="Temple not found")

    return {
        "temple_id": temple_id,
        "temple_name": str(doc.get("temple_name") or doc.get("name") or ""),
        "trust_name": str(doc.get("trust_name") or ""),
        "address": str(doc.get("address") or ""),
        "city": str(doc.get("city") or ""),
        "state": str(doc.get("state") or ""),
        "upi_id": str(doc.get("upi_id") or "").strip() or None,
        "upi_payee_name": str(doc.get("upi_payee_name") or doc.get("trust_name") or doc.get("temple_name") or ""),
        "qr_code_image_url": str(doc.get("qr_code_image_url") or "").strip() or None,
        "admin_whatsapp": str(doc.get("admin_whatsapp") or "").strip() or None,
        "upi_public_enabled": bool(doc.get("upi_public_enabled", False)),
    }


@router.get("/public/temples/{temple_id}/sevas")
@limiter.limit("30/minute")
async def mandir_public_temple_sevas(
    request: Request,
    temple_id: int,
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    app_key = resolve_app_key((x_app_key or "mandirmitra").strip())
    tenant_id = await resolve_tenant_by_temple_id(temple_id, app_key=app_key)
    if not tenant_id:
        raise HTTPException(status_code=404, detail="Temple not found")

    col = get_collection("mandir_sevas")
    docs = await col.find({
        "tenant_id": tenant_id,
        "app_key": app_key,
        "is_active": {"$ne": False},
    }).sort("seva_name", 1).to_list(length=200)

    return [
        {
            "seva_id": str(doc.get("_id") or doc.get("id") or ""),
            "seva_name": str(doc.get("seva_name") or doc.get("name") or ""),
            "description": str(doc.get("description") or ""),
            "amount": doc.get("amount"),
            "frequency": str(doc.get("frequency") or "one_time"),
            "duration_days": doc.get("duration_days"),
        }
        for doc in docs
        if doc.get("seva_name") or doc.get("name")
    ]


@router.get("/public/temples/{temple_id}/devotee/autofill/{phone}")
@limiter.limit("20/minute")
async def mandir_public_devotee_autofill(
    request: Request,
    temple_id: int,
    phone: str,
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    app_key = resolve_app_key((x_app_key or "mandirmitra").strip())
    tenant_id = await resolve_tenant_by_temple_id(temple_id, app_key=app_key)
    if not tenant_id:
        raise HTTPException(status_code=404, detail="Temple not found")

    normalized = _normalize_phone(phone)
    if not normalized:
        return {"found": False, "devotee": None}

    # Scoped by tenant_id + app_key - mobile linked to THIS temple only
    col = get_collection("mandir_devotees")
    docs = await col.find({
        "tenant_id": tenant_id,
        "app_key": app_key,
        "phone": normalized,
    }).limit(1).to_list(length=1)

    if not docs:
        return {"found": False, "devotee": None}

    doc = docs[0]
    return {
        "found": True,
        "devotee": {
            "name": str(doc.get("name") or ""),
            "city": str(doc.get("city") or ""),
            "state": str(doc.get("state") or ""),
            "pincode": str(doc.get("pincode") or ""),
            "gothra": str(doc.get("gothra") or ""),
            "nakshtra": str(doc.get("nakshtra") or ""),
            "rashi": str(doc.get("rashi") or ""),
        },
    }


@router.get("/public/location/pincode/{pincode}")
async def mandir_public_pincode_lookup(pincode: str):
    if not pincode.isdigit() or len(pincode) != 6:
        raise HTTPException(status_code=400, detail="Invalid pincode. Must be 6 digits.")

    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            resp = await client.get(f"https://api.postalpincode.in/pincode/{pincode}")
        if resp.status_code == 200:
            data = resp.json()
            if data and data[0].get("Status") == "Success":
                post_offices = data[0].get("PostOffice") or []
                if post_offices:
                    po = post_offices[0]
                    return {
                        "found": True,
                        "pincode": pincode,
                        "city": str(po.get("District") or po.get("Name") or ""),
                        "state": str(po.get("State") or ""),
                        "district": str(po.get("District") or ""),
                    }
    except Exception:
        pass

    return {"found": False, "pincode": pincode, "city": "", "state": "", "district": ""}


@router.get("/public/temples/{temple_id}/donation-categories")
async def mandir_public_donation_categories(
    temple_id: int,
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    app_key = resolve_app_key((x_app_key or "mandirmitra").strip())
    tenant_id = await resolve_tenant_by_temple_id(temple_id, app_key=app_key)
    if not tenant_id:
        raise HTTPException(status_code=404, detail="Temple not found")

    temple_doc = await get_collection("mandir_temples").find_one({"tenant_id": tenant_id, "app_key": app_key}) or {}
    return _normalize_public_donation_categories(temple_doc.get("donation_categories"))



# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: Public seva payments (create / status)
# ROUTES : POST /public/temples/{temple_id}/seva-payments  GET /public/payments/{payment_id}/status
# ════════════════════════════════════════════════════════════════════════

@router.post("/public/temples/{temple_id}/seva-payments")
@limiter.limit("10/minute")
async def mandir_public_create_seva_payment(
    temple_id: int,
    payload: dict[str, Any],
    request: Request,
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    app_key = resolve_app_key((x_app_key or "mandirmitra").strip())
    tenant_id = await resolve_tenant_by_temple_id(temple_id, app_key=app_key)
    if not tenant_id:
        raise HTTPException(status_code=404, detail="Temple not found")

    # Idempotency: if client sends idempotency_key and a matching pending record
    # exists (submitted in the last 10 minutes), return it without creating a duplicate.
    idempotency_key = str(payload.get("idempotency_key") or "").strip()
    if idempotency_key:
        from datetime import timedelta
        cutoff = (datetime.now(timezone.utc) - timedelta(minutes=10)).isoformat()
        existing = await get_collection("mandir_public_payments").find_one({
            "tenant_id": tenant_id,
            "app_key": app_key,
            "idempotency_key": idempotency_key,
            "created_at": {"$gte": cutoff},
        })
        if existing:
            temple_doc_i = await get_collection("mandir_temples").find_one({"tenant_id": tenant_id, "app_key": app_key}) or {}
            return {
                "payment_id": existing["id"][:8].upper(),
                "full_payment_id": existing["id"],
                "status": existing.get("status", "pending"),
                "payment_type": existing.get("payment_type", "seva"),
                "seva_name": existing.get("seva_name", ""),
                "amount": existing.get("amount"),
                "upi_id": str(temple_doc_i.get("upi_id") or "").strip() or None,
                "upi_payee_name": str(temple_doc_i.get("upi_payee_name") or temple_doc_i.get("trust_name") or "").strip() or None,
                "qr_code_image_url": str(temple_doc_i.get("qr_code_image_url") or "").strip() or None,
                "admin_whatsapp": str(temple_doc_i.get("admin_whatsapp") or "").strip() or None,
                "whatsapp_link": existing.get("whatsapp_link"),
                "whatsapp_message_template": existing.get("whatsapp_message_template"),
                "message": "Payment already submitted. Please complete UPI payment and send WhatsApp confirmation.",
                "_idempotent": True,
            }

    payment_type = str(payload.get("payment_type") or "seva").strip().lower()
    if payment_type not in ("seva", "donation"):
        payment_type = "seva"

    # Validate required fields
    phone_raw = str(payload.get("phone") or payload.get("mobile") or "").strip()
    normalized_phone = _normalize_phone(phone_raw)
    if not normalized_phone:
        raise HTTPException(status_code=400, detail="Valid mobile number is required")

    name = str(payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")

    if payment_type == "donation":
        category_id = str(payload.get("category_id") or "").strip()
        category_name = str(payload.get("category_name") or "").strip()
        if not category_name:
            raise HTTPException(status_code=400, detail="Donation category is required")
        seva_id = None
        seva_name = category_name
    else:
        seva_id = str(payload.get("seva_id") or "").strip()
        seva_name = str(payload.get("seva_name") or "").strip()
        if not seva_name:
            raise HTTPException(status_code=400, detail="Seva selection is required")

    now = datetime.now(timezone.utc).isoformat()

    # Upsert devotee record - scoped to this temple's tenant_id
    devotee_col = get_collection("mandir_devotees")
    devotee_update = {
        "name": name,
        "phone": normalized_phone,
        "email": str(payload.get("email") or "").strip() or None,
        "address": str(payload.get("address") or "").strip() or None,
        "city": str(payload.get("city") or "").strip() or None,
        "state": str(payload.get("state") or "").strip() or None,
        "pincode": str(payload.get("pincode") or "").strip() or None,
        "gothra": str(payload.get("gothra") or "").strip() or None,
        "nakshtra": str(payload.get("nakshtra") or "").strip() or None,
        "rashi": str(payload.get("rashi") or "").strip() or None,
        "updated_at": now,
    }
    devotee_update_clean = {k: v for k, v in devotee_update.items() if v is not None}

    await devotee_col.update_one(
        {"tenant_id": tenant_id, "app_key": app_key, "phone": normalized_phone},
        {
            "$set": devotee_update_clean,
            "$setOnInsert": {
                "id": str(uuid4()),
                "tenant_id": tenant_id,
                "app_key": app_key,
                "created_at": now,
                "verified": False,
            },
        },
        upsert=True,
    )

    # Create payment record with pending status
    payment_id = str(uuid4())
    payment_doc = {
        "id": payment_id,
        "temple_id": temple_id,
        "tenant_id": tenant_id,
        "app_key": app_key,
        "payment_type": payment_type,
        "seva_id": seva_id or None,
        "seva_name": seva_name,
        "amount": float(payload.get("amount") or 0) or None,
        "devotee_name": name,
        "devotee_phone": normalized_phone,
        "devotee_email": str(payload.get("email") or "").strip() or None,
        "gothra": str(payload.get("gothra") or "").strip() or None if payment_type == "seva" else None,
        "nakshtra": str(payload.get("nakshtra") or "").strip() or None if payment_type == "seva" else None,
        "rashi": str(payload.get("rashi") or "").strip() or None if payment_type == "seva" else None,
        "status": "pending",
        "utr_reference": None,
        "verified_at": None,
        "verified_by": None,
        "created_at": now,
        "source_ip": str(request.client.host if request.client else ""),
        "idempotency_key": idempotency_key or None,
    }

    await get_collection("mandir_public_payments").insert_one(payment_doc)

    # Formal audit log for every public payment submission
    from app.core.audit.service import log_audit_event
    try:
        await log_audit_event(
            tenant_id=tenant_id,
            user_id=f"public:{normalized_phone}",
            product="mandirmitra",
            action="public_payment_submitted",
            entity_type="mandir_public_payment",
            entity_id=payment_id,
            new_value={
                "payment_type": payment_type,
                "seva_name": seva_name,
                "amount": payload.get("amount"),
                "devotee_phone": normalized_phone,
            },
            ip_address=str(request.client.host if request.client else ""),
        )
    except Exception:
        pass  # Audit is best-effort; never block the payment flow

    # Build WhatsApp message template
    temple_doc = await get_collection("mandir_temples").find_one({"tenant_id": tenant_id, "app_key": app_key}) or {}
    admin_whatsapp = str(temple_doc.get("admin_whatsapp") or "").strip()
    upi_id = str(temple_doc.get("upi_id") or "").strip()
    amount_str = f"Rs.{payload.get('amount')}" if payload.get("amount") else ""
    payment_label = "Donation" if payment_type == "donation" else "Seva payment"
    whatsapp_message = (
        f"Namaste, I have made the {seva_name} {payment_label} {amount_str}.\n"
        f"UTR/Reference: [PASTE UTR HERE]\n"
        f"Name: {name}\n"
        f"Mobile: {normalized_phone}\n"
        f"Payment ID: {payment_id[:8].upper()}"
    )
    whatsapp_link = None
    if admin_whatsapp:
        from urllib.parse import quote
        whatsapp_link = f"https://wa.me/{admin_whatsapp.replace('+', '').replace(' ', '')}?text={quote(whatsapp_message)}"

    # Store whatsapp details on payment doc for idempotency replay
    await get_collection("mandir_public_payments").update_one(
        {"id": payment_id, "tenant_id": tenant_id, "app_key": app_key},
        {"$set": {"whatsapp_link": whatsapp_link, "whatsapp_message_template": whatsapp_message}},
    )

    return {
        "payment_id": payment_id[:8].upper(),
        "full_payment_id": payment_id,
        "status": "pending",
        "payment_type": payment_type,
        "seva_name": seva_name,
        "amount": payload.get("amount"),
        "upi_id": upi_id or None,
        "upi_payee_name": str(temple_doc.get("upi_payee_name") or temple_doc.get("trust_name") or temple_doc.get("temple_name") or "").strip() or None,
        "qr_code_image_url": str(temple_doc.get("qr_code_image_url") or "").strip() or None,
        "admin_whatsapp": admin_whatsapp or None,
        "whatsapp_link": whatsapp_link,
        "whatsapp_message_template": whatsapp_message,
        "message": "Devotee details saved. Please complete payment via UPI and send WhatsApp confirmation to the temple admin.",
    }


@router.get("/public/payments/{payment_id}/status")
async def mandir_public_payment_status(
    payment_id: str,
    temple_id: int = Query(..., ge=1),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    app_key = resolve_app_key((x_app_key or "mandirmitra").strip())
    tenant_id = await resolve_tenant_by_temple_id(temple_id, app_key=app_key)
    if not tenant_id:
        raise HTTPException(status_code=404, detail="Temple not found")

    normalized_payment_id = str(payment_id or "").strip()
    if not normalized_payment_id:
        raise HTTPException(status_code=400, detail="Payment ID is required")

    col = get_collection("mandir_public_payments")
    doc = await col.find_one({
        "tenant_id": tenant_id,
        "app_key": app_key,
        "id": normalized_payment_id,
    })
    if not doc:
        raise HTTPException(status_code=404, detail="Payment not found")
    return {
        "payment_id": str(doc.get("id") or "")[:8].upper(),
        "seva_name": doc.get("seva_name"),
        "amount": doc.get("amount"),
        "status": doc.get("status", "pending"),
        "verified_at": doc.get("verified_at"),
    }



# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: Public payments management (list / exceptions / reject / correct / verify)
# ROUTES : GET /public-payments  GET .../exceptions  PATCH .../reject|correction|verify
# ════════════════════════════════════════════════════════════════════════

@router.get("/public-payments")
async def mandir_list_public_payments(
    status: str | None = Query(default=None),
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    q: str | None = Query(default=None),
    payment_type: str | None = Query(default=None),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
    temple_id: int | None = Query(default=None),
):
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    tenant_id = await _resolve_tenant_for_mandir_request(current_user, x_tenant_id, temple_id, app_key=app_key)
    query: dict[str, Any] = {"tenant_id": tenant_id, "app_key": app_key}
    normalized_status = str(status or "").strip().lower()
    if normalized_status and normalized_status != "all":
        query["status"] = normalized_status

    col = get_collection("mandir_public_payments")
    normalized_q = str(q or "").strip().lower()
    normalized_type = str(payment_type or "").strip().lower()
    needs_in_memory_filter = any([normalized_q, normalized_type])
    fetch_limit = 500 if needs_in_memory_filter else min(limit + offset, 500)
    docs = await col.find(query).sort("created_at", -1).limit(fetch_limit).to_list(length=fetch_limit)
    filtered: list[dict[str, Any]] = []
    for doc in docs:
        doc_type = str(doc.get("payment_type") or doc.get("payment_purpose") or "").strip().lower()
        if normalized_type and doc_type != normalized_type:
            continue
        if normalized_q:
            haystack = " ".join(
                str(doc.get(key) or "")
                for key in (
                    "id",
                    "payment_id",
                    "devotee_name",
                    "name",
                    "devotee_phone",
                    "phone",
                    "seva_name",
                    "payment_type",
                    "payment_purpose",
                    "status",
                    "utr_reference",
                )
            ).lower()
            if normalized_q not in haystack:
                continue
        filtered.append(_sanitize_mongo_doc(doc))
    return filtered[offset:offset + limit]


@router.get("/public-payments/exceptions")
async def mandir_public_payment_exceptions(
    older_than_hours: int = Query(default=24, ge=1, le=720),
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    q: str | None = Query(default=None),
    reason: str | None = Query(default=None),
    status: str | None = Query(default=None),
    payment_type: str | None = Query(default=None),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    now = datetime.now(timezone.utc)
    stale_cutoff = now - timedelta(hours=older_than_hours)
    docs = await get_collection("mandir_public_payments").find(
        {"tenant_id": tenant_id, "app_key": app_key}
    ).sort("created_at", -1).limit(500).to_list(length=500)

    rows: list[dict[str, Any]] = []
    reason_counts: dict[str, int] = {}
    normalized_q = str(q or "").strip().lower()
    normalized_reason = str(reason or "").strip().lower()
    normalized_status = str(status or "").strip().lower()
    normalized_type = str(payment_type or "").strip().lower()
    for doc in docs:
        doc_status = str(doc.get("status") or "pending").strip().lower()
        amount = _safe_float(doc.get("amount"), 0.0)
        created_at = _parse_iso_datetime(doc.get("created_at"))
        reasons: list[str] = []

        if doc_status in {"failed", "error", "rejected"}:
            reasons.append(doc_status)
        if doc_status == "pending" and created_at and created_at < stale_cutoff:
            reasons.append("stale_pending")
        if amount <= 0:
            reasons.append("invalid_amount")
        if not _normalize_phone(doc.get("devotee_phone")):
            reasons.append("missing_phone")
        doc_payment_type = str(doc.get("payment_type") or "").strip().lower()
        if doc_payment_type not in {"donation", "seva"}:
            reasons.append("invalid_payment_type")
        if doc_payment_type == "seva" and not str(doc.get("seva_name") or "").strip():
            reasons.append("missing_seva")
        if doc_payment_type == "donation" and not str(doc.get("seva_name") or doc.get("category") or "").strip():
            reasons.append("missing_donation_category")

        if not reasons:
            continue
        if normalized_reason and normalized_reason not in reasons:
            continue
        if normalized_status and normalized_status != "all" and doc_status != normalized_status:
            continue
        if normalized_type and doc_payment_type != normalized_type:
            continue
        if normalized_q:
            haystack = " ".join(
                str(doc.get(key) or "")
                for key in (
                    "id",
                    "payment_id",
                    "devotee_name",
                    "name",
                    "devotee_phone",
                    "phone",
                    "seva_name",
                    "payment_type",
                    "payment_purpose",
                    "status",
                    "utr_reference",
                )
            ).lower()
            if normalized_q not in haystack:
                continue

        for reason in reasons:
            reason_counts[reason] = reason_counts.get(reason, 0) + 1

        row = _sanitize_mongo_doc(doc)
        row["exception_reasons"] = reasons
        row["age_hours"] = round(((now - created_at).total_seconds() / 3600), 2) if created_at else None
        rows.append(row)

    return {
        "summary": {
            "total": len(rows),
            "by_reason": reason_counts,
            "older_than_hours": older_than_hours,
        },
        "items": rows[offset:offset + limit],
    }


@router.patch("/public-payments/{payment_id}/reject", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
async def mandir_reject_public_payment(
    payment_id: str,
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    reason = " ".join(str(payload.get("reason") or "").strip().split())
    if len(reason) < 3:
        raise HTTPException(status_code=400, detail="Rejection reason is required")

    col = get_collection("mandir_public_payments")
    doc = await col.find_one({"id": payment_id, "tenant_id": tenant_id, "app_key": app_key})
    if not doc:
        raise HTTPException(status_code=404, detail="Payment not found")
    if str(doc.get("status") or "").strip().lower() == "verified":
        raise HTTPException(status_code=400, detail="Verified payment cannot be rejected")

    now = datetime.now(timezone.utc).isoformat()
    rejected_by = str(current_user.get("email") or current_user.get("sub") or current_user.get("id") or "admin")
    update = {
        "status": "rejected",
        "rejection_reason": reason,
        "rejected_at": now,
        "rejected_by": rejected_by,
        "updated_at": now,
    }
    await col.update_one({"id": payment_id, "tenant_id": tenant_id, "app_key": app_key}, {"$set": update})

    try:
        await log_audit_event(
            tenant_id=tenant_id,
            user_id=rejected_by,
            product="mandirmitra",
            action="public_payment_rejected",
            entity_type="mandir_public_payment",
            entity_id=payment_id,
            old_value={"status": doc.get("status")},
            new_value={"status": "rejected", "reason": reason},
        )
    except Exception:
        pass

    return {
        "status": "rejected",
        "payment_id": payment_id[:8].upper(),
        "reason": reason,
        "rejected_at": now,
    }


@router.patch("/public-payments/{payment_id}/correction", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
async def mandir_correct_public_payment(
    payment_id: str,
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    col = get_collection("mandir_public_payments")
    doc = await col.find_one({"id": payment_id, "tenant_id": tenant_id, "app_key": app_key})
    if not doc:
        raise HTTPException(status_code=404, detail="Payment not found")
    if str(doc.get("status") or "").strip().lower() == "verified":
        raise HTTPException(status_code=400, detail="Verified payment cannot be corrected")

    patch: dict[str, Any] = {}
    if "amount" in payload:
        amount = _safe_float(payload.get("amount"), 0.0)
        if amount <= 0:
            raise HTTPException(status_code=400, detail="Amount must be greater than zero")
        patch["amount"] = amount
    if "devotee_phone" in payload:
        phone = _normalize_phone(payload.get("devotee_phone"))
        if not phone:
            raise HTTPException(status_code=400, detail="Valid mobile number is required")
        patch["devotee_phone"] = phone
    if "payment_type" in payload:
        payment_type = str(payload.get("payment_type") or "").strip().lower()
        if payment_type not in {"donation", "seva"}:
            raise HTTPException(status_code=400, detail="payment_type must be donation or seva")
        patch["payment_type"] = payment_type
    if "seva_name" in payload:
        purpose = " ".join(str(payload.get("seva_name") or "").strip().split())
        if not purpose:
            raise HTTPException(status_code=400, detail="Donation/seva purpose is required")
        patch["seva_name"] = purpose

    if not patch:
        raise HTTPException(status_code=400, detail="No correction fields provided")

    now = datetime.now(timezone.utc).isoformat()
    corrected_by = str(current_user.get("email") or current_user.get("sub") or current_user.get("id") or "admin")
    patch.update({
        "corrected_at": now,
        "corrected_by": corrected_by,
        "updated_at": now,
    })
    await col.update_one({"id": payment_id, "tenant_id": tenant_id, "app_key": app_key}, {"$set": patch})
    updated = await col.find_one({"id": payment_id, "tenant_id": tenant_id, "app_key": app_key}) or {**doc, **patch}

    try:
        await log_audit_event(
            tenant_id=tenant_id,
            user_id=corrected_by,
            product="mandirmitra",
            action="public_payment_corrected",
            entity_type="mandir_public_payment",
            entity_id=payment_id,
            old_value={key: doc.get(key) for key in patch if key in doc},
            new_value={key: patch.get(key) for key in patch},
        )
    except Exception:
        pass

    return _sanitize_mongo_doc(updated)


@router.patch("/public-payments/{payment_id}/verify", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
async def mandir_verify_public_payment(
    payment_id: str,
    payload: dict[str, Any],
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    col = get_collection("mandir_public_payments")
    doc = await col.find_one({"id": payment_id, "tenant_id": tenant_id, "app_key": app_key})
    if not doc:
        raise HTTPException(status_code=404, detail="Payment not found")

    if doc.get("status") == "verified":
        raise HTTPException(status_code=400, detail="Payment already verified")

    utr_reference = _normalize_public_payment_utr_reference(
        payload.get("utr_reference") or doc.get("utr_reference") or doc.get("upi_reference_number")
    )
    payment_date = str(
        payload.get("payment_date") or datetime.now(timezone.utc).date().isoformat()
    ).strip()
    bank_account_id = payload.get("bank_account_id")  # optional explicit bank account

    payment_type = str(doc.get("payment_type") or "seva").lower()
    amount = float(doc.get("amount") or 0)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Payment amount must be greater than zero")

    devotee_name = str(doc.get("devotee_name") or "").strip() or "Unknown Devotee"
    devotee_phone = str(doc.get("devotee_phone") or "").strip() or None
    devotee_email = str(doc.get("devotee_email") or "").strip() or None
    seva_name = str(doc.get("seva_name") or "").strip()
    seva_id = str(doc.get("seva_id") or "").strip() or None

    source_record: dict[str, Any] = {}
    source_type: str
    source_id: str | None = None

    try:
        if payment_type == "seva":
            seva_payload: dict[str, Any] = {
                "amount_paid": amount,
                "payment_mode": "UPI",
                "devotee_name": devotee_name,
                "devotee_names": devotee_name,
                "devotee_phone": devotee_phone,
                "devotee_mobile": devotee_phone,
                "booking_date": payment_date,
                "seva_name": seva_name,
                "seva_id": seva_id or "",
                "upi_reference_number": utr_reference,
                "gothra": doc.get("gothra"),
                "nakshtra": doc.get("nakshtra"),
                "rashi": doc.get("rashi"),
                "notes": f"Public portal | Payment ID: {payment_id[:8].upper()}",
            }
            if bank_account_id:
                seva_payload["bank_account_id"] = bank_account_id
            source_record = await create_seva_booking(
                payload=seva_payload,
                session=session,
                current_user=current_user,
                x_tenant_id=x_tenant_id,
                x_app_key=app_key,
            )
            source_type = "seva_booking"
            source_id = str(source_record.get("id") or "").strip() or None
        else:
            # Donation — map category_name back to standard category
            _donation_cat_map = {
                "general donation": "General Donation",
                "annadanam": "Annadanam",
                "construction fund": "Construction Fund",
                "corpus fund": "Corpus Fund",
                "vastra seva": "Vastra Seva",
                "nitya puja": "Nitya Puja",
            }
            category = _donation_cat_map.get(seva_name.lower(), seva_name) or "General Donation"
            donation_payload: dict[str, Any] = {
                "amount": amount,
                "payment_mode": "UPI",
                "category": category,
                "devotee_name": devotee_name,
                "devotee_phone": devotee_phone,
                "email": devotee_email,
                "upi_reference_number": utr_reference,
                "donation_date": payment_date,
                "notes": f"Public portal | Payment ID: {payment_id[:8].upper()}",
            }
            if bank_account_id:
                donation_payload["bank_account_id"] = bank_account_id
            source_record = await create_donation(
                payload=donation_payload,
                session=session,
                current_user=current_user,
                x_tenant_id=x_tenant_id,
                x_app_key=app_key,
            )
            source_type = "donation"
            source_id = str(source_record.get("donation_id") or source_record.get("id") or "").strip() or None
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to post accounting entry: {exc}") from exc

    now = datetime.now(timezone.utc).isoformat()
    update = {
        "status": "verified",
        "verified_at": now,
        "verified_by": str(current_user.get("email") or current_user.get("sub") or "admin"),
        "utr_reference": utr_reference,
        "source_type": source_type,
        "source_id": source_id,
    }
    await col.update_one({"id": payment_id, "tenant_id": tenant_id, "app_key": app_key}, {"$set": update})

    try:
        await log_audit_event(
            tenant_id=tenant_id,
            user_id=update["verified_by"],
            product="mandirmitra",
            action="public_payment_verified",
            entity_type="mandir_public_payment",
            entity_id=payment_id,
            old_value={"status": doc.get("status")},
            new_value={
                "status": "verified",
                "utr_reference": utr_reference,
                "source_type": source_type,
                "source_id": source_id,
                "receipt_number": str(source_record.get("receipt_number") or "").strip() or None,
            },
        )
    except Exception:
        pass

    receipt_number = str(source_record.get("receipt_number") or "").strip() or None
    receipt_pdf_url = str(source_record.get("receipt_pdf_url") or "").strip() or None
    if not receipt_pdf_url and source_id:
        if source_type == "seva_booking":
            receipt_pdf_url = f"/api/v1/sevas/bookings/{source_id}/receipt/pdf"
        elif source_type == "donation":
            receipt_pdf_url = f"/api/v1/donations/{source_id}/receipt/pdf"

    return {
        "status": "verified",
        "payment_id": payment_id[:8].upper(),
        "source_type": source_type,
        "source_id": source_id,
        "receipt_number": receipt_number,
        "receipt_pdf_url": receipt_pdf_url,
        "message": "Payment verified and accounting entry posted successfully.",
    }



# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: UPI payments management (list / quick-log)
# ROUTES : GET /upi-payments  POST .../quick-log
# ════════════════════════════════════════════════════════════════════════

@router.get("/upi-payments")
async def mandir_upi_payments(
    from_date: date | None = Query(default=None),
    to_date: date | None = Query(default=None),
    limit: int = Query(default=500, ge=1, le=2000),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())

    rows = await get_collection("mandir_upi_payments").find({"tenant_id": tenant_id, "app_key": app_key}).sort("payment_datetime", -1).limit(limit).to_list(length=limit)

    filtered: list[dict[str, Any]] = []
    for row in rows:
        parsed = _parse_iso_datetime(row.get("payment_datetime") or row.get("created_at"))
        if parsed is not None:
            row_date = parsed.date()
            if from_date and row_date < from_date:
                continue
            if to_date and row_date > to_date:
                continue
        filtered.append(row)

    filtered.sort(key=lambda item: str(item.get("payment_datetime") or item.get("created_at") or ""), reverse=True)
    return [_mandir_upi_payment_view(row) for row in filtered]


@router.post("/upi-payments/quick-log", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
async def mandir_upi_quick_log(
    payload: dict[str, Any],
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_context = resolve_mandir_tenant(
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
        operation="UPI quick log",
    )
    tenant_id = tenant_context.tenant_id
    app_key = tenant_context.app_key

    amount = _safe_float(payload.get("amount"), 0.0)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than zero")

    purpose = str(payload.get("payment_purpose") or "DONATION").strip().upper()
    payment_dt = _parse_iso_datetime(payload.get("payment_datetime")) or datetime.now(timezone.utc)
    payment_datetime = payment_dt.isoformat()

    normalized_phone = _normalize_phone(payload.get("devotee_phone") or payload.get("sender_phone") or payload.get("phone"))
    devotee_name = str(payload.get("devotee_name") or payload.get("sender_name") or "").strip() or "Walk-in Devotee"
    sender_upi_id = str(payload.get("sender_upi_id") or "").strip() or None
    upi_reference_number = str(payload.get("upi_reference_number") or "").strip() or None
    notes = str(payload.get("notes") or "").strip() or None

    col = get_collection("mandir_upi_payments")
    if upi_reference_number:
        existing = await col.find_one(
            {
                "tenant_id": tenant_id,
                "app_key": app_key,
                "upi_reference_number": upi_reference_number,
            }
        )
        if existing is not None:
            return _mandir_upi_payment_view(existing)

    source_record: dict[str, Any]
    source_type: str
    source_id: str | None

    if purpose == "SEVA":
        seva_payload = {
            **payload,
            "amount_paid": amount,
            "payment_mode": "UPI",
            "devotee_name": devotee_name,
            "devotee_names": str(payload.get("devotee_names") or devotee_name),
            "devotee_phone": normalized_phone,
            "devotee_mobile": normalized_phone,
            "booking_date": str(payload.get("booking_date") or payment_dt.date().isoformat()),
            "seva_name": str(payload.get("seva_name") or "Seva Booking"),
        }
        source_record = await create_seva_booking(
            payload=seva_payload,
            session=session,
            current_user=current_user,
            x_tenant_id=x_tenant_id,
            x_app_key=x_app_key,
        )
        source_type = "seva"
        source_id = str(source_record.get("id") or "").strip() or None
    else:
        donation_category_map = {
            "ANNADANA": "Annadanam",
            "ANNADANAM": "Annadanam",
            "SPONSORSHIP": "Sponsorship",
            "OTHER": "General Donation",
            "DONATION": "General Donation",
        }
        donation_payload = {
            **payload,
            "amount": amount,
            "payment_mode": "UPI",
            "category": str(payload.get("category") or donation_category_map.get(purpose, "General Donation")),
            "devotee_name": devotee_name,
            "devotee_phone": normalized_phone,
            "phone": normalized_phone,
        }
        source_record = await create_donation(
            payload=donation_payload,
            session=session,
            current_user=current_user,
            x_tenant_id=x_tenant_id,
            x_app_key=x_app_key,
        )
        source_type = "donation"
        source_id = str(source_record.get("donation_id") or source_record.get("id") or "").strip() or None

    row_id = str(uuid4())
    receipt_number = str(source_record.get("receipt_number") or _upi_receipt_number({"id": row_id})).strip()
    now = datetime.now(timezone.utc).isoformat()
    payment_row = {
        "id": row_id,
        "tenant_id": tenant_id,
        "app_key": app_key,
        "amount": amount,
        "payment_datetime": payment_datetime,
        "payment_mode": "UPI",
        "payment_purpose": purpose,
        "devotee_name": devotee_name,
        "devotee_phone": normalized_phone,
        "sender_phone": normalized_phone,
        "sender_upi_id": sender_upi_id,
        "upi_reference_number": upi_reference_number,
        "notes": notes,
        "source_type": source_type,
        "source_id": source_id,
        "receipt_number": receipt_number,
        "created_at": now,
        "updated_at": now,
    }
    await col.insert_one(payment_row)

    return _mandir_upi_payment_view(payment_row)



# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: Users (GET list / PUT profile)
# ROUTES : GET /users  PUT /users/{user_id}
# ════════════════════════════════════════════════════════════════════════

@router.get("/users")
async def mandir_users(_current_user: dict = Depends(get_current_user), x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID")):
    tenant_id = resolve_tenant_id(_current_user, x_tenant_id)
    users = get_collection("core_users")
    docs = await users.find({"tenant_id": tenant_id, "is_active": True}).limit(200).to_list(length=200)
    return [{"user_id": d.get("user_id"), "email": d.get("email"), "full_name": d.get("full_name"), "role": d.get("role")} for d in docs]



@router.put("/users/{user_id}")
async def mandir_update_user_profile(
    user_id: str,
    payload: dict[str, Any],
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    users = get_collection("core_users")
    query = {
        "tenant_id": tenant_id,
        "$or": [
            {"user_id": user_id},
            {"id": user_id},
        ],
    }

    patch: dict[str, Any] = {}
    if "full_name" in payload:
        patch["full_name"] = str(payload.get("full_name") or "").strip()
    if "email" in payload:
        patch["email"] = str(payload.get("email") or "").strip().lower()
    if "phone" in payload:
        phone = str(payload.get("phone") or "").strip()
        patch["phone"] = phone or None
    patch["updated_at"] = datetime.now(timezone.utc).isoformat()

    await users.update_one(query, {"$set": patch}, upsert=False)
    doc = await users.find_one(query)
    if doc is None:
        raise HTTPException(status_code=404, detail="User not found")

    resolved_id = str(doc.get("user_id") or doc.get("id") or user_id)
    return {
        "id": resolved_id,
        "email": str(doc.get("email") or ""),
        "full_name": str(doc.get("full_name") or ""),
        "phone": doc.get("phone"),
        "role": doc.get("role"),
        "system_role": doc.get("system_role") or doc.get("role"),
        "role_key": doc.get("role_key"),
        "role_label": doc.get("role_label"),
        "module_permissions": doc.get("module_permissions") or {},
        "action_permissions": doc.get("action_permissions") or {},
        "is_superuser": bool(doc.get("is_superuser", False)),
        "must_change_password": bool(doc.get("must_change_password", False)),
    }


# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: Seva bookings (POST / quick-ticket / GET / receipt PDF / cancel / reschedule / approve)
# ROUTES : POST /sevas/bookings  POST .../quick-ticket  GET .../bookings  GET .../receipt/pdf  POST .../cancel  PUT .../reschedule  POST .../approve-reschedule  GET .../pending
# ════════════════════════════════════════════════════════════════════════

@router.post("/sevas/bookings", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
@router.post("/sevas/bookings/", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
async def create_seva_booking(
    payload: dict[str, Any],
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_context = resolve_mandir_tenant(
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
        operation="seva booking",
    )
    tenant_id = tenant_context.tenant_id
    app_key = tenant_context.app_key
    await _ensure_default_mandir_sql_accounts_safe(session, tenant_id, raise_on_failure=True)

    booking_id = str(uuid4())
    now = datetime.now(timezone.utc).isoformat()
    amount = _safe_float(payload.get("amount_paid") or payload.get("amount"), 0.0)
    payment_mode = str(payload.get("payment_mode") or payload.get("payment_method") or "Cash")
    seva_id = payload.get("seva_id")
    seva_name = str(payload.get("seva_name") or "Seva Booking")
    col_sevas = get_collection("mandir_sevas")
    seva_doc: dict[str, Any] | None = None
    if seva_id:
        seva_doc = await col_sevas.find_one({"id": str(seva_id), "tenant_id": tenant_id, "app_key": app_key})
        if not seva_doc:
            seva_doc = await col_sevas.find_one({"id": str(seva_id), "tenant_id": tenant_id})
        if seva_doc and seva_doc.get("name"):
            seva_name = str(seva_doc["name"])
        if seva_doc and seva_doc.get("name_kannada"):
            payload["seva_name_local"] = str(seva_doc.get("name_kannada") or "").strip()

    booking_date = _parse_booking_date(payload.get("booking_date"))
    if booking_date is None:
        raise HTTPException(status_code=400, detail="Please enter a valid booking date.")
    _validate_seva_booking_date(seva_doc, booking_date)
    await _validate_seva_booking_capacity(
        seva_doc,
        tenant_id=tenant_id,
        app_key=app_key,
        seva_id=str(seva_id),
        booking_date=booking_date,
    )

    # Compute expiry_date from seva's duration_days (for annual/subscription sevas)
    seva_expiry_date: str | None = None
    if seva_doc:
        duration_days = seva_doc.get("duration_days")
        if duration_days and int(duration_days) > 0:
            booking_date_raw = str(payload.get("booking_date") or "").strip()
            try:
                base_dt = datetime.fromisoformat(booking_date_raw)
            except Exception:
                base_dt = datetime.now(timezone.utc)
            seva_expiry_date = (base_dt + timedelta(days=int(duration_days))).date().isoformat()

    devotee_doc: dict[str, Any] | None = None
    devotee_id = str(payload.get("devotee_id") or "").strip()
    if devotee_id:
        devotee_doc = await get_collection("mandir_devotees").find_one(
            {"id": devotee_id, "tenant_id": tenant_id, "app_key": app_key}
        )
        if not devotee_doc:
            devotee_doc = await get_collection("mandir_devotees").find_one(
                {"id": devotee_id, "tenant_id": tenant_id}
            )

    devotee_phone = _normalize_phone(
        payload.get("devotee_phone") or payload.get("devotee_mobile") or payload.get("phone")
    )
    if devotee_doc is None and devotee_phone:
        devotee_doc = await _find_devotee_by_phone(tenant_id, app_key, devotee_phone)

    devotee_snapshot = _sanitize_mongo_doc(devotee_doc) if isinstance(devotee_doc, dict) else {}
    devotee_name = str(
        payload.get("devotee_names")
        or payload.get("devotee_name")
        or devotee_snapshot.get("name")
        or "Devotee"
    ).strip() or "Devotee"
    devotee_address = str(
        payload.get("devotee_address")
        or payload.get("address")
        or devotee_snapshot.get("address")
        or ""
    ).strip() or None
    devotee_phone = devotee_phone or _normalize_phone(devotee_snapshot.get("phone"))

    booking = {
        "id": booking_id,
        "tenant_id": tenant_id,
        "app_key": app_key,
        **{k: v for k, v in payload.items() if k not in ("id", "_id", "tenant_id", "app_key")},
        "payment_mode": payment_mode,
        "created_at": now,
        "updated_at": now,
        "status": "confirmed",
    }
    booking["seva_name"] = seva_name
    if payload.get("seva_name_local"):
        booking["seva_name_local"] = str(payload.get("seva_name_local") or "").strip()
    booking["devotee_id"] = devotee_id or str(booking.get("devotee_id") or devotee_snapshot.get("id") or "").strip() or None
    booking["devotee_name"] = devotee_name
    booking["devotee_names"] = devotee_name
    booking["devotee_phone"] = devotee_phone or None
    if devotee_address:
        booking["devotee_address"] = devotee_address
        booking["address"] = devotee_address
    if devotee_snapshot:
        booking["devotee"] = {
            "id": devotee_snapshot.get("id"),
            "name": devotee_name,
            "phone": devotee_phone or devotee_snapshot.get("phone"),
            "address": devotee_address or devotee_snapshot.get("address"),
            "city": devotee_snapshot.get("city"),
            "state": devotee_snapshot.get("state"),
            "pincode": devotee_snapshot.get("pincode"),
        }
    booking["receipt_number"] = await _next_receipt_number(
        tenant_id=tenant_id,
        app_key=app_key,
        receipt_kind="seva",
        receipt_date=booking.get("booking_date") or now,
    )
    booking["receipt_pdf_url"] = f"/api/v1/sevas/bookings/{booking_id}/receipt/pdf"
    if seva_expiry_date:
        booking["expiry_date"] = seva_expiry_date
        booking["reminder_count"] = 0
        booking["reminder_sent_at"] = None

    col = get_collection("mandir_seva_bookings")
    try:
        await col.insert_one(booking)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to save seva booking: {exc}") from exc

    if amount > 0:
        raw_account_id = payload.get("bank_account_id") or payload.get("payment_account_id")
        resolved_account_id = await _resolve_mandir_payment_account_id(
            session,
            tenant_id,
            raw_account_id,
            payment_mode,
        )
        if not resolved_account_id:
            await col.delete_one({"id": booking_id, "tenant_id": tenant_id, "app_key": app_key})
            raise HTTPException(status_code=400, detail="No valid cash/bank account is configured for seva posting")

        try:
            income_acc_id = await _resolve_mandir_income_account(session, tenant_id, "Seva Income - General")
            devotee_names = str(booking.get("devotee_names") or booking.get("devotee_name") or "Devotee")
            journal_payload = JournalPostRequest(
                entry_date=datetime.now(timezone.utc).date(),
                description=f"Seva Booking ({seva_name}) - {devotee_names}",
                reference=booking["receipt_number"],
                lines=[
                    JournalLineIn(account_id=resolved_account_id, debit=Decimal(str(amount)), credit=Decimal("0")),
                    JournalLineIn(account_id=income_acc_id, debit=Decimal("0"), credit=Decimal(str(amount))),
                ],
            )
            await post_journal_entry(
                session=session,
                app_key=app_key,
                tenant_id=tenant_id,
                created_by="mandir_compat_system",
                payload=journal_payload,
                idempotency_key=f"sev_{booking_id}",
            )
        except Exception as exc:
            await col.delete_one({"id": booking_id, "tenant_id": tenant_id, "app_key": app_key})
            raise HTTPException(status_code=500, detail=f"Failed to post seva journal: {exc}") from exc

    return _mandir_seva_booking_view(booking)

@router.post("/sevas/bookings/quick-ticket", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
async def create_quick_ticket(
    payload: dict[str, Any],
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())

    ticket_type = str(payload.get("ticket_type") or payload.get("purpose") or payload.get("payment_purpose") or "seva").strip().lower()
    normalized_phone = _normalize_phone(payload.get("devotee_phone") or payload.get("phone") or payload.get("mobile"))

    devotee = None
    if normalized_phone:
        try:
            devotee = await _find_devotee_by_phone(tenant_id, app_key, normalized_phone)
        except Exception:
            devotee = None

    devotee_name = str(
        payload.get("devotee_name")
        or payload.get("devotee_names")
        or (devotee or {}).get("name")
        or "Walk-in Devotee"
    ).strip() or "Walk-in Devotee"

    if ticket_type in {"donation", "annadanam", "annadana", "sponsorship", "other"}:
        category_map = {
            "annadanam": "Annadanam",
            "annadana": "Annadanam",
            "sponsorship": "Sponsorship",
            "other": "General Donation",
            "donation": "General Donation",
        }
        donation_payload = {
            **payload,
            "amount": _safe_float(payload.get("amount") or payload.get("amount_paid"), 0.0),
            "devotee_name": devotee_name,
            "devotee_phone": normalized_phone,
            "phone": normalized_phone,
            "category": str(payload.get("category") or category_map.get(ticket_type, "General Donation")),
            "payment_mode": str(payload.get("payment_mode") or payload.get("payment_method") or "Cash"),
        }
        donation = await create_donation(
            payload=donation_payload,
            session=session,
            current_user=current_user,
            x_tenant_id=x_tenant_id,
            x_app_key=x_app_key,
        )
        return {
            "ticket_type": "donation",
            "autofill_found": bool(devotee),
            "devotee": devotee,
            "receipt_number": donation.get("receipt_number"),
            "receipt_pdf_url": donation.get("receipt_pdf_url"),
            "record": donation,
        }

    seva_payload = {
        **payload,
        "amount_paid": _safe_float(payload.get("amount_paid") or payload.get("amount"), 0.0),
        "devotee_name": devotee_name,
        "devotee_names": str(payload.get("devotee_names") or devotee_name),
        "devotee_phone": normalized_phone,
        "devotee_mobile": normalized_phone,
        "payment_mode": str(payload.get("payment_mode") or payload.get("payment_method") or "Cash"),
        "booking_date": str(payload.get("booking_date") or datetime.now(timezone.utc).date().isoformat()),
    }
    booking = await create_seva_booking(
        payload=seva_payload,
        session=session,
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
    )
    return {
        "ticket_type": "seva",
        "autofill_found": bool(devotee),
        "devotee": devotee,
        "receipt_number": booking.get("receipt_number"),
        "receipt_pdf_url": booking.get("receipt_pdf_url"),
        "record": booking,
    }


@router.get("/sevas/bookings")
async def mandir_seva_bookings(
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    q: str | None = Query(default=None),
    from_date: date | None = Query(default=None),
    to_date: date | None = Query(default=None),
    status: str | None = Query(default=None),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    if from_date and to_date and from_date > to_date:
        raise HTTPException(status_code=422, detail="from_date cannot be greater than to_date")
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    col = get_collection("mandir_seva_bookings")
    fetch_limit = 500 if any([q, from_date, to_date, status]) else min(limit + offset, 500)
    docs = await col.find({"tenant_id": tenant_id, "app_key": app_key}).sort("booking_date", -1).limit(fetch_limit).to_list(length=fetch_limit)
    viewed = [_mandir_seva_booking_view(doc) for doc in docs]
    if status:
        normalized_status = str(status).strip().lower()
        viewed = [row for row in viewed if str(row.get("status") or "").strip().lower() == normalized_status]
    filtered = _mandir_filter_rows(
        viewed,
        q=q,
        from_date=from_date,
        to_date=to_date,
        date_fields=("booking_date", "created_at"),
        search_fields=("receipt_number", "devotee_name", "devotee_names", "seva_name", "seva", "upi_reference_number"),
    )
    return filtered[offset:offset + limit]


@router.get("/sevas/bookings/{booking_id}/receipt/pdf")
async def get_seva_receipt_pdf(
    booking_id: str,
    lang: str | None = Query(default=None, description="Override receipt language (kannada/hindi/tamil/telugu/malayalam)"),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())

    col = get_collection("mandir_seva_bookings")
    booking = await col.find_one({"id": str(booking_id), "tenant_id": tenant_id, "app_key": app_key})
    if booking is None:
        raise HTTPException(status_code=404, detail="Seva booking not found")

    booking_id_text = str(booking.get("id") or booking_id).strip()
    booking["id"] = booking_id_text

    resolved_seva_name = str(booking.get("seva_name") or "").strip()
    seva_id = booking.get("seva_id")
    seva_doc: dict[str, Any] | None = None
    if seva_id:
        seva_doc = await get_collection("mandir_sevas").find_one({"id": str(seva_id), "tenant_id": tenant_id, "app_key": app_key})
        if not seva_doc:
            seva_doc = await get_collection("mandir_sevas").find_one({"id": str(seva_id), "tenant_id": tenant_id})
    if seva_doc:
        if (not resolved_seva_name or resolved_seva_name.lower() in {"seva booking", "seva"}) and seva_doc.get("name"):
            booking["seva_name"] = str(seva_doc.get("name")).strip()
        if seva_doc.get("name_kannada"):
            booking["seva_name_local"] = str(seva_doc.get("name_kannada") or "").strip()

    devotee_snapshot = booking.get("devotee") if isinstance(booking.get("devotee"), dict) else {}
    if not devotee_snapshot or not any([
        booking.get("devotee_name"),
        booking.get("devotee_names"),
        booking.get("devotee_address"),
        booking.get("address"),
    ]):
        devotee_doc: dict[str, Any] | None = None
        booking_devotee_id = str(booking.get("devotee_id") or "").strip()
        if booking_devotee_id:
            devotee_doc = await get_collection("mandir_devotees").find_one(
                {"id": booking_devotee_id, "tenant_id": tenant_id, "app_key": app_key}
            )
            if not devotee_doc:
                devotee_doc = await get_collection("mandir_devotees").find_one(
                    {"id": booking_devotee_id, "tenant_id": tenant_id}
                )

        if devotee_doc is None:
            booking_phone = _normalize_phone(
                booking.get("devotee_phone") or booking.get("devotee_mobile") or (devotee_snapshot or {}).get("phone")
            )
            if booking_phone:
                devotee_doc = await _find_devotee_by_phone(tenant_id, app_key, booking_phone)

        if devotee_doc:
            devotee_snapshot = _sanitize_mongo_doc(devotee_doc)

    if devotee_snapshot:
        resolved_devotee_name = str(
            booking.get("devotee_names")
            or booking.get("devotee_name")
            or devotee_snapshot.get("name")
            or "Devotee"
        ).strip() or "Devotee"
        resolved_devotee_address = str(
            booking.get("devotee_address")
            or booking.get("address")
            or devotee_snapshot.get("address")
            or ""
        ).strip() or None
        resolved_devotee_phone = _normalize_phone(
            booking.get("devotee_phone") or booking.get("devotee_mobile") or devotee_snapshot.get("phone")
        )
        booking["devotee_id"] = str(booking.get("devotee_id") or devotee_snapshot.get("id") or "").strip() or None
        booking["devotee_name"] = resolved_devotee_name
        booking["devotee_names"] = resolved_devotee_name
        booking["devotee_phone"] = resolved_devotee_phone or None
        if resolved_devotee_address:
            booking["devotee_address"] = resolved_devotee_address
            booking["address"] = resolved_devotee_address
        booking["devotee"] = {
            "id": devotee_snapshot.get("id"),
            "name": resolved_devotee_name,
            "phone": resolved_devotee_phone or devotee_snapshot.get("phone"),
            "address": resolved_devotee_address or devotee_snapshot.get("address"),
            "city": devotee_snapshot.get("city"),
            "state": devotee_snapshot.get("state"),
            "pincode": devotee_snapshot.get("pincode"),
        }
    receipt_number = _receipt_number_for_seva(booking)
    booking["receipt_number"] = receipt_number
    booking["receipt_pdf_url"] = f"/api/v1/sevas/bookings/{booking_id_text}/receipt/pdf"

    await col.update_one(
        {"tenant_id": tenant_id, "app_key": app_key, "id": booking_id_text},
        {
            "$set": {
                "receipt_number": receipt_number,
                "receipt_pdf_url": booking["receipt_pdf_url"],
                "seva_name": booking.get("seva_name"),
                "seva_name_local": booking.get("seva_name_local"),
                "devotee_id": booking.get("devotee_id"),
                "devotee_name": booking.get("devotee_name"),
                "devotee_names": booking.get("devotee_names"),
                "devotee_phone": booking.get("devotee_phone"),
                "devotee_address": booking.get("devotee_address"),
                "address": booking.get("address"),
                "devotee": booking.get("devotee"),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
        },
        upsert=False,
    )

    temple_profile = await _resolve_temple_receipt_profile(tenant_id=tenant_id, app_key=app_key, lang=lang)

    try:
        pdf_bytes = _generate_seva_receipt_pdf_bytes(
            booking,
            temple_name=temple_profile.get("temple_name", "Temple"),
            temple_profile=temple_profile,
        )
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    safe_receipt = "".join(ch for ch in str(receipt_number) if ch.isalnum() or ch in ("-", "_")) or booking_id_text[:8]
    filename = f"seva_receipt_{safe_receipt}.pdf"

    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/sevas/bookings/{booking_id}/cancel", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
async def cancel_seva_receipt(
    booking_id: str,
    payload: dict[str, Any] | None = None,
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    return await _cancel_mandir_receipt_source(
        source_kind="seva",
        source_id=booking_id,
        collection_name="mandir_seva_bookings",
        id_field="id",
        idempotency_prefix="sev_",
        payload=payload,
        session=session,
        current_user=current_user,
        x_tenant_id=x_tenant_id,
        x_app_key=x_app_key,
    )




@router.put("/sevas/bookings/{booking_id}/reschedule", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
async def mandir_request_seva_reschedule(
    booking_id: str,
    new_date: str = Query(...),
    reason: str | None = Query(default=None),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    now = datetime.now(timezone.utc).isoformat()
    col = get_collection("mandir_seva_bookings")
    await col.update_one(
        {"id": booking_id, "tenant_id": tenant_id, "app_key": app_key},
        {
            "$set": {
                "reschedule_pending": True,
                "status": "reschedule_pending",
                "reschedule_requested_date": new_date,
                "reschedule_reason": str(reason or "").strip() or None,
                "reschedule_requested_at": now,
                "updated_at": now,
            }
        },
        upsert=False,
    )
    doc = await col.find_one({"id": booking_id, "tenant_id": tenant_id, "app_key": app_key})
    if doc is None:
        raise HTTPException(status_code=404, detail="Seva booking not found")
    return _mandir_seva_booking_view(doc)


@router.post("/sevas/bookings/{booking_id}/approve-reschedule", dependencies=_MANDIR_WRITE_ROUTE_DEPS)
async def mandir_approve_seva_reschedule(
    booking_id: str,
    approve: bool = Query(default=True),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    now = datetime.now(timezone.utc).isoformat()
    col = get_collection("mandir_seva_bookings")
    booking = await col.find_one({"id": booking_id, "tenant_id": tenant_id, "app_key": app_key})
    if booking is None:
        raise HTTPException(status_code=404, detail="Seva booking not found")

    requested_date = str(booking.get("reschedule_requested_date") or "").strip()
    patch: dict[str, Any] = {
        "reschedule_pending": False,
        "reschedule_approved": bool(approve),
        "reschedule_decided_at": now,
        "updated_at": now,
    }
    if approve and requested_date:
        patch["booking_date"] = requested_date
        patch["status"] = "confirmed"
    else:
        patch["status"] = "confirmed"

    await col.update_one(
        {"id": booking_id, "tenant_id": tenant_id, "app_key": app_key},
        {"$set": patch},
        upsert=False,
    )

    updated = await col.find_one({"id": booking_id, "tenant_id": tenant_id, "app_key": app_key})
    return _mandir_seva_booking_view(updated or booking)

@router.get("/sevas/reschedule/pending")
async def mandir_seva_reschedule_pending(
    limit: int = Query(default=100, ge=1, le=500),
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    app_key = resolve_app_key((x_app_key or current_user.get("app_key") or "mandirmitra").strip())
    col = get_collection("mandir_seva_bookings")
    q = {
        "tenant_id": tenant_id,
        "app_key": app_key,
        "$or": [{"reschedule_pending": True}, {"status": "reschedule_pending"}],
    }
    docs = await col.find(q).sort("updated_at", -1).limit(limit).to_list(length=limit)
    return [_sanitize_mongo_doc(doc) for doc in docs]


@router.get("/users/me")
async def mandir_users_me(current_user: dict = Depends(get_current_user)):
    return current_user


# ---------------------------------------------------------------------------
# Seva Reminder Config endpoints (admin — requires auth)
# ---------------------------------------------------------------------------

@router.get("/sevas/reminder-config")
async def mandir_seva_reminder_config(
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
    x_app_key: str | None = Header(default=None, alias="X-App-Key"),
):
    """List all sevas with their current reminder configuration for this temple."""
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    col = get_collection("mandir_sevas")
    docs = await col.find({"tenant_id": tenant_id, "is_active": True}).to_list(length=500)
    result = []
    for doc in docs:
        result.append({
            "seva_id": str(doc.get("id") or ""),
            "seva_name": str(doc.get("name_english") or doc.get("name") or ""),
            "amount": float(doc.get("amount") or 0),
            "frequency": str(doc.get("frequency") or "one_time"),
            "duration_days": doc.get("duration_days"),
            "reminder_enabled": bool(doc.get("reminder_enabled", False)),
            "reminder_days_before": int(doc.get("reminder_days_before") or 30),
        })
    return result


@router.patch("/sevas/{seva_id}/reminder-config", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)
async def mandir_update_seva_reminder_config(
    seva_id: str,
    payload: dict,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    """Update reminder configuration for a specific seva (admin only).

    Accepted fields: reminder_enabled (bool), reminder_days_before (int), duration_days (int).
    """
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    col = get_collection("mandir_sevas")
    doc = await col.find_one({"id": seva_id, "tenant_id": tenant_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Seva not found")

    patch: dict[str, Any] = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if "reminder_enabled" in payload:
        patch["reminder_enabled"] = bool(payload["reminder_enabled"])
    if "reminder_days_before" in payload:
        days = int(payload["reminder_days_before"])
        if days < 1 or days > 365:
            raise HTTPException(status_code=400, detail="reminder_days_before must be between 1 and 365")
        patch["reminder_days_before"] = days
    if "duration_days" in payload:
        dur = int(payload["duration_days"])
        if dur < 1:
            raise HTTPException(status_code=400, detail="duration_days must be >= 1")
        patch["duration_days"] = dur

    await col.update_one({"id": seva_id, "tenant_id": tenant_id}, {"$set": patch})
    updated = await col.find_one({"id": seva_id, "tenant_id": tenant_id})
    return {
        "seva_id": seva_id,
        "seva_name": str(updated.get("name_english") or updated.get("name") or ""),
        "reminder_enabled": bool(updated.get("reminder_enabled", False)),
        "reminder_days_before": int(updated.get("reminder_days_before") or 30),
        "duration_days": updated.get("duration_days"),
    }


@router.get("/sevas/reminders/upcoming")
async def mandir_seva_reminders_upcoming(
    days: int = 30,
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    """List all seva bookings with expiry_date within the next `days` days."""
    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    if days < 1 or days > 365:
        raise HTTPException(status_code=400, detail="days must be between 1 and 365")

    now = datetime.now(timezone.utc)
    window_end = now + timedelta(days=days)

    # Get temple display name for WhatsApp messages
    temple_doc = await get_collection("mandir_temples").find_one({"tenant_id": tenant_id}) or {}
    _temple_display_name = str(temple_doc.get("temple_name") or temple_doc.get("name") or "Temple")

    col = get_collection("mandir_seva_bookings")
    docs = await col.find(
        {
            "tenant_id": tenant_id,
            "expiry_date": {"$exists": True, "$ne": None},
            "status": {"$ne": "cancelled"},
        }
    ).sort("expiry_date", 1).to_list(length=500)

    result = []
    for b in docs:
        expiry_raw = str(b.get("expiry_date") or "")
        if not expiry_raw:
            continue
        try:
            expiry_dt = datetime.fromisoformat(expiry_raw.replace("Z", "+00:00"))
            if expiry_dt.tzinfo is None:
                expiry_dt = expiry_dt.replace(tzinfo=timezone.utc)
        except Exception:
            continue
        if expiry_dt > window_end:
            continue
        days_left = max(0, (expiry_dt.date() - now.date()).days)
        devotee_phone = str(b.get("devotee_phone") or b.get("devotee_mobile") or "")
        devotee_name = str(b.get("devotee_name") or b.get("devotee_names") or "Devotee")
        seva_name = str(b.get("seva_name") or "")
        amount = float(b.get("amount_paid") or b.get("amount") or 0)
        receipt_number = str(b.get("receipt_number") or b.get("id") or "")
        expiry_label = expiry_dt.strftime("%d %b %Y")

        # WhatsApp deep link — admin clicks to send pre-filled message to devotee
        from app.modules.mandir_compat.reminder_worker import build_whatsapp_reminder_link
        whatsapp_link = build_whatsapp_reminder_link(
            devotee_phone=devotee_phone,
            devotee_name=devotee_name,
            seva_name=seva_name,
            temple_name=_temple_display_name,
            expiry_date=expiry_label,
            amount=amount,
            days_left=days_left,
            receipt_number=receipt_number,
        )

        result.append({
            "booking_id": str(b.get("id") or ""),
            "receipt_number": receipt_number,
            "seva_id": str(b.get("seva_id") or ""),
            "seva_name": seva_name,
            "devotee_name": devotee_name,
            "devotee_phone": devotee_phone,
            "devotee_email": str(b.get("devotee_email") or ""),
            "amount": amount,
            "booking_date": str(b.get("booking_date") or ""),
            "expiry_date": expiry_raw,
            "expiry_date_label": expiry_label,
            "days_left": days_left,
            "reminder_count": int(b.get("reminder_count") or 0),
            "reminder_sent_at": b.get("reminder_sent_at"),
            "status": str(b.get("status") or "confirmed"),
            "whatsapp_link": whatsapp_link,
        })
    return result


@router.post("/sevas/reminders/trigger", dependencies=_MANDIR_ADMIN_ROUTE_DEPS)

# ════════════════════════════════════════════════════════════════════════
# SECTION: ROUTES: Seva reminders (trigger / upcoming / config)
# ROUTES : POST /sevas/reminders/trigger  GET .../upcoming  PATCH .../reminder-config  GET /users/me
# ════════════════════════════════════════════════════════════════════════

async def mandir_seva_reminders_trigger(
    payload: dict = {},
    current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    """Manually trigger the reminder job for this tenant.

    Optional payload fields:
      - seva_id (str): limit to a single seva
      - force (bool): re-send even if reminded recently (default: False)
    """
    from app.modules.mandir_compat.reminder_worker import run_reminders_once

    tenant_id = resolve_tenant_id(current_user, x_tenant_id)
    seva_id = str(payload.get("seva_id") or "").strip() or None
    force = bool(payload.get("force", False))

    result = await run_reminders_once(
        tenant_id=tenant_id,
        seva_id=seva_id,
        force=force,
    )
    return {"ok": True, "result": result}


# Refund-request routes moved to routes/refunds.py
# (docs/operations/LARGE_FILE_MODULARIZATION_PLAN.md); re-exported for route
# registration and for tests that call handlers via mandir_router.<name>.
from app.modules.mandir_compat.routes.refunds import (  # noqa: E402
    _audit_mandir_refund_event as _audit_mandir_refund_event,
    approve_mandir_refund_request as approve_mandir_refund_request,
    create_mandir_refund_request as create_mandir_refund_request,
    list_mandir_refund_requests as list_mandir_refund_requests,
    mandir_export_refunds_csv as mandir_export_refunds_csv,
    mandir_report_refunds as mandir_report_refunds,
    reject_mandir_refund_request as reject_mandir_refund_request,
    settle_mandir_refund_request as settle_mandir_refund_request,
)

# Dashboard, panchang, and devotee routes moved to routes/*.py
# (docs/operations/LARGE_FILE_MODULARIZATION_PLAN.md); side-effect import registers handlers.

from app.modules.mandir_compat.receipt_pdf import (
    _amount_to_kannada_words as _amount_to_kannada_words,
    _amount_to_words as _amount_to_words,
    _amount_words_receipt_line as _amount_words_receipt_line,
    _as_text as _as_text,
    _bilingual_label as _bilingual_label,
    _build_receipt_pdf_bytes as _build_receipt_pdf_bytes,
    _build_receipt_pdf_bytes_pillow as _build_receipt_pdf_bytes_pillow,
    _build_receipt_pdf_bytes_weasy as _build_receipt_pdf_bytes_weasy,
    _build_temple_receipt_profile as _build_temple_receipt_profile,
    _compose_receipt_address_line as _compose_receipt_address_line,
    _compose_receipt_line_description as _compose_receipt_line_description,
    _compose_receipt_party_name as _compose_receipt_party_name,
    _default_labels as _default_labels,
    _detect_script as _detect_script,
    _extract_seva_line_items as _extract_seva_line_items,
    _first_non_empty_text as _first_non_empty_text,
    _format_payment_mode_for_receipt as _format_payment_mode_for_receipt,
    _format_payment_mode_local_for_receipt as _format_payment_mode_local_for_receipt,
    _format_receipt_date as _format_receipt_date,
    _generate_donation_receipt_pdf_bytes as _generate_donation_receipt_pdf_bytes,
    _generate_seva_receipt_pdf_bytes as _generate_seva_receipt_pdf_bytes,
    _integer_to_kannada_words as _integer_to_kannada_words,
    _integer_to_words as _integer_to_words,
    _name_prefix_from_sources as _name_prefix_from_sources,
    _normalize_local_language as _normalize_local_language,
    _receipt_paragraph as _receipt_paragraph,
    _receipt_payment_line as _receipt_payment_line,
    _resolve_font_name as _resolve_font_name,
    _resolve_temple_receipt_profile as _resolve_temple_receipt_profile,
    _split_amount as _split_amount,
)

import app.modules.mandir_compat.routes.funds  # noqa: E402, F401
import app.modules.mandir_compat.routes.festivals  # noqa: E402, F401
from app.modules.mandir_compat.routes.funds import (  # noqa: E402
    approve_mandir_fund_opening_balance as approve_mandir_fund_opening_balance,
    approve_mandir_fund_transfer as approve_mandir_fund_transfer,
    cancel_mandir_fund_opening_balance as cancel_mandir_fund_opening_balance,
    cancel_mandir_fund_transfer as cancel_mandir_fund_transfer,
    create_mandir_fund as create_mandir_fund,
    create_mandir_fund_opening_balance as create_mandir_fund_opening_balance,
    create_mandir_fund_transfer as create_mandir_fund_transfer,
    list_mandir_fund_opening_balances as list_mandir_fund_opening_balances,
    list_mandir_fund_transfers as list_mandir_fund_transfers,
    list_mandir_funds as list_mandir_funds,
)
from app.modules.mandir_compat.routes.festivals import (  # noqa: E402
    create_mandir_festival as create_mandir_festival,
    list_mandir_festivals as list_mandir_festivals,
)


from app.modules.mandir_compat.helpers.legacy_coa import (
    _coerce_account_id as _coerce_account_id,
    _infer_cash_bank_nature as _infer_cash_bank_nature,
    _infer_flag as _infer_flag,
    _load_mandir_legacy_accounts as _load_mandir_legacy_accounts,
    _prepare_mandir_account_docs as _prepare_mandir_account_docs,
    _upsert_mandir_account_docs as _upsert_mandir_account_docs,
)
from app.modules.mandir_compat.helpers.account_seeding import (
    MANDIR_DEFAULT_ACCOUNTS as MANDIR_DEFAULT_ACCOUNTS,
    _dedupe_mandir_account_docs as _dedupe_mandir_account_docs,
    _ensure_default_mandir_accounts as _ensure_default_mandir_accounts,
    _ensure_default_mandir_sql_accounts as _ensure_default_mandir_sql_accounts,
    _ensure_default_mandir_sql_accounts_safe as _ensure_default_mandir_sql_accounts_safe,
    _mandir_account_view as _mandir_account_view,
    _mandir_seed_accounts as _mandir_seed_accounts,
    _sync_mandir_sql_accounts_from_seed as _sync_mandir_sql_accounts_from_seed,
)

import app.modules.mandir_compat.routes.sevas  # noqa: E402, F401
import app.modules.mandir_compat.routes.temples  # noqa: E402, F401
import app.modules.mandir_compat.routes.accounts  # noqa: E402, F401
from app.modules.mandir_compat.routes.sevas import (
    create_seva as create_seva,
    delete_seva as delete_seva,
    import_sevas as import_sevas,
    list_sevas as list_sevas,
    seva_date_availability as seva_date_availability,
    seva_dropdown_options as seva_dropdown_options,
    seva_import_template as seva_import_template,
    seva_payment_accounts as seva_payment_accounts,
    seva_priests as seva_priests,
    update_seva as update_seva,
)
from app.modules.mandir_compat.routes.temples import (
    get_current_temple as get_current_temple,
    update_current_temple as update_current_temple,
)
from app.modules.mandir_compat.routes.accounts import (
    mandir_accounts_hierarchy as mandir_accounts_hierarchy,
    mandir_accounts_import_legacy as mandir_accounts_import_legacy,
    mandir_accounts_initialize_default as mandir_accounts_initialize_default,
    mandir_accounts_list as mandir_accounts_list,
    mandir_accounts_update as mandir_accounts_update,
)


import app.modules.mandir_compat.routes.donations_read  # noqa: E402, F401
import app.modules.mandir_compat.routes.misc_compat  # noqa: E402, F401
import app.modules.mandir_compat.routes.hundi  # noqa: E402, F401
import app.modules.mandir_compat.routes.inventory  # noqa: E402, F401
import app.modules.mandir_compat.routes.reports  # noqa: E402, F401
from app.modules.mandir_compat.routes.donations_read import (
    donations_categories as donations_categories,
    donations_payment_accounts as donations_payment_accounts,
    get_donation_receipt_pdf as get_donation_receipt_pdf,
    list_donations as list_donations,
)
from app.modules.mandir_compat.routes.misc_compat import _ok as _ok
from app.modules.mandir_compat.routes.hundi import (
    approve_mandir_hundi_opening as approve_mandir_hundi_opening,
    cancel_mandir_hundi_opening as cancel_mandir_hundi_opening,
    create_mandir_hundi_master as create_mandir_hundi_master,
    create_mandir_hundi_opening as create_mandir_hundi_opening,
    mandir_hundi_masters as mandir_hundi_masters,
    mandir_hundi_openings as mandir_hundi_openings,
)
from app.modules.mandir_compat.routes.inventory import (
    _mandir_inventory_item_balance as _mandir_inventory_item_balance,
    _mandir_inventory_item_position as _mandir_inventory_item_position,
    approve_mandir_inventory_consumption as approve_mandir_inventory_consumption,
    cancel_mandir_inventory_consumption as cancel_mandir_inventory_consumption,
    create_mandir_inventory_consumption as create_mandir_inventory_consumption,
    list_mandir_inventory_consumptions as list_mandir_inventory_consumptions,
    list_mandir_inventory_movements as list_mandir_inventory_movements,
    mandir_create_inventory_item as mandir_create_inventory_item,
    mandir_delete_inventory_item as mandir_delete_inventory_item,
    mandir_inventory_items as mandir_inventory_items,
    mandir_inventory_stock_balances as mandir_inventory_stock_balances,
    mandir_inventory_summary as mandir_inventory_summary,
    mandir_update_inventory_item as mandir_update_inventory_item,
)
from app.modules.mandir_compat.routes.reports import (
    _mandir_donation_designation_report as _mandir_donation_designation_report,
    _mandir_fund_subledger_data as _mandir_fund_subledger_data,
    mandir_donations_daily_report as mandir_donations_daily_report,
    mandir_donations_export_excel as mandir_donations_export_excel,
    mandir_donations_export_pdf as mandir_donations_export_pdf,
    mandir_donations_monthly_report as mandir_donations_monthly_report,
    mandir_report_donations_category_wise as mandir_report_donations_category_wise,
    mandir_report_donations_detailed as mandir_report_donations_detailed,
    mandir_report_donations_festival_wise as mandir_report_donations_festival_wise,
    mandir_report_donations_fund_wise as mandir_report_donations_fund_wise,
    mandir_report_fund_subledger as mandir_report_fund_subledger,
    mandir_report_funds_as_of as mandir_report_funds_as_of,
    mandir_report_sevas_detailed as mandir_report_sevas_detailed,
    mandir_report_sevas_schedule as mandir_report_sevas_schedule,
)

import app.modules.mandir_compat.routes.dashboard  # noqa: E402, F401
import app.modules.mandir_compat.routes.panchang  # noqa: E402, F401
import app.modules.mandir_compat.routes.devotees  # noqa: E402, F401






















