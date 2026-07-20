"""MandirMitra legacy login and opening balance import routes.

Extracted verbatim from app/modules/mandir_compat/router.py per
docs/operations/LARGE_FILE_MODULARIZATION_PLAN.md. Pure move: logic unchanged.
"""
from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from typing import Any

from fastapi import Depends, File, Header, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.accounting.models.entities import Account
from app.accounting.schemas import JournalLineIn, JournalPostRequest
from app.accounting.service import AccountingValidationError
from app.core.auth.dependencies import get_current_user
from app.db.postgres import get_async_session
from app.modules.mandir_compat import router as mandir_router
from app.modules.mandir_compat.router import router

# SECTION: ROUTES: Login + opening balances (template / import)
# ROUTES : POST /login  POST /login/access-token  GET /opening-balances/template  POST .../import
# ════════════════════════════════════════════════════════════════════════

@router.post("/login")
@router.post("/login/access-token")
async def mandir_legacy_login(payload: dict[str, Any], x_app_key: str | None = Header(default=None, alias="X-App-Key")):
    from app.core.auth.service import login_user

    email = str(payload.get("email") or payload.get("username") or "")
    password = str(payload.get("password") or "")
    app_key = mandir_router.resolve_app_key((x_app_key or "mandirmitra").strip())
    access_token, refresh_token = await login_user(email, password, app_key=app_key)
    return {"access_token": access_token, "refresh_token": refresh_token, "token_type": "bearer"}


@router.get("/opening-balances/template")
async def mandir_opening_balances_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Opening Balances"

    headers = ["account_code", "account_name", "opening_balance_debit", "opening_balance_credit"]
    sheet.append(headers)

    header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    example_data = [
        (11001, "Cash", 50000, None),
        (11002, "Bank Account", 100000, None),
        (32001, "General Reserve", None, 150000),
        (21001, "Loan Payable", None, 50000),
    ]

    for row_data in example_data:
        sheet.append(row_data)

    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 25
    sheet.column_dimensions["C"].width = 25
    sheet.column_dimensions["D"].width = 25

    for row in sheet.iter_rows(min_row=2, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Opening_Balance_Template.xlsx"},
    )


@router.post("/opening-balances/import")
async def mandir_opening_balances_import(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_async_session),
    _current_user: dict = Depends(get_current_user),
    x_tenant_id: str | None = Header(default=None, alias="X-Tenant-ID"),
):
    try:
        tenant_id = mandir_router.resolve_tenant_id(_current_user, x_tenant_id)
        await mandir_router._ensure_default_mandir_sql_accounts_safe(session, tenant_id)

        rows = await mandir_router._parse_opening_balance_rows(file)
        if not rows:
            raise HTTPException(status_code=400, detail="Import file is empty")

        sql_accounts = await mandir_router.list_accounts(session, tenant_id=tenant_id)
        accounts_by_code: dict[str, Account] = {}

        def _score_account(acc: Account, normalized_code: str) -> tuple[int, int, int]:
            raw_code = str(acc.code or "").strip()
            return (
                1 if raw_code == normalized_code else 0,
                len(raw_code),
                int(acc.id or 0),
            )

        for acc in sql_accounts:
            normalized_code = mandir_router._normalize_mandir_account_code(acc.code, account_name=acc.name)
            if not normalized_code:
                continue
            existing = accounts_by_code.get(normalized_code)
            if existing is None or _score_account(acc, normalized_code) > _score_account(existing, normalized_code):
                accounts_by_code[normalized_code] = acc

        opening_offset_account = await mandir_router._find_or_create_opening_balance_offset_account(session, tenant_id)

        updated_rows: list[dict[str, Any]] = []
        errors: list[dict[str, Any]] = []
        skipped_count = 0

        for row_number, raw_row in enumerate(rows, start=2):
            try:
                row = {
                    str(key or "").strip().lower(): value
                    for key, value in (raw_row or {}).items()
                    if key is not None
                }

                raw_code = row.get("account_code") or row.get("legacy_code") or row.get("code")
                account_name_hint = row.get("account_name") or row.get("name")
                account_code = mandir_router._normalize_mandir_account_code(raw_code, account_name=account_name_hint)
                if not account_code:
                    raise ValueError("account_code or legacy_code is required")

                account = accounts_by_code.get(account_code)
                if account is None:
                    raise ValueError(f"Account '{account_code}' not found")

                account_type = str(account.type or "").strip().lower()
                if account_type not in {"asset", "liability", "equity"}:
                    raise ValueError("Only balance sheet accounts can have opening balances")

                debit_raw = mandir_router._parse_opening_balance_decimal(row.get("opening_balance_debit"))
                credit_raw = mandir_router._parse_opening_balance_decimal(row.get("opening_balance_credit"))
                signed_raw = mandir_router._parse_opening_balance_decimal(row.get("opening_balance"))

                if debit_raw is None and credit_raw is None and signed_raw is None:
                    raise ValueError("Provide opening_balance_debit/opening_balance_credit or opening_balance")

                debit_amount = max(debit_raw or Decimal("0"), Decimal("0"))
                credit_amount = max(credit_raw or Decimal("0"), Decimal("0"))

                if debit_raw is None and credit_raw is None and signed_raw is not None:
                    if account_type == "asset":
                        debit_amount = max(signed_raw, Decimal("0"))
                        credit_amount = max(-signed_raw, Decimal("0"))
                    else:
                        credit_amount = max(signed_raw, Decimal("0"))
                        debit_amount = max(-signed_raw, Decimal("0"))

                if debit_amount > 0 and credit_amount > 0:
                    raise ValueError("Only one side can be positive for opening balance")

                desired_net = debit_amount - credit_amount
                reference = f"OPENING-{account_code}"
                current_net = await mandir_router._current_opening_balance_net(
                    session,
                    tenant_id=tenant_id,
                    account_id=int(account.id),
                    reference=reference,
                )
                delta_net = desired_net - current_net

                if delta_net == 0:
                    skipped_count += 1
                    continue

                account_debit = delta_net if delta_net > 0 else Decimal("0")
                account_credit = -delta_net if delta_net < 0 else Decimal("0")

                payload = JournalPostRequest(
                    entry_date=date.today(),
                    reference=reference,
                    description=f"Opening balance import for {account.code} - {account.name}",
                    lines=[
                        JournalLineIn(
                            account_id=int(account.id),
                            debit=account_debit,
                            credit=account_credit,
                        ),
                        JournalLineIn(
                            account_id=int(opening_offset_account.id),
                            debit=account_credit,
                            credit=account_debit,
                        ),
                    ],
                )

                idempotency_key = f"mandir-opening-balance:{account_code}:{str(desired_net)}"
                await mandir_router.post_journal_entry(
                    session,
                    tenant_id=tenant_id,
                    created_by=str(_current_user.get("sub") or _current_user.get("email") or "system"),
                    payload=payload,
                    idempotency_key=idempotency_key,
                )

                updated_rows.append(
                    {
                        "account_code": account_code,
                        "account_name": account.name,
                        "opening_balance_debit": mandir_router._safe_float(max(desired_net, Decimal("0"))),
                        "opening_balance_credit": mandir_router._safe_float(max(-desired_net, Decimal("0"))),
                        "applied_delta": mandir_router._safe_float(delta_net),
                    }
                )
            except Exception as exc:
                errors.append({"row": row_number, "error": str(exc)})

        has_errors = len(errors) > 0
        success = len(updated_rows) > 0 and not has_errors

        if success:
            message = f"✓ Successfully imported {len(updated_rows)} opening balance(s)"
        elif len(updated_rows) > 0:
            message = f"⚠ Partial success: {len(updated_rows)} imported, {len(errors)} failed"
        else:
            message = f"✗ Import failed: All {len(rows)} row(s) had errors"

        return {
            "success": success,
            "status": "success" if success else ("partial" if len(updated_rows) > 0 else "failed"),
            "message": message,
            "processed_count": len(rows),
            "updated_count": len(updated_rows),
            "skipped_count": skipped_count,
            "error_count": len(errors),
            "updated": updated_rows,
            "errors": errors[:200],
        }
    except HTTPException:
        raise
    except Exception as exc:
        return {
            "success": False,
            "status": "failed",
            "message": f"✗ Import failed: {str(exc)}",
            "processed_count": 0,
            "updated_count": 0,
            "skipped_count": 0,
            "error_count": 1,
            "updated": [],
            "errors": [{"row": 0, "error": str(exc)}],
        }

