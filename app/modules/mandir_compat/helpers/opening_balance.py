"""MandirMitra opening balance import helpers.

Extracted verbatim from app/modules/mandir_compat/router.py per
docs/operations/LARGE_FILE_MODULARIZATION_PLAN.md. Pure move: logic unchanged.
"""
from __future__ import annotations

import csv
from decimal import Decimal
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.accounting.models.entities import Account, JournalEntry, JournalLine
from app.accounting.service import create_account
from app.modules.mandir_compat.helpers.coercions import _parse_opening_balance_decimal

async def _parse_opening_balance_rows(file: UploadFile) -> list[dict[str, Any]]:
    filename = str(file.filename or '').strip()
    if not filename:
        raise HTTPException(status_code=400, detail='File name is required')

    suffix = Path(filename).suffix.lower()
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail='Uploaded file is empty')

    if suffix == '.csv':
        try:
            text = raw.decode('utf-8-sig')
        except UnicodeDecodeError as exc:
            raise HTTPException(status_code=400, detail='Unable to decode CSV file as UTF-8') from exc
        rows = list(csv.DictReader(StringIO(text)))
        return [row for row in rows if isinstance(row, dict)]

    if suffix in {'.xlsx', '.xlsm'}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise HTTPException(
                status_code=500,
                detail='XLSX import support is unavailable on server (missing openpyxl dependency)',
            ) from exc

        try:
            workbook = load_workbook(BytesIO(raw), data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail='Unable to read XLSX workbook') from exc

        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []

        headers = [str(cell).strip() if cell is not None else '' for cell in values[0]]
        parsed_rows: list[dict[str, Any]] = []
        for row in values[1:]:
            if not any(cell not in (None, '') for cell in row):
                continue
            item: dict[str, Any] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                if index < len(row):
                    item[header] = row[index]
            if item:
                parsed_rows.append(item)
        return parsed_rows

    raise HTTPException(status_code=400, detail='Unsupported file format. Use .csv or .xlsx')


async def _find_or_create_opening_balance_offset_account(session: AsyncSession, tenant_id: str) -> Account:
    stmt = select(Account).where(
        Account.tenant_id == tenant_id,
        Account.code == '33001',
    )
    existing = (await session.execute(stmt)).scalar_one_or_none()
    if existing is not None:
        return existing

    try:
        return await create_account(
            session,
            tenant_id=tenant_id,
            code='33001',
            name='Opening Balance',
            account_type='equity',
            classification='real',
            is_cash_bank=False,
            is_receivable=False,
            is_payable=False,
        )
    except IntegrityError:
        await session.rollback()
        existing = (await session.execute(stmt)).scalar_one_or_none()
        if existing is None:
            raise
        return existing


async def _current_opening_balance_net(
    session: AsyncSession,
    *,
    tenant_id: str,
    account_id: int,
    reference: str,
) -> Decimal:
    stmt = (
        select(
            func.coalesce(func.sum(JournalLine.debit), 0).label('debit_total'),
            func.coalesce(func.sum(JournalLine.credit), 0).label('credit_total'),
        )
        .join(JournalEntry, JournalEntry.id == JournalLine.journal_id)
        .where(
            JournalEntry.tenant_id == tenant_id,
            JournalEntry.reference == reference,
            JournalLine.account_id == account_id,
        )
    )
    row = (await session.execute(stmt)).one()
    debit_total = Decimal(str(row.debit_total or 0))
    credit_total = Decimal(str(row.credit_total or 0))
    return debit_total - credit_total


