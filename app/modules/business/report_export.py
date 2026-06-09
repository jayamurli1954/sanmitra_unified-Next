"""Generic tabular report exporter — CSV / XLSX / PDF.

Any report that can be expressed as (title, columns, rows[, footer]) can be
downloaded in three real file formats with one call to ``export_report``. This
keeps every business report's export consistent and avoids per-report builders.

- columns: list of {"key", "label", "numeric": bool}. Numeric columns are
  right-aligned and number-formatted.
- rows: list of dicts keyed by column "key".
- footer: optional dict (same keys) rendered bold (e.g. a totals line).
- meta: list of (label, value) shown under the title (period, generated-on…).

Print is handled in the browser (print-friendly view); the server only produces
downloadable files.
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime
from decimal import Decimal

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas

CSV_MEDIA = "text/csv"
XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MEDIA = "application/pdf"
_MEDIA = {"csv": CSV_MEDIA, "xlsx": XLSX_MEDIA, "pdf": PDF_MEDIA}


def _cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return f"{value.quantize(Decimal('0.01'))}"
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def _pdf_text(value) -> str:
    return _cell(value).encode("latin-1", "replace").decode("latin-1")


# --------------------------------------------------------------------------- #
# Format builders
# --------------------------------------------------------------------------- #

def build_csv(*, title, columns, rows, footer=None, meta=None, org_name=None) -> bytes:
    sio = io.StringIO()
    writer = csv.writer(sio)
    if org_name:
        writer.writerow([org_name])
    writer.writerow([title])
    for label, value in (meta or []):
        writer.writerow([label, _cell(value)])
    writer.writerow([])
    writer.writerow([c["label"] for c in columns])
    for row in rows:
        writer.writerow([_cell(row.get(c["key"])) for c in columns])
    if footer:
        writer.writerow([_cell(footer.get(c["key"])) for c in columns])
    return sio.getvalue().encode("utf-8-sig")  # BOM so Excel reads UTF-8


def build_xlsx(*, title, columns, rows, footer=None, meta=None, org_name=None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Report")[:31]

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    right = Alignment(horizontal="right")

    r = 1
    if org_name:
        ws.cell(row=r, column=1, value=org_name).font = Font(bold=True, size=12)
        r += 1
    ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=14)
    r += 1
    for label, value in (meta or []):
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=_cell(value))
        r += 1
    r += 1

    header_row = r
    for ci, col in enumerate(columns, start=1):
        cell = ws.cell(row=r, column=ci, value=col["label"])
        cell.font = header_font
        cell.fill = header_fill
        if col.get("numeric"):
            cell.alignment = right
    r += 1

    def _write(row_dict, *, is_footer=False):
        nonlocal r
        for ci, col in enumerate(columns, start=1):
            raw = row_dict.get(col["key"])
            if col.get("numeric"):
                num = _to_number(raw)
                cell = ws.cell(row=r, column=ci, value=num if num is not None else _cell(raw))
                cell.alignment = right
                if num is not None:
                    cell.number_format = "#,##0.00"
            else:
                cell = ws.cell(row=r, column=ci, value=_cell(raw))
            if is_footer:
                cell.font = bold
        r += 1

    for row in rows:
        _write(row)
    if footer:
        _write(footer, is_footer=True)

    # Auto-ish column widths.
    for ci, col in enumerate(columns, start=1):
        longest = len(str(col["label"]))
        for row in rows:
            longest = max(longest, len(_cell(row.get(col["key"]))))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(longest + 2, 10), 48)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _to_number(value):
    if value is None or value == "":
        return None
    try:
        return float(Decimal(str(value)))
    except Exception:
        return None


def build_pdf(*, title, columns, rows, footer=None, meta=None, org_name=None) -> bytes:
    # Landscape when there are many/wide columns so everything fits.
    pagesize = landscape(A4) if len(columns) > 5 else A4
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=pagesize)
    width, height = pagesize
    left, right_margin, top, bottom = 36, 36, 48, 48
    usable = width - left - right_margin

    # Column x-positions proportional to a weight (numeric cols narrower-ish).
    weights = [3 if not c.get("numeric") else 2 for c in columns]
    total_w = sum(weights)
    col_w = [usable * w / total_w for w in weights]
    x_left = [left + sum(col_w[:i]) for i in range(len(columns))]
    x_right = [left + sum(col_w[:i + 1]) - 4 for i in range(len(columns))]

    def header(y):
        if org_name:
            pdf.setFont("Helvetica-Bold", 12)
            pdf.drawString(left, y, _pdf_text(org_name))
            y -= 15
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(left, y, _pdf_text(title))
        y -= 16
        pdf.setFont("Helvetica", 8)
        for label, value in (meta or []):
            pdf.drawString(left, y, _pdf_text(f"{label}: {_cell(value)}"))
            y -= 11
        y -= 3
        return _col_headers(y)

    def _col_headers(y):
        pdf.setFont("Helvetica-Bold", 8)
        for i, col in enumerate(columns):
            if col.get("numeric"):
                pdf.drawRightString(x_right[i], y, _pdf_text(col["label"])[:24])
            else:
                pdf.drawString(x_left[i], y, _pdf_text(col["label"])[:40])
        y -= 4
        pdf.setStrokeColor(colors.grey)
        pdf.line(left, y, width - right_margin, y)
        return y - 12

    def draw_row(y, row, *, is_footer=False):
        pdf.setFont("Helvetica-Bold" if is_footer else "Helvetica", 8)
        for i, col in enumerate(columns):
            val = _pdf_text(row.get(col["key"]))
            if col.get("numeric"):
                pdf.drawRightString(x_right[i], y, val[:24])
            else:
                pdf.drawString(x_left[i], y, val[:48])
        return y - 12

    y = header(height - top)
    for row in rows:
        if y < bottom + 16:
            pdf.showPage()
            y = header(height - top)
        y = draw_row(y, row)
    if footer:
        if y < bottom + 16:
            pdf.showPage()
            y = header(height - top)
        pdf.setStrokeColor(colors.grey)
        pdf.line(left, y + 6, width - right_margin, y + 6)
        draw_row(y, footer, is_footer=True)

    pdf.save()
    return buffer.getvalue()


_BUILDERS = {"csv": build_csv, "xlsx": build_xlsx, "pdf": build_pdf}


def export_report(
    fmt: str, *, title: str, columns: list[dict], rows: list[dict],
    footer: dict | None = None, meta: list | None = None, org_name: str | None = None,
    filename_base: str = "report",
) -> StreamingResponse:
    fmt = (fmt or "").lower()
    builder = _BUILDERS.get(fmt)
    if builder is None:
        raise ValueError("format must be one of: csv, xlsx, pdf")
    content = builder(title=title, columns=columns, rows=rows, footer=footer, meta=meta, org_name=org_name)
    filename = f"{filename_base}.{fmt}"
    return StreamingResponse(
        io.BytesIO(content),
        media_type=_MEDIA[fmt],
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
